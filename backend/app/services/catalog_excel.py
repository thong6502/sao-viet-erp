"""Xuất / Nhập Excel cho MỌI màn Cấu hình danh mục — MỘT cơ chế dùng chung.

Vì sao có file này: trước 29/08/2026 cơ chế Excel nằm thẳng trong
`routers/catalog_base.make_catalog_router` dưới dạng bốn tham số (`enable_import`,
`import_columns`, `import_resolve`, `export_resolve`) và chỉ tả được MỘT sheet phẳng. Năm màn bật
được, tám màn còn lại thì không — và ngay ở năm màn đó, mọi cấu hình dạng BẢNG CON (bậc bù hao,
đầu việc định mức, gói bảo trì, cặp quy đổi) hoặc rơi mất, hoặc bị nén thành một ô JSON thô mà
người khai không sửa nổi.

Nay router chỉ nhận MỘT tham số `excel_spec`; toàn bộ luật nằm ở đây và ở
`services/catalog_excel_specs.py` (khai spec cho từng màn).

BA THỨ Ở ĐÂY:

* `Cot` / `SheetCon` / `CatalogExcelSpec` — NGÔN NGỮ khai: sheet nào, cột nào, kiểu gì, dịch mã
  bằng resolver nào, field nào cố ý KHÔNG đi qua Excel (`loai_tru`).
* `xuat_excel(spec, svc)` — dựng workbook.
* `nhap_excel(spec, svc, InModel, du_lieu, …)` — đọc file, dựng KẾ HOẠCH thay đổi rồi chạy nó
  trong MỘT giao dịch. `ghi=False` là xem trước (chạy y hệt rồi rollback), `ghi=True` là chốt.

LUẬT DỮ LIỆU (chốt ở docs/spec-xuat-nhap-excel-danh-muc.md):

1. XUẤT chỉ dòng `active=True`, nhưng ĐỦ mọi cột cấu hình hiện hành. Không xuất lịch sử (nhật ký,
   phiên bản giá giấy, "lần trước công thức"), không xuất id nội bộ, mốc thời gian hệ thống, dữ
   liệu dẫn xuất, ảnh, trạng thái vận hành.
2. NHẬP là UPSERT theo `Mã`. Bản ghi KHÔNG có mặt trong sheet chính thì giữ nguyên (không xoá).
3. Cột CÓ mặt trong tiêu đề nhưng ô trống ⇒ XOÁ giá trị. Cột VẮNG mặt ⇒ giữ nguyên giá trị cũ.
4. Sheet con CÓ MẶT ⇒ thay TRỌN tập con của mọi mã cha có trong sheet chính (xoá dòng con khỏi
   file = xoá cấu hình con đó). Sheet con VẮNG MẶT ⇒ giữ nguyên dữ liệu con.
5. Dòng không đổi thì KHÔNG gọi service ⇒ không đẻ dòng nhật ký ma.
6. MỘT giao dịch cho cả file: còn một lỗi thì không ghi gì cả.
"""
from __future__ import annotations

import json
from contextlib import contextmanager
from dataclasses import dataclass, field as dc_field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Callable

from ..repositories.catalog_base import SIZE_TRAN

#: Sheet ẨN ghi loại danh mục + phiên bản định dạng — chặn nhập nhầm file của màn khác.
SHEET_META = "_meta"
#: Sheet ẨN giữ các khoá cấu hình nền KHÔNG diễn giải được (round-trip nguyên trạng).
SHEET_GIU = "_giu_nguyen"
#: Phiên bản định dạng workbook. Tăng khi hợp đồng cột đổi kiểu KHÔNG đọc ngược được.
PHIEN_BAN = "2"

#: Tiêu đề cột khoá nghiệp vụ của sheet chính — mọi màn dùng chung.
NHAN_MA = "Mã"
#: Tiêu đề cột bật/ngừng dùng (TRUE = đang dùng, FALSE = ngừng dùng).
NHAN_TRANG_THAI = "Trạng thái"
#: Tiêu đề cột thứ tự trong sheet con.
NHAN_THU_TU = "Thứ tự"


class ExcelSaiMan(Exception):
    """File không đọc được, hoặc là workbook của MÀN KHÁC → 422 ở router."""


# --------------------------------------------------------------------------------------
# Ngôn ngữ khai spec
# --------------------------------------------------------------------------------------


@dataclass
class NguCanh:
    """Thứ mà resolver cần: service của màn + Session, cộng một chỗ nhớ tạm cho tra cứu lặp.

    Resolver chạy MỘT LẦN CHO MỖI Ô (danh mục 200 dòng × 25 cột), nên mọi câu tra bảng phụ (tên
    tổ, mã bù hao, mã công đoạn…) phải đi qua `nho()` — không thì một lần xuất Excel là vài nghìn
    round-trip xuống DB.
    """

    svc: Any
    db: Any
    _bo_nho: dict[str, Any] = dc_field(default_factory=dict)

    def nho(self, khoa: str, dung: Callable[[], Any]) -> Any:
        if khoa not in self._bo_nho:
            self._bo_nho[khoa] = dung()
        return self._bo_nho[khoa]


@dataclass(frozen=True)
class Cot:
    """Một cột Excel.

    * `nhan` — TIÊU ĐỀ hiện trên file. Đây là hợp đồng với người dùng: đổi nhãn là file cũ mất cột.
    * `field` — tên khoá trong dict dữ liệu (field của `InModel`, hoặc khoá của dòng con).
    * `kieu` — `chu · so · nguyen · bool · ngay · json`. Quyết định cách ép kiểu HAI CHIỀU.
    * `doc(gt, ctx)` — Excel → giá trị lưu (dịch mã → id…). Ném `ValueError(câu tiếng Việt)` thì
      ô đó thành một dòng lỗi, các dòng khác vẫn chạy tiếp.
    * `ghi(gt, ctx)` — chiều NGƯỢC: giá trị lưu → ô Excel (id → mã đọc được). Không được ném.
    * `chi_doc` — cột ĐỐI CHIẾU cho người đọc (tên khách, tên tổ). Không bao giờ GHI vào DB: khoá
      thật là cột mã bên cạnh — để hai cột cùng ghi vào một field là mời mâu thuẫn (sửa tên mà
      quên sửa mã). Nhưng khi nhập thì có ĐỌC để đối chiếu: tên trong file phải khớp bản ghi mà mã
      đang trỏ tới, lệch là báo lỗi (xem `_lech_cot_doi_chieu`).
    * `nhan_cu` — TIÊU ĐỀ ĐỜI CŨ vẫn nhận khi nhập (không bao giờ xuất ra). Năm màn đã có Excel
      trước 29/08/2026 phát ra file với tên cột khác; người dùng còn giữ file đó trên máy.
    * `chi_nhap` — cột CHỈ ĐỌC KHI NHẬP, không xuất. Dùng cho cột đời cũ nay đã tách thành sheet
      con (danh sách nối bằng dấu phẩy). Sheet con áp SAU nên nếu file có cả hai, sheet con thắng.
    """

    nhan: str
    field: str
    kieu: str = "chu"
    doc: Callable[[Any, NguCanh], Any] | None = None
    ghi: Callable[[Any, NguCanh], Any] | None = None
    rong: int = 22
    chi_doc: bool = False
    nhan_cu: tuple[str, ...] = ()
    chi_nhap: bool = False

    def vi_tri(self, tieu_de: list[str]) -> int | None:
        """Chỉ số cột trong tiêu đề — thử nhãn hiện hành trước, rồi tới các nhãn đời cũ."""
        for nhan in (self.nhan, *self.nhan_cu):
            if nhan in tieu_de:
                return tieu_de.index(nhan)
        return None


@dataclass(frozen=True)
class SheetCon:
    """Một sheet CON — tập dòng con thuộc về một mã cha ở sheet chính.

    * `field` — field của `InModel` nhận danh sách dòng con (`bac`, `dau_viec_dinh_muc`…).
    * `rut_gon` — field này là danh sách VÔ HƯỚNG (id / tên), không phải danh sách dict: lấy đúng
      một khoá của mỗi dòng con (`routing_template`, `thay_the_ids`, `nhom_may_cho_phep`).
    * `trong_json` — field nằm LỒNG trong một cột JSON (`fields_theo_loai` → `chuan_bi_khoan`).
    * `rieng` — sheet do `spec.gop_con` tự ghép (cấu trúc hai tầng: hạng mục trong gói bảo trì,
      vật tư trong đầu việc). Nền chỉ parse rồi đưa nguyên vào tay hook.
    * `toan_cuc` — sheet KHÔNG khoá theo mã cha (danh mục Nhóm máy đi kèm màn Máy).
    * `khoa_phu` — cột khoá phụ ngoài mã cha (Mã gói bảo trì), để hook ghép hai tầng.
    * `doc_hien_co(obj, ctx)` — cách ĐỌC tập con hiện có của một bản ghi, dùng cho CẢ xuất lẫn
      so-sánh-để-biết-có-đổi-không. Không khai thì nền đọc thẳng `getattr(obj, field)`.
    * `ap_dung(ctx, obj, rows, actor_id)` — ghi tập con nằm NGOÀI `InModel` (cặp quy đổi là bảng
      riêng, nhóm máy là danh mục riêng).
    * `giu_khi_vang(obj, ctx)` — giá trị phải GÁN LẠI cho `field` khi sheet VẮNG MẶT trong file,
      dành cho field mà repo thay-trọn-bộ dù không được gửi lên. Chỉ `cong_doan.dau_viec_dinh_muc`
      cần: `CongDoanRepository._sau_gan` xoá sạch bảng con khi khoá vắng trong `data`, nên không
      gán lại là nhập một file thiếu sheet đó cũng xoá hết định mức. Các field khác nằm trong
      `repo.fields` nên `_gan` bỏ qua khi vắng — không cần khai.
    """

    ten: str
    cot: tuple[Cot, ...]
    field: str | None = None
    rut_gon: str | None = None
    trong_json: tuple[str, ...] = ()
    cot_cha: str = NHAN_MA
    khoa_phu: tuple[Cot, ...] = ()
    thu_tu: bool = True
    rieng: bool = False
    toan_cuc: bool = False
    doc_hien_co: Callable[[Any, NguCanh], list] | None = None
    ap_dung: Callable[[NguCanh, Any, list, int | None], None] | None = None
    giu_khi_vang: Callable[[Any, NguCanh], Any] | None = None


@dataclass(frozen=True)
class CatalogExcelSpec:
    """Hợp đồng workbook của MỘT màn danh mục."""

    loai: str
    """Khoá màn — ghi vào sheet `_meta`, phải khớp `ten` truyền cho `make_catalog_router`."""

    tieu_de: str
    """Tên sheet chính (≤ 31 ký tự — trần của Excel)."""

    cot: tuple[Cot, ...]
    sheets_con: tuple[SheetCon, ...] = ()

    repo_cls: Any = None
    """Lớp repository của màn — test guard đối chiếu `repo.fields` với cột Excel + `loai_tru`."""

    loai_tru: frozenset[str] = frozenset()
    """Field repo CỐ Ý không đi qua Excel, kèm lý do ở chỗ khai. Test guard đọc đúng tập này."""

    truoc_khi_ghi: Callable[[dict, "NguCanh", Any], dict] | None = None
    """Nắn dữ liệu MỘT DÒNG trước khi dựng `InModel` (Công đoạn ép `che_do_tinh`/`pricing_basis`
    theo nhóm). Nhận `(du_lieu, ctx, cu)`; `cu` = bản ghi cũ hoặc None."""

    gop_con: Callable[[dict, dict, "NguCanh"], None] | None = None
    """Ghép các sheet con `rieng=True` vào `du_lieu` (cấu trúc hai tầng). Nhận
    `(du_lieu, con_theo_sheet, ctx)` — `con_theo_sheet` chỉ chứa sheet CÓ MẶT trong file."""

    sheet_an: Callable[[Any, "NguCanh"], dict] | None = None
    """Khoá LẠ mà nền không hiểu nhưng phải round-trip (khoá ngoài `chuan_bi_khoan`/`lich_bao_tri`
    trong `fields_theo_loai` của Máy). Trả `{khoá: giá trị}`; nền ghi vào sheet ẩn `_giu_nguyen`."""

    gop_an: Callable[[dict, dict, "NguCanh"], None] | None = None
    """Chiều ngược của `sheet_an`."""

    def cot_theo_field(self) -> dict[str, Cot]:
        return {c.field: c for c in self.cot}


# --------------------------------------------------------------------------------------
# Ép kiểu hai chiều
# --------------------------------------------------------------------------------------


def _chuoi(gt: Any) -> str | None:
    if gt is None:
        return None
    s = str(gt).strip()
    return s or None


def _bool_tu_excel(gt: Any) -> bool:
    if isinstance(gt, bool):
        return gt
    s = str(gt).strip().lower()
    return s in ("true", "1", "x", "có", "co", "yes", "y", "đang dùng", "dang dung")


def _so_tu_excel(gt: Any, nhan: str) -> float:
    if isinstance(gt, bool):
        raise ValueError(f'Cột "{nhan}" phải là số.')
    if isinstance(gt, (int, float, Decimal)):
        return float(gt)
    s = str(gt).strip().replace(" ", "")
    # Excel tiếng Việt hay trả "1.234,5" — bỏ dấu nghìn rồi mới đổi dấu thập phân.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        raise ValueError(f'Cột "{nhan}" phải là số (đang là "{gt}").') from None


def _ngay_tu_excel(gt: Any, nhan: str) -> date:
    if isinstance(gt, datetime):
        return gt.date()
    if isinstance(gt, date):
        return gt
    s = str(gt).strip()
    for mau in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, mau).date()
        except ValueError:
            continue
    raise ValueError(f'Cột "{nhan}" phải là ngày (yyyy-mm-dd hoặc dd/mm/yyyy).')


def _ep_kieu_vao(cot: Cot, gt: Any) -> Any:
    """Ô Excel → giá trị Python theo `cot.kieu`. Ô trống đã được chặn TRƯỚC khi gọi."""
    if cot.kieu == "bool":
        return _bool_tu_excel(gt)
    if cot.kieu == "so":
        return _so_tu_excel(gt, cot.nhan)
    if cot.kieu == "nguyen":
        so = _so_tu_excel(gt, cot.nhan)
        if abs(so - round(so)) > 1e-9:
            raise ValueError(f'Cột "{cot.nhan}" phải là số nguyên.')
        return int(round(so))
    if cot.kieu == "ngay":
        return _ngay_tu_excel(gt, cot.nhan)
    if cot.kieu == "json":
        if isinstance(gt, (dict, list)):
            return gt
        try:
            return json.loads(str(gt))
        except (TypeError, ValueError):
            raise ValueError(f'Cột "{cot.nhan}" không phải JSON hợp lệ.') from None
    return _chuoi(gt)


def _ep_kieu_ra(cot: Cot, gt: Any) -> Any:
    """Giá trị Python → ô Excel. Giữ kiểu số/bool để Excel hiện đúng, không nhét chuỗi."""
    if gt is None:
        return None
    if cot.kieu == "bool":
        return bool(gt)
    if cot.kieu in ("so", "nguyen"):
        so = float(gt)
        return int(round(so)) if cot.kieu == "nguyen" else so
    if cot.kieu == "ngay":
        return gt.isoformat() if isinstance(gt, (date, datetime)) else str(gt)
    if cot.kieu == "json":
        return gt if isinstance(gt, str) else json.dumps(gt, ensure_ascii=False)
    if isinstance(gt, Decimal):
        return float(gt)
    if isinstance(gt, (list, tuple)):
        return ", ".join(str(x) for x in gt) or None
    return gt


def _rong(gt: Any) -> bool:
    """Ô rỗng THẬT — `0` và `False` KHÔNG rỗng (đơn giá 0 là con số người ta cố ý gõ)."""
    return gt is None or (isinstance(gt, str) and not gt.strip())


# --------------------------------------------------------------------------------------
# So sánh "có đổi gì không"
# --------------------------------------------------------------------------------------


def _chuan_de_so(gt: Any) -> Any:
    """Đưa giá trị về dạng SO SÁNH ĐƯỢC: rỗng ≡ None, Decimal ≡ float, ngày ≡ chuỗi ISO."""
    if gt is None:
        return None
    if isinstance(gt, bool):
        return gt
    if isinstance(gt, (int, float, Decimal)):
        return round(float(gt), 9)
    if isinstance(gt, (date, datetime)):
        return gt.isoformat()[:10]
    if isinstance(gt, str):
        return gt.strip() or None
    if isinstance(gt, (list, tuple)):
        return [_chuan_de_so(x) for x in gt]
    if isinstance(gt, dict):
        return {k: _chuan_de_so(v) for k, v in sorted(gt.items())}
    return gt


def _bang_nhau(a: Any, b: Any) -> bool:
    return _chuan_de_so(a) == _chuan_de_so(b)


# --------------------------------------------------------------------------------------
# Giao dịch
# --------------------------------------------------------------------------------------


@contextmanager
def mot_giao_dich(db):
    """Gộp MỌI commit lẻ của service/repo vào ĐÚNG một giao dịch, chốt (hoặc huỷ) ở cuối.

    Dùng: ``with mot_giao_dich(db) as gd: … ; gd["chot"] = True``.

    Vì sao phải chặn `commit` chứ không chỉ dựa vào `commit_on_write=False`: nền service
    (`services/catalog_base.CatalogService._chot`) gọi `repo.chot_giao_dich()`, mà hàm đó commit
    THẬT đúng khi repo để `commit_on_write=False` — tức mỗi dòng Excel là một giao dịch riêng.
    Nhập 100 dòng mà dòng 100 sai thì 99 dòng đầu đã nằm trong DB, đúng thứ luật "một file một
    giao dịch" cấm.

    Đổi `db.commit` thành `db.flush` là cách chặn được mà không phải sửa bảy service: SAVEPOINT
    không cứu nổi vì `Session.commit()` chốt luôn giao dịch NGOÀI CÙNG, không dừng ở savepoint.
    """
    _ep_mo_giao_dich(db)
    that = db.commit
    db.commit = db.flush          # type: ignore[method-assign]
    trang_thai = {"chot": False}
    try:
        yield trang_thai
    except BaseException:
        db.commit = that
        db.rollback()
        raise
    db.commit = that
    if trang_thai["chot"]:
        db.commit()
    else:
        db.rollback()


def _ep_mo_giao_dich(db) -> None:
    """Mở sẵn một giao dịch THẬT trước khi ai đó `begin_nested()`. No-op ngoài SQLite.

    Driver `sqlite3` chỉ tự phát `BEGIN` trước INSERT/UPDATE/DELETE, KHÔNG phát trước `SAVEPOINT`.
    Nên `begin_nested()` lúc chưa có giao dịch nào tạo savepoint ở mức NGOÀI CÙNG, mà theo luật
    SQLite thì `RELEASE` một savepoint ngoài cùng CHÍNH LÀ commit. `rollback()` sau đó không còn
    gì để huỷ — luật "cả file là MỘT giao dịch" âm thầm mất hiệu lực và dòng lỗi cuối file không
    kéo được các dòng trước bay theo.

    Postgres (dev/prod) không dính vì psycopg phát `BEGIN` đàng hoàng; chỉ SQLite của bộ test ghi
    thật. Vá ở ĐÂY chứ không ở `db._make_engine`: bản vá chính chủ (tắt `isolation_level` + tự
    phát `BEGIN` qua event) vỡ với `StaticPool` — cả app dùng CHUNG một connection nên hai giao
    dịch SQLAlchemy song song thành "cannot start a transaction within a transaction".
    """
    bind = getattr(db, "bind", None)
    if bind is None or bind.dialect.name != "sqlite":
        return
    conn = db.connection()
    tho = getattr(conn.connection, "driver_connection", None)
    if tho is not None and not tho.in_transaction:
        conn.exec_driver_sql("BEGIN")


def db_cua(svc) -> Any:
    """Session của service — ba khuôn service danh mục đặt nó ở ba chỗ khác nhau."""
    for duong in (("repo", "db"), ("goc", "repo", "db"), ("audit", "db")):
        o: Any = svc
        for buoc in duong:
            o = getattr(o, buoc, None)
            if o is None:
                break
        else:
            return o
    raise RuntimeError(f"Không tìm ra Session của {type(svc).__name__}.")


# --------------------------------------------------------------------------------------
# XUẤT
# --------------------------------------------------------------------------------------


def _moi_dong(svc) -> list:
    """MỌI bản ghi của màn — lật hết trang, KHÔNG lọc `active`.

    Trước 30/08/2026 chỗ này lọc `active=True` với lý do "file là ảnh của cấu hình ĐANG CHẠY".
    Lý do đó nghe xuôi nhưng làm bộ 13 file xuất ra KHÔNG tự nhập lại được sang máy khác: một dòng
    còn hiệu lực ở màn này vẫn được phép trỏ tới một dòng ĐÃ NGỪNG ở màn kia (Giấy `COUCHE-300`
    trỏ chủng loại `COUCHE` đã ngừng, Công đoạn `CD-0004` trỏ bù hao `BH-SONG-1CON` đã ngừng), mà
    dòng đã ngừng thì không có trong file của màn kia ⇒ nhập vào DB trắng là gãy tham chiếu.
    Kèm theo đó, ngừng dùng vốn là đường MỘT CHIỀU qua Excel: dòng biến mất khỏi file thì không
    còn cách nào bật lại. Xuất cả hai loại, cột `Trạng thái` nói rõ dòng nào đang tắt.
    """
    ra: list = []
    trang = 1
    while True:
        objs, total = svc.list(active=None, page=trang, size=SIZE_TRAN)
        ra.extend(objs)
        if len(objs) < SIZE_TRAN or len(ra) >= total:
            return ra
        trang += 1


def _doc_field_con(sheet: SheetCon, obj) -> Any:
    if sheet.trong_json:
        tui = getattr(obj, sheet.trong_json[0], None) or {}
        for khoa in sheet.trong_json[1:]:
            tui = (tui or {}).get(khoa) if isinstance(tui, dict) else None
        return tui or []
    return getattr(obj, sheet.field, None) or []


def doc_con(sheet: SheetCon, obj, ctx: NguCanh) -> list[dict]:
    """Tập con HIỆN CÓ của một bản ghi, chuẩn hoá thành `list[dict]` theo cột của sheet."""
    if sheet.doc_hien_co is not None:
        return list(sheet.doc_hien_co(obj, ctx) or [])
    goc = _doc_field_con(sheet, obj)
    if sheet.rut_gon:
        return [{sheet.rut_gon: v} for v in (goc or [])]
    ra: list[dict] = []
    for item in goc or []:
        if isinstance(item, dict):
            ra.append(item)
        else:      # dòng con là ORM (đầu việc định mức) — rút đúng các field của sheet
            ra.append({c.field: getattr(item, c.field, None)
                       for c in (*sheet.khoa_phu, *sheet.cot)})
    return ra


def xuat_excel(spec: CatalogExcelSpec, svc) -> bytes:
    """Workbook của màn: `_meta` (ẩn) + sheet chính + các sheet con + `_giu_nguyen` (ẩn)."""
    # lazy import: thiếu dep chỉ hỏng route này, không sập app.
    from openpyxl import Workbook
    from openpyxl.styles import Font

    ctx = NguCanh(svc=svc, db=db_cua(svc))
    objs = _moi_dong(svc)

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = SHEET_META
    ws_meta.append(["khoa", "gia_tri"])
    ws_meta.append(["loai", spec.loai])
    ws_meta.append(["phien_ban", PHIEN_BAN])
    ws_meta.sheet_state = "hidden"

    ws = wb.create_sheet(spec.tieu_de[:31])
    cot_ra = [c for c in spec.cot if not c.chi_nhap]
    _dat_tieu_de(ws, [c.nhan for c in cot_ra], [c.rong for c in cot_ra], Font)
    for obj in objs:
        ws.append([_o_xuat(c, obj, ctx) for c in cot_ra])

    for sheet in spec.sheets_con:
        _xuat_sheet_con(wb, sheet, objs, ctx, Font)

    if spec.sheet_an is not None:
        _xuat_sheet_giu(wb, spec, objs, ctx, Font)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _o_xuat(cot: Cot, obj, ctx: NguCanh) -> Any:
    gt = getattr(obj, cot.field, None)
    if cot.ghi is not None:
        gt = cot.ghi(gt, ctx)
    return _ep_kieu_ra(cot, gt)


def _dat_tieu_de(ws, nhan: list[str], rong: list[int], Font) -> None:
    ws.append(nhan)
    for o in ws[1]:
        o.font = Font(bold=True)
    for i, r in enumerate(rong, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = r


def _xuat_sheet_con(wb, sheet: SheetCon, objs: list, ctx: NguCanh, Font) -> None:
    ws = wb.create_sheet(sheet.ten[:31])
    cot_du_lieu = [c for c in (*sheet.khoa_phu, *sheet.cot) if not c.chi_nhap]
    nhan = ([] if sheet.toan_cuc else [sheet.cot_cha]) \
        + ([NHAN_THU_TU] if sheet.thu_tu else []) \
        + [c.nhan for c in cot_du_lieu]
    rong = ([] if sheet.toan_cuc else [22]) \
        + ([10] if sheet.thu_tu else []) \
        + [c.rong for c in cot_du_lieu]
    _dat_tieu_de(ws, nhan, rong, Font)

    def _o(c: Cot, dong: dict) -> Any:
        gt = dong.get(c.field)
        return _ep_kieu_ra(c, c.ghi(gt, ctx) if c.ghi is not None else gt)

    if sheet.toan_cuc:
        for i, dong in enumerate(doc_con(sheet, None, ctx), start=1):
            ws.append(([i] if sheet.thu_tu else []) + [_o(c, dong) for c in cot_du_lieu])
        return

    for obj in objs:
        for i, dong in enumerate(doc_con(sheet, obj, ctx), start=1):
            ws.append([getattr(obj, "ma", None)]
                      + ([i] if sheet.thu_tu else [])
                      + [_o(c, dong) for c in cot_du_lieu])


def _xuat_sheet_giu(wb, spec: CatalogExcelSpec, objs: list, ctx: NguCanh, Font) -> None:
    """Sheet ẨN cho khoá cấu hình nền KHÔNG diễn giải được — có nó thì round-trip không mất dữ liệu.

    Không bày ra sheet thường: đây là JSON thô, mời người khai sửa tay là mời hỏng.
    """
    ws = wb.create_sheet(SHEET_GIU)
    _dat_tieu_de(ws, [NHAN_MA, "JSON"], [22, 80], Font)
    for obj in objs:
        con_lai = spec.sheet_an(obj, ctx)
        if con_lai:
            ws.append([getattr(obj, "ma", None), json.dumps(con_lai, ensure_ascii=False)])
    ws.sheet_state = "hidden"


# --------------------------------------------------------------------------------------
# NHẬP
# --------------------------------------------------------------------------------------


@dataclass
class LoiDong:
    sheet: str
    dong: int
    cot: str
    ly_do: str


@dataclass
class KetQua:
    hop_le: bool = True
    tong_dong: int = 0
    tao_moi: int = 0
    cap_nhat: int = 0
    khong_doi: int = 0
    da_ghi: bool = False
    loi: list[LoiDong] = dc_field(default_factory=list)


def _mo_workbook(du_lieu: bytes):
    from openpyxl import load_workbook   # lazy import, cùng lý do với `xuat_excel`

    try:
        return load_workbook(BytesIO(du_lieu), read_only=True, data_only=True)
    except Exception:
        raise ExcelSaiMan("Không đọc được file — phải là .xlsx đúng mẫu.") from None


def _kiem_meta(wb, spec: CatalogExcelSpec) -> None:
    """`_meta` sai màn ⇒ từ chối. KHÔNG có `_meta` ⇒ nhận (chế độ tương thích file đời cũ)."""
    if SHEET_META not in wb.sheetnames:
        return
    doc = {str(r[0]).strip(): (str(r[1]).strip() if len(r) > 1 and r[1] is not None else "")
           for r in wb[SHEET_META].iter_rows(values_only=True) if r and r[0] is not None}
    loai = doc.get("loai")
    if loai and loai != spec.loai:
        raise ExcelSaiMan(
            f'File này là workbook của màn "{loai}", không phải "{spec.loai}". '
            "Bấm Xuất Excel ngay tại màn đang đứng rồi sửa trên file đó.")
    # Phiên bản MỚI HƠN bản đang chạy ⇒ từ chối. Cột có thể đã đổi ý nghĩa, mà đọc bừa một file
    # tương lai thì lỗi không lộ ra ở đây — nó lộ ra ở bảng giá sai mấy tuần sau.
    pb = doc.get("phien_ban")
    if pb and pb.isdigit() and int(pb) > int(PHIEN_BAN):
        raise ExcelSaiMan(
            f"File ở định dạng phiên bản {pb}, hệ đang đọc tới phiên bản {PHIEN_BAN}. "
            "Bấm Xuất Excel để lấy file đúng định dạng đang chạy.")


def _bang(ws) -> tuple[list[str], list[tuple]]:
    """`(tiêu đề đã strip, các dòng dữ liệu)` — dòng rỗng hoàn toàn bị loại ngay."""
    hang = ws.iter_rows(values_only=True)
    tieu_de = [str(h).strip() if h is not None else "" for h in next(hang, ())]
    dong = [r for r in hang if r and not all(o is None for o in r)]
    return tieu_de, dong


def _khoa_ma(ma: Any) -> str:
    return str(ma or "").strip().lower()


def _doc_o(cot: Cot, gt: Any, ctx: NguCanh) -> Any:
    gt = _ep_kieu_vao(cot, gt)
    if cot.doc is not None:
        gt = cot.doc(gt, ctx)
    return gt


def _parse_sheet_con(ws, sheet: SheetCon, ctx: NguCanh,
                     loi: list[LoiDong]) -> dict[str, list[dict]]:
    """Sheet con → `{mã cha (chuẩn hoá): [dòng con]}`. Sheet toàn cục gom vào khoá `""`."""
    tieu_de, dong = _bang(ws)
    cot_du_lieu = (*sheet.khoa_phu, *sheet.cot)
    vi_tri = {c.nhan: i for c in cot_du_lieu if (i := c.vi_tri(tieu_de)) is not None}
    i_cha = tieu_de.index(sheet.cot_cha) if sheet.cot_cha in tieu_de else None
    i_thu_tu = tieu_de.index(NHAN_THU_TU) if NHAN_THU_TU in tieu_de else None
    if i_cha is None and not sheet.toan_cuc:
        loi.append(LoiDong(sheet.ten, 1, sheet.cot_cha,
                           f'Sheet "{sheet.ten}" thiếu cột "{sheet.cot_cha}".'))
        return {}

    tam: dict[str, list[tuple[float, dict]]] = {}
    for so_dong, hang in enumerate(dong, start=2):
        cha = "" if sheet.toan_cuc else _khoa_ma(hang[i_cha] if i_cha < len(hang) else None)
        if not sheet.toan_cuc and not cha:
            loi.append(LoiDong(sheet.ten, so_dong, sheet.cot_cha, "Thiếu mã cha."))
            continue
        ban_ghi: dict = {}
        hong = False
        for c in cot_du_lieu:
            if c.nhan not in vi_tri or c.chi_doc:
                continue
            i = vi_tri[c.nhan]
            gt = hang[i] if i < len(hang) else None
            if _rong(gt):
                ban_ghi[c.field] = None
                continue
            try:
                ban_ghi[c.field] = _doc_o(c, gt, ctx)
            except ValueError as e:
                loi.append(LoiDong(sheet.ten, so_dong, c.nhan, str(e)))
                hong = True
        if hong:
            continue
        thu_tu: float = so_dong
        if i_thu_tu is not None and i_thu_tu < len(hang) and not _rong(hang[i_thu_tu]):
            try:
                thu_tu = _so_tu_excel(hang[i_thu_tu], NHAN_THU_TU)
            except ValueError:
                thu_tu = so_dong
        tam.setdefault(cha, []).append((thu_tu, ban_ghi))

    return {cha: [d for _, d in sorted(ds, key=lambda x: x[0])] for cha, ds in tam.items()}


def _gan_con(sheet: SheetCon, du_lieu: dict, dong_con: list[dict]) -> None:
    """Đưa tập con đã parse vào `du_lieu` (nhánh sheet THƯỜNG — không `rieng`, không `ap_dung`)."""
    if sheet.rut_gon:
        gia_tri: Any = [d.get(sheet.rut_gon) for d in dong_con
                        if d.get(sheet.rut_gon) not in (None, "")]
    else:
        gia_tri = [dict(d) for d in dong_con]
    if sheet.trong_json:
        tui = du_lieu.get(sheet.trong_json[0])
        tui = dict(tui) if isinstance(tui, dict) else {}
        cho = tui
        for khoa in sheet.trong_json[1:-1]:
            con = cho.get(khoa)
            cho[khoa] = dict(con) if isinstance(con, dict) else {}
            cho = cho[khoa]
        cho[sheet.trong_json[-1]] = gia_tri
        du_lieu[sheet.trong_json[0]] = tui
        return
    du_lieu[sheet.field] = gia_tri


def _ap_sheet_con(spec: CatalogExcelSpec, thong: dict, con_theo_sheet: dict, ma: str, cu,
                  ctx: NguCanh) -> dict[str, list[dict]]:
    """Đưa mọi sheet con vào `thong`. Trả về các sheet `rieng` để `spec.gop_con` tự ghép.

    Ba luật gặp nhau ở đây:

    * sheet CÓ MẶT ⇒ thay trọn tập con của mã cha này (dòng con bị xoá khỏi file = xoá thật);
    * sheet VẮNG MẶT ⇒ giữ nguyên. Với hầu hết field thì "giữ nguyên" là KHÔNG gán gì (`_gan` bỏ
      qua khoá vắng), riêng field có `giu_khi_vang` phải gán LẠI giá trị cũ — xem `SheetCon`;
    * túi JSON (`fields_theo_loai`) mà chỉ MỘT sheet con có mặt: phải mồi nguyên túi cũ vào trước,
      nếu không khoá của sheet vắng mặt (và mọi khoá lạ) bị ghi đè thành túi rỗng.
    """
    from copy import deepcopy

    goc_json = {s.trong_json[0] for s in spec.sheets_con
                if s.trong_json and s.ten in con_theo_sheet}
    for goc in goc_json:
        if cu is not None and goc not in thong:
            thong[goc] = deepcopy(getattr(cu, goc, None) or {})

    rieng: dict[str, list[dict]] = {}
    for sheet in spec.sheets_con:
        if sheet.toan_cuc or sheet.ap_dung is not None:
            continue
        if sheet.ten not in con_theo_sheet:
            if sheet.giu_khi_vang is not None and cu is not None and sheet.field:
                thong[sheet.field] = sheet.giu_khi_vang(cu, ctx)
            continue
        dong_con = con_theo_sheet[sheet.ten].get(_khoa_ma(ma), [])
        if sheet.rieng:
            rieng[sheet.ten] = dong_con
        else:
            _gan_con(sheet, thong, dong_con)
    return rieng


def _truong_cua_sheet(spec: CatalogExcelSpec) -> set[str]:
    """Field do SHEET CON làm chủ — so-sánh-có-đổi-không của chúng đi đường riêng."""
    return {s.trong_json[0] if s.trong_json else s.field
            for s in spec.sheets_con if s.field or s.trong_json}


def nhap_excel(spec: CatalogExcelSpec, svc, InModel, du_lieu: bytes, *,
               actor_id: int | None, ghi: bool,
               bat_loi: tuple[type[Exception], ...] = ()) -> KetQua:
    """Đọc file → dựng kế hoạch → chạy trong MỘT giao dịch. `ghi=False` = xem trước (rollback).

    `bat_loi` — họ exception NGHIỆP VỤ của màn (`CatalogError` + lớp riêng). KHÔNG bắt `Exception`
    trần: lỗi lập trình phải nổ ra 500 chứ không nằm im thành một dòng "lỗi dữ liệu".
    """
    wb = _mo_workbook(du_lieu)
    _kiem_meta(wb, spec)

    ctx = NguCanh(svc=svc, db=db_cua(svc))
    kq = KetQua()

    # -- sheet chính -------------------------------------------------------------------
    ten_chinh = spec.tieu_de[:31]
    ws_chinh = wb[ten_chinh] if ten_chinh in wb.sheetnames else _sheet_dau(wb)
    if ws_chinh is None:
        raise ExcelSaiMan(f'File thiếu sheet "{ten_chinh}".')
    tieu_de, dong_chinh = _bang(ws_chinh)
    if NHAN_MA not in tieu_de:
        raise ExcelSaiMan(f'Sheet chính thiếu cột "{NHAN_MA}" — không biết cập nhật dòng nào.')
    i_ma = tieu_de.index(NHAN_MA)
    co_mat = {c.nhan: i for c in spec.cot if (i := c.vi_tri(tieu_de)) is not None}

    # -- sheet con CÓ MẶT (vắng mặt = giữ nguyên dữ liệu con) --------------------------
    con_theo_sheet: dict[str, dict[str, list[dict]]] = {}
    for sheet in spec.sheets_con:
        ten = sheet.ten[:31]
        if ten in wb.sheetnames:
            con_theo_sheet[sheet.ten] = _parse_sheet_con(wb[ten], sheet, ctx, kq.loi)
    giu_nguyen = _doc_sheet_giu(wb)
    do_sheet = _truong_cua_sheet(spec)

    # -- dựng kế hoạch từng dòng -------------------------------------------------------
    ke_hoach: list[tuple[int, str, Any, dict, set[str]]] = []
    da_gap: set[str] = set()
    for so_dong, hang in enumerate(dong_chinh, start=2):
        kq.tong_dong += 1
        ma = _chuoi(hang[i_ma] if i_ma < len(hang) else None)
        if not ma:
            kq.loi.append(LoiDong(ten_chinh, so_dong, NHAN_MA, "Thiếu mã."))
            continue
        if _khoa_ma(ma) in da_gap:
            kq.loi.append(LoiDong(ten_chinh, so_dong, NHAN_MA, f'Mã "{ma}" xuất hiện hai lần.'))
            continue
        da_gap.add(_khoa_ma(ma))

        cu = svc.find_by_ma(ma)
        thong, hong, tu_cot = _doc_dong_chinh(spec, hang, co_mat, ctx, cu, ten_chinh, so_dong,
                                              kq.loi)
        if hong:
            continue
        thong["ma"] = ma

        rieng = _ap_sheet_con(spec, thong, con_theo_sheet, ma, cu, ctx)
        if rieng and spec.gop_con is not None:
            spec.gop_con(thong, rieng, ctx)
        if spec.gop_an is not None and _khoa_ma(ma) in giu_nguyen:
            spec.gop_an(thong, giu_nguyen[_khoa_ma(ma)], ctx)
        if spec.truoc_khi_ghi is not None:
            try:
                thong = spec.truoc_khi_ghi(thong, ctx, cu)
            except ValueError as e:
                kq.loi.append(LoiDong(ten_chinh, so_dong, "—", str(e)))
                continue

        ke_hoach.append((so_dong, ma, cu, thong, do_sheet - tu_cot))

    # -- chạy ---------------------------------------------------------------------------
    db = ctx.db
    with mot_giao_dich(db) as gd:
        cho_ap: list[tuple[int, str, Any]] = []
        for so_dong, ma, cu, thong, bo_qua in ke_hoach:
            diem = db.begin_nested()
            try:
                loai_thay_doi, obj = _ghi_mot_dong(spec, svc, InModel, ctx, ma, cu, thong,
                                                   con_theo_sheet, bo_qua, actor_id)
                diem.commit()
            except (ValueError, *bat_loi) as e:
                diem.rollback()
                kq.loi.append(LoiDong(ten_chinh, so_dong, "—", str(e)))
                continue
            if obj is not None:
                cho_ap.append((so_dong, ma, obj))
            if loai_thay_doi == "tao":
                kq.tao_moi += 1
            elif loai_thay_doi == "sua":
                kq.cap_nhat += 1
            else:
                kq.khong_doi += 1

        # Sheet `ap_dung` (cặp quy đổi của Đơn vị) chạy SAU khi MỌI dòng chính đã ghi. Nó tra mã
        # ở NGAY TRONG danh mục này — `ram → to` mà `to` nằm dưới `ram` trong file thì tra lúc ghi
        # dòng `ram` là chưa có. Sắp thứ tự dòng lúc xuất không cứu được: người ta chèn dòng mới
        # vào cuối file, và cả file là MỘT giao dịch nên nhập hai lượt cũng không xong.
        for so_dong, ma, obj in cho_ap:
            diem = db.begin_nested()
            try:
                _chay_ap_dung(spec, ctx, obj, ma, con_theo_sheet, actor_id)
                diem.commit()
            except (ValueError, *bat_loi) as e:
                diem.rollback()
                kq.loi.append(LoiDong(ten_chinh, so_dong, "—", str(e)))

        # Sheet TOÀN CỤC (danh mục Nhóm máy đi kèm màn Máy) — không khoá theo mã cha nên chạy
        # một lượt sau cùng.
        for sheet in spec.sheets_con:
            if not (sheet.toan_cuc and sheet.ap_dung is not None):
                continue
            if sheet.ten not in con_theo_sheet:
                continue
            diem = db.begin_nested()
            try:
                sheet.ap_dung(ctx, None, con_theo_sheet[sheet.ten].get("", []), actor_id)
                diem.commit()
            except (ValueError, *bat_loi) as e:
                diem.rollback()
                kq.loi.append(LoiDong(sheet.ten, 0, "—", str(e)))

        kq.hop_le = not kq.loi
        # Xem trước KHÔNG bao giờ ghi; có lỗi cũng không ghi — một file là một giao dịch.
        gd["chot"] = bool(ghi and kq.hop_le)
        kq.da_ghi = gd["chot"]
    return kq


def _sheet_dau(wb):
    """Sheet dữ liệu đầu tiên — đường TƯƠNG THÍCH cho file đời cũ (sheet tên khác, không `_meta`)."""
    for ten in wb.sheetnames:
        if ten not in (SHEET_META, SHEET_GIU):
            return wb[ten]
    return None


def _doc_sheet_giu(wb) -> dict[str, dict]:
    if SHEET_GIU not in wb.sheetnames:
        return {}
    ra: dict[str, dict] = {}
    _, dong = _bang(wb[SHEET_GIU])
    for hang in dong:
        ma = _khoa_ma(hang[0] if hang else None)
        if not ma or len(hang) < 2 or _rong(hang[1]):
            continue
        try:
            doc = json.loads(str(hang[1]))
        except (TypeError, ValueError):
            continue
        if isinstance(doc, dict):
            ra[ma] = doc
    return ra


def _doc_dong_chinh(spec: CatalogExcelSpec, hang, co_mat: dict[str, int], ctx: NguCanh, cu,
                    ten_sheet: str, so_dong: int,
                    loi: list[LoiDong]) -> tuple[dict, bool, set[str]]:
    """Một dòng sheet chính → `(dict field, có lỗi, field do CỘT CHÍNH ghi)`.

    Cột CÓ mặt mà ô trống ⇒ xoá giá trị tường minh. Tập field thứ ba dùng cho so-sánh-có-đổi-không:
    field mà sheet con làm chủ thì so bằng đường sheet con, TRỪ khi chính cột đời cũ ở sheet chính
    vừa ghi vào nó (xem `Cot.chi_nhap`).
    """
    thong: dict = {}
    tu_cot: set[str] = set()
    hong = False
    for c in spec.cot:
        if c.nhan not in co_mat or c.chi_doc or c.field == "ma":
            continue
        i = co_mat[c.nhan]
        gt = hang[i] if i < len(hang) else None
        if _rong(gt):
            # Cột VẮNG mặt không vào tới đây (`co_mat` lọc trước) — nên chỗ này đúng là "người ta
            # xoá trắng ô". Với dòng TẠO MỚI thì để mặc định của schema lo, đừng ép None.
            if cu is not None:
                thong[c.field] = False if c.kieu == "bool" else None
                tu_cot.add(c.field)
            continue
        try:
            thong[c.field] = _doc_o(c, gt, ctx)
            tu_cot.add(c.field)
        except ValueError as e:
            loi.append(LoiDong(ten_sheet, so_dong, c.nhan, str(e)))
            hong = True
    if _lech_cot_doi_chieu(spec, hang, co_mat, ctx, thong, tu_cot, ten_sheet, so_dong, loi):
        hong = True
    return thong, hong, tu_cot


def _lech_cot_doi_chieu(spec: CatalogExcelSpec, hang, co_mat: dict[str, int], ctx: NguCanh,
                        thong: dict, tu_cot: set[str], ten_sheet: str, so_dong: int,
                        loi: list[LoiDong]) -> bool:
    """Cột TÊN (`chi_doc`) phải khớp bản ghi mà cột MÃ bên cạnh đang trỏ tới.

    Mã của Phòng ban / Khách hàng do máy cấp theo thứ tự tạo, nên `PB008` ở máy nguồn và `PB008` ở
    máy đích rất dễ là hai phòng khác nhau. Trước 30/08/2026 cột tên bị bỏ qua HOÀN TOÀN khi nhập:
    file ghi `PB008 · Tổ Chế bản`, DB đích có `PB008 · Giao hàng`, dữ liệu vào NHẦM phòng mà không
    một dòng cảnh báo. Cột tên vẫn KHÔNG được ghi vào DB — nó chỉ làm chốt chặn.
    """
    ma_theo_field = {c.field: c for c in spec.cot if not c.chi_doc and c.nhan in co_mat}
    lech = False
    for c in spec.cot:
        if not c.chi_doc or c.ghi is None or c.nhan not in co_mat:
            continue
        i = co_mat[c.nhan]
        gt = hang[i] if i < len(hang) else None
        # Ô tên trống, hoặc chính ô mã trống/hỏng ⇒ không có gì để đối chiếu.
        if _rong(gt) or c.field not in tu_cot or thong.get(c.field) is None:
            continue
        that = c.ghi(thong[c.field], ctx)
        if not that or str(gt).strip().lower() == str(that).strip().lower():
            continue
        cot_ma = ma_theo_field.get(c.field)
        o_ma = hang[co_mat[cot_ma.nhan]] if cot_ma is not None else None
        loi.append(LoiDong(
            ten_sheet, so_dong, c.nhan,
            f'"{c.nhan}" trong file là "{str(gt).strip()}", nhưng '
            f'"{cot_ma.nhan if cot_ma else "mã"}" = {o_ma} ở hệ này lại là "{that}". '
            "Mã do máy cấp theo thứ tự tạo nên hai máy dễ lệch nhau — sửa lại mã cho khớp tên."))
        lech = True
    return lech


def _ghi_mot_dong(spec: CatalogExcelSpec, svc, InModel, ctx: NguCanh, ma: str, cu, thong: dict,
                  con_theo_sheet: dict, bo_qua: set[str],
                  actor_id: int | None) -> tuple[str, Any]:
    """Ghi MỘT dòng. Trả `("tao" | "sua" | "khong_doi", bản ghi cần chạy `ap_dung` hoặc None)`.

    `bo_qua` — field KHÔNG so ở vòng vô hướng vì sheet con đã so rồi (xem `_doc_dong_chinh`).

    Sheet `ap_dung` KHÔNG chạy ở đây — người gọi gom lại chạy sau, xem `nhap_excel`.
    """
    if cu is None:
        obj = svc.create(_dung_payload(InModel, thong), actor_id=actor_id)
        return "tao", obj

    doi_chinh = _co_doi(cu, thong, bo_qua)
    doi_con = _co_doi_con(spec, cu, ma, thong, con_theo_sheet, ctx)
    if not doi_chinh and not doi_con:
        return "khong_doi", None

    svc.update(cu.id, _dung_payload(InModel, {**_mac_dinh_tu_ban_ghi(InModel, cu), **thong}),
               actor_id=actor_id)
    return "sua", cu


def _dung_payload(InModel, thong: dict) -> dict:
    from pydantic import ValidationError

    try:
        return InModel(**thong).model_dump(exclude_unset=True)
    except ValidationError as e:
        raise ValueError(_cau_loi_pydantic(e)) from None


#: Kiểu giá trị được phép mồi lại từ bản ghi cũ ở `_mac_dinh_tu_ban_ghi` (xem docstring ở đó).
_VO_HUONG = (str, int, float, bool, Decimal, date, datetime)


def _mac_dinh_tu_ban_ghi(InModel, obj) -> dict:
    """Ảnh chụp bản ghi cũ làm NỀN cho payload cập nhật: field bắt buộc + mọi ô VÔ HƯỚNG.

    Vì sao cần field BẮT BUỘC: file có thể chỉ mang vài cột (`Mã` + `Tên`), trong khi `InModel` đòi
    đủ field bắt buộc (`nhom` của Công đoạn, `gsm` của Giấy). Không mồi giá trị cũ vào thì mọi lần
    nhập một phần đều ăn "field required" — trái với luật đã chốt: "cột vắng mặt thì giữ nguyên".

    Vì sao mồi cả ô KHÔNG bắt buộc: `_validate` của service soi `data` như một BẢN KHAI ĐẦY ĐỦ (màn
    hình luôn gửi cả form), nên nó kiểm CHÉO giữa các ô. Công đoạn là ca thật: định mức đầu việc
    được `giu_khi_vang` gán lại khi file thiếu sheet con, nhưng `department_id` thì không có trong
    file ⇒ "Muốn khai định mức đầu việc phải chọn tổ phụ trách", dù người ta chỉ đổi mỗi cái tên.
    Mồi lại nguyên trạng ⇒ ghi đúng giá trị đang có ⇒ không đẻ dòng nhật ký "thay đổi" ma.

    Chỉ mồi ô VÔ HƯỚNG: field dạng bảng con / danh sách là việc của sheet con và `giu_khi_vang` —
    kéo collection của ORM vào payload là ghi đè bậy đúng thứ sheet con vừa dựng.
    """
    ra: dict = {}
    for ten, f in InModel.model_fields.items():
        if not hasattr(obj, ten):
            continue
        gt = getattr(obj, ten)
        if not f.is_required() and gt is not None and not isinstance(gt, _VO_HUONG):
            continue
        ra[ten] = float(gt) if isinstance(gt, Decimal) else gt
    return ra


def _cau_loi_pydantic(e) -> str:
    cot = ", ".join(str(x["loc"][0]) for x in e.errors() if x.get("loc"))
    return f"Dữ liệu không hợp lệ ({cot})." if cot else "Dữ liệu không hợp lệ."


def _co_doi(cu, thong: dict, bo_qua: set[str]) -> bool:
    """Có field VÔ HƯỚNG nào ở sheet chính khác giá trị đang lưu không."""
    for khoa, moi in thong.items():
        if khoa in bo_qua or khoa == "ma":
            continue
        if not _bang_nhau(getattr(cu, khoa, None), moi):
            return True
    return False


def _cat_theo_cot(sheet: SheetCon, dong: list[dict]) -> list[dict]:
    """Chỉ giữ các khoá SHEET tả được — so hai bên trên cùng một tập cột, không lệch."""
    khoa = [c.field for c in (*sheet.khoa_phu, *sheet.cot)]
    return [{k: d.get(k) for k in khoa} for d in dong]


def _co_doi_con(spec: CatalogExcelSpec, cu, ma: str, thong: dict, con_theo_sheet: dict,
                ctx: NguCanh) -> bool:
    """Tập con nào khác so với đang lưu — so trên đúng các cột sheet tả được."""
    for sheet in spec.sheets_con:
        if sheet.toan_cuc or sheet.ten not in con_theo_sheet:
            continue
        moi = con_theo_sheet[sheet.ten].get(_khoa_ma(ma), [])
        if not _bang_nhau(_cat_theo_cot(sheet, doc_con(sheet, cu, ctx)),
                          _cat_theo_cot(sheet, moi)):
            return True
    return False


def _chay_ap_dung(spec: CatalogExcelSpec, ctx: NguCanh, obj, ma: str, con_theo_sheet: dict,
                  actor_id: int | None) -> None:
    """Ghi các tập con nằm NGOÀI `InModel` (cặp quy đổi của Đơn vị)."""
    for sheet in spec.sheets_con:
        if sheet.ap_dung is None or sheet.toan_cuc or sheet.ten not in con_theo_sheet:
            continue
        sheet.ap_dung(ctx, obj, con_theo_sheet[sheet.ten].get(_khoa_ma(ma), []), actor_id)
