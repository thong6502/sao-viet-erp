"""Báo cáo KCS — tổng hợp theo filter + scope, và xuất Excel hai sheet (§5.7, mg 0250).

TÁCH RIÊNG khỏi `kcs.py` (chỉ đọc, không viết) để không làm file đó phình thêm. Endpoint JSON
(`bao_cao_kcs`) và endpoint Excel (`xuat_excel_kcs`) dùng CHUNG một hàm lấy dòng dữ liệu
(`_hang_kcs_theo_scope`) — bắt buộc để "cùng filter trả cùng tổng" (§9 mục 10) không thể lệch
khi một bên sửa mà quên bên kia.

`kcs_department_id` trên `SanXuatKcsBatch` CHỈ được set cho batch `dot_xuat` (kiêm nhiệm); batch
`routing` để NULL — tổ sở hữu thật của nó là `cong_viec.department_id` (tổ đang chạy việc, vì
bước KCS routing đứng sẵn trong routing của CHÍNH tổ đó). `_to_kcs_hieu_luc` gộp hai nhánh này —
cùng khái niệm "tổ hiệu lực" mà `_gate_dieu_chinh` (Task 6) đã dùng cho gate ghi, ở đây dùng cho
lọc/scope đọc.
"""
from __future__ import annotations

from datetime import date, datetime
from zoneinfo import ZoneInfo

from sqlalchemy import select
from sqlalchemy.orm import Session

from ...models.cong_doan import CongDoan
from ...models.department import Department
from ...models.lsx import Lsx
from ...models.order import Order
from ...models.san_xuat import SanXuatCongViec
from ...models.san_xuat_kcs import (
    KCS_DAT,
    KCS_DAT_MOT_PHAN,
    KCS_KHONG_DAT,
    KCS_LOAI_DOT_XUAT,
    KCS_LOAI_ROUTING,
    SanXuatKcsBatch,
    SanXuatKcsLoi,
)
from ...models.user import User
from ...repositories.san_xuat_kcs_repo import SanXuatKcsRepository
from ...services.rbac_service import AuthorizationService
from ..gio_xuong import lich_hien_thi
from .board import _to_thay_duoc
from .thuc_thi import _aware

_VN_TZ = ZoneInfo("Asia/Bangkok")
_LOAI_LABEL = {KCS_LOAI_ROUTING: "Routing", KCS_LOAI_DOT_XUAT: "Đột xuất"}
_KET_LUAN_LABEL = {KCS_DAT: "Đạt", KCS_DAT_MOT_PHAN: "Đạt một phần", KCS_KHONG_DAT: "Không đạt"}
_CHUA_GAN_TO = "Chưa gán"
_CHUA_PHAN_LOAI = "Chưa phân loại"


def _to_kcs_hieu_luc(cv: SanXuatCongViec, kcs: SanXuatKcsBatch) -> int | None:
    """Tổ SỞ HỮU kết quả KCS — routing lấy tổ đang chạy việc, đột xuất lấy tổ đi kiểm (mục 3.4)."""
    if kcs.loai == KCS_LOAI_DOT_XUAT:
        return kcs.kcs_department_id
    return cv.department_id


def _ve_gio_vn(dt: datetime | None) -> datetime | None:
    """Quy datetime aware/naive-UTC về giờ VN, bỏ tzinfo (wall-clock) — đúng convention `_naive`
    đã dùng khắp `services/san_xuat/*.py` cho đầu ra hiển thị."""
    if dt is None:
        return None
    return _aware(dt).astimezone(_VN_TZ).replace(tzinfo=None)


def _ngay_vn(dt: datetime | None) -> date | None:
    v = _ve_gio_vn(dt)
    return v.date() if v else None


def _hang_kcs_theo_scope(
    db: Session,
    user: User,
    authz: AuthorizationService,
    *,
    tu: date | None = None,
    den: date | None = None,
    kcs_department_id: int | None = None,
    lsx_id: int | None = None,
    tu_khoa: str | None = None,
    cong_doan_id: int | None = None,
    loai: str | None = None,
    nhom_loi_id: int | None = None,
) -> list[tuple[SanXuatKcsBatch, SanXuatCongViec]]:
    """Danh sách (batch, công việc) đã lọc filter + scope — NGUỒN DUY NHẤT cho cả `bao_cao_kcs`
    và `xuat_excel_kcs` (§9 mục 10: hai đầu ra phải cùng tổng)."""
    _tos, ids_thay_duoc = _to_thay_duoc(db, user, authz)

    stmt = (
        select(SanXuatKcsBatch, SanXuatCongViec)
        .join(SanXuatCongViec, SanXuatKcsBatch.cong_viec_id == SanXuatCongViec.id)
        .order_by(SanXuatKcsBatch.bat_dau, SanXuatKcsBatch.id)
    )
    if loai:
        stmt = stmt.where(SanXuatKcsBatch.loai == loai)
    if lsx_id:
        stmt = stmt.where(SanXuatCongViec.lsx_id == lsx_id)
    if cong_doan_id:
        # Không FK trực tiếp từ SanXuatCongViec tới danh mục CongDoan (công đoạn neo LỎNG qua
        # lsx_cong_doan_id/bai_ghep_cong_doan_id — id KHÔNG ổn định, replace_routing tái sinh id,
        # join trực tiếp có thể mất dấu batch cũ). So khớp theo TÊN snapshot thay vào — cùng cách
        # công đoạn được hiển thị ở mọi nơi khác trong module (Ruling 3).
        cd = db.get(CongDoan, cong_doan_id)
        ten_cd = cd.ten if cd is not None else "\0__khong_khop__"
        stmt = stmt.where(SanXuatCongViec.ten_cong_doan == ten_cd)
    if nhom_loi_id:
        stmt = stmt.where(
            SanXuatKcsBatch.id.in_(
                select(SanXuatKcsLoi.kcs_batch_id).where(SanXuatKcsLoi.nhom_loi_id == nhom_loi_id)
            )
        )
    if tu_khoa:
        kw = f"%{tu_khoa.strip()}%"
        stmt = stmt.where(
            SanXuatCongViec.lsx_id.in_(
                select(Lsx.id)
                .join(Order, Lsx.order_id == Order.id, isouter=True)
                .where((Lsx.ma.ilike(kw)) | (Order.order_no.ilike(kw)))
            )
        )

    out: list[tuple[SanXuatKcsBatch, SanXuatCongViec]] = []
    for kcs, cv in db.execute(stmt).all():
        to_id = _to_kcs_hieu_luc(cv, kcs)
        if to_id not in ids_thay_duoc:
            continue
        if kcs_department_id and to_id != kcs_department_id:
            continue
        d = _ngay_vn(kcs.bat_dau)
        if tu and (d is None or d < tu):
            continue
        if den and (d is None or d > den):
            continue
        out.append((kcs, cv))
    return out


def _checklist_rows_cho_batch(kcs: SanXuatKcsBatch, cv: SanXuatCongViec) -> list[dict]:
    """1 dòng/tiêu chí/kết quả — JOIN `checklist_json` (câu trả lời, key `thu_tu`) với
    `cv.kcs_tieu_chi_json` (tên/mã/bắt buộc, cùng key `thu_tu`) — hai field JSON KHÁC NHAU
    (mục 3.7), không phải cùng một cột."""
    if not kcs.checklist_json or not cv.kcs_tieu_chi_json:
        return []
    tieu_chi = {
        t.get("thu_tu"): t for t in cv.kcs_tieu_chi_json if isinstance(t, dict)
    }
    out = []
    for kq in kcs.checklist_json:
        if not isinstance(kq, dict):
            continue
        tc = tieu_chi.get(kq.get("thu_tu")) or {}
        out.append({
            "kcs_batch_id": kcs.id,
            "thoi_diem": lich_hien_thi(kcs.bat_dau),
            "ma": tc.get("ma"),
            "ten": tc.get("ten"),
            "bat_buoc": tc.get("bat_buoc"),
            "dat": bool(kq.get("dat")),
            "ghi_chu": kq.get("ghi_chu"),
        })
    return out


def _xep_hang_loi(db: Session, batch_ids: list[int]) -> dict:
    """Bảng xếp hạng nhóm lỗi / công đoạn / tổ — SUM(so_luong) giảm dần (mục 2 dùng TỔNG số
    lượng chứ không phải trung bình tỷ lệ; mục 3 áp cùng nguyên tắc TỔNG cho xếp hạng). Tổ: LOẠI
    HẲN các dòng `to_chiu_id IS NULL` (mục 4 — không gán bừa vào "chưa xác định")."""
    if not batch_ids:
        return {"nhom_loi": [], "cong_doan": [], "to": []}
    rows = list(
        db.scalars(select(SanXuatKcsLoi).where(SanXuatKcsLoi.kcs_batch_id.in_(batch_ids)))
    )
    ten_nhom = SanXuatKcsRepository(db).nhan_ly_do({l.nhom_loi_id for l in rows if l.nhom_loi_id})
    cv_ids = {l.cong_doan_ref_id for l in rows if l.cong_doan_ref_id}
    ten_cong_doan_map = (
        {c.id: c.ten_cong_doan for c in db.scalars(
            select(SanXuatCongViec).where(SanXuatCongViec.id.in_(cv_ids))
        )}
        if cv_ids else {}
    )
    to_ids = {l.to_chiu_id for l in rows if l.to_chiu_id}
    to_ten_map = (
        {d.id: d.name for d in db.scalars(select(Department).where(Department.id.in_(to_ids)))}
        if to_ids else {}
    )

    nhom_acc: dict[int | None, float] = {}
    cd_acc: dict[str, float] = {}
    to_acc: dict[int, float] = {}
    for l in rows:
        sl = float(l.so_luong or 0)
        nhom_acc[l.nhom_loi_id] = nhom_acc.get(l.nhom_loi_id, 0.0) + sl
        if l.cong_doan_ref_id and l.cong_doan_ref_id in ten_cong_doan_map:
            key = ten_cong_doan_map[l.cong_doan_ref_id]
            cd_acc[key] = cd_acc.get(key, 0.0) + sl
        if l.to_chiu_id:  # mục 4: bỏ hẳn khỏi xếp hạng tổ nếu chưa xác định trách nhiệm
            to_acc[l.to_chiu_id] = to_acc.get(l.to_chiu_id, 0.0) + sl

    nhom_loi = sorted(
        (
            {
                "nhom_loi_id": k,
                "ten": (ten_nhom.get(k) if k else None) or _CHUA_PHAN_LOAI,
                "tong_so_luong": v,
            }
            for k, v in nhom_acc.items()
        ),
        key=lambda r: r["tong_so_luong"], reverse=True,
    )
    cong_doan = sorted(
        ({"ten_cong_doan": k, "tong_so_luong": v} for k, v in cd_acc.items()),
        key=lambda r: r["tong_so_luong"], reverse=True,
    )
    to = sorted(
        (
            {"to_id": k, "ten": to_ten_map.get(k, ""), "tong_so_luong": v}
            for k, v in to_acc.items()
        ),
        key=lambda r: r["tong_so_luong"], reverse=True,
    )
    return {"nhom_loi": nhom_loi, "cong_doan": cong_doan, "to": to}


def bao_cao_kcs(
    db: Session, user: User, authz: AuthorizationService, **filters,
) -> dict:
    """Tổng hợp cho dashboard (§6.2 KPI + biểu đồ). `**filters` = đúng 7 tham số của
    `_hang_kcs_theo_scope` (tu/den/kcs_department_id/lsx_id/tu_khoa/cong_doan_id/loai/
    nhom_loi_id) — router forward nguyên `Query()` params vào đây."""
    hang = _hang_kcs_theo_scope(db, user, authz, **filters)
    batch_ids = [kcs.id for kcs, _cv in hang]

    tong_nhan = sum(float(kcs.so_luong_nhan) for kcs, _cv in hang)
    tong_dat = sum(float(kcs.so_luong_dat) for kcs, _cv in hang)
    tong_loi = sum(float(kcs.so_luong_khong_dat) for kcs, _cv in hang)
    ty_le_dat = (tong_dat / tong_nhan) if tong_nhan else None  # mục 2: tổng/tổng, KHÔNG trung bình

    theo_ngay: dict[date, dict[str, float]] = {}
    for kcs, _cv in hang:
        d = _ngay_vn(kcs.bat_dau)
        if d is None:
            continue
        acc = theo_ngay.setdefault(d, {"tong_nhan": 0.0, "tong_dat": 0.0, "tong_loi": 0.0})
        acc["tong_nhan"] += float(kcs.so_luong_nhan)
        acc["tong_dat"] += float(kcs.so_luong_dat)
        acc["tong_loi"] += float(kcs.so_luong_khong_dat)
    theo_ngay_list = [
        {"ngay": d, **acc} for d, acc in sorted(theo_ngay.items())
    ]

    xep_hang = _xep_hang_loi(db, batch_ids)
    return {
        "tong_luot": len(hang),
        "tong_nhan": tong_nhan,
        "tong_dat": tong_dat,
        "tong_loi": tong_loi,
        "ty_le_dat": ty_le_dat,
        "theo_ngay": theo_ngay_list,
        "nhom_loi": xep_hang["nhom_loi"],
        "cong_doan": xep_hang["cong_doan"],
        "to": xep_hang["to"],
    }


def _ten_file_xlsx(tu: date | None, den: date | None) -> str:
    t = tu.isoformat() if tu else "toanbo"
    d = den.isoformat() if den else "toanbo"
    return f"bao-cao-kcs-{t}_{d}.xlsx"


def xuat_excel_kcs(
    db: Session, user: User, authz: AuthorizationService, **filters,
) -> tuple[bytes, str]:
    """Workbook 2 sheet (§5.7) — dùng CHUNG `_hang_kcs_theo_scope` với `bao_cao_kcs` (§9 mục 10).
    Trả `(bytes, filename)`."""
    from io import BytesIO

    from openpyxl import Workbook  # lazy import: thiếu dep chỉ hỏng route này, không sập app
    from openpyxl.styles import Font

    hang = _hang_kcs_theo_scope(db, user, authz, **filters)
    batch_ids = [kcs.id for kcs, _cv in hang]

    lsx_ids = {cv.lsx_id for _kcs, cv in hang if cv.lsx_id}
    lsx_ma = (
        {l.id: l.ma for l in db.scalars(select(Lsx).where(Lsx.id.in_(lsx_ids)))}
        if lsx_ids else {}
    )
    to_ids = {tid for kcs, cv in hang if (tid := _to_kcs_hieu_luc(cv, kcs))}
    to_ten = (
        {d.id: d.name for d in db.scalars(select(Department).where(Department.id.in_(to_ids)))}
        if to_ids else {}
    )
    nguoi_ids = {kcs.created_by for kcs, _cv in hang if kcs.created_by}
    nguoi_ten = (
        {u.id: u.name for u in db.scalars(select(User).where(User.id.in_(nguoi_ids)))}
        if nguoi_ids else {}
    )

    repo = SanXuatKcsRepository(db)
    loi_by_batch = repo.cac_loi_nhieu(batch_ids)
    ten_nhom = repo.nhan_ly_do(
        {l.nhom_loi_id for ls in loi_by_batch.values() for l in ls if l.nhom_loi_id}
    )
    loi_ids_all = [l.id for ls in loi_by_batch.values() for l in ls]
    anh_by_loi = repo.anh_cua_loi_nhieu(loi_ids_all)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Kết quả KCS"
    headers1 = [
        "Mã kết quả", "Thời điểm", "Loại", "Tổ KCS", "Mã LSX", "Công đoạn", "Số nhận",
        "Số đạt", "Số không đạt", "Đơn vị", "Kết luận", "Ghi chú", "Người ghi",
        "Số lỗi ghi nhận", "Nhóm lỗi", "Tổ chịu trách nhiệm", "URL ảnh",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    _COL_SO_LUONG_1 = {7, 8, 9}  # Số nhận / Số đạt / Số không đạt
    for kcs, cv in hang:
        loi_list = loi_by_batch.get(kcs.id, [])
        nhom_ten_set = sorted({
            (ten_nhom.get(l.nhom_loi_id) if l.nhom_loi_id else None) or _CHUA_PHAN_LOAI
            for l in loi_list
        })
        to_chiu_ten = sorted({
            (to_ten.get(l.to_chiu_id) if l.to_chiu_id else None) or _CHUA_GAN_TO
            for l in loi_list
        })
        anh_urls = [a.file_url for l in loi_list for a in anh_by_loi.get(l.id, [])]
        ws1.append([
            kcs.id, _ve_gio_vn(kcs.bat_dau), _LOAI_LABEL.get(kcs.loai, kcs.loai),
            to_ten.get(_to_kcs_hieu_luc(cv, kcs), ""),
            lsx_ma.get(cv.lsx_id, "") if cv.lsx_id else "",
            cv.ten_cong_doan, float(kcs.so_luong_nhan), float(kcs.so_luong_dat),
            float(kcs.so_luong_khong_dat), kcs.don_vi,
            _KET_LUAN_LABEL.get(kcs.ket_luan, kcs.ket_luan), kcs.ghi_chu or "",
            nguoi_ten.get(kcs.created_by, ""), len(loi_list),
            "; ".join(nhom_ten_set), "; ".join(to_chiu_ten) if loi_list else "",
            "; ".join(anh_urls),
        ])
        r = ws1.max_row
        for c in _COL_SO_LUONG_1:
            ws1.cell(row=r, column=c).number_format = "#,##0.###"
        ws1.cell(row=r, column=2).number_format = "dd/mm/yyyy hh:mm"
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("Chi tiết checklist")
    ws2.append(["Mã kết quả", "Thời điểm", "Mã tiêu chí", "Tên tiêu chí", "Bắt buộc", "Đạt", "Ghi chú"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for kcs, cv in hang:
        for cr in _checklist_rows_cho_batch(kcs, cv):
            ws2.append([
                cr["kcs_batch_id"], _ve_gio_vn(cr["thoi_diem"]), cr["ma"] or "", cr["ten"] or "",
                "Có" if cr["bat_buoc"] else ("Không" if cr["bat_buoc"] is not None else ""),
                "Có" if cr["dat"] else "Không", cr["ghi_chu"] or "",
            ])
            ws2.cell(row=ws2.max_row, column=2).number_format = "dd/mm/yyyy hh:mm"
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), _ten_file_xlsx(filters.get("tu"), filters.get("den"))
