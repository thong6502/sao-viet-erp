"""Business service for Thu mua MVP."""
from __future__ import annotations

import io
import re
import secrets
import string
import unicodedata
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from zoneinfo import ZoneInfo

from sqlalchemy import select

from ..models.role import SCOPE_ALL, SCOPE_DEPARTMENT, SCOPE_OWN
# Soft ref (đọc-only) sang module Kho: biết đợt giao nào ĐÃ sinh yêu cầu NHẬP để chặn nhập trùng.
from ..models.stock_request import REQ_CANCELLED, StockRequest
from ..models.purchase import (
    DEPARTMENT_PURCHASE_SOURCE_TYPES,
    DPR_CANCELLED,
    DPR_DONE,
    DPR_IN_PURCHASE,
    DPR_OPEN,
    DPR_PENDING_APPROVAL,
    DepartmentPurchaseRequest,
    PR_APPROVED,
    PR_CANCELLED,
    PR_DRAFT,
    PR_PARTIALLY_RECEIVED,
    PR_PENDING,
    PR_PURCHASED,
    PR_RECEIVED,
    PR_REJECTED,
    CHANGE_BY_MAY,
    CHANGE_BY_NGUOI,
    DOC_PMH,
    DOC_YCMH,
    PURCHASE_ATTACHMENT_KINDS,
    PURCHASE_ATTACHMENT_HOP_DONG,
    PurchaseAttachment,
    PurchaseDelivery,
    PurchaseDeliveryLine,
    SUPPLIER_ACTIVE,
    SUPPLIER_INACTIVE,
    SUPPLIER_STATUSES,
    PurchaseRequest,
    SOURCE_CONG_NGHE,
    SOURCE_GIA_CONG_NGOAI,
    SOURCE_KHAC,
    SOURCE_KHO,
    SOURCE_KINH_DOANH,
    SOURCE_SAN_XUAT,
    Supplier,
)
from ..models.accounting import (
    PAYMENT_RECEIPT_RECEIVED,
    PAYMENT_STAGE_ADVANCE,
    PAYMENT_VOUCHER_CANCELLED,
    PAYMENT_VOUCHER_PAID,
)
from ..models.vat_lieu_kho import HANG_LOAI
from ..repositories.audit_repo import AuditLogRepository
from ..repositories.purchase_repo import (
    DepartmentPurchaseRequestLineInput,
    DepartmentPurchaseRequestRepository,
    PurchaseRequestLineInput,
    PurchaseRequestRepository,
    PurchaseStatusHistoryRepository,
    SupplierItemInput,
    SupplierRepository,
)
from ..repositories.rbac_repo import DepartmentRepository
from ..repositories.user_repo import UserRepository
from ..storage import get_storage, key_from_url, make_key, url_from_key
from .danh_gia_ncc import DanhGiaNcc, tu_tong_hop
from .rbac_service import AuthorizationService


# Kho file của mua hàng. ⚠️ Tiền tố này PHẢI có trong `_PREFIX_PERMISSION` (`routers/files.py`) —
# bảng đó fail-MỞ: tiền tố không khai thì chỉ cần đăng nhập là đọc được, tức hợp đồng NCC lộ cho
# toàn công ty.
PURCHASE_ATTACHMENT_SUBDIR = "mua-hang"
MAX_PURCHASE_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_PURCHASE_ATTACHMENTS = 30


# Những module được cấp quyền ĐỌC danh sách YCMH. Router dựng cổng quyền từ đúng danh sách này
# (`DEPARTMENT_REQUEST_READERS`), nên hai nơi không thể lệch nhau.
#
# Dùng cả cho việc co danh sách về phòng ban: ai có scope `all` ở BẤT KỲ module nào trong đây thì
# thấy YCMH toàn công ty. Chỉ hỏi mỗi `thu_mua` là sai — kế toán (SEAM-25) truy vết YCMH nguồn từ
# PMH/Phiếu chi mà KHÔNG hề có quyền `thu_mua`, `scope_for` trả None nên bị co về phòng Kế toán và
# nhìn thấy RỖNG. Người chỉ có scope phòng ban thì vẫn bị co như cũ.
DEPARTMENT_REQUEST_READER_MODULES = (
    # Khoá RIÊNG của màn Yêu cầu mua hàng (tách 10/08/2026) — phải đứng đầu: ai được cấp đúng màn
    # này thì đọc được, không cần mượn quyền của phân hệ khác.
    "yeu_cau_mua_hang",
    "thu_mua",
    "bao_gia",
    "kho",
    "san_xuat",
    "dm_giay",
    "ke_toan",
)


# Những module được cấp quyền ĐỌC phiếu mua hàng. Ai có scope `all` ở BẤT KỲ module nào trong đây
# thì thấy phiếu của toàn công ty.
#
# ⚠️ Phải có `ke_toan`: màn Đơn mua hàng của kế toán (`/api/accounting/inbox`) gọi CHUNG
# `list_requests`, mà kế toán KHÔNG có quyền `thu_mua` ⇒ `scope_for` trả None ⇒ bị co về "của
# mình" ⇒ thấy RỖNG. Đúng cái bẫy đã sập với YCMH — đừng lặp lại.
PURCHASE_REQUEST_READER_MODULES = ("thu_mua", "ke_toan")


# Thang bậc tiến độ của một phiếu mua, dùng để SUY trạng thái yêu cầu mua hàng (xem
# `_tinh_lai_trang_thai_ycmh`).
#
# BỊ TỪ CHỐI = bậc 1: YCMH vẫn bị GIỮ bởi đúng PMH đó. Thu mua sửa PMH bị từ chối rồi gửi lại,
# tuyệt đối không lập PMH thứ hai từ cùng YCMH — nếu thả về bậc 0, giao diện hiện lại nút "Tạo
# đơn" và API cũng cho tạo trùng. ĐÃ HUỶ mới là bậc 0 vì quan hệ mua đó đã kết thúc thật.
#
# NHÁP = bậc 1 (Chờ duyệt), KHÔNG phải 0. Trông lệch nhưng đúng với hệ: `_replace_sources`
# (purchase_repo.py) đẩy yêu cầu sang "Chờ duyệt" ngay khi thu mua TẠO phiếu, kể cả phiếu còn nháp
# — đó là cách hệ GIỮ CHỖ yêu cầu để người thứ hai không lập phiếu chồng lên. Để bậc 0 thì suy ra
# một đằng, repo ghi một nẻo, và chỗ giữ chỗ vỡ.
#
# ⚠️ Việc repo tự đặt trạng thái nghiệp vụ là SAI TẦNG. Gỡ được thì phải tách "giữ chỗ" ra khỏi
# "trạng thái" — ngoài phạm vi đợt này, ghi lại đây để đợt sau còn biết đường.
#
# GIAO MỘT PHẦN = bậc 2 (Đang mua), cùng bậc `approved`/`purchased`: hàng mới về một phần thì yêu
# cầu CHƯA xong. Cho nó bậc 3 là bộ phận đề nghị nhìn thấy "Hoàn tất" rồi thôi không hỏi nữa, trong
# khi hai phần ba đơn còn nằm ở kho NCC — đúng kiểu báo lạc quan mà luật "lấy bậc thấp nhất" ở dưới
# sinh ra để tránh.
_BAC_PHIEU = {
    PR_REJECTED: 1,
    PR_CANCELLED: 0,
    PR_DRAFT: 1,
    PR_PENDING: 1,
    PR_APPROVED: 2,
    PR_PURCHASED: 2,
    PR_PARTIALLY_RECEIVED: 2,
    PR_RECEIVED: 3,
}

_BAC_SANG_TRANG_THAI = {
    0: DPR_OPEN,
    1: DPR_PENDING_APPROVAL,
    2: DPR_IN_PURCHASE,
    3: DPR_DONE,
}


class PurchaseError(Exception):
    pass


# Nhãn NGẮN của trạng thái đơn mua — chỉ dùng để ghép câu chặn "món này đang nằm ở đơn nào".
# Không phải bảng nhãn của giao diện (nhãn UI ở FE); ở đây chỉ cần đủ để người đọc câu lỗi biết
# phải đi xử lý ở đâu.
_NHAN_PMH_NGAN = {
    PR_DRAFT: "còn nháp",
    PR_PENDING: "chờ duyệt",
    PR_APPROVED: "đã duyệt",
    PR_REJECTED: "bị từ chối, thu mua đang sửa",
    PR_PURCHASED: "đã đặt hàng",
    PR_PARTIALLY_RECEIVED: "đã về một phần",
    PR_RECEIVED: "đã nhận đủ",
}


class PurchaseValidationError(PurchaseError):
    pass


class PurchaseNotFound(PurchaseError):
    pass


class PurchaseConflict(PurchaseError):
    pass


class PurchaseForbidden(PurchaseError):
    pass


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _business_today() -> date:
    return datetime.now(ZoneInfo("Asia/Bangkok")).date()


def _money_round(value: Decimal) -> int:
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _doc_mat_hang(get) -> tuple[str | None, int | None]:
    """Đọc cặp `(hang_loai, hang_id)` của một dòng hàng (mg 0174).

    PHẢI ĐỦ CẶP hoặc TRỐNG CẢ HAI. Nửa cặp thì **báo lỗi**, không im lặng bỏ: `hang_loai='giay'`
    mà thiếu id là một con trỏ hỏng, và hậu quả của nó im re — dòng mua mất liên kết mặt hàng, bảng
    cân đối không cộng lô đó vào "hàng đang về", kế hoạch giục mua thêm một lô giấy nữa. Đúng cái
    bug mg 0174 sinh ra để chữa.

    "Không đoán" nghĩa là không tự bịa nửa còn thiếu — KHÔNG có nghĩa là nuốt lỗi. Cùng cách xử lý
    với dòng mặt hàng NCC (`_clean_supplier_items`), để một khái niệm chỉ có một luật.
    """
    loai = (get("hang_loai") or "").strip() or None
    raw_id = get("hang_id")
    hid = int(raw_id) if raw_id not in (None, "") else None
    if (loai is None) != (hid is None):
        raise PurchaseValidationError(
            "Mặt hàng gốc phải có cả loại lẫn mã — chọn lại ở ô Vật tư."
        )
    if loai is None:
        return None, None
    if loai not in HANG_LOAI:
        raise PurchaseValidationError(f'Loại mặt hàng "{loai}" không hợp lệ.')
    return loai, hid


def _purchase_line_amounts(
    *,
    quantity: float,
    unit_price: int,
    discount_percent: float,
    vat_percent: float,
) -> tuple[int, int, int, int]:
    gross = _money_round(Decimal(str(quantity)) * Decimal(unit_price))
    discount = _money_round(Decimal(gross) * Decimal(str(discount_percent)) / Decimal(100))
    taxable = max(0, gross - discount)
    vat = _money_round(Decimal(taxable) * Decimal(str(vat_percent)) / Decimal(100))
    return gross, discount, vat, taxable + vat


def da_giao_theo_dong(row) -> dict[int, float] | None:
    """Σ số đã giao của TỪNG DÒNG ĐẶT, gom từ mọi đợt giao của phiếu.

    Trả **None** khi phiếu chưa có đợt giao nào — đó là tín hiệu cho mọi hàm gọi phải lùi về LUẬT
    CŨ (`received_quantity ?? quantity`). Đừng đổi thành `{}`: dict rỗng nghĩa là "có đợt giao mà
    chưa nhận món nào" ⇒ mọi đơn cũ tụt về nhận 0 và công nợ biến mất sạch.
    """
    dot = getattr(row, "deliveries", None)
    if not dot:
        return None
    out: dict[int, float] = {}
    for d in dot:
        for dl in d.lines:
            key = dl.purchase_request_line_id
            out[key] = out.get(key, 0.0) + float(dl.quantity)
    return out


def qty_thuc_nhan(line, da_giao: dict[int, float] | None = None) -> float:
    """Số THỰC NHẬN của một dòng phiếu mua.

    Hai đường, theo đúng thứ tự:

    1. Phiếu **có đợt giao** ⇒ số thực nhận là Σ các đợt. `received_quantity` trên dòng khi đó là
       cột DORMANT — không ai gõ nữa, và cũng không được đọc, nếu không là hai nguồn sự thật.
    2. Phiếu **chưa có đợt giao nào** (mọi phiếu lập trước 06/08/2026) ⇒ luật cũ: `received_quantity`
       NULL = chưa ai khai = coi như nhận đủ số đặt.

    Nhánh 2 chính là cầu tương thích ngược: đơn cũ giữ nguyên từng đồng sau khi lên bản mới, không
    cần backfill dữ liệu.
    """
    if da_giao is not None:
        return float(da_giao.get(line.id, 0.0))
    return float(line.quantity if line.received_quantity is None else line.received_quantity)


def phan_bo_du_dot(row) -> dict[int, dict]:
    """Chia số nhận của TỪNG đợt thành phần TÍNH TIỀN và phần DƯ (0đ), theo LUỸ KẾ.

    Chủ chốt 28/08/2026: *"cho điền tự do, nếu điền vượt thì hệ thống tự tính phần dư ra cho dễ"*
    — NCC giao 1000 cái cho đơn đặt 500 mà giá vẫn giữ nguyên (tặng thêm) là chuyện có thật. Trước
    đó `_clean_dot_lines` chặn cứng, nên 500 cái kia nằm trong kho mà sổ không ghi nổi; lời khuyên
    cũ ("sửa số đặt rồi duyệt lại") lại càng sai vì nâng số đặt lên 1000 là công nợ tăng gấp đôi.

    LUẬT: phần TÍNH TIỀN luôn được lấp TRƯỚC, theo thứ tự đợt (`delivery_date`, `seq_no`); phần
    vượt số đặt rơi vào DƯ, giá 0đ. Hai hệ quả cố ý:

      • Tổng nợ của đơn dừng đúng ở giá trị đơn đã duyệt — gõ thêm bao nhiêu cũng không nhích.
      • Người ghi đợt KHÔNG có cần gạt nào để biến hàng phải trả tiền thành hàng miễn phí: họ chỉ
        khai được "về bao nhiêu", còn chia thì máy chia. Đây là chỗ bản thiết kế đầu (một ô "số
        tặng" gõ tay ở đợt giao) bị chính chủ bác — ô đó cho phép ghi 500 trả tiền + 500 tặng cho
        một lô mua đủ 1000, chênh lệch thanh toán ngoài sổ.

    ⚠️ Hệ KHÔNG biết phần dư có thật là hàng tặng hay không — nó chỉ biết "nhận nhiều hơn đặt và
    không tính tiền". Vì vậy con số dư phải được PHƠI RA ở giao diện (đợt giao + dòng đơn), để ca
    NCC thực sự có tính tiền phần dư bị bắt lúc đối chiếu hoá đơn, chứ không im lặng ghi thiếu nợ.

    Trả `{delivery_id: {"amount": int, "lines": {delivery_line_id: {"tinh_tien", "du"}}}}`.
    """
    line_by_id = {line.id: line for line in row.lines}
    dots = sorted(
        getattr(row, "deliveries", []) or [], key=lambda d: (d.delivery_date, d.seq_no)
    )
    luy_ke: dict[int, float] = {}
    out: dict[int, dict] = {}
    for d in dots:
        tien = 0
        chi_tiet: dict[int, dict] = {}
        for dl in d.lines:
            line = line_by_id.get(dl.purchase_request_line_id)
            nhan = float(dl.quantity)
            if line is None:
                # Dòng đợt trỏ tới dòng đặt đã bị xoá (dữ liệu lỗi): bỏ qua tiền, thà thiếu một
                # dòng còn hơn nổ cả màn công nợ. Vẫn khai báo để giao diện không mất dòng.
                chi_tiet[dl.id] = {"tinh_tien": 0.0, "du": nhan}
                continue
            dat = float(line.quantity)
            truoc = luy_ke.get(line.id, 0.0)
            sau = truoc + nhan
            luy_ke[line.id] = sau
            tinh_tien = max(0.0, min(sau, dat) - min(truoc, dat))
            chi_tiet[dl.id] = {"tinh_tien": tinh_tien, "du": max(0.0, nhan - tinh_tien)}
            if tinh_tien <= 0:
                continue
            _, _, _, thanh_tien = _purchase_line_amounts(
                quantity=tinh_tien,
                unit_price=int(line.expected_unit_price),
                discount_percent=float(line.discount_percent or 0),
                vat_percent=float(line.vat_percent or 0),
            )
            tien += thanh_tien
        out[d.id] = {"amount": tien, "lines": chi_tiet}
    return out


def gia_tri_cac_dot(row) -> dict[int, int]:
    """`{delivery_id: thành tiền}` — vỏ mỏng của `phan_bo_du_dot` cho chỗ chỉ cần tiền.

    KHÔNG còn hàm tính tiền cho MỘT đợt đứng riêng: từ 28/08/2026 tiền của một đợt phụ thuộc các
    đợt TRƯỚC nó (phần tính tiền lấp trước), nên hỏi "đợt này bao nhiêu tiền" mà không đưa cả đơn
    là câu hỏi không có đáp án đúng.
    """
    return {did: v["amount"] for did, v in phan_bo_du_dot(row).items()}


# TIỀN CỦA ĐỢT: MÁY TÍNH, KHÔNG AI GÕ TAY (chủ chốt 07/08/2026). 06/08 từng mở ô "Số tiền theo hoá
# đơn" cho gõ tay; chủ đảo lại ngay hôm sau — *"không cho sửa nữa, dựa vào số lượng thực tế tính ra
# tiền luôn"*. Ô tiền gõ tay đẻ ra đúng cái lệch mà chính chủ bắt được: chi tiết PMH hiện 1.000.000
# (số khai) còn ngoài bảng 1.100.000 (số tính) — hai con số cho cùng một đợt.
#
# Cột `purchase_deliveries.amount` thành DORMANT: giữ lại vì dự án không có Alembic và xoá là mất
# dữ liệu, nhưng KHÔNG ĐỌC và KHÔNG GHI nữa. Đừng đọc lại nó nếu chưa hỏi chủ.
def phan_bo_tien_dot(row) -> tuple[list[dict], int, int]:
    """Phân tiền đã chi của một phiếu mua về TỪNG ĐỢT GIAO.

    MỘT nguồn duy nhất cho phép phân bổ này — màn Mua hàng (trần lập phiếu, hiển thị đợt) và màn
    Công nợ (bảng nợ theo đợt) đều gọi vào đây. Hai bên tự phân lấy là hai bên lệch, mà lệch ở đây
    thì một đợt có thể biến mất khỏi công nợ ở màn này mà vẫn còn ở màn kia.

    Mỗi đợt trả về `{delivery, amount, paid, coc_bu, con_no}`:
      - `paid`   = tiền trả ĐÍCH DANH đợt (`payment_vouchers.delivery_id` trỏ đúng đợt) — cột này
                   phải khớp sao kê NCC theo từng đợt.
      - `coc_bu` = phần CỌC của cả đơn chiếu xuống, GIAO TRƯỚC BÙ TRƯỚC. Không phải ai đó trả riêng
                   cho đợt này.
      - `con_no` = giá trị − paid − coc_bu.

    Trả kèm `(tổng cọc, cọc chưa dùng hết)`. Phần trả THỪA của một đợt chảy ngược vào cọc chung để
    tổng luôn khớp `purchase_money`. Đơn không có đợt nào ⇒ `([], 0, 0)`."""
    dots = sorted(
        getattr(row, "deliveries", []) or [], key=lambda d: (d.delivery_date, d.seq_no)
    )
    if not dots:
        return [], 0, 0
    tien_dot = gia_tri_cac_dot(row)

    tra_theo_dot: dict[int, int] = {}
    coc = 0
    for v in row.payment_vouchers:
        if v.status != PAYMENT_VOUCHER_PAID:
            continue
        did = getattr(v, "delivery_id", None)
        if did is None:
            coc += int(v.amount_vnd)
        else:
            tra_theo_dot[did] = tra_theo_dot.get(did, 0) + int(v.amount_vnd)
    # Tiền NỘP LẠI hoàn lại phần ứng trước ⇒ trừ vào cọc chung, không trừ vào đợt nào.
    coc -= sum(
        int(r.amount_vnd)
        for v in row.payment_vouchers
        for r in v.receipts
        if r.status == PAYMENT_RECEIPT_RECEIVED
    )

    out: list[dict] = []
    for d in dots:
        gia_tri = tien_dot.get(d.id, 0)
        tra = tra_theo_dot.get(d.id, 0)
        if tra > gia_tri:
            coc += tra - gia_tri
            tra = gia_tri
        out.append(
            {
                "delivery": d,
                "amount": gia_tri,
                "paid": tra,
                "coc_bu": 0,
                "con_no": max(0, gia_tri - tra),
            }
        )
    con_coc = max(0, coc)
    for m in out:
        bu = min(m["con_no"], con_coc)
        m["coc_bu"] = bu
        m["con_no"] -= bu
        con_coc -= bu
    return out, coc, con_coc


def han_tra_dot(delivery, supplier, ngay_chot: date | None) -> date | None:
    """Hạn trả của một đợt giao. Thang bốn bậc, bậc trên thắng bậc dưới.

    1. **NGÀY CHỐT CÔNG NỢ của đơn** + `credit_days` (chủ chốt 28/08/2026). NCC báo một mốc chốt
       cho cả đơn rồi cho nợ N ngày kể từ mốc đó — chốt 31/8 + 30 ngày ⇒ phải trả trước 30/9.
       Đồng hồ chạy từ mốc CHỐT, nên hàng giao 05/8 và 28/8 cùng hạn 30/9; luật cũ (bậc 2) báo
       cái giao 05/8 quá hạn từ 04/9, sớm hơn thoả thuận thật 26 ngày.
    2. Ngày hóa đơn + `credit_days` — luật cũ, giữ nguyên cho đơn chưa báo ngày chốt.
    3. `due_date` — hạn chốt thủ công / dữ liệu cũ, khi chưa đủ dữ liệu để suy.
    4. Ngày giao + `credit_days`.

    `credit_days` NULL = NCC CHƯA ĐẶT hạn ⇒ trả None = đợt này **không bao giờ vào cột Quá hạn**.
    Vì thế màn Công nợ phải đẩy đợt không-có-hạn lên ĐẦU kèm badge, đúng nếp chống giấu nợ đã áp
    cho phiếu chi thiếu hạn trước đây — im lặng ở đây nghĩa là một món nợ không ai canh.

    ⚠️ `ngay_chot` là tham số BẮT BUỘC, cố ý không cho mặc định None: nó quyết hạn trả của cả
    đơn, mà mọi chỗ gọi đều đang cầm sẵn `row`. Để mặc định là mở đường cho một call-site quên
    truyền rồi âm thầm tính theo luật cũ — sai lệch kiểu đó không ai thấy tới lúc đối chiếu NCC.
    """
    so_ngay = getattr(supplier, "credit_days", None) if supplier is not None else None
    if ngay_chot is not None and so_ngay is not None:
        return ngay_chot + timedelta(days=int(so_ngay))
    if delivery.invoice_date is not None and so_ngay is not None:
        return delivery.invoice_date + timedelta(days=int(so_ngay))
    if delivery.due_date is not None:
        return delivery.due_date
    if so_ngay is not None:
        return delivery.delivery_date + timedelta(days=int(so_ngay))
    return None


def purchase_money(row) -> dict:
    """MỘT nguồn số cho mọi con tiền của một phiếu mua hàng.

    Cả `_to_request_out` (màn Mua hàng / Đơn mua hàng) lẫn màn **Công nợ phải trả** đều gọi hàm
    này. Để hai bên tự cộng lấy là để hai bên lệch nhau — mà lệch TIỀN thì không ai phát hiện cho
    tới lúc ngồi đối chiếu với NCC.

    Hai tổng đi song song, đừng lẫn:
    - `total` — giá trị **đơn đặt**, theo `quantity`. Vẫn là con số in trên đơn.
    - `received_total` — giá trị hàng **thực nhận**, theo `received_quantity`.

    - `gia_tri_da_giao` — giá trị hàng **ĐÃ VỀ**, cộng theo từng ĐỢT GIAO. Đây là số đẻ ra công nợ:
      hàng về tới đâu nợ tới đó.

    Công nợ (chủ chốt 06/08/2026 — docs/prd-mua-hang-cong-no.md §5.3):

        cong_no = max(0, gia_tri_da_giao − net_paid)

    Đúng công thức bên nghiệp vụ đọc ra — *nợ − cọc − đã trả* — vì phiếu ĐẶT CỌC cũng là một phiếu
    chi nên cọc đã nằm sẵn trong `net_paid`. Cọc vì thế **tự khấu trừ ngay từ đợt giao đầu tiên**.

    Ba lỗi mà công thức này chữa (đều đã xảy ra trên bản trước):
      (a) GIẤU NỢ — giao 1/3 đợt, đơn còn `purchased`, màn công nợ hiện 0đ. Nay đợt 1 về là nợ hiện.
      (b) THỪA NỢ — bấm "Đã nhận hàng" sớm thì ghi nợ đủ 100% khi hàng mới về 1/3.
      (c) NỢ ẢO — tạm ứng 10tr, mua 8,5tr, nộp lại 1,5tr từng ra "còn nợ 1,5tr" trong khi tiền đã về
          két. Nay: 8,5 − (10 − 1,5) = 0. Vì công nợ đo theo HÀNG ĐÃ VỀ chứ không theo giá trị đơn.

    Hai TRẦN lập phiếu chi, khác nhau theo loại phiếu:
      - `outstanding_amount` (= công nợ) — trần của phiếu THANH TOÁN. Không cho chi quá phần nợ đã
        phát sinh, nếu không kế toán trả tiền cho hàng chưa về.
      - `tran_dat_coc` — trần của phiếu ĐẶT CỌC/ứng trước = **CỌC DỰ KIẾN đã khai trên phiếu mua**
        trừ phần cọc đã chi (chủ chốt 09/08/2026). Chưa khai cọc dự kiến ⇒ trần 0 ⇒ không lập được
        phiếu cọc nào.

        Vì sao đổi gốc: trước 09/08 trần cọc lấy theo GIÁ TRỊ ĐƠN, nghĩa là đơn 500tr thì kế toán
        ứng trước được tới 500tr trong khi hai bên chỉ thoả thuận cọc 50tr — hệ không có gì để đối
        chiếu với thoả thuận. Nay số cọc phải được KHAI TRƯỚC, và chi bao nhiêu lần cũng được miễn
        tổng không vượt số đã khai.
    """
    da_giao = da_giao_theo_dong(row)
    total = 0
    received_total = 0
    for line in row.lines:
        unit_price = int(line.expected_unit_price)
        discount_percent = float(line.discount_percent or 0)
        vat_percent = float(line.vat_percent or 0)
        _, _, _, line_total = _purchase_line_amounts(
            quantity=float(line.quantity),
            unit_price=unit_price,
            discount_percent=discount_percent,
            vat_percent=vat_percent,
        )
        _, _, _, line_received = _purchase_line_amounts(
            quantity=qty_thuc_nhan(line, da_giao),
            unit_price=unit_price,
            discount_percent=discount_percent,
            vat_percent=vat_percent,
        )
        total += line_total
        received_total += line_received

    if da_giao is None:
        # Phiếu CŨ (chưa có đợt giao nào): giữ nguyên luật trước 06/08/2026 — chỉ đơn đã bấm "Đã
        # nhận hàng" mới sinh nợ. Không backfill đợt giao cho dữ liệu cũ nên nhánh này phải ở lại.
        gia_tri_da_giao = received_total if row.status == PR_RECEIVED else 0
    else:
        gia_tri_da_giao = sum(gia_tri_cac_dot(row).values())

    paid_amount = sum(
        int(voucher.amount_vnd)
        for voucher in row.payment_vouchers
        if voucher.status == PAYMENT_VOUCHER_PAID
    )
    # Tiền ĐÃ THU về (phiếu thu received) làm giảm số đã chi thực;
    # paid_amount giữ số thô để UI hiện tách bạch "đã chi X, đã thu Y".
    receipt_received_amount = sum(
        int(receipt.amount_vnd)
        for voucher in row.payment_vouchers
        for receipt in voucher.receipts
        if receipt.status == PAYMENT_RECEIPT_RECEIVED
    )
    net_paid = paid_amount - receipt_received_amount
    outstanding_amount = max(0, gia_tri_da_giao - net_paid)
    # Trần CỌC = cọc dự kiến − cọc ĐÃ CHI (chỉ đếm phiếu đặt cọc còn hiệu lực, không đếm phiếu
    # thanh toán). Cố ý KHÔNG trừ `net_paid`: tiền thanh toán cho hàng đã về không liên quan gì tới
    # hạn mức cọc — trừ vào đây là càng trả tiền hàng càng hết quyền ứng trước.
    coc_du_kien = int(getattr(row, "deposit_expected", 0) or 0)
    coc_da_chi = sum(
        int(v.amount_vnd)
        for v in row.payment_vouchers
        if v.status == PAYMENT_VOUCHER_PAID and v.payment_stage == PAYMENT_STAGE_ADVANCE
    )
    tran_dat_coc = max(0, coc_du_kien - coc_da_chi)
    if gia_tri_da_giao > 0 and net_paid >= gia_tri_da_giao:
        payment_status = "paid"
    elif net_paid > 0:
        payment_status = "partial"
    else:
        payment_status = "unpaid"
    return {
        "total": total,
        "received_total": received_total,
        "gia_tri_da_giao": gia_tri_da_giao,
        "paid_amount": paid_amount,
        "receipt_received_amount": receipt_received_amount,
        "net_paid": net_paid,
        "outstanding_amount": outstanding_amount,
        "tran_dat_coc": tran_dat_coc,
        "payment_status": payment_status,
    }


CAM_TRONG_TEN_SHEET = set(chr(c) for c in (58, 92, 47, 63, 42, 91, 93))  # : \ / ? * [ ]


class PurchaseService:
    def __init__(
        self,
        suppliers: SupplierRepository,
        department_requests: DepartmentPurchaseRequestRepository,
        requests: PurchaseRequestRepository,
        users: UserRepository,
        departments: DepartmentRepository,
        audit: AuditLogRepository,
        authz: AuthorizationService,
        lich_su: PurchaseStatusHistoryRepository,
        hang=None,
        giu_cho=None,
    ) -> None:
        self.lich_su = lich_su
        self.suppliers = suppliers
        self.department_requests = department_requests
        self.requests = requests
        self.users = users
        self.departments = departments
        self.audit = audit
        self.authz = authz
        # `VatLieuKhoService` — tra danh mục gốc + quy đổi đơn vị, để bảng giá NCC gắn được về
        # mặt hàng và so được giá. None → bỏ qua phần gắn (giữ tương thích với test cũ).
        self.hang = hang
        # `GiuChoService` — TUỲ CHỌN (30/08/2026), cùng nếp với `hang`. Vắng thì PMH chạy y như
        # trước (test cũ không phải kéo theo cả bảng cân đối); có mặt thì đợt giao đổi / huỷ đơn /
        # đóng đơn / mở lại đơn tự đối lại phần giữ HỨA đã bám đúng dòng phiếu — xem
        # `_doi_soat_giu_cho`.
        self.giu_cho = giu_cho

    # --- suppliers ---------------------------------------------------------

    def so_gia_ncc(self, hang_loai: str, hang_id: int) -> dict:
        """Các NCC bán MỘT mặt hàng, giá quy về ĐƠN VỊ GỐC để so ngang.

        Vì sao phải quy đổi mới so được: NCC A báo 1.020.000 đ/ram, NCC B báo 24.500 đ/kg — hai
        con số này không so trực tiếp được. Đưa cả hai về đ/<đơn vị gốc> thì mới biết ai rẻ.

        Dòng không quy đổi được KHÔNG bị loại (người dùng vẫn cần thấy NCC đó có bán) nhưng xếp
        CUỐI kèm lý do — xếp hạng một dòng chưa biết giá thật là mời chọn nhầm.
        """
        from .vat_lieu_kho_service import VatLieuKhoError

        info = self.hang.don_vi_cua_mat_hang(hang_loai, hang_id)
        # Mã đơn vị (cai/thung…) → TÊN có dấu (cái/thùng) cho hiển thị; NCC lưu unit dạng mã.
        dv_ten = {d.ma: d.ten for d in self.hang.don_vi.all_active()}
        rows: list[dict] = []
        for sup, it in self.suppliers.items_for_hang(hang_loai, hang_id):
            r = {
                "supplier_id": sup.id, "supplier_name": sup.name,
                "supplier_item_id": it.id, "unit": it.unit,
                "unit_ten": dv_ten.get((it.unit or "").strip().lower(), None),
                "unit_price": int(it.unit_price or 0),
                "vat_percent": float(it.vat_percent or 0),
                "gia_quy_doi": None, "gia_quy_doi_vat": None,
                "dien_giai": None, "ly_do": None,
            }
            try:
                # 1 đơn vị NCC bán bằng `he_so_ve_goc` đơn vị gốc → giá/đơn-vị-gốc = giá ÷ hệ số.
                qd = self.hang.quy_ve_goc(hang_loai, hang_id, it.unit, 1)
                hs = qd["he_so_ve_goc"]
                if hs > 0:
                    gia = int(round(int(it.unit_price or 0) / hs))
                    r["gia_quy_doi"] = gia
                    r["gia_quy_doi_vat"] = int(round(gia * (1 + float(it.vat_percent or 0) / 100)))
                    r["dien_giai"] = qd["dien_giai"]
            except VatLieuKhoError as e:
                r["ly_do"] = str(e)
            rows.append(r)
        # `float("inf")` cho dòng chưa quy đổi được → luôn nằm cuối, không lẫn vào xếp hạng.
        rows.sort(key=lambda x: (x["gia_quy_doi"] if x["gia_quy_doi"] is not None else float("inf")))
        return {
            "hang_loai": hang_loai, "hang_id": hang_id,
            "hang_ma": info.get("ma"), "hang_ten": info.get("ten"),
            "don_vi_goc": info.get("don_vi_goc"), "don_vi_goc_ten": info.get("don_vi_goc_ten"),
            "items": rows,
        }

    def list_suppliers(
        self,
        *,
        q: str | None = None,
        status: str | None = None,
        supplier_group: str | None = None,
        rating_min: float | None = None,
        sort: str = "name",
        page: int = 1,
        size: int = 20,
    ) -> tuple[list[tuple[Supplier, DanhGiaNcc]], int]:
        """Danh sách NCC, mỗi dòng kèm SỔ ĐIỂM đã tính sẵn (xem `services/danh_gia_ncc.py`).

        Sao đi kèm ngay trong lượt này chứ không phải một cú gọi riêng: cột sao nằm trên BẢNG danh
        sách, tách ra là 500 lượt đi DB cho một màn.
        """
        rows, total = self.suppliers.list(
            q=q,
            status=status,
            supplier_group=supplier_group,
            rating_min=rating_min,
            sort=sort,
            page=page,
            size=size,
        )
        return [(sup, tu_tong_hop(tho)) for sup, tho in rows], total

    def danh_gia_ncc(self, supplier_id: int) -> DanhGiaNcc:
        """Sổ điểm của MỘT nhà cung cấp — cho các cửa chỉ trả về một dòng (tạo · sửa · bật/tắt).

        Chưa có đơn nào đủ điều kiện ⇒ `DanhGiaNcc.rating is None` ("Chưa đánh giá"), KHÔNG phải 0.
        """
        return tu_tong_hop(self.suppliers.danh_gia_mot(supplier_id))

    def get_supplier(self, supplier_id: int) -> Supplier:
        supplier = self.suppliers.get_by_id(supplier_id)
        if supplier is None:
            raise PurchaseNotFound("Không tìm thấy nhà cung cấp.")
        return supplier

    def _clean_supplier_values(self, **values) -> dict:
        name = (values.get("name") or "").strip()
        if not name:
            raise PurchaseValidationError("Tên nhà cung cấp không được trống.")
        tax_code = (values.get("tax_code") or "").strip()
        phone = (values.get("phone") or "").strip()
        email = (values.get("email") or "").strip()
        address = (values.get("address") or "").strip()
        contact_name = (values.get("contact_name") or "").strip()
        supplier_group = (values.get("supplier_group") or "").strip()
        required = [
            (tax_code, "Mã số thuế"),
            (phone, "Số điện thoại"),
            (email, "Email"),
            (address, "Địa chỉ"),
            (contact_name, "Người liên hệ"),
            (supplier_group, "Nhóm nhà cung cấp"),
        ]
        missing = [label for value, label in required if not value]
        if missing:
            raise PurchaseValidationError(
                "Nhà cung cấp thiếu thông tin bắt buộc: " + ", ".join(missing) + "."
            )
        # ĐIỆN THOẠI phải đủ 10 CHỮ SỐ (chủ chốt 15/08/2026). Gõ thiếu/thừa một số thì gọi không
        # được, mà cái sai đó chỉ lộ ra đúng lúc cần gọi gấp cho nhà cung cấp.
        # Bỏ dấu cách / chấm / gạch / ngoặc trước khi đếm — người ta hay gõ "090 123 4567", chặn
        # cách viết đó là bắt gõ lại một con số vốn đã đúng. Lưu lại dạng đã gọn.
        phone = re.sub(r"[\s.\-()]", "", phone)
        if not phone.isdigit():
            raise PurchaseValidationError(
                "Số điện thoại chỉ được gồm chữ số, và phải đủ 10 số (ví dụ 0901234567)."
            )
        if len(phone) != 10:
            raise PurchaseValidationError(
                f"Số điện thoại phải đủ 10 chữ số (ví dụ 0901234567) — đang nhập {len(phone)} số."
            )
        # EMAIL phải có @, có phần trước và phần sau, không dấu cách. Cố tình KHÔNG bắt đúng chuẩn
        # RFC: mọi biểu thức "đúng chuẩn" đều dài, khó đọc, và vẫn chặn oan vài địa chỉ có thật.
        # Chỗ này chỉ cần chặn cái sai rõ ràng — thiếu @ thì thư không bao giờ tới.
        if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
            raise PurchaseValidationError(
                "Email phải có dạng ten@tencongty.vn — thiếu @ hoặc thiếu phần đuôi thì "
                "thư gửi đi không tới nơi."
            )
        status = values.get("status") or SUPPLIER_ACTIVE
        if status not in SUPPLIER_STATUSES:
            raise PurchaseValidationError("Trạng thái nhà cung cấp không hợp lệ.")
        items = self._clean_supplier_items(values.get("items") or [])
        # HẠN MỨC (tiền) và ĐỊNH MỨC (số ngày cho nợ) — chủ chốt 06/08/2026.
        # `credit_days` giữ được NULL: NULL = CHƯA ĐẶT hạn ⇒ đợt giao của NCC này không vào cột Quá
        # hạn. Ép về 0 là "trả ngay" ⇒ qua một đêm cả bảng công nợ đỏ rực vì một quyết định không ai
        # ra. Hai thứ khác nhau, đừng gộp.
        han_muc = int(values.get("credit_limit") or 0)
        if han_muc < 0:
            raise PurchaseValidationError("Hạn mức công nợ không được âm.")
        so_ngay_raw = values.get("credit_days")
        so_ngay = None if so_ngay_raw is None or so_ngay_raw == "" else int(so_ngay_raw)
        if so_ngay is not None and so_ngay < 0:
            raise PurchaseValidationError("Số ngày cho nợ không được âm.")
        return {
            "name": name,
            "tax_code": tax_code,
            "phone": phone,
            "email": email,
            "address": address,
            "contact_name": contact_name,
            "supplier_group": supplier_group,
            "payment_terms": (values.get("payment_terms") or "").strip() or None,
            "credit_limit": han_muc,
            "credit_days": so_ngay,
            "status": status,
            "note": (values.get("note") or "").strip() or None,
            "items": items,
        }

    def _clean_supplier_items(self, raw_items) -> list[SupplierItemInput]:
        items: list[SupplierItemInput] = []
        for raw in raw_items or []:
            get = raw.get if isinstance(raw, dict) else lambda key, default=None: getattr(raw, key, default)
            item_name = (get("item_name") or "").strip()
            unit = (get("unit") or "").strip()
            raw_price = get("unit_price", 0) or 0
            raw_vat = get("vat_percent", 0) or 0
            note = (get("note") or "").strip() or None
            if not item_name and not unit and not raw_price and not raw_vat and not note:
                continue
            if not item_name:
                raise PurchaseValidationError("Ten mat hang nha cung cap khong duoc trong.")
            if not unit:
                raise PurchaseValidationError("Don vi tinh mat hang nha cung cap khong duoc trong.")
            unit_price = int(raw_price)
            if unit_price <= 0:
                raise PurchaseValidationError("Don gia mat hang nha cung cap phai lon hon 0.")
            vat_percent = float(raw_vat)
            if vat_percent < 0 or vat_percent > 100:
                raise PurchaseValidationError("VAT mat hang nha cung cap phai tu 0 den 100.")
            hang_loai = (get("hang_loai") or "").strip() or None
            hang_id = get("hang_id") or None
            if (hang_loai is None) != (hang_id is None):
                raise PurchaseValidationError(
                    "Mat hang goc phai co ca loai lan ma — chon lai o o Vat tu."
                )
            if hang_loai is not None:
                # Đơn vị NCC bán phải nằm trong tập đổi được của mặt hàng, nếu không thì cột "giá
                # quy về đơn vị gốc" vĩnh viễn trống và dòng này không bao giờ so giá được.
                self._kiem_don_vi_ncc(hang_loai, int(hang_id), unit)
            items.append(
                SupplierItemInput(
                    hang_loai=hang_loai,
                    hang_id=int(hang_id) if hang_id else None,
                    item_name=item_name,
                    unit=unit,
                    unit_price=unit_price,
                    vat_percent=vat_percent,
                    note=note,
                )
            )
        return items

    def _kiem_don_vi_ncc(self, hang_loai: str, hang_id: int, unit: str) -> None:
        from .vat_lieu_kho_service import VatLieuKhoError

        if self.hang is None:
            return
        try:
            self.hang.quy_ve_goc(hang_loai, hang_id, unit, 1)
        except VatLieuKhoError as e:
            raise PurchaseValidationError(str(e)) from None

    @staticmethod
    def _chan_coc_vuot_tong(coc: int, tong: int, *, khi: str) -> None:
        """CỌC DỰ KIẾN KHÔNG ĐƯỢC LỚN HƠN TỔNG DỰ KIẾN CỦA ĐƠN (chủ chốt 15/08/2026).

        Cọc là ứng TRƯỚC một phần của chính đơn này. Khai cọc 10tr cho đơn 2tr là con số vô nghĩa,
        mà nó không nằm yên: `tran_dat_coc` = cọc dự kiến − cọc đã chi, nên số khai thừa thành
        HẠN MỨC CHI THẬT — kế toán lập được phiếu cọc 10tr cho đơn 2tr, tiền ra khỏi két rồi mới
        có người hỏi.

        Chặn CẢ HAI ĐẦU: lúc khai cọc, và lúc sửa dòng hàng làm tổng tụt xuống dưới cọc đã khai.
        Bịt mỗi đầu trên thì chỉ cần khai cọc lúc đơn còn to rồi sửa đơn nhỏ lại là lách được."""
        if coc > tong:
            raise PurchaseValidationError(
                f"Cọc dự kiến ({coc:,}đ) đang lớn hơn tổng dự kiến của đơn ({tong:,}đ) {khi}. "
                "Cọc là ứng trước một phần của chính đơn này nên không thể vượt giá trị đơn."
                if tong > 0 else
                f"Chưa khai được cọc ({coc:,}đ): đơn chưa có dòng hàng nào nên tổng dự kiến đang "
                "là 0đ. Khai hàng trước rồi mới đặt cọc."
            )

    def _chan_trung_mst(self, tax_code, *, bo_qua_id: int | None = None) -> None:
        """MÃ SỐ THUẾ KHÔNG ĐƯỢC TRÙNG (chủ chốt 12/08/2026).

        MST là định danh pháp lý của doanh nghiệp. Hai hồ sơ cùng MST = một nhà cung cấp bị nhập
        hai lần: công nợ chẻ đôi, đối chiếu hoá đơn ra hai kết quả, và không ai biết phiếu chi nên
        gắn vào hồ sơ nào.

        BỎ TRỐNG THÌ KHÔNG CHẶN — hộ kinh doanh nhỏ nhiều nơi không có MST, và trước nay vẫn khai
        được. Chặn cả ô rỗng là khoá mất nhóm NCC đó."""
        ma = (tax_code or "").strip()
        if not ma:
            return
        trung = self.suppliers.find_by_tax_code(ma)
        if trung is not None and trung.id != bo_qua_id:
            raise PurchaseConflict(
                f"Mã số thuế {ma} đã dùng cho nhà cung cấp \"{trung.name}\". "
                "Mỗi mã số thuế chỉ thuộc một nhà cung cấp — kiểm lại xem có phải trùng hồ sơ không."
            )

    def create_supplier(self, *, actor, **values) -> Supplier:
        cleaned = self._clean_supplier_values(**values)
        existing = self.suppliers.find_by_name(cleaned["name"])
        if existing is not None:
            raise PurchaseConflict("Nhà cung cấp đã tồn tại.")
        self._chan_trung_mst(cleaned.get("tax_code"))
        supplier = self.suppliers.create(**cleaned)
        self.audit.create(
            actor_user_id=actor.id,
            action="create_supplier",
            target=f"supplier:{supplier.id}",
            detail=supplier.name,
        )
        return supplier

    def update_supplier(self, supplier_id: int, *, actor, **values) -> Supplier:
        supplier = self.get_supplier(supplier_id)
        cleaned = self._clean_supplier_values(**values)
        existing = self.suppliers.find_by_name(cleaned["name"])
        if existing is not None and existing.id != supplier.id:
            raise PurchaseConflict("Nhà cung cấp đã tồn tại.")
        self._chan_trung_mst(cleaned.get("tax_code"), bo_qua_id=supplier.id)
        supplier = self.suppliers.update(supplier, **cleaned)
        self.audit.create(
            actor_user_id=actor.id,
            action="update_supplier",
            target=f"supplier:{supplier.id}",
            detail=supplier.name,
        )
        return supplier

    def toggle_supplier_active(self, supplier_id: int, *, actor) -> Supplier:
        supplier = self.get_supplier(supplier_id)
        next_status = SUPPLIER_INACTIVE if supplier.status == SUPPLIER_ACTIVE else SUPPLIER_ACTIVE
        supplier = self.suppliers.update(supplier, status=next_status)
        self.audit.create(
            actor_user_id=actor.id,
            action="toggle_supplier",
            target=f"supplier:{supplier.id}",
            detail=f"{supplier.name} -> {supplier.status}",
        )
        return supplier

    # --- vật tư NCC: mẫu · xuất · đọc file Excel -------------------------

    #: Nhãn cột trong file. Đổi ở đây là đổi CẢ mẫu tải về lẫn bộ dò tiêu đề khi nhập.
    COT_VAT_TU = ("Tên hàng*", "Đơn vị*", "Đơn giá*", "VAT %", "Ghi chú")
    #: Trần dòng/file. Không phải giới hạn kỹ thuật — file to hơn mức này gần như luôn là dán nhầm
    #: bảng giá của nhiều NCC vào một sheet, cho chạy tiếp là nhập rác.
    TRAN_DONG_IMPORT = 500

    @staticmethod
    def _khoa_vat_tu(item_name: str, unit: str) -> tuple[str, str]:
        """Khoá TRÙNG = tên + đơn vị, bỏ hoa/thường và khoảng trắng thừa.

        Cùng tên cùng ĐVT mà hai giá thì form phiếu mua không biết chọn dòng nào — nên hai dòng như
        vậy phải gộp làm một, không đẻ dòng thứ hai."""
        return (" ".join(item_name.split()).lower(), " ".join(unit.split()).lower())

    @staticmethod
    def _chuan_hoa_tieu_de(text: str) -> str:
        bo_dau = unicodedata.normalize("NFD", (text or "").strip())
        bo_dau = "".join(ch for ch in bo_dau if unicodedata.category(ch) != "Mn")
        return "".join(ch for ch in bo_dau.lower() if ch.isalnum())

    def _wb_vat_tu(self, rows: list[dict], *, ten_sheet: str):
        from openpyxl import Workbook  # lazy import: thiếu dep chỉ hỏng endpoint này, không sập app
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        # Tên sheet Excel: tối đa 31 ký tự, và cấm : \\ / ? * [ ]
        ws.title = "".join(ch for ch in ten_sheet if ch not in CAM_TRONG_TEN_SHEET)[:31] or "Vat tu"
        ws.append(list(self.COT_VAT_TU))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([
                row.get("item_name") or "",
                row.get("unit") or "",
                int(row.get("unit_price") or 0),
                float(row.get("vat_percent") or 0),
                row.get("note") or "",
            ])
        for idx, width in enumerate((38, 12, 16, 10, 34), start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        ws.freeze_panes = "A2"
        return wb

    @staticmethod
    def _xuat_bytes(wb) -> bytes:
        from io import BytesIO

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def mau_vat_tu_xlsx(self) -> bytes:
        """File mẫu: tiêu đề + 2 dòng ví dụ + một dòng nhắc luật nhập.

        Ví dụ nằm NGAY trong file chứ không chỉ ở tài liệu: mở file ra là thấy đơn giá gõ số trơn
        (2200), không dấu phân cách — chỗ sai nhiều nhất khi nhập từ Excel."""
        wb = self._wb_vat_tu(
            [
                {"item_name": "Giấy Duplex 350gsm", "unit": "tờ", "unit_price": 2200,
                 "vat_percent": 8, "note": "Khổ 79x109"},
                {"item_name": "Keo cán màng", "unit": "kg", "unit_price": 80000,
                 "vat_percent": 10, "note": ""},
            ],
            ten_sheet="Mau vat tu",
        )
        ws = wb.active
        ws.append([])
        ws.append([
            "Xoá 2 dòng ví dụ trước khi nhập. Đơn giá gõ số trơn (2200), không dấu phân cách. "
            "Trùng Tên hàng + Đơn vị với dòng đã có ⇒ CẬP NHẬT dòng đó. "
            f"Tối đa {self.TRAN_DONG_IMPORT} dòng/file, mỗi file cho MỘT nhà cung cấp."
        ])
        return self._xuat_bytes(wb)

    def xuat_vat_tu_xlsx(self, supplier_id: int) -> tuple[bytes, str]:
        supplier = self.get_supplier(supplier_id)
        rows = [
            {
                "item_name": item.item_name,
                "unit": item.unit,
                "unit_price": int(item.unit_price or 0),
                "vat_percent": float(item.vat_percent or 0),
                "note": item.note or "",
            }
            for item in supplier.items
        ]
        wb = self._wb_vat_tu(rows, ten_sheet=supplier.name)
        # Tên file bỏ dấu: vẫn còn trình duyệt/OS làm hỏng tên có dấu khi tải về.
        ten = unicodedata.normalize("NFD", supplier.name)
        ten = "".join(ch for ch in ten if unicodedata.category(ch) != "Mn")
        ten = "".join(ch if ch.isalnum() else "-" for ch in ten).strip("-")[:60] or "ncc"
        return self._xuat_bytes(wb), f"vat-tu-{ten}.xlsx"

    def doc_vat_tu_xlsx(self, data: bytes) -> dict:
        """Đọc file Excel thành danh sách mặt hàng — CỐ Ý KHÔNG ghi DB.

        Bảng giá nằm trong form sửa NCC, người dùng bấm *Lưu nhà cung cấp* mới là lúc dữ liệu vào
        sổ. Ghi thẳng ở đây thì chính cú lưu form đó (đang giữ danh sách CŨ) sẽ xoá mất phần vừa
        nhập — mất dữ liệu mà không ai hiểu vì sao.

        Dòng hỏng KHÔNG huỷ cả file: trả về dòng lành + danh sách lỗi kèm SỐ DÒNG EXCEL. File 200
        dòng sai dòng 197 mà bắt nhập lại từ đầu là hành người dùng."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            raise PurchaseValidationError(
                "Không đọc được file. Cần file Excel .xlsx — tải file mẫu để lấy đúng định dạng."
            ) from None
        try:
            ws = wb.active
            raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

        raw_rows = [r for r in raw_rows if any(str(c).strip() for c in r if c is not None)]
        if not raw_rows:
            raise PurchaseValidationError("File rỗng.")

        # Dò tiêu đề theo nhãn ĐÃ BỎ DẤU — người dùng hay xoá dấu * hoặc gõ không dấu.
        mong_doi = [self._chuan_hoa_tieu_de(c) for c in self.COT_VAT_TU]
        header = [self._chuan_hoa_tieu_de(str(c) if c is not None else "") for c in raw_rows[0]]
        if mong_doi[0] not in header or mong_doi[1] not in header:
            raise PurchaseValidationError(
                'Thiếu cột "Tên hàng" hoặc "Đơn vị" — tải file mẫu để lấy đúng tiêu đề.'
            )
        vi_tri = {ten: header.index(ten) for ten in mong_doi if ten in header}
        body = raw_rows[1:]
        if not body:
            raise PurchaseValidationError("File chỉ có tiêu đề, không có dòng dữ liệu.")
        if len(body) > self.TRAN_DONG_IMPORT:
            raise PurchaseValidationError(
                f"File có {len(body)} dòng, vượt trần {self.TRAN_DONG_IMPORT} dòng. "
                "Tách nhỏ file rồi nhập lại."
            )

        def o(row: list, ten: str) -> str:
            idx = vi_tri.get(ten)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        items: list[dict] = []
        errors: list[dict] = []
        da_gap: dict[tuple[str, str], int] = {}
        for offset, row in enumerate(body):
            so_dong = offset + 2  # +1 dòng tiêu đề, +1 vì Excel đếm từ 1
            ten_hang = o(row, mong_doi[0])
            don_vi = o(row, mong_doi[1])
            gia_raw = o(row, mong_doi[2])
            vat_raw = o(row, mong_doi[3])
            ghi_chu = o(row, mong_doi[4])
            if not ten_hang and not don_vi and not gia_raw:
                continue
            if not ten_hang:
                errors.append({"row": so_dong, "message": "Thiếu tên hàng."})
                continue
            if not don_vi:
                errors.append({"row": so_dong, "message": f'"{ten_hang}": thiếu đơn vị.'})
                continue
            try:
                # Chấp cả "2.200" · "2,200" · "2200.0" — Excel trả về đủ kiểu tuỳ ô định dạng gì.
                don_gia = int(round(float(gia_raw.replace(".", "").replace(",", "").replace(" ", ""))))
            except ValueError:
                errors.append(
                    {"row": so_dong, "message": f'"{ten_hang}": đơn giá "{gia_raw}" không phải số.'}
                )
                continue
            if don_gia <= 0:
                errors.append({"row": so_dong, "message": f'"{ten_hang}": đơn giá phải lớn hơn 0.'})
                continue
            try:
                vat = float((vat_raw or "0").replace(",", ".").replace("%", "").strip() or 0)
            except ValueError:
                errors.append(
                    {"row": so_dong, "message": f'"{ten_hang}": VAT "{vat_raw}" không phải số.'}
                )
                continue
            if vat < 0 or vat > 100:
                errors.append({"row": so_dong, "message": f'"{ten_hang}": VAT phải từ 0 đến 100.'})
                continue

            khoa = self._khoa_vat_tu(ten_hang, don_vi)
            item = {
                "item_name": ten_hang[:255],
                "unit": don_vi[:32],
                "unit_price": don_gia,
                "vat_percent": vat,
                "note": (ghi_chu or "")[:2000] or None,
            }
            # Trùng NGAY TRONG file: dòng dưới đè dòng trên, và nói rõ đè dòng nào — im lặng thì
            # người dùng tưởng bị mất dòng.
            cu = da_gap.get(khoa)
            if cu is not None:
                errors.append({
                    "row": so_dong,
                    "message": f'"{ten_hang}" ({don_vi}) trùng dòng {cu} trong file — '
                               "lấy dòng dưới cùng.",
                })
                for i, da_co in enumerate(items):
                    if self._khoa_vat_tu(da_co["item_name"], da_co["unit"]) == khoa:
                        items[i] = item
                        break
            else:
                da_gap[khoa] = so_dong
                items.append(item)

        return {"items": items, "errors": errors, "total_rows": len(body)}

    def list_supplier_item_catalog(self) -> list[dict]:
        return self.suppliers.list_item_catalog()

    # --- department purchase requests -------------------------------------

    def list_department_requests(
        self,
        *,
        actor,
        q: str | None = None,
        status: str | None = None,
        source_type: str | None = None,
        sort: str = "-created_at",
        page: int = 1,
        size: int = 20,
    ) -> tuple[list[dict], int]:
        pham_vi = self.authz.scope_for(actor, "yeu_cau_mua_hang")
        rows, total = self.department_requests.list(
            q=q,
            status=status,
            source_type=source_type,
            requesting_department_id=actor.department_id,
            filter_by_department=not self._sees_all_department_requests(actor),
            # `own` = ĐÚNG yêu cầu do chính mình gửi. Trước 11/08/2026 `own` rơi xuống nhánh lọc
            # theo phòng nên thấy luôn yêu cầu của đồng nghiệp — đo được 1 dòng do người khác tạo.
            requested_by_user_id=actor.id if pham_vi == SCOPE_OWN else None,
            sort=sort,
            page=page,
            size=size,
        )
        return [self._to_department_request_out(row) for row in rows], total

    def _sees_all_department_requests(self, actor) -> bool:
        """Có nhìn được YCMH của TOÀN công ty không.

        HAI vai khác hẳn nhau, đừng gộp:

        · **Người XỬ LÝ** — ai có module `thu_mua` — thấy YCMH của MỌI phòng ban. Đó là HỘP VIỆC
          của họ: công việc của thu mua chính là biến đơn của phòng khác thành phiếu mua. Không
          phụ thuộc scope: nhân viên thu mua scope `own` (chỉ thấy PHIẾU MUA của mình) vẫn phải
          thấy đủ yêu cầu gửi đến, nếu không thì ngồi nhìn màn hình trống.
        · **Người ĐỀ NGHỊ** (bao_gia · kho · san_xuat · dm_giay) — chỉ thấy yêu cầu của
          phòng mình, trừ khi được cấp scope `all`.

        ⚠️ Ngày 04/08/2026 tôi hạ scope `thu_mua` của nhân viên mua hàng xuống `own` để họ chỉ
        thấy PHIẾU MUA của mình — và làm mù luôn hộp việc này, vì lúc đó cả hai danh sách cùng đọc
        một ô scope. Hai thứ khác nhau thì phải hỏi hai câu khác nhau.

        ⚠️ 11/08/2026 — GỠ LỐI TẮT `thu_mua`. Trước đó dòng đầu hàm là:

            if self.authz.scope_for(actor, "thu_mua") is not None:
                return True

        `scope_for` trả về chuỗi phạm vi khi vai CÓ dòng quyền `thu_mua`, **bất kể phạm vi là gì**
        — nên chỉ cần được cấp màn Mua hàng (dù `own`) là ô chọn phạm vi ở màn Yêu cầu mua hàng bị
        bỏ qua sạch, thấy yêu cầu của MỌI phòng. Chủ chốt báo đúng chỗ này; đo lại: vai
        `yeu_cau_mua_hang` phạm vi `own` thấy 1 dòng, cấp thêm `thu_mua` là thành 2 dòng đủ cả hai
        phòng.

        Lối tắt đó SINH RA khi YCMH chưa có khoá riêng. Từ 10/08 nó có khoá `yeu_cau_mua_hang`,
        nên cách đúng là cấp thẳng phạm vi `all` trên khoá đó cho bộ phận mua hàng — hiện rõ trên
        ma trận, gỡ được. Migration 0183 làm việc cấp bù đó.

        LUẬT NAY, đúng hai nhánh:

        1. Vai CÓ dòng `yeu_cau_mua_hang` ⇒ **chỉ nghe phạm vi của chính khoá đó**. Ô nào của màn
           nào thì màn đó nghe — đúng ý "cứ cấp quyền là được phép, không cấp là không được".
           Không còn chuyện đặt phạm vi ở đây mà một khoá khác đè lên.
        2. Vai KHÔNG có dòng đó (DB cũ chưa qua migration) ⇒ giữ luật cũ để không ai mù hộp việc.
        """
        pham_vi = self.authz.scope_for(actor, "yeu_cau_mua_hang")
        if pham_vi is not None:
            return pham_vi == SCOPE_ALL
        return any(
            self.authz.scope_for(actor, module) == SCOPE_ALL
            for module in DEPARTMENT_REQUEST_READER_MODULES
        )

    def _can_view_department_request(self, row: DepartmentPurchaseRequest, actor) -> bool:
        if self._sees_all_department_requests(actor):
            return True
        return row.requesting_department_id == actor.department_id

    def _department_request(self, request_id: int) -> DepartmentPurchaseRequest:
        row = self.department_requests.get_by_id(request_id)
        if row is None:
            raise PurchaseNotFound("Khong tim thay phieu yeu cau mua tu phong ban.")
        return row

    def get_department_request(self, request_id: int, *, actor) -> dict:
        row = self._department_request(request_id)
        if not self._can_view_department_request(row, actor):
            raise PurchaseNotFound("Khong tim thay phieu yeu cau mua tu phong ban.")
        return self._to_department_request_out(row)

    @staticmethod
    def _gop_noi_dung(content: str | None, purpose: str | None, note: str | None) -> str:
        """MỘT ô nội dung, nhận cả đường gọi cũ (chủ chốt 07/08/2026).

        Client mới gửi `content`. Client cũ còn gửi `purpose` (+ `note`) — nối lại thay vì báo lỗi:
        bắt mọi nơi gọi API đổi cùng lúc với giao diện là chuyện không xảy ra được."""
        moi = (content or "").strip()
        if moi:
            return moi
        cu = " — ".join(x for x in [(purpose or "").strip(), (note or "").strip()] if x)
        return cu

    def _clean_department_request_header(
        self,
        *,
        source_type: str | None,
        purpose: str | None = None,
        content: str | None = None,
        needed_date: date | None = None,
    ) -> tuple[str, str, date]:
        cleaned_source_type = (source_type or "").strip()
        if cleaned_source_type not in DEPARTMENT_PURCHASE_SOURCE_TYPES:
            raise PurchaseValidationError("Nguon yeu cau mua khong hop le.")
        cleaned_purpose = (purpose or "").strip()
        if not cleaned_purpose:
            raise PurchaseValidationError("Nội dung / mục đích yêu cầu mua không được trống.")
        if needed_date is None:
            raise PurchaseValidationError("Ngay can hang la thong tin bat buoc.")
        if needed_date < _business_today():
            raise PurchaseValidationError("Ngay can hang khong duoc nho hon hom nay.")
        return cleaned_source_type, cleaned_purpose, needed_date

    def _source_type_for_actor(self, actor) -> str:
        if actor.department_id is None:
            return SOURCE_KHAC
        department = self.departments.get_by_id(actor.department_id)
        if department is None:
            return SOURCE_KHAC
        normalized = unicodedata.normalize("NFD", department.name or "")
        name = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn").lower()
        if "kho" in name:
            return SOURCE_KHO
        if "san xuat" in name:
            return SOURCE_SAN_XUAT
        if "cong nghe" in name:
            return SOURCE_CONG_NGHE
        if "gia cong" in name:
            return SOURCE_GIA_CONG_NGOAI
        if "kinh doanh" in name or "sale" in name:
            return SOURCE_KINH_DOANH
        return SOURCE_KHAC

    def _clean_department_lines(self, raw_lines) -> list[DepartmentPurchaseRequestLineInput]:
        if not raw_lines:
            raise PurchaseValidationError("Yeu cau mua phai co it nhat mot dong vat tu.")
        lines: list[DepartmentPurchaseRequestLineInput] = []
        for line in raw_lines:
            get = line.get if isinstance(line, dict) else lambda key, default=None: getattr(line, key, default)
            item_name = (get("item_name") or "").strip()
            if not item_name:
                raise PurchaseValidationError("Tên vật tư không được trống.")
            hang_loai, hang_id = _doc_mat_hang(get)
            # Dòng ĐÃ GẮN MẶT HÀNG GỐC (mg 0174) thì thôi kiểm theo TÊN: cặp `(hang_loai, hang_id)`
            # là bằng chứng mạnh hơn hẳn — món đó đang nằm trong danh mục gốc, không phải chữ gõ
            # tay. Giữ lại phép so tên ở đây là chặn oan đúng luồng vừa dựng: bảng cân đối gửi
            # "Couché 150 79×109" (tên danh mục) trong khi NCC khai "Couche 150" ⇒ không lập nổi
            # yêu cầu mua, mà lý do báo ra lại là "vật tư chưa có trong danh mục" — sai và khó hiểu.
            # Dòng KHÔNG gắn mặt hàng gốc: chỉ còn chấp nhận nếu tên trùng bảng giá NCC — đó là
            # đường lùi cho DỮ LIỆU CŨ (phiếu lập trước khi ô chọn danh mục ra đời) và cho client
            # cũ. Giao diện hiện tại luôn gửi kèm cặp `(hang_loai, hang_id)`: ô Vật tư ở màn Yêu
            # cầu mua hàng là combobox tra thẳng danh mục Giấy + Vật tư khác, không gõ tự do.
            if hang_loai is not None:
                if not self.suppliers.has_active_item_for_hang(hang_loai, hang_id):
                    raise PurchaseValidationError(
                        "Vật tư chưa được nhà cung cấp nào khai bán. Vui lòng khai mặt hàng "
                        "trong Nhà cung cấp trước khi tạo yêu cầu mua."
                    )
            elif not self.suppliers.has_active_item(item_name):
                raise PurchaseValidationError(
                    "Vật tư phải được nhà cung cấp đang hoạt động khai bán trước khi tạo yêu cầu mua."
                )
            unit = (get("unit") or "").strip()
            if not unit:
                raise PurchaseValidationError("Đơn vị tính không được trống.")
            quantity = float(get("quantity"))
            if quantity <= 0:
                raise PurchaseValidationError("So luong phai lon hon 0.")
            lines.append(
                DepartmentPurchaseRequestLineInput(
                    item_name=item_name,
                    unit=unit,
                    quantity=quantity,
                    expected_unit_price=0,
                    note=(get("note") or "").strip() or None,
                    hang_loai=hang_loai,
                    hang_id=hang_id,
                )
            )
        return lines

    def _new_department_request_code(self) -> str:
        today = datetime.now().strftime("%y%m%d")
        alphabet = string.ascii_uppercase + string.digits
        for _ in range(20):
            rand = "".join(secrets.choice(alphabet) for _ in range(4))
            code = f"YCMH-{today}-{rand}"
            if self.department_requests.get_by_code(code) is None:
                return code
        raise PurchaseConflict("Khong sinh duoc ma yeu cau mua duy nhat, vui long thu lai.")

    def create_department_request(
        self,
        *,
        source_type: str | None,
        related_document_type: str | None,
        related_document_code: str | None,
        purpose: str | None = None,
        content: str | None = None,
        needed_date: date | None = None,
        note: str | None = None,
        lines=None,
        actor=None,
    ) -> dict:
        if not self.can_create_department_request(actor):
            raise PurchaseForbidden("Ban khong co quyen tao yeu cau mua hang cho bo phan.")
        source_type = self._source_type_for_actor(actor)
        # MỘT ô nội dung (07/08/2026). `purpose` + `note` của client cũ được nối lại ở đây.
        noi_dung = self._gop_noi_dung(content, purpose, note)
        source_type, noi_dung, needed_date = self._clean_department_request_header(
            source_type=source_type, purpose=noi_dung, needed_date=needed_date
        )
        row = self.department_requests.create(
            code=self._new_department_request_code(),
            source_type=source_type,
            requesting_department_id=actor.department_id,
            requested_by_user_id=actor.id,
            related_document_type=(related_document_type or "").strip() or None,
            related_document_code=(related_document_code or "").strip() or None,
            purpose=noi_dung[:500],
            content=noi_dung,
            needed_date=needed_date,
            lines=self._clean_department_lines(lines),
        )
        self.audit.create(
            actor_user_id=actor.id,
            action="create_department_purchase_request",
            target=f"department_purchase_request:{row.id}",
            detail=row.code,
        )
        return self._to_department_request_out(row)

    def can_create_department_request(self, actor) -> bool:
        """Được LẬP yêu cầu mua hàng hay không — HỎI ĐÚNG MỘT THỨ: ô quyền của màn đó.

        ⚠️ ĐÃ BỎ (10/08/2026) hai đường tắt cũ, theo yêu cầu chủ chốt "cứ cấp quyền là được phép,
        không cấp là không được":

        1. `thu_mua/request` — mượn quyền của MÀN KHÁC. Bật một ô ở màn Mua hàng lại mở cửa màn Yêu
           cầu mua hàng: nhìn ma trận quyền không đoán được, đúng chỗ tester kêu "quyền không ăn khớp".
        2. `department.head_user_id == actor.id` — QUYỀN NGẦM THEO CHỨC DANH. Ai đang là trưởng
           phòng thì tự động lập được yêu cầu chi tiền, không cần ai cấp và KHÔNG BỎ ĐI ĐƯỢC: trên
           ma trận quyền không có ô nào để tắt. Đổi trưởng phòng là quyền tự chuyển người theo.

        Migration 0177 cấp bù `yeu_cau_mua_hang` cho vai của các trưởng phòng đang tại vị, nên không
        ai mất đường làm việc — khác ở chỗ từ nay quyền đó HIỆN RÕ trên ma trận và gỡ được.
        """
        return self.authz.can(actor, "yeu_cau_mua_hang", "create")

    def update_department_request(
        self,
        request_id: int,
        *,
        source_type: str | None,
        related_document_type: str | None,
        related_document_code: str | None = None,
        purpose: str | None = None,
        content: str | None = None,
        needed_date: date | None = None,
        note: str | None = None,
        lines=None,
        actor=None,
    ) -> dict:
        row = self._department_request(request_id)
        if row.status != DPR_OPEN:
            raise PurchaseConflict("Chi yeu cau chua tao phieu mua moi duoc sua.")
        if row.requested_by_user_id != actor.id:
            raise PurchaseForbidden("Chi nguoi tao yeu cau moi duoc sua.")
        noi_dung = self._gop_noi_dung(content, purpose, note)
        source_type, noi_dung, needed_date = self._clean_department_request_header(
            source_type=row.source_type, purpose=noi_dung, needed_date=needed_date
        )
        saved = self.department_requests.update(
            row,
            purpose=noi_dung[:500],
            content=noi_dung,
            needed_date=needed_date,
            lines=self._clean_department_lines(lines),
        )
        self.audit.create(
            actor_user_id=actor.id,
            action="update_department_purchase_request",
            target=f"department_purchase_request:{row.id}",
            detail=row.code,
        )
        return self._to_department_request_out(saved)

    def cancel_department_request(self, request_id: int, *, reason: str | None, actor) -> dict:
        row = self._department_request(request_id)
        if row.status != DPR_OPEN:
            raise PurchaseConflict("Chi yeu cau dang cho mua moi duoc huy.")
        can_cancel_any = self.authz.can(actor, "yeu_cau_mua_hang", "cancel")
        if row.requested_by_user_id != actor.id and not can_cancel_any:
            raise PurchaseForbidden("Chi nguoi tao yeu cau hoac admin moi duoc huy.")
        # Lý do vào cột RIÊNG. Trước đây `row.note = reason` GHI ĐÈ mất ghi chú người lập.
        row.reject_reason = (reason or "").strip() or None
        self._dat_trang_thai(row, DPR_CANCELLED, doc_type=DOC_YCMH, actor=actor, ly_do=reason)
        saved = self.department_requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="cancel_department_purchase_request",
            target=f"department_purchase_request:{row.id}",
            detail=row.code,
        )
        return self._to_department_request_out(saved)

    def _phieu_dang_giu_dong(self, row) -> dict[int, PurchaseRequest]:
        """Dòng nào của yêu cầu đang bị một ĐƠN MUA CÒN SỐNG nắm giữ.

        Còn sống = mọi trạng thái TRỪ `cancelled`. Đơn BỊ TỪ CHỐI vẫn tính là đang giữ: luồng hiện
        tại bắt thu mua sửa lại chính đơn đó rồi gửi duyệt lại (xem `_tinh_lai_trang_thai_ycmh`),
        nên món hàng vẫn đang trên đường đi — bỏ nó khỏi yêu cầu lúc này là đơn kia mua về một thứ
        không ai còn cần.

        Một dòng có thể qua nhiều đơn (đơn cũ huỷ, đơn mới lập) ⇒ giữ đơn có tiến độ CAO NHẤT, để
        câu chặn gọi đúng tên cái đơn đang thực sự chạy."""
        giu: dict[int, PurchaseRequest] = {}
        for link in getattr(row, "purchase_links", []):
            phieu = link.purchase_request
            if phieu is None or phieu.status == PR_CANCELLED:
                continue
            for pl in phieu.lines:
                sid = getattr(pl, "department_request_line_id", None)
                if sid is None:
                    continue
                cu = giu.get(sid)
                if cu is None or _BAC_PHIEU.get(phieu.status, 0) > _BAC_PHIEU.get(cu.status, 0):
                    giu[sid] = phieu
        return giu

    @staticmethod
    def _cau_chan_huy_dong(phieu) -> str:
        nhan = _NHAN_PMH_NGAN.get(phieu.status, phieu.status)
        return (
            f"Món này đang nằm ở đơn mua {phieu.code} ({nhan}). "
            "Xử lý ở đơn mua đó trước rồi mới bỏ được món khỏi yêu cầu."
        )

    def cancel_department_request_line(
        self, request_id: int, line_id: int, *, reason: str | None, actor
    ) -> dict:
        """Huỷ MỘT MÓN trong yêu cầu, giữ nguyên các món còn lại (mg 0233).

        Chủ chốt 24/08/2026: *"phải quản tới từng món hàng, đừng quản tới cấp chứng từ nữa"*.
        Trước đây chỉ có `cancel_department_request` — huỷ là huỷ cả phiếu, và chỉ huỷ được khi
        phiếu còn `open`, nên yêu cầu 5 dòng mà thu mua đã lập đơn cho 3 dòng thì 2 dòng thừa mắc
        kẹt vĩnh viễn.

        Ở đây KHÔNG chặn theo trạng thái phiếu cha — chặn theo TỪNG MÓN: món nào chưa bị đơn mua
        nào nắm thì bỏ được, kể cả khi phiếu cha đang "Đang mua" vì món khác. Huỷ hết món thì phiếu
        cha mới thành `cancelled`."""
        row = self._department_request(request_id)
        if row.status == DPR_CANCELLED:
            raise PurchaseConflict("Yêu cầu này đã huỷ.")
        can_cancel_any = self.authz.can(actor, "yeu_cau_mua_hang", "cancel")
        if row.requested_by_user_id != actor.id and not can_cancel_any:
            raise PurchaseForbidden("Chi nguoi tao yeu cau hoac admin moi duoc huy.")
        ly_do = (reason or "").strip()
        if not ly_do:
            # Bỏ một món giữa chừng là việc người khác sẽ hỏi lại ("sao không mua nữa?") ⇒ bắt ghi
            # lý do ngay, đừng để phải đi lục chat.
            raise PurchaseValidationError("Nhập lý do bỏ món này khỏi yêu cầu.")
        line = next((l for l in row.lines if l.id == line_id), None)
        if line is None:
            raise PurchaseValidationError("Không tìm thấy món này trong yêu cầu.")
        if line.cancelled_at is not None:
            raise PurchaseConflict("Món này đã huỷ rồi.")
        phieu = self._phieu_dang_giu_dong(row).get(line.id)
        if phieu is not None:
            raise PurchaseConflict(self._cau_chan_huy_dong(phieu))

        line.cancelled_at = _now()
        line.cancelled_by_user_id = actor.id
        line.cancel_reason = ly_do
        con_song = [l for l in row.lines if l.cancelled_at is None]
        if not con_song:
            # Món cuối cùng ⇒ cả yêu cầu khép lại. Lý do của món cuối cũng là lý do của phiếu.
            row.reject_reason = ly_do
            self._dat_trang_thai(row, DPR_CANCELLED, doc_type=DOC_YCMH, actor=actor, ly_do=ly_do)
        else:
            # Bỏ một món có thể làm đổi trạng thái phiếu: yêu cầu 2 món, món chưa ai mua bị bỏ ⇒
            # phần còn lại đang ở đơn đã duyệt ⇒ phiếu nhảy từ "Chờ mua" sang "Đang mua".
            self._tinh_lai_trang_thai_ycmh(row)
        saved = self.department_requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="cancel_department_purchase_request_line",
            target=f"department_purchase_request_line:{line.id}",
            detail=f"{row.code} · {line.item_name} · {ly_do}",
        )
        return self._to_department_request_out(saved)

    # --- purchase requests -------------------------------------------------

    def list_requests(
        self,
        *,
        q: str | None = None,
        status: str | None = None,
        supplier_id: int | None = None,
        created_from: date | None = None,
        created_to: date | None = None,
        needed_from: date | None = None,
        needed_to: date | None = None,
        expected_receipt_from: date | None = None,
        expected_receipt_to: date | None = None,
        deposit_status: str | None = None,
        sort: str = "-created_at",
        page: int = 1,
        size: int = 20,
        actor=None,
        exclude_statuses: list[str] | None = None,
    ) -> tuple[list[dict], int]:
        # PHẠM VI NHÌN (chủ 04/08/2026: "nhân viên chỉ thấy đơn của tôi thôi, trưởng bộ phận hoặc
        # giám đốc mới thấy cả"). Trước đây hàm này KHÔNG nhận `actor` — ai có `thu_mua:read` là
        # thấy phiếu của cả công ty, bất kể vai được khai scope gì.
        creator_ids = None if actor is None else self._creator_ids_theo_scope(actor)
        rows, total = self.requests.list(
            q=q,
            status=status,
            supplier_id=supplier_id,
            created_from=created_from,
            created_to=created_to,
            needed_from=needed_from,
            needed_to=needed_to,
            expected_receipt_from=expected_receipt_from,
            expected_receipt_to=expected_receipt_to,
            deposit_status=deposit_status,
            sort=sort,
            page=page,
            size=size,
            creator_ids=creator_ids, exclude_statuses=exclude_statuses,
        )
        return [self._to_request_out(r) for r in rows], total

    def _purchase_scope(self, actor) -> str | None:
        """Phạm vi phiếu mua actor được nhìn: `all` | `department` | `own`.

        `all` nếu có scope `all` ở BẤT KỲ module nào trong `PURCHASE_REQUEST_READER_MODULES` —
        không thì lấy scope rộng nhất trong số các module actor có. Người không có module nào →
        `own` (chỉ phiếu của chính mình), đó là mức chặt nhất chứ không phải "thấy hết"."""
        rong_dan = {SCOPE_OWN: 0, SCOPE_DEPARTMENT: 1, SCOPE_ALL: 2}
        best = SCOPE_OWN
        for module in PURCHASE_REQUEST_READER_MODULES:
            sc = self.authz.scope_for(actor, module)
            if sc and rong_dan.get(sc, 0) > rong_dan[best]:
                best = sc
        return best

    def _creator_ids_theo_scope(self, actor) -> list[int] | None:
        """Những `created_by_user_id` mà actor được nhìn. **None = không lọc** (thấy toàn công ty).

        Tách riêng để DANH SÁCH và BADGE dùng CHUNG một phép lọc. Badge mà tự đếm lấy theo luật
        khác danh sách thì con số nó báo không mở ra xem được: báo 5 mà màn hiện 1: người dùng
        thôi tin badge, và badge thành thứ trang trí. Đây cũng chính là ràng buộc "đếm theo phạm
        vi của NGƯỜI XEM" — nhân viên scope `own` phải đếm đúng việc của mình, không đếm cả công
        ty."""
        scope = self._purchase_scope(actor)
        if scope == SCOPE_ALL:
            return None
        if scope == SCOPE_DEPARTMENT:
            return self._nguoi_cung_phong(actor)
        return [actor.id]

    def _nguoi_cung_phong(self, actor) -> list[int]:
        """id của những người CÙNG PHÒNG BAN với actor — dùng cho scope `department`.

        `purchase_requests` không có cột phòng ban, chỉ có `created_by_user_id`, nên phải quy về
        "người tạo thuộc phòng ban của tôi"."""
        if getattr(actor, "department_id", None) is None:
            return [actor.id]
        return [u.id for u in self.users.list_by_department(actor.department_id)] or [actor.id]

    # --- badge thông báo Thu mua (notify-summary) --------------------------

    def dem_dot_giao_qua_han(self) -> int:
        """Số ĐỢT GIAO đã quá hạn trả mà CÒN NỢ.

        ⚠️ Cố ý đếm bằng PYTHON, dù mọi con số khác của badge đều COUNT ở DB. `con_no` của một đợt
        KHÔNG phải một cột: nó là kết quả của một chuỗi phép tính có trạng thái —
        (1) giá trị đợt cộng từ SL nhận × đơn giá/CK/VAT trên dòng đặt (`phan_bo_du_dot`),
        (2) phiếu chi không trỏ đợt nào là CỌC CHUNG, trả thừa một đợt chảy ngược vào cọc, tiền
            nộp lại trừ vào cọc,
        (3) cọc chiếu xuống các đợt theo thứ tự GIAO TRƯỚC BÙ TRƯỚC.
        Viết lại bằng SQL là dựng nguồn sự thật THỨ HAI cho tiền — hai chỗ lệch nhau thì badge và
        màn Công nợ nói hai kiểu, mà lệch tiền chỉ lòi ra lúc đối chiếu với NCC. Nên dùng lại đúng
        `phan_bo_tien_dot` + `han_tra_dot` mà màn Công nợ đang dùng.

        `list_for_payables` đã lọc hẹp ở SQL còn 4 trạng thái có thể nợ và eager-load sẵn quan hệ
        (không N+1). Tập này vẫn lớn dần theo thời gian — khi nào chậm thì cắt bằng mốc ngày ở
        repo, đừng bỏ phần Python này đi.

        Đợt không có hạn (`han_tra_dot` trả None vì NCC chưa khai `credit_days`) KHÔNG vào đây —
        đúng luật màn Công nợ; nó được canh bằng badge 'Chưa đặt hạn' ở màn đó.
        """
        hom_nay = _business_today()
        so_dot = 0
        for row in self.requests.list_for_payables():
            phan_bo, _coc, _coc_du = phan_bo_tien_dot(row)
            for m in phan_bo:
                if m["con_no"] <= 0:
                    continue
                han = han_tra_dot(m["delivery"], row.supplier, row.debt_cutoff_date)
                if han is not None and han < hom_nay:
                    so_dot += 1
        return so_dot

    def notify_summary(self, *, actor) -> dict:
        """Ba con số nuôi badge Thu mua trên sidebar. Mỗi con số ĐẾM THEO PHẠM VI NGƯỜI XEM.

        HAI phép lọc KHÁC NHAU, đừng gộp lại cho gọn:

        · `ycmh_cho_lap_phieu` theo `_sees_all_department_requests` — YCMH là HỘP VIỆC của thu mua,
          nhân viên scope `own` vẫn phải thấy đủ yêu cầu mọi phòng gửi tới. Đếm bằng `_purchase_scope`
          ở đây là lặp lại đúng sự cố 04/08/2026: hạ scope xuống `own` làm mù hộp việc, badge báo 0
          trong khi màn YCMH của họ đầy việc.
        · `pmh_bi_tu_choi` theo `_purchase_scope` — đây đúng là PHIẾU MUA do chính họ lập, mà luật
          phiếu mua là "nhân viên chỉ thấy đơn của tôi, trưởng bộ phận/giám đốc mới thấy cả".

        `dot_giao_qua_han` là số CÔNG NỢ: chỉ trả cho người có `ke_toan:read`. Người chỉ có quyền
        thu mua nhận 0 — không rò tình hình nợ NCC cho người không được xem, và con số đó cũng
        không cộng vào badge của họ (FE cộng thẳng ba số nên trả 0 là đủ, không cần luật riêng ở
        giao diện). Đây cũng là lý do không gộp cả ba vào một câu đếm chung.
        """
        ycmh = self.department_requests.count_open(
            requesting_department_id=getattr(actor, "department_id", None),
            filter_by_department=not self._sees_all_department_requests(actor),
        )
        pmh = self.requests.count_rejected_pending_correction(
            creator_ids=self._creator_ids_theo_scope(actor)
        )
        qua_han = self.dem_dot_giao_qua_han() if self.authz.can(actor, "ke_toan", "read") else 0
        return {
            "ycmh_cho_lap_phieu": ycmh,
            "pmh_bi_tu_choi": pmh,
            "dot_giao_qua_han": qua_han,
        }

    def _co_duoc_xem(self, row: PurchaseRequest, actor) -> bool:
        scope = self._purchase_scope(actor)
        if scope == SCOPE_ALL:
            return True
        if scope == SCOPE_DEPARTMENT:
            return row.created_by_user_id in set(self._nguoi_cung_phong(actor))
        return row.created_by_user_id == actor.id

    def get_request(self, request_id: int, *, actor=None) -> dict:
        row = self._request(request_id)
        # Lọc ở DANH SÁCH mà để CHI TIẾT mở là vô nghĩa — biết id là đọc được hết. Báo 404 chứ
        # không 403: không xác nhận cho người ngoài biết phiếu đó có tồn tại hay không.
        if actor is not None and not self._co_duoc_xem(row, actor):
            raise PurchaseNotFound("Không tìm thấy phiếu yêu cầu mua hàng.")
        return self._to_request_out(row)

    def _request(self, request_id: int) -> PurchaseRequest:
        """Tra phiếu theo id, KHÔNG kiểm phạm vi.

        ⚠️ Chỉ dùng cho đường đọc đã tự kiểm (`get_request`) hoặc chỗ gọi nội bộ. Mọi hàm GHI phải
        đi qua `_request_ghi` — xem lý do ở đó."""
        row = self.requests.get_by_id(request_id)
        if row is None:
            raise PurchaseNotFound("Không tìm thấy phiếu yêu cầu mua hàng.")
        return row

    def _request_ghi(self, request_id: int, actor) -> PurchaseRequest:
        """Cửa DUY NHẤT cho mọi hàm ghi lên phiếu mua. **Ghi không được rộng hơn đọc.**

        Trước 05/08/2026 chỉ đường ĐỌC kiểm phạm vi; toàn bộ đường GHI tra thẳng `_request` nên
        cổng quyền ở router chỉ hỏi *"có quyền `thu_mua:update` không"*, KHÔNG hỏi phiếu đó của ai.
        Nhân viên không nhìn thấy phiếu đồng nghiệp trong danh sách (nhận 404) nhưng gọi thẳng theo
        id thì **sửa, xoá, gửi duyệt, đánh dấu đã nhận hàng** đều được — mà "đã nhận hàng" thì ĐẺ RA
        CÔNG NỢ trên bàn kế toán. Id là số chạy, đoán được.

        Trớ trêu là `cancel` đã có chốt sở hữu riêng từ 04/08, còn `delete` thì không: huỷ phiếu
        người khác bị chặn mà XOÁ HẲN lại được.

        Báo 404 chứ không 403 — giống hệt đường đọc, để không xác nhận cho người ngoài biết phiếu đó
        có tồn tại. Trả 403 là tự khai ra "có phiếu này nhưng anh không được đụng"."""
        row = self._request(request_id)
        if not self._co_duoc_xem(row, actor):
            raise PurchaseNotFound("Không tìm thấy phiếu yêu cầu mua hàng.")
        return row

    def _require_supplier_active(self, supplier_id: int | None) -> None:
        if supplier_id is None:
            raise PurchaseValidationError("Nhà cung cấp là thông tin bắt buộc.")
        supplier = self.suppliers.get_by_id(supplier_id)
        if supplier is None:
            raise PurchaseValidationError("Nhà cung cấp không tồn tại.")
        if supplier.status != SUPPLIER_ACTIVE:
            raise PurchaseValidationError("Nhà cung cấp đang ngừng hợp tác.")

    def _clean_request_header(
        self,
        *,
        supplier_id: int | None,
        purpose: str | None = None,
        content: str | None = None,
        needed_date: date | None = None,
    ) -> tuple[int, str, date]:
        if supplier_id is None:
            raise PurchaseValidationError("Nhà cung cấp là thông tin bắt buộc.")
        cleaned_purpose = (purpose or "").strip()
        if not cleaned_purpose:
            raise PurchaseValidationError("Nội dung / mục đích mua hàng không được trống.")
        if needed_date is None:
            raise PurchaseValidationError("Ngày cần hàng là thông tin bắt buộc.")
        if needed_date < _business_today():
            raise PurchaseValidationError("Ngay can hang khong duoc nho hon hom nay.")
        return supplier_id, cleaned_purpose, needed_date

    def _clean_expected_receipt_date(
        self,
        *,
        needed_date: date,
        expected_receipt_date: date | None,
    ) -> date | None:
        if expected_receipt_date is None:
            return None
        if expected_receipt_date < _business_today():
            raise PurchaseValidationError("Ngay du kien nhan hang khong duoc nho hon hom nay.")
        # KHÔNG chặn "nhận sớm hơn ngày cần" (chủ 03/08/2026). Nhận hàng TRƯỚC ngày cần chính là
        # trường hợp MONG MUỐN — bắt `nhận ≥ cần` là cấm đúng cái tốt, ép mọi kế hoạch phải về sát
        # hạn hoặc trễ. Ràng buộc duy nhất còn lại: không nhận vào ngày đã qua.
        return expected_receipt_date

    @staticmethod
    def _chot_noi_dong(cleaned_lines, source_requests) -> None:
        """Dòng phiếu chỉ được trỏ về dòng của CHÍNH yêu cầu nguồn nó gắn — VÀ kế thừa mặt hàng gốc.

        Không chốt thì phiếu trỏ được sang dòng của yêu cầu khác, và chi tiết YCMH hiện nhầm tiến
        độ của người khác — im lặng, không báo lỗi.

        Chạy RIÊNG sau khi đã gỡ yêu cầu nguồn, không nhét vào `_clean_lines`: nhét vào thì phải
        gỡ nguồn trước khi làm sạch dòng, và thứ tự báo lỗi đảo ngược — người dùng gõ thiếu đơn vị
        tính lại nhận câu "yêu cầu không còn ở trạng thái chờ mua".

        KẾ THỪA `(hang_loai, hang_id)` (mg 0174): đây là chỗ DUY NHẤT vừa cầm dòng phiếu vừa cầm
        dòng yêu cầu nguồn, nên cũng là chỗ duy nhất nối được mặt hàng gốc mà không phải đoán từ
        tên hàng. Thiếu bước này thì bảng cân đối vật tư không thấy "hàng đang về" của chính lô
        giấy nó vừa bảo đi mua ⇒ nó giục mua thêm lần nữa. Client CÓ gửi thì tôn trọng client
        (thu mua đổi mặt hàng khi lập phiếu là hợp lệ); chỉ điền khi đang trống.
        """
        dong_nguon = {line.id: line for src in source_requests for line in getattr(src, "lines", [])}
        for line in cleaned_lines:
            src_line_id = getattr(line, "department_request_line_id", None)
            if src_line_id is None:
                continue
            if src_line_id not in dong_nguon:
                raise PurchaseValidationError(
                    f'Dòng "{line.item_name}" trỏ tới một dòng không thuộc yêu cầu nguồn của phiếu này.'
                )
            goc = dong_nguon[src_line_id]
            if getattr(goc, "cancelled_at", None) is not None:
                raise PurchaseValidationError(
                    f'Dòng "{goc.item_name}" đã bị bộ phận yêu cầu huỷ — không lập phiếu mua cho nó nữa.'
                )
            if getattr(line, "hang_loai", None) is None and getattr(line, "hang_id", None) is None:
                line.hang_loai = getattr(goc, "hang_loai", None)
                line.hang_id = getattr(goc, "hang_id", None)

    def _clean_lines(
        self, raw_lines, *, supplier_id: int | None = None
    ) -> list[PurchaseRequestLineInput]:
        """Làm sạch dòng hàng của PHIẾU MUA.

        `supplier_id` bắt buộc trên đường thật (create/update): mỗi dòng phải nằm trong danh mục
        mặt hàng ĐANG BẬT của CHÍNH NCC đó. Trước 04/08/2026 chỗ này không kiểm gì — chọn NCC A
        rồi ghi mặt hàng chỉ NCC B bán thì phiếu vẫn tạo được, im lặng, tới lúc gửi đơn mới vỡ.
        Để `None` là bỏ qua kiểm — chỉ dành cho chỗ gọi không có NCC (nếu sau này có).

        Cái nối dòng ↔ dòng (`department_request_line_id`) chỉ được ĐỌC ở đây, chốt tính hợp lệ
        nằm ở `_chot_noi_dong` — chạy sau khi đã gỡ yêu cầu nguồn, để không đảo thứ tự báo lỗi.
        """
        if not raw_lines:
            raise PurchaseValidationError("Phiếu phải có ít nhất một dòng hàng.")
        lines: list[PurchaseRequestLineInput] = []
        for line in raw_lines:
            get = line.get if isinstance(line, dict) else lambda key, default=None: getattr(line, key, default)
            item_name = (get("item_name") or "").strip()
            if not item_name:
                raise PurchaseValidationError("Tên hàng không được trống.")
            if supplier_id is not None and not self.suppliers.supplier_sells(supplier_id, item_name):
                raise PurchaseValidationError(
                    f'Nhà cung cấp này không bán "{item_name}". '
                    "Chọn nhà cung cấp khác cho dòng này, hoặc khai mặt hàng đó cho họ trước."
                )
            unit = (get("unit") or "").strip()
            if not unit:
                raise PurchaseValidationError("Đơn vị tính không được trống.")
            quantity = float(get("quantity"))
            expected_unit_price = int(get("expected_unit_price"))
            discount_percent = float(get("discount_percent") or 0)
            vat_percent = float(get("vat_percent") or 0)
            if quantity <= 0:
                raise PurchaseValidationError("Số lượng phải lớn hơn 0.")
            if expected_unit_price <= 0:
                raise PurchaseValidationError("Đơn giá dự kiến phải lớn hơn 0.")
            if discount_percent < 0 or discount_percent > 100:
                raise PurchaseValidationError("Giảm giá (%) phải trong khoảng 0 đến 100.")
            if vat_percent < 0 or vat_percent > 100:
                raise PurchaseValidationError("Thuế GTGT (%) phải trong khoảng 0 đến 100.")
            raw_src = get("department_request_line_id")
            src_line_id = int(raw_src) if raw_src not in (None, "") else None
            hang_loai, hang_id = _doc_mat_hang(get)
            lines.append(
                PurchaseRequestLineInput(
                    item_name=item_name,
                    unit=unit,
                    quantity=quantity,
                    expected_unit_price=expected_unit_price,
                    discount_percent=discount_percent,
                    vat_percent=vat_percent,
                    note=(get("note") or "").strip() or None,
                    department_request_line_id=src_line_id,
                    hang_loai=hang_loai,
                    hang_id=hang_id,
                )
            )
        return lines

    def _new_purchase_code(self) -> str:
        # Prefix DMH (Đơn mua hàng) — đổi từ PMH 25/08/2026 theo yêu cầu chủ. CHỈ áp cho đơn MỚI;
        # đơn cũ đã lưu mã "PMH-..." giữ nguyên (mã đã in/tham chiếu ở phiếu chi, không đổi ngược).
        today = datetime.now().strftime("%y%m%d")
        alphabet = string.ascii_uppercase + string.digits
        for _ in range(20):
            rand = "".join(secrets.choice(alphabet) for _ in range(4))
            code = f"DMH-{today}-{rand}"
            if self.requests.get_by_code(code) is None:
                return code
        raise PurchaseConflict("Không sinh được mã phiếu duy nhất, vui lòng thử lại.")

    def _resolve_source_requests(
        self,
        source_request_ids,
        *,
        allow_in_purchase: bool = True,
        allowed_reserved_ids: set[int] | None = None,
    ) -> list[DepartmentPurchaseRequest]:
        allowed_reserved_ids = allowed_reserved_ids or set()
        ids: list[int] = []
        seen: set[int] = set()
        for raw_id in source_request_ids or []:
            source_id = int(raw_id)
            if source_id not in seen:
                ids.append(source_id)
                seen.add(source_id)
        if not ids:
            raise PurchaseValidationError("Phieu mua phai gan it nhat mot yeu cau mua tu phong ban.")
        if len(ids) != 1:
            raise PurchaseValidationError("Moi phieu mua hang chi duoc gan 1 yeu cau mua tu phong ban.")
        rows = self.department_requests.get_many(ids)
        by_id = {row.id: row for row in rows}
        missing = [str(source_id) for source_id in ids if source_id not in by_id]
        if missing:
            raise PurchaseValidationError("Yeu cau mua khong ton tai: " + ", ".join(missing) + ".")
        blocked = [
            row.code
            for row in rows
            if row.status in (DPR_DONE, DPR_CANCELLED)
            or (
                row.status in (DPR_PENDING_APPROVAL, DPR_IN_PURCHASE)
                and row.id not in allowed_reserved_ids
                and not allow_in_purchase
            )
        ]
        if blocked:
            raise PurchaseValidationError(
                "Yeu cau mua khong con o trang thai cho mua: " + ", ".join(blocked) + "."
            )
        return [by_id[source_id] for source_id in ids]

    def create_request(
        self,
        *,
        supplier_id: int | None,
        purpose: str | None = None,
        content: str | None = None,
        needed_date: date | None = None,
        expected_receipt_date: date | None = None,
        note: str | None,
        lines,
        source_request_ids,
        actor,
    ) -> dict:
        noi_dung = self._gop_noi_dung(content, purpose, note)
        supplier_id, noi_dung, needed_date = self._clean_request_header(
            supplier_id=supplier_id, purpose=noi_dung, needed_date=needed_date
        )
        expected_receipt_date = self._clean_expected_receipt_date(
            needed_date=needed_date, expected_receipt_date=expected_receipt_date
        )
        self._require_supplier_active(supplier_id)
        cleaned_lines = self._clean_lines(lines, supplier_id=supplier_id)
        source_requests = self._resolve_source_requests(source_request_ids, allow_in_purchase=False)
        self._chot_noi_dong(cleaned_lines, source_requests)
        # Chụp trạng thái YCMH TRƯỚC khi repo giữ chỗ — xem `_ghi_lich_su_giu_cho`.
        truoc_giu_cho = {sr.id: sr.status for sr in source_requests}
        row = self.requests.create(
            code=self._new_purchase_code(),
            supplier_id=supplier_id,
            purpose=noi_dung[:500],
            content=noi_dung,
            needed_date=needed_date,
            expected_receipt_date=expected_receipt_date,
            created_by_user_id=actor.id,
            note=(note or "").strip() or None,
            lines=cleaned_lines,
            source_requests=source_requests,
        )
        self._ghi_lich_su_giu_cho(truoc_giu_cho, source_requests)
        self.audit.create(
            actor_user_id=actor.id,
            action="create_purchase_request",
            target=f"purchase_request:{row.id}",
            detail=row.code,
        )
        return self._to_request_out(row)

    def create_requests_batch(
        self,
        *,
        purpose=None,
        content=None,
        needed_date=None,
        expected_receipt_date: date | None = None,
        note: str | None,
        lines,
        source_request_ids,
        actor,
    ) -> list[dict]:
        """Tạo N phiếu mua từ một danh sách dòng ĐÃ GÁN nhà cung cấp — nhóm theo NCC.

        Một phiếu mua là thoả thuận với MỘT nhà cung cấp, nên yêu cầu chứa hàng của nhiều nơi thì
        buộc phải tách. Gộp vào một lời gọi vì hai lý do:

        1. Giao diện KHÔNG gọi API tạo phiếu nhiều lần được: tạo phiếu đầu là yêu cầu nguồn bị
           giữ chỗ ngay (`_replace_sources` đẩy sang `pending_approval`), lần thứ hai cho NCC khác
           sẽ bị `_resolve_source_requests` chặn. Ở đây khai luôn tập yêu cầu đó là "đã biết,
           cho phép" nên cả mẻ dùng chung được.
        2. Hỏng thì hỏng CẢ MẺ (`create_many` một commit), không để lại phiếu mồ côi giữ chỗ.

        Thứ tự phiếu ra theo thứ tự NCC XUẤT HIỆN LẦN ĐẦU trong danh sách dòng — người dùng nhìn
        bảng từ trên xuống thì phiếu cũng ra theo thứ tự đó, không nhảy lung tung.
        """
        raw = list(lines or [])
        if not raw:
            raise PurchaseValidationError("Phai co it nhat mot dong hang.")

        # Nhóm theo NCC, giữ thứ tự xuất hiện.
        nhom: dict[int, list] = {}
        for line in raw:
            get = line.get if isinstance(line, dict) else lambda k, d=None: getattr(line, k, d)
            sid = get("supplier_id")
            if not sid:
                raise PurchaseValidationError(
                    f'Dong "{(get("item_name") or "").strip()}" chua chon nha cung cap.'
                )
            nhom.setdefault(int(sid), []).append(line)

        # Yêu cầu nguồn: giải MỘT lần cho cả mẻ. `allowed_reserved_ids` để chính các yêu cầu này
        # không tự chặn nhau khi phiếu thứ hai trở đi gắn lại cùng tập.
        source_requests = self._resolve_source_requests(source_request_ids, allow_in_purchase=False)
        source_ids = {row.id for row in source_requests}

        # Kiểm TẤT CẢ trước khi dựng bất cứ thứ gì — vỡ ở nhóm thứ ba thì hai nhóm đầu cũng
        # không được tạo.
        items: list[dict] = []
        for supplier_id, group in nhom.items():
            sid, noi_dung_me, cleaned_needed = self._clean_request_header(
                supplier_id=supplier_id,
                purpose=self._gop_noi_dung(content, purpose, note),
                needed_date=needed_date,
            )
            cleaned_receipt = self._clean_expected_receipt_date(
                needed_date=cleaned_needed, expected_receipt_date=expected_receipt_date
            )
            self._require_supplier_active(sid)
            cleaned_group = self._clean_lines(group, supplier_id=sid)
            self._chot_noi_dong(cleaned_group, source_requests)
            items.append(dict(
                code=self._new_purchase_code(),
                supplier_id=sid,
                purpose=noi_dung_me[:500],
                content=noi_dung_me,
                needed_date=cleaned_needed,
                expected_receipt_date=cleaned_receipt,
                created_by_user_id=actor.id,
                note=(note or "").strip() or None,
                lines=cleaned_group,
                source_requests=self._resolve_source_requests(
                    source_ids, allow_in_purchase=False, allowed_reserved_ids=source_ids
                ),
            ))

        rows = self.requests.create_many(items)
        for row in rows:
            self.audit.create(
                actor_user_id=actor.id,
                action="create_purchase_request",
                target=f"purchase_request:{row.id}",
                detail=row.code,
            )
        return [self._to_request_out(row) for row in rows]

    def update_request(
        self,
        request_id: int,
        *,
        actor,
        supplier_id,
        source_request_ids,
        purpose=None,
        content=None,
        needed_date=None,
        expected_receipt_date=None,
        note=None,
        lines=None,
    ) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status not in (PR_DRAFT, PR_REJECTED):
            raise PurchaseConflict("Chỉ phiếu nháp hoặc bị từ chối mới được sửa.")
        noi_dung = self._gop_noi_dung(content, purpose, note)
        supplier_id, noi_dung, needed_date = self._clean_request_header(
            supplier_id=supplier_id, purpose=noi_dung, needed_date=needed_date
        )
        expected_receipt_date = self._clean_expected_receipt_date(
            needed_date=needed_date, expected_receipt_date=expected_receipt_date
        )
        self._require_supplier_active(supplier_id)
        cleaned_lines = self._clean_lines(lines, supplier_id=supplier_id)
        nguon = self._resolve_source_requests(
            source_request_ids,
            allow_in_purchase=False,
            allowed_reserved_ids={link.department_request_id for link in row.sources},
        )
        self._chot_noi_dong(cleaned_lines, nguon)
        # Tính tổng MỚI trước khi ghi: `update_header_and_lines` lưu luôn, raise sau đó là dữ liệu
        # đã đổi rồi mới báo lỗi.
        tong_moi = sum(
            _purchase_line_amounts(
                quantity=float(ln.quantity),
                unit_price=int(ln.expected_unit_price),
                discount_percent=float(ln.discount_percent or 0),
                vat_percent=float(ln.vat_percent or 0),
            )[3]
            for ln in cleaned_lines
        )
        self._chan_coc_vuot_tong(
            int(row.deposit_expected or 0), int(tong_moi),
            khi="sau khi sửa — hạ cọc dự kiến xuống trước rồi sửa hàng",
        )
        row = self.requests.update_header_and_lines(
            row,
            supplier_id=supplier_id,
            purpose=noi_dung[:500],
            content=noi_dung,
            needed_date=needed_date,
            expected_receipt_date=expected_receipt_date,
            note=(note or "").strip() or None,
            lines=cleaned_lines,
            source_requests=nguon,
        )
        self.audit.create(
            actor_user_id=actor.id,
            action="update_purchase_request",
            target=f"purchase_request:{row.id}",
            detail=row.code,
        )
        return self._to_request_out(row)

    def delete_request(self, request_id: int, *, actor) -> None:
        row = self._request_ghi(request_id, actor)
        if row.status != PR_DRAFT:
            raise PurchaseConflict("Chỉ phiếu nháp mới được xóa.")
        code = row.code
        # Suy LẠI bỏ qua chính phiếu sắp xoá — sau `delete` thì quan hệ đã rời, không đọc được nữa.
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request, bo_qua_phieu_id=row.id)
        self.requests.delete(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="delete_purchase_request",
            target=f"purchase_request:{request_id}",
            detail=code,
        )

    def submit(self, request_id: int, *, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status not in (PR_DRAFT, PR_REJECTED):
            raise PurchaseConflict("Chỉ phiếu nháp hoặc bị từ chối mới được gửi duyệt.")
        self._dat_trang_thai(row, PR_PENDING, doc_type=DOC_PMH, actor=actor)
        row.submitted_at = _now()
        row.approved_by_user_id = None
        row.approved_at = None
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self.audit.create(actor_user_id=actor.id, action="submit_purchase_request", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    def approve(self, request_id: int, *, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status != PR_PENDING:
            raise PurchaseConflict("Chỉ phiếu đang chờ duyệt mới được duyệt.")
        # TÁCH VAI (chủ 04/08/2026: "thu mua làm gì có quyền duyệt"): ai đề xuất chi tiền thì
        # không được là người đồng ý chi. Chốt ở ĐÂY chứ không chỉ ở phân quyền, vì phân quyền là
        # cấu hình — bật lại lúc nào cũng được ở màn Phân quyền mà không ai hay.
        #
        # KHÔNG miễn cho giám đốc (chủ chốt). Hệ quả: giám đốc tự lập phiếu thì phải người khác
        # duyệt. Muốn nới thì thêm một ô miễn trừ, đừng gỡ chốt.
        if row.created_by_user_id is not None and row.created_by_user_id == actor.id:
            raise PurchaseForbidden(
                "Nguoi lap phieu khong duoc tu duyet phieu cua chinh minh."
            )
        self._dat_trang_thai(row, PR_APPROVED, doc_type=DOC_PMH, actor=actor)
        row.approved_by_user_id = actor.id
        row.approved_at = _now()
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self.audit.create(actor_user_id=actor.id, action="approve_purchase_request", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    def reject(self, request_id: int, *, reason: str | None, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status != PR_PENDING:
            raise PurchaseConflict("Chỉ phiếu đang chờ duyệt mới được từ chối.")
        self._dat_trang_thai(row, PR_REJECTED, doc_type=DOC_PMH, actor=actor, ly_do=reason)
        row.approved_by_user_id = actor.id
        row.approved_at = _now()
        row.reject_reason = (reason or "").strip() or row.reject_reason
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self.audit.create(actor_user_id=actor.id, action="reject_purchase_request", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    def mark_purchased(self, request_id: int, *, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status != PR_APPROVED:
            raise PurchaseConflict("Chỉ phiếu đã duyệt mới được đánh dấu đã mua.")
        self._dat_trang_thai(row, PR_PURCHASED, doc_type=DOC_PMH, actor=actor)
        saved = self.requests.save(row)
        self.audit.create(actor_user_id=actor.id, action="mark_purchase_request_purchased", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    def _ghi_lich_su_giu_cho(self, truoc: dict[int, str], sau_khi) -> None:
        """Ghi lịch sử cho YCMH bị **repo** đổi trạng thái.

        `_replace_sources` (purchase_repo) tự đặt `source.status` để GIỮ CHỖ yêu cầu ngay khi thu
        mua lập phiếu — kể cả phiếu còn nháp. Đó là chỗ đổi trạng thái DUY NHẤT không đi qua
        `_dat_trang_thai`, và nó là SAI TẦNG (repo không được quyết nghiệp vụ) — ghi chú này đã có
        từ trước ở `_BAC_PHIEU`, gỡ hẳn thì phải tách "giữ chỗ" khỏi "trạng thái", ngoài phạm vi
        đợt này.

        Nên: chụp trạng thái TRƯỚC khi gọi repo, so lại SAU, cái nào đổi thì ghi một dòng `may`.
        Không có vế này thì lịch sử YCMH mất hẳn bước "bị giữ chỗ" — nhìn vào tưởng nó tự nhảy từ
        Chờ mua sang Chờ duyệt mà không ai làm gì."""
        for row in sau_khi:
            cu = truoc.get(row.id)
            if cu is None or cu == row.status:
                continue
            self.lich_su.them(
                doc_type=DOC_YCMH,
                doc_id=row.id,
                from_status=cu,
                to_status=row.status,
                changed_by_user_id=None,
                source=CHANGE_BY_MAY,
                reason="Thu mua lập phiếu — yêu cầu được giữ chỗ",
            )

    def _dat_trang_thai(
        self,
        row,
        moi: str,
        *,
        doc_type: str,
        actor=None,
        ly_do: str | None = None,
    ) -> None:
        """CỬA DUY NHẤT để đổi trạng thái YCMH/PMH. Mọi chỗ khác phải đi qua đây.

        Vì sao một cửa: trước đợt này có **13 chỗ** gán thẳng `row.status = ...`. Rải lệnh ghi lịch
        sử ra 13 chỗ thì chắc chắn sót — đúng bệnh `_tinh_lai_trang_thai_ycmh` đã mắc (6/7 mốc quên
        suy lại, YCMH treo sai trạng thái hàng tháng không ai biết).

        `actor=None` ⇒ MÁY tự suy (`source='may'`), dùng cho trạng thái SUY RA của YCMH và của phiếu
        mua theo đợt giao. Không ghi tên người vào đó: không ai bấm thật.

        **KHÔNG ĐỔI THÌ KHÔNG GHI.** YCMH được suy lại ở mọi thao tác chạm phiếu con; suy ra trùng
        trạng thái cũ mà vẫn ghi thì mỗi cú bấm đẻ một dòng rác."""
        cu = row.status
        if cu == moi:
            return
        row.status = moi
        self.lich_su.them(
            doc_type=doc_type,
            doc_id=row.id,
            from_status=cu,
            to_status=moi,
            changed_by_user_id=actor.id if actor is not None else None,
            source=CHANGE_BY_NGUOI if actor is not None else CHANGE_BY_MAY,
            reason=(ly_do or "").strip() or None,
        )

    def _tinh_lai_trang_thai_ycmh(self, source, *, bo_qua_phieu_id: int | None = None) -> None:
        """SUY trạng thái yêu cầu mua hàng từ tập phiếu con. Gọi ở MỌI mốc đụng tới phiếu.

        Trước 05/08/2026 chỉ mốc *nhận hàng* biết suy; sáu chỗ còn lại set thẳng nên "ai bấm sau
        thì ghi đè". Hệ quả thấy được trên màn:
        - Duyệt MỘT phiếu là yêu cầu thành "Đang mua", dù phiếu kia còn nằm chờ giám đốc ⇒ bộ phận
          tưởng cả yêu cầu đã được duyệt.
        - Luật cũ kéo phiếu bị từ chối về "Chờ mua", làm lộ lại nút Tạo đơn và cho sinh PMH trùng
          nguồn thay vì buộc sửa phiếu cũ.
        - Một phiếu đã nhận không được che mất phiếu khác còn bị từ chối trong cùng yêu cầu.

        Luật: xếp bậc, lấy bậc **THẤP NHẤT**. Báo bi quan thì cùng lắm bộ phận đi hỏi; báo lạc
        quan thì họ ngồi chờ hàng không bao giờ tới.

        Phiếu **bị từ chối** vẫn tính là bậc 1: YCMH tiếp tục bị giữ bởi PMH cũ để Thu mua sửa và
        gửi duyệt lại. Chỉ phiếu **đã huỷ** mới tính là bậc 0 (Chờ mua), vì lúc đó quan hệ mua đã
        kết thúc và YCMH mới thực sự được phép lập đơn khác.

        Suy theo **DÒNG** khi phiếu có nối `department_request_line_id` — chính xác hơn hẳn: phiếu
        cũ bị huỷ và phiếu mới đã về hàng cùng phủ một dòng thì dòng đó tính là đã xong. Dữ liệu cũ
        không có nối thì lùi về suy theo PHIẾU.

        Yêu cầu đã HUỶ thì không đụng — huỷ là quyết định của người, không phải trạng thái suy ra.

        DÒNG đã huỷ cũng không tính (mg 0233): một dòng bỏ đi thì nó không được kéo cả yêu cầu về
        "Chờ mua" nữa — trước đây nó kéo, vì dòng không có phiếu nào phủ luôn tính bậc 0.
        """
        if source is None or source.status == DPR_CANCELLED:
            return
        dong_song = [
            line for line in getattr(source, "lines", [])
            if getattr(line, "cancelled_at", None) is None
        ]
        if getattr(source, "lines", None) and not dong_song:
            # Huỷ hết dòng ⇒ yêu cầu coi như huỷ. Đường huỷ dòng đã tự đặt trạng thái này rồi;
            # đây là lưới an toàn cho các mốc khác gọi vào (duyệt/huỷ phiếu con).
            self._dat_trang_thai(source, DPR_CANCELLED, doc_type=DOC_YCMH)
            return
        phieu = [
            link.purchase_request
            for link in getattr(source, "purchase_links", [])
            if link.purchase_request is not None
            and (bo_qua_phieu_id is None or link.purchase_request.id != bo_qua_phieu_id)
        ]
        if not phieu:
            self._dat_trang_thai(source, DPR_OPEN, doc_type=DOC_YCMH)
            return

        bac_theo_dong: dict[int, int] = {}
        for p in phieu:
            for line in p.lines:
                src_line_id = getattr(line, "department_request_line_id", None)
                if src_line_id is None:
                    continue
                # Dữ liệu cũ có thể còn một dòng đi qua nhiều phiếu; lấy tiến độ CAO NHẤT của
                # dòng đó để không làm mất tiến độ đã nhận. Luồng mới không cho lập phiếu thay thế
                # khi phiếu cũ chỉ bị từ chối — phải sửa và gửi lại chính phiếu cũ.
                bac_theo_dong[src_line_id] = max(
                    bac_theo_dong.get(src_line_id, 0), _BAC_PHIEU.get(p.status, 0)
                )
        if bac_theo_dong:
            # Dòng nào chưa có phiếu nào phủ thì vẫn là "Chờ mua" — chính là chỗ hay bị bỏ sót.
            # Chỉ đếm DÒNG CÒN SỐNG: dòng đã huỷ không còn chờ ai mua nữa.
            bac = min(
                bac_theo_dong.get(line.id, 0) for line in dong_song
            ) if dong_song else min(bac_theo_dong.values())
        else:
            bac = min(_BAC_PHIEU.get(p.status, 0) for p in phieu)
        self._dat_trang_thai(source, _BAC_SANG_TRANG_THAI[bac], doc_type=DOC_YCMH)

    # `_moi_phieu_da_ve_hang` đã GỠ 05/08/2026: `_tinh_lai_trang_thai_ycmh` nuốt trọn việc của nó
    # và làm đúng thêm hai ca mà nó bỏ sót — phiếu BỊ TỪ CHỐI (nó chỉ loại phiếu huỷ nên yêu cầu
    # treo vĩnh viễn) và suy theo DÒNG khi có nối dòng ↔ dòng.

    def _ap_so_thuc_nhan(self, row: PurchaseRequest, khai: list[dict] | None) -> None:
        """Ghi số thực nhận vào các dòng của phiếu.

        `khai` là None hoặc rỗng ⇒ KHÔNG đụng gì, các dòng giữ `received_quantity` cũ (thường là
        NULL = nhận đủ). Nhờ vậy đường gọi cũ `mark_received(id, actor=...)` chạy y như trước.

        Không cho khai NHIỀU hơn số đặt: giá trị thực nhận là trần lập phiếu chi, khai vống lên là
        chi vượt giá trị đơn giám đốc đã duyệt mà không qua duyệt lại. NCC giao dư thật thì sửa đơn,
        không phải sửa ở đây."""
        if not khai:
            return
        theo_id = {line.id: line for line in row.lines}
        for item in khai:
            line_id = item.get("line_id")
            line = theo_id.get(line_id)
            if line is None:
                raise PurchaseValidationError(f"Dòng {line_id} không thuộc phiếu này.")
            raw = item.get("received_quantity")
            if raw is None:
                line.received_quantity = None
                continue
            qty = float(raw)
            if qty < 0:
                raise PurchaseValidationError("Số thực nhận không được âm.")
            if qty > float(line.quantity):
                raise PurchaseValidationError(
                    f"'{line.item_name}': nhận {qty:g} nhiều hơn số đặt {float(line.quantity):g}. "
                    "Nhận dư thì sửa đơn rồi duyệt lại, không khai vống ở đây."
                )
            line.received_quantity = qty

    # --- đợt giao ----------------------------------------------------------

    _TRANG_THAI_GHI_DOT = (PR_PURCHASED, PR_PARTIALLY_RECEIVED)

    def _suy_trang_thai_nhan_hang(self, row: PurchaseRequest) -> None:
        """SUY trạng thái nhận hàng từ các đợt giao. Gọi sau MỌI lần đụng đợt giao.

        Chỉ chạy khi phiếu đang ở `purchased` / `partially_received`. Phiếu đã `received` thì ĐỨNG
        YÊN: nó tới đó bằng một quyết định (giao đủ, hoặc người bấm "Đóng đơn" vì NCC không giao
        nữa) — để hàm này tự lùi về "giao một phần" là xoá mất quyết định đó, và món nợ vừa chốt
        xong lại hiện lên bàn kế toán. Đường ra khỏi `received` chỉ có một: "Mở lại đơn", có bắt
        lý do và ghi nhật ký."""
        if row.status not in self._TRANG_THAI_GHI_DOT:
            return
        da_giao = da_giao_theo_dong(row)
        if da_giao is None:
            self._dat_trang_thai(row, PR_PURCHASED, doc_type=DOC_PMH)
            return
        du_het = all(
            da_giao.get(line.id, 0.0) >= float(line.quantity) - 1e-9 for line in row.lines
        )
        self._dat_trang_thai(
            row, PR_RECEIVED if du_het else PR_PARTIALLY_RECEIVED, doc_type=DOC_PMH
        )

    def _dot_giao(self, row: PurchaseRequest, delivery_id: int) -> PurchaseDelivery:
        for d in row.deliveries:
            if d.id == delivery_id:
                return d
        raise PurchaseNotFound("Không tìm thấy đợt giao.")

    def _chan_neu_dot_da_co_phieu_chi(self, row: PurchaseRequest, delivery_id: int) -> None:
        """Đợt đã có phiếu chi thì cấm sửa/xoá — tiền đã ra thì không được đổi số hàng dưới chân nó.

        Không kiểm trạng thái phiếu chi: từ 06/08/2026 phiếu chi lập ra là đã chi, nên tồn tại một
        phiếu gắn đợt này nghĩa là tiền đã rời két. Phiếu đã HUỶ thì không tính."""
        vuong = [
            v.code
            for v in row.payment_vouchers
            if getattr(v, "delivery_id", None) == delivery_id
            and v.status != PAYMENT_VOUCHER_CANCELLED
        ]
        if vuong:
            raise PurchaseConflict(
                f"Đợt giao này đã có phiếu chi ({', '.join(vuong)}) — huỷ phiếu chi trước rồi mới sửa."
            )

    def _clean_dot_lines(self, row: PurchaseDelivery | None, request: PurchaseRequest, raw_lines) -> list[dict]:
        """Chuẩn hoá + kiểm dòng của một đợt giao.

        Hai chốt:
        - Mỗi mặt hàng chỉ một dòng trong một đợt (khớp UNIQUE ở DB, nhưng phải báo lỗi tử tế chứ
          không để IntegrityError bắn lên 500).
        - **CHO khai vượt số đặt** (chủ chốt 28/08/2026) — NCC giao thêm mà giá giữ nguyên là
          chuyện có thật. Chặn cứng như trước thì hàng nằm trong kho mà sổ không ghi nổi. Không sợ
          "bơm nợ chưa từng phát sinh" nữa vì tiền KHÔNG còn đi theo số nhận: `phan_bo_du_dot` lấp
          phần tính tiền trước rồi cho phần vượt giá 0đ, nên tổng nợ luôn dừng ở giá trị đơn đã
          duyệt. Trần tiền `_chan_tong_dot_vuot_don` vẫn đứng đó làm lưới cuối.
          (Đường CŨ `_ap_so_thuc_nhan` — khai nhận thiếu khi KHÔNG có đợt giao — vẫn giữ trần
          "không quá số đặt": ở đó không có phép chia luỹ kế nào, nhận vượt không có nghĩa tiền.)
        """
        if not raw_lines:
            raise PurchaseValidationError("Đợt giao phải có ít nhất một dòng hàng.")
        line_by_id = {line.id: line for line in request.lines}
        bo_qua = row.id if row is not None else None
        da_giao_khac: dict[int, float] = {}
        for d in request.deliveries:
            if bo_qua is not None and d.id == bo_qua:
                continue
            for dl in d.lines:
                da_giao_khac[dl.purchase_request_line_id] = (
                    da_giao_khac.get(dl.purchase_request_line_id, 0.0) + float(dl.quantity)
                )

        seen: set[int] = set()
        out: list[dict] = []
        for item in raw_lines:
            line_id = item.get("purchase_request_line_id")
            line = line_by_id.get(line_id)
            if line is None:
                raise PurchaseValidationError(f"Dòng {line_id} không thuộc phiếu mua này.")
            if line_id in seen:
                raise PurchaseValidationError(
                    f"'{line.item_name}' bị khai hai lần trong cùng một đợt giao."
                )
            seen.add(line_id)
            qty = float(item.get("quantity") or 0)
            if qty <= 0:
                raise PurchaseValidationError(
                    f"'{line.item_name}': số nhận của đợt phải lớn hơn 0. "
                    "Không nhận món nào thì bỏ dòng đó ra khỏi đợt."
                )
            out.append(
                {
                    "purchase_request_line_id": line_id,
                    "quantity": qty,
                    "note": (item.get("note") or "").strip() or None,
                }
            )
        return out

    @staticmethod
    def _clean_dot_header(values: dict) -> dict:
        ngay = values.get("delivery_date")
        if ngay is None:
            raise PurchaseValidationError("Đợt giao phải có ngày giao.")
        if ngay > _business_today():
            raise PurchaseValidationError(
                "Ngày giao không được ở tương lai — kiểm lại xem có gõ nhầm không."
            )
        han = values.get("due_date")
        if han is not None and han < ngay:
            raise PurchaseValidationError("Hạn trả không được trước ngày giao.")
        ngay_hd = values.get("invoice_date")
        if ngay_hd is not None and ngay_hd > _business_today():
            raise PurchaseValidationError("Ngày hóa đơn không được ở tương lai.")
        # KHÔNG nhận số tiền: tiền của đợt do máy tính từ đơn giá đã chốt trên phiếu
        # (chủ chốt 07/08/2026 — xem `phan_bo_du_dot`). Cột `amount` để dormant.
        return {
            "delivery_date": ngay,
            "due_date": han,
            "invoice_number": (values.get("invoice_number") or "").strip() or None,
            "invoice_date": ngay_hd,
            "note": (values.get("note") or "").strip() or None,
        }

    def ghi_dot_giao(self, request_id: int, *, lines: list[dict], actor, **values) -> dict:
        """Ghi một đợt giao mới cho phiếu mua."""
        row = self._request_ghi(request_id, actor)
        if row.status not in self._TRANG_THAI_GHI_DOT:
            raise PurchaseConflict(
                "Chỉ phiếu ĐÃ MUA hoặc ĐANG GIAO MỘT PHẦN mới ghi được đợt giao."
            )
        header = self._clean_dot_header(values)
        cleaned = self._clean_dot_lines(None, row, lines)
        seq = max((d.seq_no for d in row.deliveries), default=0) + 1
        dot = PurchaseDelivery(
            seq_no=seq, created_by_user_id=actor.id, **header
        )
        dot.lines = [PurchaseDeliveryLine(**c) for c in cleaned]
        row.deliveries.append(dot)
        self._sau_khi_doi_dot(row)
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="create_purchase_delivery",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {self._tom_tat_dot_giao(dot, row)}",
        )
        return self._to_request_out(saved)

    def sua_dot_giao(self, request_id: int, delivery_id: int, *, lines: list[dict] | None, actor, **values) -> dict:
        row = self._request_ghi(request_id, actor)
        dot = self._dot_giao(row, delivery_id)
        self._chan_neu_dot_da_co_phieu_chi(row, delivery_id)
        header = self._clean_dot_header(values)
        for k, v in header.items():
            setattr(dot, k, v)
        if lines is not None:
            cleaned = self._clean_dot_lines(dot, row, lines)
            # Cập nhật TẠI CHỖ theo dòng gốc. Gán cả collection bằng object mới khiến SQLAlchemy
            # có thể INSERT dòng thay thế trước khi DELETE dòng cũ, đụng UNIQUE
            # (delivery_id, purchase_request_line_id) và nổ 500 khi sửa một đợt đã có hàng.
            hien_tai = {line.purchase_request_line_id: line for line in dot.lines}
            incoming_ids = {item["purchase_request_line_id"] for item in cleaned}
            for item in cleaned:
                line_id = item["purchase_request_line_id"]
                existing = hien_tai.get(line_id)
                if existing is None:
                    dot.lines.append(PurchaseDeliveryLine(**item))
                else:
                    existing.quantity = item["quantity"]
                    existing.note = item["note"]
            for existing in list(dot.lines):
                if existing.purchase_request_line_id not in incoming_ids:
                    dot.lines.remove(existing)
        self._sau_khi_doi_dot(row)
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="update_purchase_delivery",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {self._tom_tat_dot_giao(dot, row)}",
        )
        return self._to_request_out(saved)

    def xoa_dot_giao(self, request_id: int, delivery_id: int, *, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        dot = self._dot_giao(row, delivery_id)
        self._chan_neu_dot_da_co_phieu_chi(row, delivery_id)
        seq = dot.seq_no
        tom_tat = self._tom_tat_dot_giao(dot, row)
        row.deliveries.remove(dot)
        self._sau_khi_doi_dot(row)
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="delete_purchase_delivery",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {tom_tat}",
        )
        return self._to_request_out(saved)

    def gan_hoa_don(
        self,
        request_id: int,
        *,
        delivery_ids: list[int],
        invoice_number: str | None,
        invoice_date: date | None,
        actor,
    ) -> dict:
        """Gán MỘT hoá đơn cho NHIỀU đợt giao.

        Ca thật: NCC giao ba đợt rồi mới xuất một hoá đơn chung. Không có thao tác này thì kế toán
        phải mở sửa từng đợt và gõ lại cùng một số ba lần — gõ lệch một ký tự là hệ hiểu thành ba
        hoá đơn khác nhau."""
        row = self._request_ghi(request_id, actor)
        if not delivery_ids:
            raise PurchaseValidationError("Chưa chọn đợt giao nào để gán hóa đơn.")
        so_hd = (invoice_number or "").strip() or None
        if invoice_date is not None and invoice_date > _business_today():
            raise PurchaseValidationError("Ngày hóa đơn không được ở tương lai.")
        for did in delivery_ids:
            dot = self._dot_giao(row, did)
            dot.invoice_number = so_hd
            dot.invoice_date = invoice_date
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="assign_purchase_invoice",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — HĐ {so_hd or '(gỡ)'} cho {len(delivery_ids)} đợt",
        )
        return self._to_request_out(saved)

    # --- hạn mức công nợ NCC -----------------------------------------------

    def han_muc_ncc(self, supplier_id: int | None) -> dict:
        """Nợ hiện tại của một NCC so với hạn mức đã khai.

        **CẢNH BÁO MỀM, không chặn** (chủ chốt 06/08/2026, Đ6) — cùng nếp với rào chiết khấu/biên
        của khách hàng: hệ nói cho người biết, người quyết. Chặn cứng ở đây là đúng lúc gấp nhất
        (hết giấy, phải mua ngay) thì hệ khoá đường mua.

        `credit_limit = 0` nghĩa là KHÔNG đặt hạn mức ⇒ không bao giờ vượt. Đừng đổi thành "hạn mức
        0đ ⇒ mua gì cũng vượt": mọi NCC cũ đều đang để 0."""
        trong = {
            # `payment_terms` đi kèm từ 28/08/2026: màn Đơn mua hàng bên Kế toán bày cả ba điều
            # kiện với NCC ở một chỗ, mà nó là trường DUY NHẤT trong ba cái không nằm sẵn ở đây.
            # Nhét vào đây thay vì gọi thêm `/api/suppliers/{id}`: cùng một câu hỏi ("trả ông này
            # thế nào") thì đừng bắt màn hình hỏi hai lần.
            "payment_terms": None,
            "credit_limit": 0,
            "credit_days": None,
            "no_hien_tai": 0,
            "vuot_han_muc": False,
            "vuot_bao_nhieu": 0,
        }
        if supplier_id is None:
            return trong
        supplier = self.suppliers.get_by_id(supplier_id)
        if supplier is None:
            return trong
        han_muc = int(getattr(supplier, "credit_limit", 0) or 0)
        no = sum(
            purchase_money(r)["outstanding_amount"]
            for r in self.requests.list_for_payables(supplier_id=supplier_id)
        )
        return {
            "payment_terms": getattr(supplier, "payment_terms", None),
            "credit_limit": han_muc,
            "credit_days": getattr(supplier, "credit_days", None),
            "no_hien_tai": no,
            "vuot_han_muc": han_muc > 0 and no > han_muc,
            "vuot_bao_nhieu": max(0, no - han_muc) if han_muc > 0 else 0,
        }

    # --- hợp đồng & đính kèm ------------------------------------------------

    def cap_nhat_hop_dong(
        self,
        request_id: int,
        *,
        contract_number: str | None,
        deposit_expected: int | None,
        debt_cutoff_date=None,
        actor,
    ) -> dict:
        """Số hợp đồng + cọc dự kiến.

        Tách khỏi `update_request` (chỉ sửa được phiếu nháp/bị từ chối) vì hợp đồng thường ký SAU
        khi phiếu đã duyệt — bắt sửa ở màn nháp là không bao giờ điền được.

        `deposit_expected` KHÔNG vào công thức công nợ. Nó chỉ là con số để kế toán đối chiếu với
        phiếu chi đặt cọc thực tế, và để điền sẵn số tiền khi lập phiếu cọc. Cho nó vào công thức
        là trừ cọc HAI LẦN.

        **CỌC KHOÁ SAU KHI DUYỆT** (chủ chốt 06/08/2026): cọc là một phần của khoản chi mà người
        duyệt đã đồng ý. Cho sửa sau khi duyệt là đổi con số đã ký mà không ai duyệt lại — đúng cái
        lỗ mà trần lập phiếu chi sinh ra để bịt. Số hợp đồng và ảnh đính kèm thì KHÔNG khoá: hợp
        đồng thường ký sau khi duyệt, khoá luôn là không bao giờ điền được."""
        row = self._request_ghi(request_id, actor)
        if row.status == PR_CANCELLED:
            raise PurchaseConflict("Phiếu đã huỷ thì không sửa được thông tin hợp đồng.")
        so_hd = (contract_number or "").strip() or None
        coc = int(deposit_expected or 0)
        if coc < 0:
            raise PurchaseValidationError("Cọc dự kiến không được âm.")
        self._chan_coc_vuot_tong(coc, int(purchase_money(row)["total"]), khi="")
        if coc != int(row.deposit_expected or 0) and row.status not in (
            PR_DRAFT,
            PR_PENDING,
            PR_REJECTED,
        ):
            raise PurchaseConflict(
                "Đơn đã duyệt thì không sửa được cọc dự kiến — đó là con số người duyệt đã đồng ý. "
                "Cần đổi thì lùi phiếu về nháp rồi duyệt lại."
            )
        row.contract_number = so_hd
        row.deposit_expected = coc
        # NGÀY CHỐT CÔNG NỢ đi cùng số hợp đồng: NCC báo sau khi đơn đã lập, nên KHÔNG khoá theo
        # duyệt như cọc. Nó không đổi số tiền nào — chỉ đổi HẠN trả, mà hạn thì NCC có quyền báo
        # muộn hoặc dời. Khoá lại là ép kế toán canh một cái hạn họ biết là sai.
        row.debt_cutoff_date = debt_cutoff_date
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="update_purchase_contract",
            target=f"purchase_request:{row.id}",
            detail=(
                f"{row.code} — HĐ {so_hd or '(trống)'} · cọc dự kiến {coc:,}đ"
                f" · chốt công nợ {debt_cutoff_date or '(chưa báo)'}"
            ),
        )
        return self._to_request_out(saved)

    def them_dinh_kem(
        self,
        request_id: int,
        *,
        delivery_id: int | None,
        kind: str,
        file_name: str | None,
        content_type: str | None,
        data: bytes,
        actor,
    ) -> dict:
        """Đính ảnh/PDF vào phiếu mua (hợp đồng) hoặc vào một đợt giao (hoá đơn, biên bản).

        Bytes vào `mua-hang/<pmh_id>/`. Tiền tố `mua-hang` PHẢI có trong `_PREFIX_PERMISSION` của
        `routers/files.py`, nếu không thì file đọc được bởi bất kỳ ai đăng nhập — bảng đó fail-MỞ."""
        row = self._request_ghi(request_id, actor)
        if kind not in PURCHASE_ATTACHMENT_KINDS:
            raise PurchaseValidationError("Loại tài liệu không hợp lệ.")
        if delivery_id is not None:
            self._dot_giao(row, delivery_id)  # ném 404 nếu đợt không thuộc phiếu này
        ct = (content_type or "").lower()
        if not (ct.startswith("image/") or ct == "application/pdf"):
            raise PurchaseValidationError("Chỉ nhận ảnh (image/*) hoặc PDF.")
        if not data:
            raise PurchaseValidationError("Tệp rỗng.")
        if len(data) > MAX_PURCHASE_ATTACHMENT_BYTES:
            raise PurchaseValidationError("Tệp vượt quá 10 MB.")
        if len(row.attachments) >= MAX_PURCHASE_ATTACHMENTS:
            raise PurchaseValidationError(
                f"Mỗi phiếu mua tối đa {MAX_PURCHASE_ATTACHMENTS} tài liệu."
            )
        key, safe_name = make_key(PURCHASE_ATTACHMENT_SUBDIR, row.id, file_name)
        get_storage().save(key, data, content_type)
        row.attachments.append(
            PurchaseAttachment(
                delivery_id=delivery_id,
                kind=kind,
                file_name=safe_name,
                file_url=url_from_key(key),
                file_type=content_type,
                uploaded_by=actor.id,
            )
        )
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="add_purchase_attachment",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {kind}: {file_name}",
        )
        return self._to_request_out(saved)

    def xoa_dinh_kem(self, request_id: int, attachment_id: int, *, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        found = next((a for a in row.attachments if a.id == attachment_id), None)
        if found is None:
            raise PurchaseNotFound("Không tìm thấy tài liệu.")
        ten = found.file_name
        key = key_from_url(found.file_url)
        row.attachments.remove(found)
        saved = self.requests.save(row)
        # Dọn bytes SAU khi hàng DB đã đi — xoá file trước mà commit hỏng thì mất file, còn hàng DB
        # trỏ vào chỗ trống. `delete()` của storage là best-effort, không raise.
        if key:
            get_storage().delete(key)
        self.audit.create(
            actor_user_id=actor.id,
            action="delete_purchase_attachment",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {ten}",
        )
        return self._to_request_out(saved)

    def _chan_tong_dot_vuot_don(self, row: PurchaseRequest) -> None:
        """TỔNG tiền các đợt giao không được vượt giá trị ĐƠN ĐẶT (chủ chốt 06/08/2026).

        Số tiền từng đợt gõ tay theo hoá đơn nên không còn bị đơn giá trói — nhưng cái trần thì
        vẫn phải còn: giá trị đơn là con số giám đốc đã DUYỆT. Cho tổng hoá đơn vượt nó là chi
        vượt mức đã duyệt mà không qua duyệt lại, và công nợ phình lên bằng một con số không ai ký.

        NCC xuất hoá đơn cao hơn thật thì đường đúng là sửa đơn rồi duyệt lại, không phải nhét
        chênh lệch vào đợt giao.

        Gọi ở MỌI mốc đụng tập đợt giao (thêm/sửa/xoá) — đặt ở đây thay vì ở từng hàm để không có
        cửa nào lọt."""
        money_total = purchase_money(row)["total"]
        tong = sum(gia_tri_cac_dot(row).values())
        if tong > money_total:
            raise PurchaseValidationError(
                f"Tổng tiền các đợt giao ({tong:,}đ) vượt giá trị đơn đã duyệt "
                f"({money_total:,}đ). Hóa đơn cao hơn đơn thì sửa đơn rồi duyệt lại, "
                "đừng nhét chênh lệch vào đợt giao."
            )

    def _sau_khi_doi_dot(self, row: PurchaseRequest) -> None:
        """Suy lại trạng thái phiếu VÀ trạng thái YCMH nguồn sau khi tập đợt giao đổi."""
        self._chan_tong_dot_vuot_don(row)
        self._suy_trang_thai_nhan_hang(row)
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        self._doi_soat_giu_cho(row)

    def _doi_soat_giu_cho(self, row: PurchaseRequest) -> None:
        """Đợt giao vừa đổi (ghi/sửa/xoá), hoặc đơn vừa đóng/huỷ/mở lại ⇒ giữ chỗ đối lại phần
        hứa đã bám PHIẾU này. TUỲ CHỌN — vắng `giu_cho` thì bỏ qua, PMH chạy y như trước
        30/08/2026."""
        if self.giu_cho is not None:
            self.giu_cho.doi_soat_dang_ve_don(row.id)

    def dong_don(self, request_id: int, *, reason: str | None, actor) -> dict:
        """ĐÓNG ĐƠN — NCC không giao nữa, chốt số thực nhận = số đã giao.

        Khác "giao đủ thì tự thành Đã nhận hàng": đây là ca giao THIẾU rồi thôi. Vì nó cắt phần hàng
        chưa về ra khỏi công nợ (nợ chỉ còn theo số đã giao), nó là quyết định về TIỀN ⇒ đòi
        `thu_mua:approve` + bắt lý do + vào nhật ký, cùng lằn ranh đã áp cho `undo_received`."""
        # GỘP VỀ Ô "Thao tác" (`thu_mua:update`) ngày 12/08/2026 — chủ chốt test rồi kết luận ô
        # riêng `manage_status` không đáng có: "quyền Sửa / đảo trạng thái đơn sau khi nhận hàng
        # vô dụng, bỏ đi được không". Ba việc nó gác (sửa số nhận · mở lại đơn · đóng đơn) đều là
        # việc thường ngày của chính người lập phiếu, tách ra chỉ thêm một ô phải nhớ tick.
        # ⚠️ Vẫn KHÔNG dùng chung với "Duyệt / từ chối PMH"
        # nữa. Đóng đơn là chốt một đơn hàng chưa về đủ (NCC không giao nữa): việc của bộ phận mua
        # hàng, không phải quyết định chi tiền. Câu báo lỗi cũ ghi "quyền duyệt" nên đọc lên tưởng
        # đúng, thật ra sai nghĩa.
        if not self.authz.can(actor, "thu_mua", "update"):
            raise PurchaseForbidden(
                "Cần quyền “Sửa / đảo trạng thái đơn sau khi nhận hàng” mới được đóng đơn."
            )
        ly_do = (reason or "").strip()
        if not ly_do:
            raise PurchaseValidationError("Phải ghi lý do đóng đơn khi hàng chưa về đủ.")
        row = self._request_ghi(request_id, actor)
        if row.status not in self._TRANG_THAI_GHI_DOT:
            raise PurchaseConflict("Chỉ phiếu đã mua / đang giao một phần mới đóng đơn được.")
        # BẮT BUỘC đã có ít nhất một đợt giao. Đóng một đơn CHƯA giao đợt nào sẽ đẩy nó sang
        # `received` với 0 đợt ⇒ `purchase_money` rơi vào nhánh phiếu-cũ và ghi nợ NGUYÊN GIÁ TRỊ
        # ĐƠN dù không món hàng nào về — đẻ ra một món nợ ma trên bàn kế toán.
        #
        # Giao diện chỉ hiện nút này ở "Giao một phần" nên bấm tay không tới được, nhưng id là số
        # chạy: gọi thẳng API là qua. Chốt phải nằm ở service.
        if not row.deliveries:
            raise PurchaseConflict(
                "Đơn chưa nhận đợt nào — không có gì để chốt. Không mua nữa thì HUỶ đơn."
            )
        self._dat_trang_thai(row, PR_RECEIVED, doc_type=DOC_PMH, actor=actor, ly_do=ly_do)
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self._doi_soat_giu_cho(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="close_purchase_request",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {ly_do}",
        )
        return self._to_request_out(saved)

    def mark_received(self, request_id: int, *, received_lines: list[dict] | None = None, actor) -> dict:
        """ĐƯỜNG CŨ — đánh dấu "Đã nhận hàng" cho phiếu KHÔNG theo dõi theo đợt giao.

        Giữ lại nguyên vẹn cho phiếu lập trước 06/08/2026 (và cho đơn nhỏ giao một lần, không ai
        muốn khai đợt). Phiếu ĐÃ có đợt giao thì trạng thái là số SUY RA — không cho gán tay nữa,
        nếu không sẽ có hai nguồn sự thật và số công nợ nhảy tuỳ ai bấm sau."""
        row = self._request_ghi(request_id, actor)
        if row.status != PR_PURCHASED:
            raise PurchaseConflict("Chỉ phiếu đã mua mới được đánh dấu đã nhận hàng.")
        if row.deliveries:
            raise PurchaseConflict(
                "Đơn này đang theo dõi theo đợt giao — ghi nốt đợt còn lại, "
                "hoặc dùng 'Đóng đơn' nếu nhà cung cấp không giao nữa."
            )
        self._ap_so_thuc_nhan(row, received_lines)
        self._dat_trang_thai(row, PR_RECEIVED, doc_type=DOC_PMH, actor=actor)
        # Một yêu cầu có thể tách thành NHIỀU phiếu (mỗi NCC một phiếu) ⇒ chỉ "Xong" khi MỌI phần
        # của nó đã về. Phiếu giấy về trước mà báo Xong thì bộ phận tưởng đủ hàng trong khi băng
        # keo còn chưa ai mua.
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self.audit.create(actor_user_id=actor.id, action="mark_purchase_request_received", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    def update_received_quantities(self, request_id: int, *, received_lines: list[dict], actor) -> dict:
        """Sửa lại số thực nhận SAU khi phiếu đã ở 'Đã nhận hàng'.

        Dùng cho ca NCC giao làm nhiều đợt: đợt 1 khai 60, đợt 2 về thì sửa lên 100. (Lịch sử từng
        đợt về ngày nào thì hàm này KHÔNG lưu — chờ phiếu nhập kho.)

        Đòi `thu_mua:approve` chứ không phải `update`: sửa số này là **đổi số nợ đã ghi trên màn kế
        toán**, ngang với việc lùi trạng thái. Vào nhật ký để còn truy được ai sửa.

        Chốt an toàn: hạ số thực nhận xuống DƯỚI số ĐÃ CHI là chặn — nếu không thì tiền đã ra vượt
        cả giá trị hàng, công nợ bị kẹp về 0 và số liệu im lặng sai.

        ⚠️ Chỉ dùng cho phiếu KHÔNG theo dõi theo đợt giao. Phiếu có đợt giao thì số thực nhận là
        Σ các đợt (số dẫn xuất) — sửa ở đây sẽ bị nhánh dẫn xuất ghi đè trong im lặng, nên chặn."""
        if not self.authz.can(actor, "thu_mua", "update"):
            raise PurchaseForbidden(
                "Cần quyền “Sửa / đảo trạng thái đơn sau khi nhận hàng” mới được sửa số thực nhận."
            )
        row = self._request_ghi(request_id, actor)
        if row.status != PR_RECEIVED:
            raise PurchaseConflict("Chỉ phiếu đã nhận hàng mới sửa được số thực nhận.")
        if row.deliveries:
            raise PurchaseConflict(
                "Đơn này theo dõi theo đợt giao — sửa số nhận ở đúng đợt giao đó, không sửa ở đây."
            )
        self._ap_so_thuc_nhan(row, received_lines)
        money = purchase_money(row)
        if money["received_total"] < money["net_paid"]:
            raise PurchaseConflict(
                f"Giá trị thực nhận ({money['received_total']:,}đ) thấp hơn số đã chi "
                f"({money['net_paid']:,}đ). Huỷ bớt phiếu chi trước rồi sửa lại."
            )
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="update_purchase_request_received_quantities",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — thực nhận {money['received_total']:,}đ / đặt {money['total']:,}đ",
        )
        return self._to_request_out(saved)

    def undo_received(self, request_id: int, *, reason: str | None, actor) -> dict:
        """Lùi 'Đã nhận hàng' về 'Đã mua'.

        Trước 05/08/2026 KHÔNG có đường nào ra khỏi `received`: `mark_received` chỉ gán vào, còn
        `cancel()` chặn thẳng phiếu đã nhận. Bấm nhầm là kẹt vĩnh viễn — trước đây không ai thấy,
        nhưng từ khi có màn Công nợ phải trả thì mỗi lần bấm nhầm là đẻ một món nợ ảo nằm chình ình
        trên bàn kế toán.

        Ba việc, thiếu việc nào cũng hỏng:

        1. **Tiền đã ra thì không lùi.** Có phiếu chi ĐÃ CHI nghĩa là tiền rời két rồi; quay lại
           khai 'chưa nhận hàng' là nói dối sổ quỹ. Huỷ phiếu chi trước rồi mới lùi được.
        2. Lật phiếu về `partially_received` nếu còn đợt giao, ngược lại về `purchased` — đây là
           đường DUY NHẤT ra khỏi `received`, nên nó phải trả phiếu về đúng chỗ mà `_suy_trang_thai
           _nhan_hang` sẽ tính ra, không thì bấm xong trạng thái lại nhảy ở lần ghi đợt kế tiếp.
        3. **Tính LẠI trạng thái YCMH nguồn** qua `_tinh_lai_trang_thai_ycmh`. Quên vế này thì
           YCMH đứng nguyên 'Xong' trong khi phiếu đã lùi — bộ phận đề nghị nhìn vào tưởng đủ hàng.
           Suy lại chứ KHÔNG kéo mù: một YCMH tách ra nhiều phiếu theo NCC, lùi một phiếu chưa chắc
           đã hết 'Xong'.

        Quyền: `thu_mua:approve` (trưởng bộ phận / giám đốc). Nút này XOÁ một món nợ khỏi màn kế
        toán nên không phải việc nhân viên tự quyết — cùng lằn ranh đã áp cho `cancel`."""
        if not self.authz.can(actor, "thu_mua", "update"):
            raise PurchaseForbidden(
                "Cần quyền “Sửa / đảo trạng thái đơn sau khi nhận hàng” mới được lùi trạng thái."
            )
        ly_do = (reason or "").strip()
        if not ly_do:
            raise PurchaseValidationError("Phải ghi lý do lùi trạng thái đã nhận hàng.")
        row = self._request_ghi(request_id, actor)
        if row.status != PR_RECEIVED:
            raise PurchaseConflict("Chỉ phiếu đã nhận hàng mới lùi được.")
        if any(v.status == PAYMENT_VOUCHER_PAID for v in row.payment_vouchers):
            raise PurchaseConflict(
                "Đơn đã có phiếu chi ĐÃ CHI — xử lý phần tiền trước khi lùi trạng thái."
            )
        self._dat_trang_thai(
            row,
            PR_PARTIALLY_RECEIVED if row.deliveries else PR_PURCHASED,
            doc_type=DOC_PMH,
            actor=actor,
            ly_do=ly_do,
        )
        self._sau_khi_doi_dot(row)
        saved = self.requests.save(row)
        self.audit.create(
            actor_user_id=actor.id,
            action="undo_purchase_request_received",
            target=f"purchase_request:{row.id}",
            detail=f"{row.code} — {ly_do}",
        )
        return self._to_request_out(saved)

    def cancel(self, request_id: int, *, reason: str | None, actor) -> dict:
        row = self._request_ghi(request_id, actor)
        if row.status in (PR_RECEIVED, PR_CANCELLED):
            raise PurchaseConflict("Phiếu đã nhận hàng hoặc đã hủy thì không thể hủy tiếp.")
        # Hàng đã về một phần thì huỷ là xoá luôn món nợ của phần đã về. Đường đúng là "Đóng đơn":
        # chốt nợ theo số đã giao rồi thôi, phần chưa giao tự rơi ra khỏi công nợ.
        if row.status == PR_PARTIALLY_RECEIVED:
            raise PurchaseConflict(
                "Đơn đã nhận một phần — dùng 'Đóng đơn' để chốt theo số đã giao, không huỷ."
            )
        # Huỷ phiếu ĐÃ GỬI DUYỆT là quyết định của người duyệt, không phải của thu mua (chủ chốt
        # 04/08/2026). Người chỉ có `cancel` thì chỉ được dọn phiếu NHÁP DO CHÍNH MÌNH lập — giữ
        # được việc tự dọn nháp mà không cho giết phiếu đang nằm trên bàn giám đốc.
        #
        # Phân biệt theo NĂNG LỰC (`approve`) chứ không theo tên vai: tên vai đổi lúc nào cũng
        # được, mà đổi xong thì luật viết theo tên vai câm lặng thất hiệu.
        # Huỷ phiếu ĐÃ GỬI DUYỆT là quyết định của NGƯỜI DUYỆT — mà ô duyệt nay nằm ở khoá
        # `ke_toan` (dời 11/08/2026, nút Duyệt/Từ chối chỉ có ở màn Đơn mua hàng bên Kế toán).
        # Người lập vẫn tự dọn được phiếu NHÁP của chính mình, không cần ô nào.
        if not self.authz.can(actor, "ke_toan", "approve"):
            if row.status != PR_DRAFT or row.created_by_user_id != actor.id:
                raise PurchaseForbidden(
                    "Chi huy duoc phieu nhap do chinh minh lap. "
                    "Phieu da gui duyet thi nguoi duyet quyet."
                )
        if any(voucher.status == PAYMENT_VOUCHER_PAID for voucher in row.payment_vouchers):
            raise PurchaseConflict("Phiếu đã có chứng từ thanh toán nên không thể hủy.")
        self._dat_trang_thai(row, PR_CANCELLED, doc_type=DOC_PMH, actor=actor, ly_do=reason)
        row.reject_reason = (reason or "").strip() or row.reject_reason
        for link in row.sources:
            self._tinh_lai_trang_thai_ycmh(link.department_request)
        saved = self.requests.save(row)
        self._doi_soat_giu_cho(row)
        self.audit.create(actor_user_id=actor.id, action="cancel_purchase_request", target=f"purchase_request:{row.id}", detail=row.code)
        return self._to_request_out(saved)

    # --- output helpers ----------------------------------------------------

    @staticmethod
    def _tom_tat_dot_giao(dot: PurchaseDelivery, row: PurchaseRequest) -> str:
        """Mô tả ngắn của một đợt để audit còn đọc được cả sau khi đợt bị xoá."""
        line_by_id = {line.id: line for line in row.lines}
        hang = []
        for delivery_line in dot.lines:
            line = line_by_id.get(delivery_line.purchase_request_line_id)
            if line is None:
                continue
            hang.append(
                f"{line.item_name} {float(delivery_line.quantity):g} {line.unit}"
            )
        hang_text = "; ".join(hang) if hang else "không còn dòng hàng"
        return f"Đợt {dot.seq_no} ngày {dot.delivery_date}: {hang_text}"

    def _lich_su_out(self, doc_type: str, doc_id: int) -> list[dict]:
        """Lịch sử trạng thái, MỚI NHẤT TRƯỚC.

        Danh sách rỗng = phiếu lập TRƯỚC 07/08/2026 (không backfill — bịa ra ngày giờ không ai
        biết). Màn hình phải nói rõ "chưa ghi nhận", đừng để trống trơn như thể mất dữ liệu."""
        return [
            {
                "id": h.id,
                "from_status": h.from_status,
                "to_status": h.to_status,
                "source": h.source,
                "changed_by_name": self._user_name(h.changed_by_user_id),
                "reason": h.reason,
                "created_at": h.created_at,
            }
            for h in self.lich_su.cua(doc_type, doc_id)
        ]

    def _lich_su_don_mua_out(self, row: PurchaseRequest) -> list[dict]:
        """Ghép lịch sử đổi trạng thái với các mốc đợt giao thành một timeline.

        Không ghi đợt giao thành một ``status`` giả: thêm đợt 2 trong khi đơn vẫn "Giao một
        phần" vẫn là việc phải truy vết, nhưng không phải đổi trạng thái. Audit đã lưu đúng các
        sự kiện này, nên chỉ cần đưa chúng ra màn hình thay vì tạo thêm bảng lịch sử thứ ba.
        """
        events: list[dict] = []
        for history in self.lich_su.cua(DOC_PMH, row.id):
            events.append(
                {
                    "id": f"status-{history.id}",
                    "event_type": "status",
                    "title": "Thay đổi trạng thái",
                    "detail": None,
                    "actor_name": self._user_name(history.changed_by_user_id),
                    "source": history.source,
                    "from_status": history.from_status,
                    "to_status": history.to_status,
                    "reason": history.reason,
                    "created_at": history.created_at,
                }
            )

        delivery_actions = {
            "create_purchase_delivery": ("delivery_created", "Ghi nhận đợt giao"),
            "update_purchase_delivery": ("delivery_updated", "Cập nhật đợt giao"),
            "delete_purchase_delivery": ("delivery_deleted", "Xóa đợt giao"),
            "assign_purchase_invoice": ("invoice_assigned", "Gán hóa đơn cho đợt giao"),
        }
        for log in self.audit.list_by_target(f"purchase_request:{row.id}", limit=200):
            info = delivery_actions.get(log.action)
            if info is None:
                continue
            event_type, title = info
            events.append(
                {
                    "id": f"audit-{log.id}",
                    "event_type": event_type,
                    "title": title,
                    "detail": log.detail or None,
                    "actor_name": self._user_name(log.actor_user_id),
                    "source": "nguoi" if log.actor_user_id is not None else "may",
                    "from_status": None,
                    "to_status": None,
                    "reason": None,
                    "created_at": log.created_at,
                }
            )
        return sorted(events, key=lambda event: event["created_at"], reverse=True)

    def _user_name(self, user_id: int | None) -> str | None:
        if user_id is None:
            return None
        u = self.users.get_by_id(user_id)
        return u.name if u is not None else None

    def _to_request_out(self, row: PurchaseRequest) -> dict:
        # Mọi con TIỀN lấy từ `purchase_money` — dùng chung với màn Công nợ phải trả, không cộng lại
        # ở đây. Vòng lặp dưới chỉ dựng phần HIỂN THỊ của từng dòng.
        money = purchase_money(row)
        total = money["total"]
        lines = []
        # Mã/tên mặt hàng gốc (mg 0174) cho các dòng có liên kết danh mục — tra 1 lượt, tránh N+1.
        _pairs = [(l.hang_loai, l.hang_id) for l in row.lines if l.hang_loai and l.hang_id]
        _hmap = self.hang.map_theo_cap(_pairs) if (_pairs and self.hang is not None) else {}
        for line in row.lines:
            qty = float(line.quantity)
            unit_price = int(line.expected_unit_price)
            discount_percent = float(line.discount_percent or 0)
            vat_percent = float(line.vat_percent or 0)
            _, discount_amount, vat_amount, line_total = _purchase_line_amounts(
                quantity=qty,
                unit_price=unit_price,
                discount_percent=discount_percent,
                vat_percent=vat_percent,
            )
            lines.append(
                {
                    "id": line.id,
                    "item_name": line.item_name,
                    "unit": line.unit,
                    "quantity": qty,
                    "received_quantity": (
                        None if line.received_quantity is None else float(line.received_quantity)
                    ),
                    "expected_unit_price": unit_price,
                    "discount_percent": discount_percent,
                    "discount_amount": discount_amount,
                    "vat_percent": vat_percent,
                    "vat_amount": vat_amount,
                    "line_total": line_total,
                    "note": line.note,
                    "hang_loai": line.hang_loai,
                    "hang_id": line.hang_id,
                    "hang_ma": getattr(_hmap.get((line.hang_loai, line.hang_id)), "ma", None),
                    "hang_ten": getattr(_hmap.get((line.hang_loai, line.hang_id)), "ten", None),
                }
            )
        sources = []
        for link in row.sources:
            source = link.department_request
            sources.append(
                {
                    "id": link.id,
                    "department_request_id": link.department_request_id,
                    "code": source.code if source is not None else link.source_code_snapshot,
                    "status": source.status if source is not None else None,
                    "source_type": source.source_type if source is not None else None,
                    "purpose": source.purpose if source is not None else None,
                    # Ô GỘP của YCMH — Thu mua lập phiếu mua thì lấy NGUYÊN VĂN chỗ này. Dùng
                    # `purpose` sẽ cụt ở 500 ký tự (cột đó chỉ còn giữ bản cắt cho phiếu cũ).
                    "content": source.content if source is not None else None,
                    "needed_date": source.needed_date if source is not None else None,
                    "requesting_department_name": (
                        source.requesting_department.name
                        if source is not None and source.requesting_department is not None
                        else None
                    ),
                    "requested_by_name": (
                        source.requested_by.name
                        if source is not None and source.requested_by is not None
                        else None
                    ),
                }
            )
        line_by_id = {line.id: line for line in row.lines}
        # Chia luỹ kế MỘT lần cho cả đơn — tiền của đợt phụ thuộc các đợt trước nó.
        chia_dot = phan_bo_du_dot(row)
        _pb, _coc, _du = phan_bo_tien_dot(row)
        phan_bo = {m["delivery"].id: m for m in _pb}
        # Phiếu ĐẶT CỌC đã lập cho đơn này. Form lập phiếu chi cần biết để CẢNH BÁO khi kế toán
        # sắp lập phiếu cọc thứ hai — chủ chốt 06/08/2026: cảnh báo, KHÔNG chặn. Ứng thêm là ca có
        # thật (cọc 30% rồi NCC đòi ứng thêm 20%), và mỗi lần tiền rời két phải là một chứng từ
        # riêng; sửa phiếu cũ lên số to hơn là làm phiếu không khớp lần chi thật.
        coc_phieu = []
        for v in row.payment_vouchers:
            if v.status != PAYMENT_VOUCHER_PAID:
                continue
            if v.payment_stage == "advance":
                coc_phieu.append(
                    {
                        "code": v.code,
                        "doc_no": v.doc_no,
                        "amount": int(v.amount_vnd),
                        "voucher_date": v.voucher_date,
                    }
                )
        # Đợt giao nào ĐÃ sinh yêu cầu NHẬP (chưa hủy) → chặn nhập kho trùng (nút đổi "Đã nhập kho").
        _dot_ids = [d.id for d in getattr(row, "deliveries", []) or []]
        _nhap_map: dict[int, tuple[int, str]] = {}
        if _dot_ids:
            for sr in self.requests.db.execute(
                select(StockRequest.id, StockRequest.ma, StockRequest.purchase_delivery_id).where(
                    StockRequest.purchase_delivery_id.in_(_dot_ids),
                    StockRequest.trang_thai != REQ_CANCELLED,
                )
            ).all():
                _nhap_map[sr.purchase_delivery_id] = (sr.id, sr.ma)
        deliveries = [
            {
                "id": d.id,
                "seq_no": d.seq_no,
                # Liên thông Kho: đợt đã có yêu cầu nhập (chưa hủy) chưa? + trỏ tới yêu cầu đó.
                "da_nhap_kho": d.id in _nhap_map,
                "stock_request_id": _nhap_map.get(d.id, (None, None))[0],
                "stock_request_ma": _nhap_map.get(d.id, (None, None))[1],
                "delivery_date": d.delivery_date,
                "due_date": han_tra_dot(d, row.supplier, row.debt_cutoff_date),
                # NULL = NCC chưa khai số ngày cho nợ ⇒ đợt này không bao giờ vào cột Quá hạn.
                # Cờ này để màn hình đẩy nó lên đầu kèm badge thay vì để nó chìm nghỉm.
                "chua_dat_han": han_tra_dot(d, row.supplier, row.debt_cutoff_date) is None,
                "invoice_number": d.invoice_number,
                "invoice_date": d.invoice_date,
                "note": d.note,
                "amount": chia_dot.get(d.id, {}).get("amount", 0),
                "paid_amount": phan_bo.get(d.id, {}).get("paid", 0),
                # Cọc của cả đơn chiếu xuống đợt này, và phần CÒN NỢ sau khi trừ cả hai.
                # `con_no` chính là TRẦN lập phiếu chi thanh toán cho đợt — form phải bám nó, không
                # được bám công nợ cả đơn (lỗi 07/08/2026: trả thừa cho đợt 2 xoá sổ nợ của đợt 1).
                "coc_bu": phan_bo.get(d.id, {}).get("coc_bu", 0),
                "con_no": phan_bo.get(d.id, {}).get("con_no", 0),
                # Ai ghi đợt này, lúc nào. Đợt giao đẻ ra công nợ nên phải truy được người khai —
                # bảng ĐÃ lưu sẵn hai cột này, thiếu mỗi việc phơi ra.
                "created_by_name": self._user_name(d.created_by_user_id),
                "created_at": d.created_at,
                "lines": [
                    {
                        "id": dl.id,
                        "purchase_request_line_id": dl.purchase_request_line_id,
                        "item_name": (
                            line_by_id[dl.purchase_request_line_id].item_name
                            if dl.purchase_request_line_id in line_by_id
                            else "(dòng đã bị xoá)"
                        ),
                        "unit": (
                            line_by_id[dl.purchase_request_line_id].unit
                            if dl.purchase_request_line_id in line_by_id
                            else ""
                        ),
                        "quantity": float(dl.quantity),
                        # SỐ NHẬN tách làm hai: phần sinh tiền và phần DƯ (0đ). Phơi cả hai chứ
                        # không chỉ phơi tổng — người đọc phải thấy ngay "700 cái, 500 trong đó
                        # không tính tiền", nếu không thì ca NCC có tính tiền phần dư sẽ trôi
                        # lọt tới tận lúc đối chiếu hoá đơn.
                        "quantity_tinh_tien": float(
                            chia_dot.get(d.id, {}).get("lines", {}).get(dl.id, {}).get(
                                "tinh_tien", float(dl.quantity)
                            )
                        ),
                        "quantity_du": float(
                            chia_dot.get(d.id, {}).get("lines", {}).get(dl.id, {}).get("du", 0.0)
                        ),
                        "note": dl.note,
                    }
                    for dl in d.lines
                ],
            }
            for d in row.deliveries
        ]
        attachments = [
            {
                "id": a.id,
                "delivery_id": a.delivery_id,
                "kind": a.kind,
                "file_name": a.file_name,
                "file_url": a.file_url,
                "file_type": a.file_type,
                "uploaded_by_name": self._user_name(a.uploaded_by),
                "uploaded_at": a.uploaded_at,
            }
            for a in getattr(row, "attachments", [])
        ]
        return {
            "id": row.id,
            "code": row.code,
            "status": row.status,
            "supplier_id": row.supplier_id,
            "supplier_name": row.supplier.name if row.supplier else None,
            "purpose": row.purpose,
            "needed_date": row.needed_date,
            "expected_receipt_date": row.expected_receipt_date,
            "debt_cutoff_date": row.debt_cutoff_date,
            # Số ngày cho nợ của NCC đi kèm để màn Thu mua tự suy được hạn trả ngay tại ô gõ
            # ngày chốt ("chốt 31/8 + 30 ngày ⇒ 30/9"). Không có nó thì người khai gõ xong một
            # cái ngày rồi phải sang màn khác mới biết hậu quả.
            "supplier_credit_days": (
                getattr(row.supplier, "credit_days", None) if row.supplier is not None else None
            ),
            "created_by_user_id": row.created_by_user_id,
            "created_by_name": self._user_name(row.created_by_user_id),
            "submitted_at": row.submitted_at,
            "approved_by_user_id": row.approved_by_user_id,
            "approved_by_name": self._user_name(row.approved_by_user_id),
            "approved_at": row.approved_at,
            "note": row.note,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
            "content": row.content or row.purpose,
            "reject_reason": row.reject_reason,
            "status_history": self._lich_su_out(DOC_PMH, row.id),
            "activity_history": self._lich_su_don_mua_out(row),
            "contract_number": row.contract_number,
            "deposit_expected": int(row.deposit_expected or 0),
            "total_estimate": total,
            # Giá trị hàng thực nhận TÍNH THEO ĐƠN GIÁ. Đây KHÔNG phải số đẻ ra công nợ kể từ
            # 06/08/2026 — công nợ bám `gia_tri_da_giao` (tổng tiền hoá đơn các đợt). Giữ lại vì
            # đơn KHÔNG theo dõi đợt giao vẫn dùng nó, và vì nó là số để đối chiếu với hoá đơn.
            "received_total": money["received_total"],
            # Giá trị hàng ĐÃ VỀ — số đẻ ra công nợ. Khác `received_total` ở chỗ nó chỉ đếm phần
            # thực sự đã giao; đơn chưa giao đợt nào thì bằng 0 dù đơn to bao nhiêu.
            "gia_tri_da_giao": money["gia_tri_da_giao"],
            "paid_amount": money["paid_amount"],
            "receipt_received_amount": money["receipt_received_amount"],
            "net_paid": money["net_paid"],
            # = CÔNG NỢ của phiếu, và cũng là trần lập phiếu chi THANH TOÁN.
            "outstanding_amount": money["outstanding_amount"],
            # Trần lập phiếu ĐẶT CỌC — tính theo giá trị đơn đặt vì cọc là chi khi hàng chưa về.
            "tran_dat_coc": money["tran_dat_coc"],
            # Các phiếu ĐẶT CỌC đã lập (đã chi). Form phiếu chi dùng để cảnh báo khi sắp lập
            # phiếu cọc thứ hai — cảnh báo, không chặn.
            "coc_da_lap": coc_phieu,
            "coc_da_chi": sum(c["amount"] for c in coc_phieu),
            "payment_status": money["payment_status"],
            "payment_voucher_count": len(row.payment_vouchers),
            "sources": sources,
            "lines": lines,
            "deliveries": deliveries,
            "attachments": attachments,
        }

    def _tinh_trang_tung_dong(self, row: DepartmentPurchaseRequest) -> dict[int, dict]:
        """Với mỗi DÒNG của yêu cầu: nó đã vào phiếu nào, của NCC nào, tới đâu rồi.

        Trước 05/08/2026 chi tiết yêu cầu KHÔNG hề nhắc tới phiếu mua nào — bộ phận vào xem cũng
        không biết yêu cầu của mình đã thành phiếu nào, ai bán, tới đâu.

        Một dòng có thể đi qua NHIỀU phiếu (bị từ chối rồi thu mua lập lại) ⇒ lấy phiếu có tiến độ
        CAO NHẤT, không phải phiếu mới nhất: phiếu lập lại đã về hàng thì dòng đó xong, dù phiếu cũ
        vẫn nằm đó ở trạng thái bị từ chối.

        Chỉ đọc qua `department_request_line_id`. KHÔNG ghép bù bằng tên hàng cho dữ liệu cũ — thà
        để trống và nói "chưa rõ" còn hơn đoán sai mà trông như thật."""
        theo_dong: dict[int, dict] = {}
        for link in getattr(row, "purchase_links", []):
            phieu = link.purchase_request
            if phieu is None:
                continue
            # Σ theo đợt giao — tính MỘT LẦN cho mỗi phiếu, không lặp trong vòng dòng.
            da_giao = da_giao_theo_dong(phieu)
            for pl in phieu.lines:
                src_line_id = getattr(pl, "department_request_line_id", None)
                if src_line_id is None:
                    continue
                bac = _BAC_PHIEU.get(phieu.status, 0)
                cu = theo_dong.get(src_line_id)
                if cu is not None and cu["_bac"] >= bac:
                    continue
                theo_dong[src_line_id] = {
                    "_bac": bac,
                    "purchase_request_id": phieu.id,
                    "purchase_code": phieu.code,
                    "purchase_status": phieu.status,
                    "supplier_name": phieu.supplier.name if phieu.supplier else None,
                    "ordered_quantity": float(pl.quantity),
                    "ordered_unit": pl.unit,
                    # ĐI QUA `qty_thuc_nhan`, đừng đọc thẳng `pl.received_quantity`: cột đó
                    # DORMANT với mọi phiếu có đợt giao (từ 06/08/2026). Đọc thẳng là chi tiết
                    # yêu cầu báo "chưa nhận gì" trong khi hàng đã về mấy đợt — bộ phận tưởng thu
                    # mua chưa làm, gọi điện giục nhầm.
                    #
                    # `None` chỉ còn đúng MỘT nghĩa: phiếu CHƯA có đợt giao nào VÀ chưa ai khai số
                    # thực nhận ⇒ giao diện hiểu là "chưa có tin", không phải "nhận 0".
                    "received_quantity": (
                        qty_thuc_nhan(pl, da_giao)
                        if (da_giao is not None or pl.received_quantity is not None)
                        else None
                    ),
                }
        for muc in theo_dong.values():
            muc.pop("_bac", None)
        return theo_dong

    def _to_department_request_out(self, row: DepartmentPurchaseRequest) -> dict:
        total = 0
        lines = []
        tinh_trang = self._tinh_trang_tung_dong(row)
        dang_giu = self._phieu_dang_giu_dong(row)
        so_dong_huy = 0
        for line in row.lines:
            qty = float(line.quantity)
            unit_price = int(line.expected_unit_price)
            line_total = int(round(qty * unit_price))
            da_huy = line.cancelled_at is not None
            if da_huy:
                so_dong_huy += 1
            else:
                # Món đã bỏ KHÔNG cộng vào tiền dự kiến — nó không còn là tiền sắp chi.
                total += line_total
            phieu_giu = None if da_huy else dang_giu.get(line.id)
            lines.append(
                {
                    "id": line.id,
                    "hang_loai": line.hang_loai,
                    "hang_id": line.hang_id,
                    "item_name": line.item_name,
                    "unit": line.unit,
                    "quantity": qty,
                    "expected_unit_price": unit_price,
                    "line_total": line_total,
                    "note": line.note,
                    # None = chưa vào phiếu nào, HOẶC phiếu lập trước 05/08/2026 (chưa có nối
                    # dòng ↔ dòng). Giao diện phải nói rõ hai ca đó, đừng hiện như nhau.
                    "fulfilment": tinh_trang.get(line.id),
                    "cancelled_at": line.cancelled_at,
                    "cancelled_by_name": self._user_name(line.cancelled_by_user_id),
                    "cancel_reason": line.cancel_reason,
                    # Luật "món này bỏ được không" thuộc về MÁY CHỦ. Trả sẵn cả câu lý do để giao
                    # diện KHOÁ nút kèm lời giải thích (chủ chốt 20/08/2026: "đừng ẩn nút — khoá
                    # và nói lý do") mà không phải chép lại luật ở FE rồi lệch nhau.
                    # Chỉ nói về TÌNH TRẠNG MÓN; quyền của người đang xem thì FE tự AND thêm.
                    "can_cancel": not da_huy and phieu_giu is None,
                    "cancel_block_reason": (
                        self._cau_chan_huy_dong(phieu_giu) if phieu_giu is not None else None
                    ),
                }
            )
        so_dong_song = len(row.lines) - so_dong_huy
        phieu_con = [
            link.purchase_request
            for link in getattr(row, "purchase_links", [])
            if link.purchase_request is not None
        ]
        # Ưu tiên trạng thái cần người dùng hành động. `status` vẫn giữ vai trò khóa luồng; trường
        # này chỉ giúp bảng nói đúng việc Thu mua phải làm tiếp theo.
        if row.status in (DPR_OPEN, DPR_IN_PURCHASE, DPR_DONE, DPR_CANCELLED):
            progress_status = row.status
        elif any(p.status == PR_REJECTED for p in phieu_con):
            progress_status = "needs_correction"
        elif any(p.status == PR_DRAFT for p in phieu_con):
            progress_status = "drafting"
        else:
            progress_status = row.status
        # HUỶ MỘT PHẦN đè lên nhãn tiến độ (chủ chốt 24/08/2026: có món bị bỏ là phải thấy ngay).
        # Tiến độ phần còn lại KHÔNG mất — nó ở `progress_status`, giao diện in thành dòng chữ nhỏ
        # dưới huy hiệu. Huỷ HẾT món thì `status` đã là `cancelled`, rơi vào nhánh trên.
        workflow_status = (
            "partially_cancelled" if (so_dong_huy and so_dong_song) else progress_status
        )

        return {
            "id": row.id,
            "code": row.code,
            "status": row.status,
            "workflow_status": workflow_status,
            "progress_status": progress_status,
            "cancelled_line_count": so_dong_huy,
            "active_line_count": so_dong_song,
            "source_type": row.source_type,
            "requesting_department_id": row.requesting_department_id,
            "requesting_department_name": (
                row.requesting_department.name if row.requesting_department is not None else None
            ),
            "requested_by_user_id": row.requested_by_user_id,
            "requested_by_name": row.requested_by.name if row.requested_by is not None else None,
            "related_document_type": row.related_document_type,
            "related_document_code": row.related_document_code,
            "purpose": row.purpose,
            # Ô GỘP. `purpose` giữ lại cho client cũ; cái đi vào giao diện mới là `content`.
            "content": row.content or row.purpose,
            "reject_reason": row.reject_reason,
            "status_history": self._lich_su_out(DOC_YCMH, row.id),
            "needed_date": row.needed_date,
            "note": row.note,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
            "total_estimate": total,
            "lines": lines,
            # Các phiếu mua sinh ra từ yêu cầu này. Vẫn cần dù đã có `fulfilment` theo dòng: yêu
            # cầu lập trước 05/08/2026 không có nối dòng ↔ dòng nên `fulfilment` rỗng, còn đây thì
            # luôn có — ít nhất bộ phận biết yêu cầu của mình đã thành phiếu nào, ai bán.
            "purchase_requests": sorted(
                (
                    {
                        "id": link.purchase_request.id,
                        "code": link.purchase_request.code,
                        "status": link.purchase_request.status,
                        "supplier_name": (
                            link.purchase_request.supplier.name
                            if link.purchase_request.supplier
                            else None
                        ),
                    }
                    for link in getattr(row, "purchase_links", [])
                    if link.purchase_request is not None
                ),
                key=lambda p: p["code"],
            ),
        }
