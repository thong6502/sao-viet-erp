"""Xuất SỔ TỔNG HỢP CÔNG NỢ ra .xlsx — ĐÚNG KHUÔN MISA (docs/prd-bao-cao-cong-no.md §5.5).

Chốt 03/09/2026: file xuất phải **dán thẳng vào bộ hồ sơ kế toán đang dùng**, không phải sửa tay.
Vì thế mọi con số dưới đây bóc từ chính bản xuất của họ (`TONG_HOP_CONG_NO_PHAI_THU.xlsx`), không
phải tự đặt cho đẹp — đổi một cái là file lệch khuôn và mất luôn lý do tồn tại.

KHÔNG đính file mẫu của họ vào repo làm template: file đó chứa tên + công nợ của 202 khách hàng
thật. Dựng lại từ 0 bằng spec ở đây.

Khuôn:

    A1:I1  TỔNG HỢP CÔNG NỢ PHẢI THU          Times New Roman 14 đậm, canh giữa
    A2:I2  Tài khoản: 131; Loại tiền: …       Times New Roman 11 đậm, canh giữa
    A3:A4  Mã KH   B3:B4 Tên KH   C3:C4 TK    ┐
    D3:E3  Số dư đầu kỳ                        ├ Microsoft Sans Serif 8, canh giữa, có viền
    F3:G3  Số phát sinh   H3:I3 Số dư cuối kỳ ┘
    dòng 4: Nợ | Có × 3 cụm
    dòng 5+: dữ liệu
    dòng cuối: "Số dòng = N" + tổng 6 cột tiền
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: Định dạng số của MISA: số âm ĐỎ, trong ngoặc. Báo cáo mình không bao giờ ra số âm (âm thì nhảy
#: sang cột bên kia — luật sổ Nợ/Có) nên nhánh đỏ không bao giờ kích hoạt. Vẫn giữ cho khớp khuôn.
DINH_DANG_TIEN = "#,##0_);[Red](#,##0)"

#: Cột A-C theo mẫu MISA; D–I (6 cột tiền) đặt 17.1 (bằng cột D) để đủ chỗ cho số tiền lớn, tránh bị tràn thành '########'.
RONG_COT = {
    "A": 17.1,
    "B": 30.0,
    "C": 14.3,
    "D": 17.1,
    "E": 17.1,
    "F": 17.1,
    "G": 17.1,
    "H": 17.1,
    "I": 17.1,
}

_FONT_TIEU_DE = ("Times New Roman", 14, True)
_FONT_KY = ("Times New Roman", 11, True)
_FONT_BANG = ("Microsoft Sans Serif", 8, False)

_COT_TIEN = ("dau_no", "dau_co", "ps_no", "ps_co", "cuoi_no", "cuoi_co")


def _ngay_vn(d: date) -> str:
    """`dd/mm/yyyy`, ĐỀU TAY.

    Bản của MISA ghi `Từ ngày 01/01/2026 đến ngày 29/8/2026` — `01/01` có số 0 đằng trước mà
    `29/8` thì không. Đó là lỗi vặt của họ, cố ý KHÔNG bắt chước.
    """
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def xuat_xlsx(bao_cao: dict) -> bytes:
    """`bao_cao` = kết quả `tong_hop_phai_thu` / `tong_hop_phai_tra` → nội dung file .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"

    def _font(spec) -> Font:
        ten, co, dam = spec
        return Font(name=ten, size=co, bold=dam)

    mong = Side(style="thin")
    vien = Border(left=mong, right=mong, top=mong, bottom=mong)
    giua = Alignment(horizontal="center", vertical="center", wrap_text=True)
    trai = Alignment(horizontal="left", vertical="center")
    phai = Alignment(horizontal="right", vertical="center")

    # --- dòng 1-2: tiêu đề + kỳ ---
    ws.merge_cells("A1:I1")
    ws["A1"] = bao_cao["tieu_de"]
    ws["A1"].font = _font(_FONT_TIEU_DE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = (
        f"Tài khoản: {bao_cao['tk']}; Loại tiền: Tổng hợp; "
        f"Từ ngày {_ngay_vn(bao_cao['tu_ngay'])} đến ngày {_ngay_vn(bao_cao['den_ngay'])}"
    )
    ws["A2"].font = _font(_FONT_KY)
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- dòng 3-4: tiêu đề bảng hai tầng ---
    for o, chu in (
        ("A3", bao_cao["nhan_ma"]),
        ("B3", bao_cao["nhan_ten"]),
        ("C3", "TK công nợ"),
        ("D3", "Số dư đầu kỳ"),
        ("F3", "Số phát sinh"),
        ("H3", "Số dư cuối kỳ"),
    ):
        ws[o] = chu
    for o in ("A3:A4", "B3:B4", "C3:C4", "D3:E3", "F3:G3", "H3:I3"):
        ws.merge_cells(o)
    for cot, chu in zip("DEFGHI", ("Nợ", "Có") * 3):
        ws[f"{cot}4"] = chu
    for hang in (3, 4):
        for cot in "ABCDEFGHI":
            o = ws[f"{cot}{hang}"]
            o.font = _font(_FONT_BANG)
            o.alignment = giua
            o.border = vien

    # --- dòng 5+: dữ liệu ---
    hang = 5
    for d in bao_cao["items"]:
        ws[f"A{hang}"] = d.get("ma") or ""
        ws[f"B{hang}"] = d.get("ten") or ""
        ws[f"C{hang}"] = ""  # Để trống cột TK công nợ theo yêu cầu
        for cot, khoa in zip("DEFGHI", _COT_TIEN):
            ws[f"{cot}{hang}"] = int(d.get(khoa) or 0)
        _to_hang(ws, hang, _font(_FONT_BANG), vien, trai, phai)
        hang += 1

    # --- dòng chân: "Số dòng = N" + tổng ---
    tong = bao_cao["tong"]
    ws[f"A{hang}"] = f"Số dòng = {tong.get('so_dong', len(bao_cao['items']))}"
    for cot, khoa in zip("DEFGHI", _COT_TIEN):
        ws[f"{cot}{hang}"] = int(tong.get(khoa) or 0)
    _to_hang(ws, hang, _font(_FONT_BANG), vien, trai, phai)

    for cot, rong in RONG_COT.items():
        ws.column_dimensions[cot].width = rong
    # Khoá hai dòng tiêu đề bảng: sổ 275 dòng mà cuộn xuống mất tiêu đề thì không biết cột nào là
    # Nợ cột nào là Có. Bản của MISA không có, nhưng đây là thứ chỉ giúp người đọc chứ không đổi
    # khuôn — mở file bằng Excel vẫn ra đúng bảng đó.
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_hang(ws, hang: int, font, vien, trai, phai) -> None:
    """Một hàng dữ liệu: chữ canh trái, tiền canh phải + định dạng số, cả hàng có viền."""
    for cot in "ABCDEFGHI":
        o = ws[f"{cot}{hang}"]
        o.font = font
        o.border = vien
        if cot in "ABC":
            o.alignment = trai
        else:
            o.alignment = phai
            o.number_format = DINH_DANG_TIEN


def ten_file(bao_cao: dict) -> str:
    """`tong-hop-cong-no-phai-thu-2026-01-01-den-2026-08-29.xlsx`.

    Có KỲ trong tên: kế toán tải mỗi tháng một lần, mà `bao-cao.xlsx` thì tải ba lần là ba file
    `bao-cao (1).xlsx` không phân biệt nổi.
    """
    ben = "phai-thu" if bao_cao["tk"] == "131" else "phai-tra"
    return (
        f"tong-hop-cong-no-{ben}-{bao_cao['tu_ngay']:%Y-%m-%d}"
        f"-den-{bao_cao['den_ngay']:%Y-%m-%d}.xlsx"
    )
