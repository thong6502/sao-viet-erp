"""Router — Phiếu nhập/xuất kho + lô + ngưỡng tồn (spec-kho-de-nghi §5–§7, §9).

Điểm phân quyền quan trọng: mọi số TIỀN (`don_gia`, `thanh_tien`, `gia_von`,
`don_gia_nhap`) chỉ được điền khi người gọi có `can_view_cost`. Thiếu quyền thì không
những ẩn cột mà còn KHÔNG TÍNH — số không lọt ra qua response, kể cả khi mở DevTools.
Bản in 01-VT/02-VT dùng chính response này nên tự động bỏ 2 cột tiền.
"""
from __future__ import annotations

import json
from typing import Annotated

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps import get_authorization_service, require_permission
from ..models.stock_voucher import VOUCHER_NHAP
from ..models.user import User
from ..repositories.audit_repo import AuditLogRepository
from ..repositories.document_sequence_repo import DocumentSequenceRepository
from ..repositories.kho_hang_repo import KhoHangRepository
from ..repositories.stock_lot_repo import StockLotRepository, StockThresholdRepository
from ..repositories.don_vi_do_repo import DonViDoRepository
from ..repositories.stock_request_repo import StockRequestRepository
from ..repositories.vat_lieu_kho_repo import VatLieuKhoRepository
from ..repositories.stock_voucher_repo import StockVoucherRepository
from ..repositories.rbac_repo import DepartmentRepository
from ..repositories.user_repo import UserRepository
from ..schemas.stock import (
    AllocationLineOut,
    AllocationOut,
    DieuChinhLichSuRow,
    DieuChinhXuatIn,
    DieuChuyenIn,
    DieuChuyenOut,
    MaterialHistoryOut,
    MaterialXuatRow,
    StockLotOut,
    StockLotViTriIn,
    StockThresholdIn,
    StockThresholdOut,
    StockVoucherAttachmentListOut,
    StockVoucherAttachmentOut,
    StockVoucherCancel,
    StockVoucherViTriIn,
    StockVoucherCreate,
    StockVoucherLineOut,
    StockVoucherOut,
    StockVoucherPage,
)
from ..services.qr_token import sign_scan
from ..services.vat_lieu_kho_service import HANG_NHAN, VatLieuKhoService
from ..services.rbac_service import AuthorizationService
from ..services.sequence_service import SequenceService
from ..services.stock_request_service import StockRequestService
from ..services.stock_voucher_service import StockVoucherError, StockVoucherService

from ..services.delivery_notify import bao_tai_xe_kho_lap_phieu

router = APIRouter(prefix="/api/kho/phieu", tags=["kho-phieu"])
# Ngưỡng tồn để PREFIX RIÊNG, không nhét dưới /phieu: `/phieu/nguong` là path 1 đoạn nên sẽ
# bị `/phieu/{voucher_id}` nuốt trước (FastAPI khớp theo thứ tự khai báo) → 422.
threshold_router = APIRouter(prefix="/api/kho/nguong-ton", tags=["kho-nguong"])
# Điều chuyển kho — prefix RIÊNG (không nhét dưới /phieu) để không bị `/phieu/{voucher_id}` nuốt.
dieu_chuyen_router = APIRouter(prefix="/api/kho/dieu-chuyen", tags=["kho-dieu-chuyen"])
MODULE = "kho"
# Action nhật ký khi ĐIỀU CHỈNH phiếu xuất (SX dùng ít hơn) — đọc lại cho "Lịch sử điều chỉnh".
_ACTION_DIEU_CHINH_XUAT = "kho_xuat_dieu_chinh"


def _hang_service(db: Session) -> VatLieuKhoService:
    return VatLieuKhoService(VatLieuKhoRepository(db), DonViDoRepository(db))


def get_service(db: Annotated[Session, Depends(get_db)]) -> StockVoucherService:
    sequence = SequenceService(DocumentSequenceRepository(db))
    requests = StockRequestRepository(db)
    lots = StockLotRepository(db)
    hang = _hang_service(db)
    request_service = StockRequestService(
        requests, lots, StockThresholdRepository(db), sequence, hang=hang)
    return StockVoucherService(
        StockVoucherRepository(db), requests, lots, sequence, request_service, hang,
        giu_cho=_giu_cho_service(db),
    )


def _giu_cho_service(db: Session):
    """`GiuChoService` cho đường GHI SỔ — kho phải biết phần nào đã có chủ.

    Dựng TẠI ĐÂY chứ không nhét vào `deps.py`: cùng lối `get_service` của router này, và giữ cho
    `StockVoucherService` không phải biết cách lắp cả bảng cân đối (nó nhận sẵn qua tham số, vắng
    thì chạy như cũ).
    """
    from ..repositories.bai_ghep_repo import BaiGhepRepository
    from ..repositories.don_vi_do_repo import DonViDoRepository
    from ..repositories.lsx_repo import LsxRepository
    from ..repositories.purchase_repo import PurchaseRequestRepository, SupplierRepository
    from ..services.giu_cho_service import GiuChoService
    from ..services.ke_hoach_vat_tu_service import KeHoachVatTuService

    kh = KeHoachVatTuService(
        db, lsx_repo=LsxRepository(db), bai_ghep_repo=BaiGhepRepository(db),
        hang=_hang_service(db), lots=StockLotRepository(db),
        requests=StockRequestRepository(db), purchases=PurchaseRequestRepository(db),
        suppliers=SupplierRepository(db), don_vi=DonViDoRepository(db),
    )
    return GiuChoService(db, kh)


Service = Annotated[StockVoucherService, Depends(get_service)]
Db = Annotated[Session, Depends(get_db)]
Authz = Annotated[AuthorizationService, Depends(get_authorization_service)]


def _err(e: StockVoucherError) -> HTTPException:
    return HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


def _can_see_cost(v, user, authz, db) -> bool:
    """Ai thấy GIÁ VỐN của phiếu: người có quyền `view_cost`, HOẶC chính NGƯỜI TẠO yêu cầu gốc
    (chủ 10/08/2026: người tạo xem được phiếu sinh từ yêu cầu của họ — không ẩn giá)."""
    if authz.can(user, MODULE, "view_cost"):
        return True
    req = StockRequestRepository(db).get(v.request_id) if getattr(v, "request_id", None) else None
    return req is not None and req.nguoi_tao_id == user.id


def _serialize(v, *, svc: StockVoucherService, db: Session, can_view_cost: bool,
               hang_map: dict | None = None, lot_map: dict | None = None,
               req_map: dict | None = None) -> StockVoucherOut:
    """`hang_map`/`lot_map`/`req_map` dựng SẴN theo cả trang (list) để tránh N+1; gọi lẻ 1 phiếu
    thì để None, hàm tự nạp mỗi map 1 query cho các dòng của phiếu đó."""
    users = UserRepository(db)
    khos = KhoHangRepository(db)
    requests = StockRequestRepository(db)
    if req_map is None:
        req_map = requests.by_ids_with_lines([v.request_id])
    if hang_map is None:
        hang_map = svc.hang.map_theo_cap([(ln.hang_loai, ln.hang_id) for ln in v.lines])
    if lot_map is None:
        lot_map = svc.lots.by_ids([ln.lot_id for ln in v.lines])

    req = req_map.get(v.request_id)
    req_lines = req.lines if req is not None else []
    line_dvt = {ln.id: ln.dvt for ln in req_lines}
    # SL yêu cầu theo dòng yêu cầu (request_line_id) — nối vào mỗi dòng phiếu để đối chiếu
    # "yêu cầu vs thực nhận/xuất". Đọc-nối, không lưu cột.
    line_sl_de_nghi = {ln.id: float(ln.sl_de_nghi) for ln in req_lines}
    goc_map = svc.hang.don_vi_goc_map([(ln.hang_loai, ln.hang_id) for ln in v.lines])

    lines: list[StockVoucherLineOut] = []
    gia_von_total = 0
    if v.loai == VOUCHER_NHAP:
        # NHẬP: mỗi dòng = 1 lô sắp tạo, giá lấy TRÊN DÒNG → giữ nguyên 1 dòng/lô. Đơn giá và
        # `so_luong` cùng ở đơn vị NGƯỜI KHAI nên nhân thẳng với nhau.
        for ln in v.lines:
            key = (ln.hang_loai, ln.hang_id)
            m = hang_map.get(key)
            lot = lot_map.get(ln.lot_id) if ln.lot_id else None
            unit = int(ln.don_gia or 0)
            thanh_tien = int(round(unit * float(ln.so_luong)))
            gia_von_total += thanh_tien
            lines.append(StockVoucherLineOut(
                id=ln.id,
                request_line_id=ln.request_line_id,
                hang_loai=ln.hang_loai,
                hang_id=ln.hang_id,
                hang_ma=getattr(m, "ma", None),
                hang_ten=getattr(m, "ten", None),
                dvt=line_dvt.get(ln.request_line_id),
                lot_id=ln.lot_id,
                ma_lo=getattr(lot, "ma_lo", None),
                sl_de_nghi=line_sl_de_nghi.get(ln.request_line_id),
                so_luong=float(ln.so_luong),
                sl_goc=float(ln.sl_goc),
                don_vi_goc=goc_map.get(key),
                ghi_chu=ln.ghi_chu,
                hsd=ln.hsd,   # HSD đích danh theo lô (phiếu điều chuyển hiện được); không phải tiền → không ẩn
                vi_tri=ln.vi_tri,   # vị trí cất lô — phiếu điều chuyển khai/hiện per-lô
                don_gia=unit if can_view_cost else None,
                thanh_tien=thanh_tien if can_view_cost else None,
            ))
    else:
        # XUẤT: đích danh trừ lô per-lô GIỮ NGUYÊN ở DB; chỉ ở tầng ĐỌC gộp các dòng lô lẻ theo
        # `request_line_id` (giữ thứ tự xuất hiện) → 1 dòng/mặt hàng với ĐƠN GIÁ BÌNH QUÂN gia
        # quyền. `lot_id`/`ma_lo` = None vì đã gộp nhiều lô.
        #
        # TIỀN TÍNH THEO `sl_goc`, KHÔNG theo `so_luong`: giá lô (`don_gia_nhap`) là giá trên ĐƠN
        # VỊ GỐC, còn `so_luong` là đơn vị người khai. Nhân nhầm thì lệch đúng bằng hệ số quy đổi
        # (khai 10 ram sẽ ra tiền của 10 kg) mà không có gì báo.
        groups: dict[int, list] = {}
        order: list[int] = []
        for ln in v.lines:
            k = ln.request_line_id
            if k not in groups:
                groups[k] = []
                order.append(k)
            groups[k].append(ln)
        for k in order:
            grp = groups[k]
            first = grp[0]
            key = (first.hang_loai, first.hang_id)
            m = hang_map.get(key)
            qty = sum(float(ln.so_luong) for ln in grp)          # hiển thị: đơn vị người khai
            qty_goc = sum(float(ln.sl_goc) for ln in grp)        # tiền: đơn vị gốc
            raw_cost = sum(
                int(round(
                    int(getattr(lot_map.get(ln.lot_id), "don_gia_nhap", 0) or 0) * float(ln.sl_goc)
                ))
                for ln in grp
            )
            blended = int(round(raw_cost / qty_goc)) if qty_goc > 0 else 0
            thanh_tien = int(round(blended * qty_goc))  # tổng thành tiền QUA đơn giá bình quân
            gia_von_total += thanh_tien
            ghi_chu = next((ln.ghi_chu for ln in grp if (ln.ghi_chu or "").strip()), None)
            lines.append(StockVoucherLineOut(
                id=first.id,
                request_line_id=k,
                hang_loai=first.hang_loai,
                hang_id=first.hang_id,
                hang_ma=getattr(m, "ma", None),
                hang_ten=getattr(m, "ten", None),
                dvt=line_dvt.get(k),
                lot_id=None,
                ma_lo=None,
                sl_de_nghi=line_sl_de_nghi.get(k),
                so_luong=qty,
                sl_goc=qty_goc,
                don_vi_goc=goc_map.get(key),
                ghi_chu=ghi_chu,
                don_gia=blended if can_view_cost else None,
                thanh_tien=thanh_tien if can_view_cost else None,
            ))
    kho = khos.get(v.kho_id)
    lap = users.get_by_id(v.nguoi_lap_id) if v.nguoi_lap_id else None
    ghi_so_u = users.get_by_id(v.nguoi_ghi_so_id) if getattr(v, "nguoi_ghi_so_id", None) else None
    # Ai yêu cầu / ai duyệt lấy từ yêu cầu gốc (phiếu ứng theo yêu cầu đã duyệt).
    de_nghi_u = users.get_by_id(req.nguoi_tao_id) if req and req.nguoi_tao_id else None
    duyet_u = (
        users.get_by_id(req.nguoi_duyet_id)
        if req and getattr(req, "nguoi_duyet_id", None)
        else None
    )
    return StockVoucherOut(
        id=v.id, ma=v.ma, loai=v.loai,
        request_id=v.request_id, request_ma=getattr(req, "ma", None),
        loai_kho=getattr(req, "loai_kho", None),
        kho_id=v.kho_id, kho_ten=getattr(kho, "ten", None),
        ngay=v.ngay, nguoi_lap_id=v.nguoi_lap_id, nguoi_lap_ten=getattr(lap, "name", None),
        nguoi_de_nghi_ten=getattr(de_nghi_u, "name", None),
        nguoi_duyet_ten=getattr(duyet_u, "name", None),
        nguoi_ghi_so_ten=getattr(ghi_so_u, "name", None),
        nguoi_giao_nhan=v.nguoi_giao_nhan, ghi_chu=v.ghi_chu,
        trang_thai=v.trang_thai, ghi_so_luc=v.ghi_so_luc, created_at=v.created_at,
        dieu_chuyen=bool(getattr(v, "dieu_chuyen", False)),
        lines=lines,
        gia_von=gia_von_total if can_view_cost else None,
    )


@router.get("", response_model=StockVoucherPage)
def list_vouchers(
    svc: Service, db: Db, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
    loai: str | None = Query(default=None),
    trang_thai: str | None = Query(default=None),
    request_id: int | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    q: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=200),
) -> StockVoucherPage:
    rows, total = svc.vouchers.list(
        loai=loai, trang_thai=trang_thai, request_id=request_id,
        kho_id=kho_id, q=q, page=page, size=size,
    )
    can_view_cost = authz.can(user, MODULE, "view_cost")
    # Người TẠO yêu cầu lọc phiếu theo yêu cầu của MÌNH → cho thấy giá (không ẩn), như GET phiếu.
    if not can_view_cost and request_id is not None:
        req0 = StockRequestRepository(db).get(request_id)
        if req0 is not None and req0.nguoi_tao_id == user.id:
            can_view_cost = True
    # Nạp SẴN mã hàng / lô / đề nghị của cả trang trong vài query (tránh N+1 trong _serialize).
    hang_map = _hang_service(db).map_theo_cap(
        [(ln.hang_loai, ln.hang_id) for v in rows for ln in v.lines])
    lot_map = svc.lots.by_ids([ln.lot_id for v in rows for ln in v.lines])
    req_map = StockRequestRepository(db).by_ids_with_lines([v.request_id for v in rows])
    return StockVoucherPage(
        items=[
            _serialize(v, svc=svc, db=db, can_view_cost=can_view_cost,
                       hang_map=hang_map, lot_map=lot_map, req_map=req_map)
            for v in rows
        ],
        total=total,
    )


@router.get("/{voucher_id}", response_model=StockVoucherOut)
def get_voucher(
    voucher_id: int, svc: Service, db: Db, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> StockVoucherOut:
    v = svc.vouchers.get_with_lines(voucher_id)
    if v is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy phiếu")
    return _serialize(v, svc=svc, db=db, can_view_cost=_can_see_cost(v, user, authz, db))


@router.post("", response_model=StockVoucherOut, status_code=status.HTTP_201_CREATED)
def create_voucher(
    payload: StockVoucherCreate, svc: Service, db: Db, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> StockVoucherOut:
    try:
        v = svc.create(
            user=user, request_id=payload.request_id, kho_id=payload.kho_id, ma=payload.ma,
            lines=[ln.model_dump() for ln in payload.lines],
            ngay=payload.ngay, nguoi_giao_nhan=payload.nguoi_giao_nhan, ghi_chu=payload.ghi_chu,
        )
    except StockVoucherError as e:
        raise _err(e) from None
    # Yêu cầu của GIAO HÀNG thì báo tài xế "kho soạn xong, tới lấy được" — mốc đó chỉ kho biết.
    # Toàn bộ logic ở `services/delivery_notify`; hàm đó tự nuốt lỗi nên không cần `try` ở đây và
    # không đường nào làm hỏng việc lập phiếu.
    bao_tai_xe_kho_lap_phieu(db, payload.request_id, getattr(v, "ma", None))
    return _serialize(v, svc=svc, db=db, can_view_cost=authz.can(user, MODULE, "view_cost"))


@router.post("/{voucher_id}/ghi-so", response_model=StockVoucherOut)
def post_voucher(
    voucher_id: int, svc: Service, db: Db, authz: Authz,
    # ĐÃ GỘP quyền (bỏ SoD): người có quyền lập phiếu (create) tự ghi sổ luôn — không tách "thủ kho
    # lập" & "QL/kế toán chốt sổ" nữa (yêu cầu vận hành: tạo phiếu + ghi sổ cùng 1 quyền).
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> StockVoucherOut:
    """Ghi sổ — điểm DUY NHẤT tồn kho thay đổi."""
    try:
        v = svc.post(voucher_id, user)
    except StockVoucherError as e:
        raise _err(e) from None
    return _serialize(v, svc=svc, db=db, can_view_cost=authz.can(user, MODULE, "view_cost"))


@router.post("/{voucher_id}/huy", response_model=StockVoucherOut)
def cancel_voucher(
    voucher_id: int, payload: StockVoucherCancel, svc: Service, db: Db, authz: Authz,
    # ĐÃ GỘP quyền: hủy phiếu = cùng quyền lập phiếu (create), không tách người lập & người chốt sổ.
    # Chỉ hủy được phiếu CHƯA ghi sổ. BẮT BUỘC lý do → yêu cầu chuyển 'Đã hủy' kèm lý do (kết thúc).
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> StockVoucherOut:
    try:
        v = svc.cancel(voucher_id, ly_do=payload.ly_do)
    except StockVoucherError as e:
        raise _err(e) from None
    return _serialize(v, svc=svc, db=db, can_view_cost=authz.can(user, MODULE, "view_cost"))


@router.post("/{voucher_id}/dieu-chinh-xuat", response_model=StockVoucherOut)
def dieu_chinh_xuat(
    voucher_id: int, payload: DieuChinhXuatIn, svc: Service, db: Db, authz: Authz,
    # Điều chỉnh phiếu XUẤT đã ghi sổ khi SX dùng ÍT hơn (xuất 10 → 7): trả dư về lô, giảm 'đã cấp'
    # của yêu cầu. Cùng quyền vận hành phiếu (create) — thủ kho là người biết SX dùng bao nhiêu.
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> StockVoucherOut:
    """Sửa giảm số đã xuất của phiếu XUẤT đã ghi sổ (SX dùng không hết) — trả phần dư về lô nguồn."""
    try:
        v, changes = svc.dieu_chinh_xuat(
            voucher_id, {ln.line_id: ln.so_luong_moi for ln in payload.lines}, user)
    except StockVoucherError as e:
        raise _err(e) from None
    # Điều chỉnh phiếu ĐÃ ghi sổ = thao tác nhạy cảm (đụng sổ) → ghi nhật ký để truy vết (ai · lúc nào
    # · đổi gì · VÌ SAO). detail = JSON {noi_dung, ly_do} để lịch sử tách 2 cột.
    AuditLogRepository(db).create(
        actor_user_id=user.id, action=_ACTION_DIEU_CHINH_XUAT,
        target=f"voucher:{voucher_id}",
        detail=json.dumps({"noi_dung": "; ".join(changes), "ly_do": payload.ly_do},
                          ensure_ascii=False),
    )
    return _serialize(v, svc=svc, db=db, can_view_cost=authz.can(user, MODULE, "view_cost"))


@router.get("/{voucher_id}/lich-su-dieu-chinh", response_model=list[DieuChinhLichSuRow])
def lich_su_dieu_chinh(
    voucher_id: int, db: Db,
    _: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> list[DieuChinhLichSuRow]:
    """Lịch sử ĐIỀU CHỈNH của 1 phiếu xuất (ai · bộ phận · lúc nào · đổi gì) — mới nhất trước."""
    users = UserRepository(db)
    depts = DepartmentRepository(db)
    out: list[DieuChinhLichSuRow] = []
    for r in AuditLogRepository(db).list_by_target(f"voucher:{voucher_id}"):
        if r.action != _ACTION_DIEU_CHINH_XUAT:
            continue
        # detail = JSON {noi_dung, ly_do}; bản ghi CŨ (trước khi bắt lý do) là chuỗi thường → về noi_dung.
        try:
            d = json.loads(r.detail or "{}")
            noi_dung, ly_do = d.get("noi_dung"), d.get("ly_do")
        except (ValueError, TypeError):
            noi_dung, ly_do = r.detail or None, None
        u = users.get_by_id(r.actor_user_id) if r.actor_user_id else None
        dept = depts.get_by_id(u.department_id) if (u and u.department_id) else None
        out.append(DieuChinhLichSuRow(
            thoi_diem=r.created_at,
            nguoi_ten=getattr(u, "name", None),
            bo_phan_ten=getattr(dept, "name", None),
            chi_tiet=noi_dung,
            ly_do=ly_do,
        ))
    return out


@router.patch("/lo/{lot_id}/vi-tri")
def update_lot_vi_tri(
    lot_id: int, payload: StockLotViTriIn, svc: Service,
    # Vị trí VẬT LÝ (kệ/ô) do người CẦM HÀNG (thủ kho) quản → gate `create`.
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
):
    """Sửa VỊ TRÍ cất lô trong kho (sửa từ drawer Lịch sử của sản phẩm)."""
    try:
        lot = svc.set_lot_vi_tri(lot_id, payload.vi_tri)
    except StockVoucherError as e:
        raise _err(e) from None
    return {"id": lot.id, "vi_tri": lot.vi_tri}


@router.patch("/{voucher_id}/vi-tri")
def set_voucher_vi_tri(
    voucher_id: int, payload: StockVoucherViTriIn, svc: Service,
    # Khai VỊ TRÍ cất lô cho DÒNG phiếu NHẬP còn NHÁP — dùng cho phiếu ĐIỀU CHUYỂN đích dựng sẵn
    # (không có form nhập vị trí). Ghi sổ chép sang lô. Đã ghi sổ → sửa ở lô (`/lo/{id}/vi-tri`).
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
):
    """Khai vị trí cất lô cho dòng phiếu nhập còn nháp (trước khi ghi sổ)."""
    try:
        svc.set_draft_line_vi_tri(voucher_id, [ln.model_dump() for ln in payload.lines])
    except StockVoucherError as e:
        raise _err(e) from None
    return {"ok": True}


# --- Lô & gợi ý phân bổ ------------------------------------------------------

@router.get("/lo/goi-y", response_model=AllocationOut)
def suggest_allocation(
    svc: Service, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
    hang_loai: str = Query(...),
    hang_id: int = Query(..., gt=0),
    kho_id: int = Query(...),
    so_luong: float = Query(..., gt=0, description="Số theo ĐƠN VỊ GỐC của mặt hàng"),
) -> AllocationOut:
    """Gợi ý lấy hàng từ lô nào (FEFO → FIFO). Thủ kho sửa được — giá xuất là ĐÍCH DANH."""
    rows, thieu = svc.suggest_allocation((hang_loai, hang_id), kho_id, so_luong)
    can_view_cost = authz.can(user, MODULE, "view_cost")
    return AllocationOut(
        lines=[
            AllocationLineOut(
                lot_id=r["lot_id"], ma_lo=r["ma_lo"], ngay_nhap=r["ngay_nhap"],
                hsd=r["hsd"], sl_con_lai=r["sl_con_lai"], so_luong=r["so_luong"],
                don_gia_nhap=r["don_gia_nhap"] if can_view_cost else None,
            )
            for r in rows
        ],
        thieu=thieu,
    )

# --- Điều chuyển kho -----------------------------------------------------------

@dieu_chuyen_router.post("", response_model=DieuChuyenOut, status_code=status.HTTP_201_CREATED)
def dieu_chuyen(
    payload: DieuChuyenIn, svc: Service, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> DieuChuyenOut:
    """Ấn điều chuyển 1 HAY NHIỀU mặt hàng kho nguồn → kho đích (gộp vào MỘT yêu cầu). Server tự tạo
    phiếu XUẤT nguồn NHÁP + YÊU CẦU ĐIỀU CHUYỂN (NHẬP) ở đích + phiếu NHẬP đích DỰNG SẴN (khoá giá
    vốn + HSD đích danh theo TỪNG LÔ). Kho đích chỉ ghi sổ phiếu nhập → trừ nguồn + cộng đích cùng nhịp."""
    try:
        res = svc.dieu_chuyen(
            user=user, kho_nguon_id=payload.kho_nguon_id, kho_den_id=payload.kho_den_id,
            items=[it.model_dump() for it in payload.items], ghi_chu=payload.ghi_chu,
        )
    except StockVoucherError as e:
        raise _err(e) from None
    yc, px, pn = res["yeu_cau"], res["phieu_xuat"], res["phieu_nhap"]
    return DieuChuyenOut(
        yeu_cau_id=yc.id, yeu_cau_ma=yc.ma,
        phieu_xuat_id=px.id, phieu_xuat_ma=px.ma,
        phieu_nhap_id=pn.id, phieu_nhap_ma=pn.ma,
        kho_nguon_id=payload.kho_nguon_id, kho_den_id=payload.kho_den_id,
        so_dong=len(payload.items),
        gia_von=res["gia_von"] if authz.can(user, MODULE, "view_cost") else None,
    )


@dieu_chuyen_router.post("/{req_id}/huy", status_code=status.HTTP_204_NO_CONTENT)
def huy_dieu_chuyen(
    req_id: int, payload: StockVoucherCancel, svc: Service,
    # Hủy CẢ phiếu điều chuyển khi CHƯA ghi sổ: hủy 2 phiếu con (nháp) + đóng 2 yêu cầu. Cùng quyền
    # lập phiếu (create). Đã ghi sổ → 400 (sửa sai bằng điều chuyển ngược). `req_id` = đầu mối phiếu
    # điều chuyển = yêu cầu NHẬP đích.
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
):
    """Hủy phiếu điều chuyển (spec-phieu-dieu-chuyen §5) — chỉ khi chưa ghi sổ."""
    try:
        svc.huy_dieu_chuyen(req_id, payload.ly_do)
    except StockVoucherError as e:
        raise _err(e) from None


@dieu_chuyen_router.get("/by-voucher/{voucher_id}")
def dc_request_by_voucher(
    voucher_id: int, svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> dict[str, int]:
    """Từ 1 phiếu điều chuyển (nhập-đích hoặc xuất-nguồn) → id yêu cầu điều chuyển ĐÍCH, để FE mở
    mặt tiền PHIẾU ĐIỀU CHUYỂN thay vì mẫu nhập/xuất. 404 nếu không phải phiếu điều chuyển."""
    rid = svc.dc_request_id_for_voucher(voucher_id)
    if rid is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không phải phiếu điều chuyển.")
    return {"request_id": rid}


# --- Xuất Excel Báo cáo Kho ---
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_response(content: bytes, filename: str) -> Response:
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _build_stock_xlsx(rows) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao ton kho"

    headers = [
        "Mã Kho", "Tên Kho", "Mã Vật Tư", "Tên Vật Tư", "Nhóm Vật Tư",
        "Đơn Vị Tính", "Mã Lô", "Ngày Nhập", "Hạn Sử Dụng", "Vị Trí Kệ/Ô",
        "Trạng Thái Lô", "Số Lượng Ban Đầu", "Số Lượng Còn Lại", "Đơn Giá Nhập", "Thành Tiền"
    ]
    ws.append(headers)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    group_map = {
        "paper": "Giấy",
        "ink": "Mực",
        "film": "Màng/Film",
        "glue": "Keo",
        "packaging": "Bao bì",
        "auxiliary": "Phụ trợ/Khác"
    }

    status_map = {
        "available": "Sẵn sàng",
        "hold": "Giữ chỗ",
        "qc_wait": "Chờ KCS",
        "defect": "Lỗi",
        "empty": "Hết hàng"
    }

    for r_num, row_data in enumerate(rows, 2):
        lot, mat, kho = row_data

        group_vn = group_map.get(mat.material_group, mat.material_group or "")
        status_vn = status_map.get(lot.trang_thai, lot.trang_thai or "")

        qty_rem = float(lot.sl_con_lai)
        qty_orig = float(lot.sl_ban_dau)
        price = int(lot.don_gia_nhap or 0)
        value = int(round(qty_rem * price))

        hsd_str = lot.hsd.strftime("%d/%m/%Y") if lot.hsd else ""
        ngay_nhap_str = lot.ngay_nhap.strftime("%d/%m/%Y") if lot.ngay_nhap else ""

        values = [
            kho.ma,
            kho.ten,
            mat.code,
            mat.name,
            group_vn,
            mat.unit,
            lot.ma_lo,
            ngay_nhap_str,
            hsd_str,
            lot.vi_tri or "",
            status_vn,
            qty_orig,
            qty_rem,
            price,
            value
        ]

        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=r_num, column=col_num, value=val)
            cell.border = thin_border

            if col_num in [1, 3, 6, 7, 8, 9, 10, 11]:
                cell.alignment = center_align
            elif col_num in [2, 4, 5]:
                cell.alignment = left_align
            else:
                cell.alignment = right_align

            if col_num in [12, 13]:
                cell.number_format = '#,##0.00'
            elif col_num in [14, 15]:
                cell.number_format = '#,##0'

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/lo/export.xlsx")
def export_stock_xlsx(
    db: Db,
    user: Annotated[User, Depends(require_permission(MODULE, "export"))],
    kho_id: int | None = Query(default=None),
    con_hang: bool = Query(default=True),
) -> Response:
    from sqlalchemy import select
    from ..models.kho_hang import KhoHang
    from ..models.material import Material
    from ..models.stock_lot import StockLot

    stmt = (
        select(StockLot, Material, KhoHang)
        .join(Material, StockLot.material_id == Material.id)
        .join(KhoHang, StockLot.kho_id == KhoHang.id)
    )
    if kho_id is not None:
        stmt = stmt.where(StockLot.kho_id == kho_id)
    if con_hang:
        stmt = stmt.where(StockLot.sl_con_lai > 0)

    stmt = stmt.order_by(KhoHang.ten.asc(), Material.name.asc(), StockLot.ngay_nhap.asc())
    rows = db.execute(stmt).all()

    filename = "bao-cao-ton-kho.xlsx"
    if kho_id is not None:
        kho = db.get(KhoHang, kho_id)
        if kho:
            filename = f"bao-cao-ton-kho-{kho.ma}.xlsx"

    return _xlsx_response(_build_stock_xlsx(rows), filename)


@router.get("/lo/danh-sach", response_model=list[StockLotOut])
def list_lots(
    svc: Service, db: Db, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
    hang_loai: str | None = Query(default=None),
    hang_id: int | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    con_hang: bool = Query(default=True),
) -> list[StockLotOut]:
    can_view_cost = authz.can(user, MODULE, "view_cost")
    hang = (hang_loai, hang_id) if (hang_loai and hang_id) else None
    lots = svc.lots.list_lots(hang=hang, kho_id=kho_id, con_hang=con_hang)
    # Nạp SẴN mọi mặt hàng của các lô trong 1 lượt (tránh N+1).
    hang_map = svc.hang.map_theo_cap([(lot.hang_loai, lot.hang_id) for lot in lots])
    # Mã phiếu NHẬP sinh ra từng lô — để hiển thị lô THEO MÃ PHIẾU (đợt hàng vào kho) thay mã lô
    # kỹ thuật. Nạp 1 lượt tránh N+1.
    voucher_ma_map = svc.vouchers.ma_by_ids(
        list({lot.voucher_id for lot in lots if lot.voucher_id is not None})
    )
    # Mã đơn vị (to/cai/kem…) → TÊN có dấu (tờ/cái/bản kẽm) cho HIỂN THỊ — 1 truy vấn cho cả trang.
    dv_ten = {d.ma: d.ten for d in svc.hang.don_vi.all_active()}
    out = []
    for lot in lots:
        row = StockLotOut.model_validate(lot)
        m = hang_map.get((lot.hang_loai, lot.hang_id))
        row.hang_ma = getattr(m, "ma", None)
        row.hang_ten = getattr(m, "ten", None)
        row.hang_anh = getattr(m, "anh_url", None)
        row.dvt = getattr(m, "don_vi_gia", None)
        row.dvt_ten = dv_ten.get(row.dvt, row.dvt)
        row.voucher_ma = voucher_ma_map.get(lot.voucher_id) if lot.voucher_id else None
        # Thủ kho chọn lô nhưng KHÔNG thấy giá (spec §6).
        row.don_gia_nhap = int(lot.don_gia_nhap or 0) if can_view_cost else None
        out.append(row)
    return out


@router.get("/mat-hang/{hang_loai}/{hang_id}/lich-su", response_model=MaterialHistoryOut)
def material_history(
    hang_loai: str, hang_id: int, svc: Service, db: Db, authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
    kho_id: int = Query(...),
) -> MaterialHistoryOut:
    """Lịch sử NHẬP (mọi lô, kể cả đã hết) + XUẤT (dòng phiếu xuất đã ghi sổ) của 1 mặt hàng
    tại 1 kho — cho popup màn Tồn kho, tách theo dõi nhập/xuất riêng. Giá vốn ẩn nếu thiếu
    `can_view_cost` (đường path nhiều đoạn nên không đụng route `/{voucher_id}`)."""
    can_view_cost = authz.can(user, MODULE, "view_cost")
    hang = (hang_loai, hang_id)
    m = svc.hang.map_theo_cap([hang]).get(hang)
    dvt = getattr(m, "don_vi_gia", None)

    # NHẬP = mọi lô của mặt hàng tại kho (con_hang=False để giữ cả lô đã xuất hết), FIFO theo ngày.
    lots = svc.lots.list_lots(hang=hang, kho_id=kho_id, con_hang=False)
    # Mã phiếu NHẬP của từng lô (hiển thị lô THEO PHIẾU thay mã lô kỹ thuật) — nạp 1 lượt, tránh N+1.
    voucher_ids = list({lot.voucher_id for lot in lots if lot.voucher_id is not None})
    voucher_ma_map = svc.vouchers.ma_by_ids(voucher_ids)
    # Lô sinh từ phiếu ĐIỀU CHUYỂN → gắn cờ để FE tách tab "Chuyển kho" (nhận về) khỏi Nhập thường.
    dc_voucher_ids = svc.vouchers.dieu_chuyen_by_ids(voucher_ids)
    # SL yêu cầu của từng lô NHẬP (nối lô → dòng phiếu NHẬP → dòng yêu cầu) — nạp 1 lượt, tránh N+1.
    sl_de_nghi_map = svc.vouchers.sl_de_nghi_by_lot(lots)
    nhap: list[StockLotOut] = []
    for lot in lots:
        row = StockLotOut.model_validate(lot)
        row.hang_ma = getattr(m, "ma", None)
        row.hang_ten = getattr(m, "ten", None)
        row.dvt = dvt
        row.voucher_ma = voucher_ma_map.get(lot.voucher_id) if lot.voucher_id else None
        row.don_gia_nhap = int(lot.don_gia_nhap or 0) if can_view_cost else None
        # SL yêu cầu KHÔNG phải tiền → luôn hiện (không gate theo can_view_cost). Kèm ĐƠN VỊ người
        # xin (có thể khác đơn vị gốc của lô) để lịch sử ghi rõ đơn vị cột "SL yêu cầu".
        sl_dvt = sl_de_nghi_map.get(lot.id)
        row.sl_de_nghi = sl_dvt[0] if sl_dvt else None
        row.dvt_yeu_cau = sl_dvt[1] if sl_dvt else None
        row.dieu_chuyen = lot.voucher_id in dc_voucher_ids if lot.voucher_id else False
        nhap.append(row)

    # XUẤT = dòng phiếu xuất đã ghi sổ (đích danh lô); giá vốn = giá lô, ẩn nếu thiếu quyền.
    # `sl_de_nghi` nối qua dòng phiếu xuất → dòng yêu cầu (không phải tiền → luôn hiện).
    xuat = [
        MaterialXuatRow(
            ngay=r["ngay"], voucher_id=r["voucher_id"], voucher_ma=r["voucher_ma"],
            lot_id=r["lot_id"], ma_lo=r["ma_lo"], so_luong=r["so_luong"],
            sl_de_nghi=r["sl_de_nghi"], dvt_yeu_cau=r["dvt_yeu_cau"],
            don_gia=r["don_gia"] if can_view_cost else None,
            dieu_chuyen=r["dieu_chuyen"],
        )
        for r in svc.vouchers.xuat_history(hang, kho_id)
    ]

    return MaterialHistoryOut(
        hang_loai=hang_loai,
        hang_id=hang_id,
        hang_ma=getattr(m, "ma", None),
        hang_ten=getattr(m, "ten", None),
        dvt=dvt,
        on_hand=svc.lots.on_hand(hang, kho_id),
        nhap=nhap, xuat=xuat,
    )


@router.get("/mat-hang/{hang_loai}/{hang_id}/qr-token")
def material_qr_token(
    hang_loai: str, hang_id: int,
    _: Annotated[User, Depends(require_permission(MODULE, "read"))],
    kho_id: int = Query(...),
) -> dict[str, str]:
    """Mã QR đã ký cho tem dán kệ (in tem CẦN đăng nhập; trang quét công khai thì KHÔNG).
    FE nhúng vào link `#s=<token>` rồi vẽ QR. Path nhiều đoạn nên không đụng route `/{voucher_id}`."""
    return {"token": sign_scan(kho_id, hang_loai, hang_id)}


# --- Đính kèm hóa đơn/chứng từ gốc --------------------------------------------

@router.get("/{voucher_id}/attachments", response_model=StockVoucherAttachmentListOut)
def list_voucher_attachments(
    voucher_id: int, svc: Service,
    _: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> StockVoucherAttachmentListOut:
    try:
        items = svc.list_attachments(voucher_id)
    except StockVoucherError as e:
        raise _err(e) from None
    return StockVoucherAttachmentListOut(
        items=[StockVoucherAttachmentOut(**it) for it in items]
    )


@router.post(
    "/{voucher_id}/attachments",
    response_model=StockVoucherAttachmentOut,
    status_code=status.HTTP_201_CREATED,
)
def upload_voucher_attachment(
    voucher_id: int, svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
    file: UploadFile = File(...),
) -> StockVoucherAttachmentOut:
    data = file.file.read()
    try:
        return StockVoucherAttachmentOut(**svc.add_attachment(
            voucher_id, actor=user, file_name=file.filename,
            content_type=file.content_type, data=data,
        ))
    except StockVoucherError as e:
        raise _err(e) from None


@router.delete(
    "/{voucher_id}/attachments/{attachment_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
)
def delete_voucher_attachment(
    voucher_id: int, attachment_id: int, svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> Response:
    try:
        svc.delete_attachment(voucher_id, attachment_id, actor=user)
    except StockVoucherError as e:
        raise _err(e) from None
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# --- Ngưỡng tồn ---------------------------------------------------------------

@threshold_router.get("", response_model=list[StockThresholdOut])
def list_thresholds(
    db: Db, _: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> list[StockThresholdOut]:
    return [
        StockThresholdOut.model_validate(t)
        for t in StockThresholdRepository(db).list_active()
    ]


@threshold_router.put("", response_model=StockThresholdOut)
def upsert_threshold(
    payload: StockThresholdIn, db: Db,
    _: Annotated[User, Depends(require_permission(MODULE, "set_threshold"))],
) -> StockThresholdOut:
    """Khai ngưỡng. Bỏ trống `nguong_can_ton` thì để NULL — service tự suy ra
    `nguong_ton × 1.3` lúc so sánh, khỏi phải backfill khi đổi hệ số."""
    if payload.nguong_can_ton is not None and payload.nguong_can_ton < payload.nguong_ton:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ngưỡng cận tồn phải lớn hơn hoặc bằng ngưỡng tồn.",
        )
    if payload.nguong_toi_da is not None and payload.nguong_toi_da < payload.nguong_ton:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ngưỡng tối đa phải lớn hơn hoặc bằng ngưỡng tồn.",
        )
    obj = StockThresholdRepository(db).upsert(
        hang=(payload.hang_loai, payload.hang_id), kho_id=payload.kho_id,
        nguong_ton=payload.nguong_ton, nguong_can_ton=payload.nguong_can_ton,
        nguong_toi_da=payload.nguong_toi_da, canh_bao=payload.canh_bao,
    )
    return StockThresholdOut.model_validate(obj)
