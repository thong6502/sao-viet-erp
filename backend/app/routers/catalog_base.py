"""Nền chung cho ROUTER các màn Cấu hình danh mục.

Vì sao có file này: tám router danh mục viết đi viết lại đúng một khuôn — `GET ""` (lọc + cắt
trang), `GET "/{id}"`, `POST ""`, `PUT "/{id}"`, `DELETE "/{id}"`, cộng một hàm `_err()` dịch
exception sang mã HTTP **giống hệt nhau ở BẢY file**. Chỗ khác nhau chỉ là schema, module quyền,
một bộ lọc riêng, và vài màn có `facets` / cách dựng dòng riêng.

HAI THỨ Ở ĐÂY, DÙNG ĐƯỢC RỜI NHAU:

* `loi_http(e)` — thay bảy bản `_err()`. Nó soi theo LỚP CƠ SỞ CHUNG
  (`services/catalog_base.CatalogNotFound`…) chứ không theo lớp riêng của từng danh mục, nên
  router thủ công cũng gọi được.
* `make_catalog_router(...)` — sinh trọn bộ CRUD cho màn nào vừa khuôn.

⚠️ THỨ TỰ ROUTE. Route TĨNH (`/ho`, `/quy-doi`, `/dau-viec`, `/trang-thai`) phải nằm TRƯỚC
`/{item_id}`, nếu không FastAPI khớp `"ho"` vào `{item_id}` rồi ăn 422 vì không ép được sang int.
Vì thế factory GẮN THẲNG vào router của màn theo đúng thứ tự được gọi (không `include_router`
một router con): file gọi khai các route tĩnh của mình TRƯỚC, rồi mới gọi factory ở CUỐI file.

⚠️ TÊN ROUTE. Mọi route đều đặt `name=f"..._{ten}"` — factory chạy nhiều lần mà quên tham số hoá
tên là trùng `operation_id` và OpenAPI không dựng được (test `test_openapi_dung_duoc` bắt ca này).
"""
from __future__ import annotations

from inspect import Parameter, Signature
from typing import Any, Callable

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from pydantic import BaseModel, ValidationError

from ..deps import require_any_permission, require_permission
from ..models.user import User
from ..repositories.cong_thuc_lich_su_repo import CongThucLichSuRepository
from ..schemas.cong_thuc_lich_su import CongThucLichSuOut
from ..services.catalog_base import (
    CatalogDuplicate, CatalogError, CatalogInUse, CatalogNotFound,
)


class MaGoiYOut(BaseModel):
    """Mã kế tiếp cho form khai mới — `{"ma": "KHO-0007"}`."""

    ma: str


class ImportExcelLoi(BaseModel):
    """Một dòng lỗi khi nhập Excel — không đoán số, chỉ ra đúng dòng/cột/lý do."""

    dong: int
    cot: str
    ly_do: str


class ImportExcelOut(BaseModel):
    """Kết quả nhập Excel — dòng lỗi KHÔNG chặn các dòng khác."""

    tong_dong: int
    thanh_cong: int
    loi: list[ImportExcelLoi] = []


class _ActiveIn(BaseModel):
    """Thân của `PATCH /{id}/active` — ĐÚNG một khoá, cố ý.

    Không dùng lại `InModel` ở đây: bật/tắt một dòng thì người dùng không sửa gì khác, mà `InModel`
    lại đòi đủ field bắt buộc. Xem chú thích ở chỗ khai route.
    """

    active: bool


def loi_http(e: Exception) -> HTTPException:
    """Exception nghiệp vụ của danh mục → mã HTTP. MỘT bản cho cả tám màn.

    * 404 — không tìm thấy.
    * 409 — XUNG ĐỘT TRẠNG THÁI: trùng mã, hoặc còn ràng buộc nên không xoá được. Không phải 422:
      dữ liệu client gửi lên chẳng sai gì cả.
    * 422 — dữ liệu khai sai (mặc định).

    Soi theo lớp CƠ SỞ chung nên mỗi danh mục vẫn giữ lớp exception + câu báo lỗi riêng của nó
    (xem `services/catalog_base`) — router cũ viết `except KhoHangNotFound` vẫn bắt được y nguyên.
    """
    if isinstance(e, CatalogNotFound):
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    if isinstance(e, (CatalogDuplicate, CatalogInUse)):
        return HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(e))
    return HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e))


def _tham_so(ServiceDep, doc, loc: str | None, co_active: bool) -> list[Parameter]:
    """Chữ ký của handler `list` — dựng bằng tay vì bộ lọc KHÁC NHAU theo từng màn.

    Không thể viết một `def` cố định: màn Máy không có cột `active`, còn bộ lọc riêng thì mỗi màn
    một tên (`nhom` · `tinh_trang` · `loai_may` · `structural_type` · `ho`). Gán `__signature__`
    là cách FastAPI đọc tham số (`get_typed_signature` gọi `inspect.signature`), nên tham số hiện
    ra ở OpenAPI đúng như viết tay.
    """
    K = Parameter.KEYWORD_ONLY
    ps = [
        Parameter("svc", K, annotation=ServiceDep),
        Parameter("_", K, default=Depends(doc), annotation=User),
        Parameter("q", K, default=Query(default=None), annotation=str | None),
    ]
    if loc:
        ps.append(Parameter(loc, K, default=Query(default=None), annotation=str | None))
    if co_active:
        ps.append(Parameter("active", K, default=Query(default=None), annotation=bool | None))
    ps += [
        Parameter("page", K, default=Query(default=1, ge=1), annotation=int),
        # Trần 200 khớp `repositories/catalog_base.SIZE_TRAN` — chặn client gõ `?size=99999`.
        Parameter("size", K, default=Query(default=50, ge=1, le=200), annotation=int),
    ]
    return ps


def make_catalog_router(
    router: APIRouter,
    *,
    ten: str,
    goc: str = "",
    ServiceDep: Any,
    InModel: type[BaseModel],
    RowModel: type[BaseModel],
    ListModel: type[BaseModel],
    module: str,
    doc: Callable | None = None,
    loc: str | None = None,
    co_active: bool = True,
    facets: Callable[[Any, dict], dict] | None = None,
    dung_rows: Callable[[Any, list], list] | None = None,
    ma_goi_y: bool = False,
    loi_khac: tuple[type[Exception], ...] = (),
    enable_clone: bool = False,
    cong_thuc_truong: str | None = None,
    enable_import: bool = False,
    import_columns: dict[str, str] | None = None,
    import_resolve: Callable[[dict, Any], dict] | None = None,
) -> APIRouter:
    """GẮN trọn bộ CRUD của một danh mục vào `router` của màn. Trả lại chính `router` đó.

    Gắn thẳng (không `include_router`) vì THỨ TỰ khai route là thứ tự khớp của FastAPI — xem cảnh
    báo ở docstring module.

    * `router` — router của màn, đã có `prefix` (`/api/kho`…).
    * `goc` — đường dẫn con, cho màn có danh mục THỨ HAI trên cùng router (`"/quy-doi"`).
    * `ten` — slug đặt tên route (`list_kho_hang`…). PHẢI duy nhất toàn app.
    * `ServiceDep` — `Annotated[Svc, Depends(get_service)]` của màn.
    * `doc` — dependency ĐỌC. Mặc định `require_permission(module, "read")`; màn nào là danh mục
      THAM CHIẾU thì truyền OR-gate rộng hơn (xem `deps.require_any_permission`).
    * `loc` — TÊN bộ lọc riêng của màn (`nhom`, `tinh_trang`…). Chỉ một, và luôn là chuỗi.
    * `co_active` — màn Máy đặt `False`: `may_thiet_bi` KHÔNG có cột `active`.
    * `facets(svc, kw)` — số trên tab lọc; `kw` là bộ lọc ĐÃ BỎ bộ lọc riêng (tab đang không được
      chọn vẫn phải khoe số của nó).
    * `dung_rows(svc, objs) -> list[RowModel]` — cách dựng dòng trả về. Mặc định
      `RowModel.model_validate`. Màn nào cần điền thêm (tên đơn vị, chip quy đổi) thì truyền hàm
      riêng — dùng cho CẢ list, get, create, update nên không có chỗ nào lệch nhau.
    * `ma_goi_y` — mở `GET /ma-goi-y`. Chỉ bật cho danh mục có `ma_prefix` ở repo.
    * `loi_khac` — lớp exception NGOÀI họ `Catalog*` mà handler cũng phải bắt.
    * `enable_clone` — mở `POST /{item_id}/clone`, gác bằng quyền `clone` riêng (không dùng chung
      `create` — nhân bản là thao tác khác, vai được tạo mới chưa chắc được nhân bản hàng cũ).
    * `cong_thuc_truong` — tên cột công thức của danh mục này (`"cong_thuc_luong"` hoặc
      `"cong_thuc_san_luong"`). Bật thì mỗi dòng trả về có thêm `<truong>_truoc` +
      `<truong>_sua_luc` (giá trị NGAY TRƯỚC lần sửa gần nhất, đọc từ `cong_thuc_lich_su`), và mở
      thêm `GET /{item_id}/lich-su-cong-thuc` cho lịch sử đầy đủ. Xem
      `services/nhat_ky_danh_muc._ghi_lich_su_cong_thuc` — nơi ghi vào bảng đó.
    * `enable_import` — mở `GET /mau-excel` (tải file mẫu) + `POST /import-excel` (nhập). CHỈ TẠO
      MỚI — mã trùng dòng đã có trong DB là LỖI của riêng dòng đó, không ghi đè, không chặn các
      dòng khác. Đòi truyền `import_columns`.
    * `import_columns` — ánh xạ TIÊU ĐỀ cột Excel → tên field của `InModel` (vd `{"Mã": "ma", "Tên":
      "ten"}`) — dùng để dựng cả file mẫu lẫn đọc file nhập. Cột thừa/thiếu tiêu đề bị bỏ qua.
    * `import_resolve(du_lieu, svc) -> dict` — chỗ cắm cho danh mục cần DỊCH một cột Excel dạng
      chữ (tên tổ…) sang FK id trước khi dựng `InModel` (vd Công việc khoán bắt buộc `department_id`
      lúc tạo). Ném `ValueError(câu lỗi)` thì dòng đó rơi vào `loi`, các dòng khác vẫn chạy tiếp.
    """
    doc = doc or require_permission(module, "read")
    req_create = require_permission(module, "create")
    req_update = require_permission(module, "update")
    req_delete = require_permission(module, "delete")
    # Bắt ĐÚNG họ exception nghiệp vụ của danh mục — KHÔNG bắt `Exception` trần, không thì lỗi lập
    # trình (AttributeError, TypeError) cũng hoá 422 và nằm im.
    BAT = (CatalogError, *loi_khac)

    def _rows(svc, objs) -> list:
        rows = dung_rows(svc, objs) if dung_rows else [RowModel.model_validate(o) for o in objs]
        # "Lần trước công thức" (mục 3+7) — 1 truy vấn cho cả trang, giống hệt mẫu `gan_ten_don_vi`.
        if cong_thuc_truong and rows and getattr(svc, "audit", None) is not None:
            moi_nhat = CongThucLichSuRepository(svc.audit.db).moi_nhat_nhieu(
                ten, [r.id for r in rows], cong_thuc_truong)
            for r in rows:
                m = moi_nhat.get(r.id)
                setattr(r, f"{cong_thuc_truong}_truoc", m.gia_tri_cu if m else None)
                setattr(r, f"{cong_thuc_truong}_sua_luc", m.sua_luc if m else None)
        return rows

    def _mot(svc, obj):
        return _rows(svc, [obj])[0]

    # -- GET "" : danh sách -------------------------------------------------------------
    def _list(**kw):
        svc = kw.pop("svc")
        kw.pop("_", None)
        page, size = kw["page"], kw["size"]
        rows, total = svc.list(**kw)
        them = {}
        if facets is not None:
            # `facets` KHÔNG nhận bộ lọc riêng: tab đang không được chọn vẫn phải khoe số của nó.
            them["facets"] = facets(svc, {k: v for k, v in kw.items()
                                          if k not in ("page", "size", loc)})
        return ListModel(items=_rows(svc, rows), total=total, page=page, size=size, **them)

    _list.__signature__ = Signature(_tham_so(ServiceDep, doc, loc, co_active))
    _list.__name__ = f"list_{ten}"
    router.get(goc or "", response_model=ListModel, name=f"list_{ten}")(_list)

    # -- GET "/ma-goi-y" : PHẢI khai trước "/{item_id}" ---------------------------------
    if ma_goi_y:
        def _ma_goi_y(svc, _=Depends(doc)) -> MaGoiYOut:
            """Mã kế tiếp cho bản ghi mới.

            Trước 15/08/2026 frontend tự ĐOÁN tiền tố bằng cách dò chuỗi trong URL rồi bắn HAI
            request để mò mã lớn nhất (`frontend/src/pages/danh-muc/maGoiY.ts`) — luật vốn thuộc
            về danh mục mà nằm ở màn, và mã lớn nhất nằm ở TRANG CUỐI nên đoán trên trang đang
            xem là ra mã ĐÃ CÓ.
            """
            return MaGoiYOut(ma=svc.ma_goi_y())
        _ma_goi_y.__annotations__["svc"] = ServiceDep
        _ma_goi_y.__annotations__["_"] = User
        _ma_goi_y.__name__ = f"ma_goi_y_{ten}"
        router.get(f"{goc}/ma-goi-y", response_model=MaGoiYOut, name=f"ma_goi_y_{ten}")(_ma_goi_y)

    # -- GET "/mau-excel" + POST "/import-excel" : PHẢI khai trước "/{item_id}" ---------
    if enable_import:
        def _mau_excel(_=Depends(doc)) -> Response:
            """File mẫu — chỉ có dòng tiêu đề, đúng thứ tự cột `import_columns`."""
            from io import BytesIO

            # lazy import: thiếu dep chỉ hỏng route này, không sập app (mẫu `export_employees_xlsx`)
            from openpyxl import Workbook
            from openpyxl.styles import Font

            wb = Workbook()
            ws = wb.active
            ws.title = ten[:31]
            headers = list(import_columns.keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for idx in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 22
            buf = BytesIO()
            wb.save(buf)
            return Response(
                content=buf.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="mau-{ten}.xlsx"'},
            )
        _mau_excel.__annotations__["_"] = User
        _mau_excel.__name__ = f"mau_excel_{ten}"
        router.get(f"{goc}/mau-excel", name=f"mau_excel_{ten}")(_mau_excel)

        def _import_excel(svc, user=Depends(req_create), file: UploadFile = File(...)):
            """Nhập Excel — CHỈ TẠO MỚI. Dòng lỗi không chặn các dòng khác, không đoán số."""
            from io import BytesIO

            from openpyxl import load_workbook  # lazy import, cùng lý do trên

            try:
                wb = load_workbook(BytesIO(file.file.read()), read_only=True, data_only=True)
            except Exception:
                raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY,
                                     "Không đọc được file — phải là .xlsx đúng mẫu.") from None
            hang = wb.active.iter_rows(values_only=True)
            tieu_de = [str(h).strip() if h is not None else "" for h in next(hang, ())]
            cot_theo_field = {
                field: tieu_de.index(nhan) for nhan, field in import_columns.items() if nhan in tieu_de
            }
            tong_dong = 0
            thanh_cong = 0
            loi: list[ImportExcelLoi] = []
            for so_dong, hang_du_lieu in enumerate(hang, start=2):
                if not hang_du_lieu or all(o is None for o in hang_du_lieu):
                    continue
                tong_dong += 1
                du_lieu = {}
                for field, idx in cot_theo_field.items():
                    if idx < len(hang_du_lieu):
                        gt = hang_du_lieu[idx]
                        if isinstance(gt, str):
                            gt = gt.strip()
                        if gt not in (None, ""):
                            du_lieu[field] = gt
                if import_resolve:
                    try:
                        du_lieu = import_resolve(du_lieu, svc)
                    except ValueError as e:
                        loi.append(ImportExcelLoi(dong=so_dong, cot="?", ly_do=str(e)))
                        continue
                try:
                    payload = InModel(**du_lieu)
                except ValidationError as e:
                    cot = ", ".join(str(err["loc"][0]) for err in e.errors() if err["loc"])
                    loi.append(ImportExcelLoi(dong=so_dong, cot=cot or "?", ly_do="Dữ liệu không hợp lệ."))
                    continue
                try:
                    svc.create(payload.model_dump(exclude_unset=True), actor_id=user.id)
                    thanh_cong += 1
                except BAT as e:
                    loi.append(ImportExcelLoi(dong=so_dong, cot="ma", ly_do=str(e)))
            return ImportExcelOut(tong_dong=tong_dong, thanh_cong=thanh_cong, loi=loi)
        _import_excel.__annotations__["svc"] = ServiceDep
        _import_excel.__annotations__["user"] = User
        _import_excel.__name__ = f"import_excel_{ten}"
        router.post(f"{goc}/import-excel", response_model=ImportExcelOut,
                    name=f"import_excel_{ten}")(_import_excel)

    # -- GET "/{item_id}" ---------------------------------------------------------------
    def _get(item_id: int, svc, _=Depends(doc)):
        try:
            obj = svc.get(item_id)
        except BAT as e:
            raise loi_http(e) from None
        return _mot(svc, obj)
    _get.__annotations__["svc"] = ServiceDep
    _get.__annotations__["_"] = User
    _get.__name__ = f"get_{ten}"
    router.get(goc + "/{item_id}", response_model=RowModel, name=f"get_{ten}")(_get)

    # -- POST "" ------------------------------------------------------------------------
    def _create(payload, svc, user=Depends(req_create)):
        try:
            obj = svc.create(payload.model_dump(exclude_unset=True), actor_id=user.id)
        except BAT as e:
            raise loi_http(e) from None
        return _mot(svc, obj)
    _create.__annotations__["payload"] = InModel
    _create.__annotations__["svc"] = ServiceDep
    _create.__annotations__["user"] = User
    _create.__name__ = f"create_{ten}"
    router.post(goc or "", response_model=RowModel, status_code=status.HTTP_201_CREATED,
                name=f"create_{ten}")(_create)

    # -- PUT "/{item_id}" ---------------------------------------------------------------
    def _update(item_id: int, payload, svc, user=Depends(req_update)):
        try:
            obj = svc.update(item_id, payload.model_dump(exclude_unset=True), actor_id=user.id)
        except BAT as e:
            raise loi_http(e) from None
        return _mot(svc, obj)
    _update.__annotations__["payload"] = InModel
    _update.__annotations__["svc"] = ServiceDep
    _update.__annotations__["user"] = User
    _update.__name__ = f"update_{ten}"
    router.put(goc + "/{item_id}", response_model=RowModel, name=f"update_{ten}")(_update)

    # -- PATCH "/{item_id}/active" : BẬT / NGỪNG dùng ------------------------------------
    #
    # VÌ SAO PHẢI CÓ ROUTE RIÊNG thay vì `PUT /{item_id}` với mỗi `{"active": false}`:
    # `_update` nhận `InModel` — schema ĐẦY ĐỦ, có field bắt buộc (`ma`, `ten`, `nhom`…). Gửi mỗi
    # một khoá là Pydantic chặn ở cổng với 422 "field required", service KHÔNG bao giờ chạy tới.
    # Đó chính là lỗi đã làm nút "Ngừng dùng" / "Bật lại" bấm-không-ăn ở CẢ BỐN danh mục xoá mềm
    # (đo 15/08/2026: cong_doan thiếu ma·ten·nhom · bu_hao thiếu ma·ten · khuon_be thiếu ten ·
    # loai_san_pham thiếu ma·ten·structural_type). Màn nuốt 422 thành "Request failed" nên bấm
    # xong không thấy gì xảy ra.
    #
    # Quyền: nhận CẢ `update` lẫn `delete` — cùng một nút này đi từ hai cửa (hộp thoại Xóa dùng
    # quyền `delete`, nút "Bật lại" dùng `update`), gác một bên là bịt mất cửa kia.
    if co_active:
        req_bat_tat = require_any_permission((module, "update"), (module, "delete"))

        def _dat_active(item_id: int, payload: _ActiveIn, svc, user=Depends(req_bat_tat)):
            try:
                obj = svc.dat_active(item_id, payload.active, actor_id=user.id)
            except BAT as e:
                raise loi_http(e) from None
            return _mot(svc, obj)
        _dat_active.__annotations__["svc"] = ServiceDep
        _dat_active.__annotations__["user"] = User
        _dat_active.__name__ = f"dat_active_{ten}"
        router.patch(goc + "/{item_id}/active", response_model=RowModel,
                     name=f"dat_active_{ten}")(_dat_active)

    # -- POST "/{item_id}/clone" : Nhân bản -----------------------------------------------
    if enable_clone:
        req_clone = require_permission(module, "clone")

        def _clone(item_id: int, svc, user=Depends(req_clone)):
            try:
                obj = svc.clone(item_id, actor_id=user.id)
            except BAT as e:
                raise loi_http(e) from None
            return _mot(svc, obj)
        _clone.__annotations__["svc"] = ServiceDep
        _clone.__annotations__["user"] = User
        _clone.__name__ = f"clone_{ten}"
        router.post(goc + "/{item_id}/clone", response_model=RowModel,
                    status_code=status.HTTP_201_CREATED, name=f"clone_{ten}")(_clone)

    # -- GET "/{item_id}/lich-su-cong-thuc" : lịch sử ĐẦY ĐỦ, không chỉ "lần trước" ------
    if cong_thuc_truong:
        def _lich_su_cong_thuc(item_id: int, svc, _=Depends(doc)) -> list[CongThucLichSuOut]:
            try:
                svc.get(item_id)  # 404 nếu không có — cùng khuôn `_get`, tránh lộ lịch sử id ma.
            except BAT as e:
                raise loi_http(e) from None
            if getattr(svc, "audit", None) is None:
                return []
            rows = CongThucLichSuRepository(svc.audit.db).liet_ke(ten, item_id, cong_thuc_truong)
            return [CongThucLichSuOut.model_validate(r) for r in rows]
        _lich_su_cong_thuc.__annotations__["svc"] = ServiceDep
        _lich_su_cong_thuc.__annotations__["_"] = User
        _lich_su_cong_thuc.__name__ = f"lich_su_cong_thuc_{ten}"
        router.get(goc + "/{item_id}/lich-su-cong-thuc", response_model=list[CongThucLichSuOut],
                   name=f"lich_su_cong_thuc_{ten}")(_lich_su_cong_thuc)

    # -- DELETE "/{item_id}" ------------------------------------------------------------
    def _delete(item_id: int, svc, user=Depends(req_delete)):
        try:
            svc.delete(item_id, actor_id=user.id)
        except BAT as e:
            raise loi_http(e) from None
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    _delete.__annotations__["svc"] = ServiceDep
    _delete.__annotations__["user"] = User
    _delete.__name__ = f"delete_{ten}"
    router.delete(goc + "/{item_id}", status_code=status.HTTP_204_NO_CONTENT,
                  response_class=Response, name=f"delete_{ten}")(_delete)

    return router
