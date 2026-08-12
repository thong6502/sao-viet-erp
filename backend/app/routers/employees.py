"""Hồ sơ nhân sự (nhan_su) routes — lát #1.

Thin HTTP shell over EmployeeService. Every route is guarded by
`require_permission('nhan_su', <action>)`; list/detail narrow to the caller's data scope
(own/department/all) resolved from their role. Stage changes (trạng thái / điều chuyển /
nâng bậc) go through `/transitions` so each writes a Quá trình công tác event.
"""
from __future__ import annotations

from datetime import date, timedelta
from typing import Annotated

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)

from ..deps import (
    CurrentUser,
    get_audit_repository,
    get_authorization_service,
    get_department_repository,
    get_employee_service,
    get_payroll_service,
    get_role_repository,
    get_user_repository,
    require_any_permission,
    require_permission,
)
from ..models.user import User
from ..repositories.audit_repo import AuditLogRepository
from ..repositories.rbac_repo import DepartmentRepository, RoleRepository
from ..repositories.user_repo import UserRepository
from ..schemas.employee import (
    AccountIn,
    AssignShiftBulkIn,
    AssignShiftBulkOut,
    AssignShiftIn,
    AttachmentOut,
    AttachmentsOut,
    DepartmentOption,
    DuplicateRef,
    EmployeeActivityOut,
    EmployeeActivityRowOut,
    EmployeeCreate,
    EmployeeCreateOut,
    EmployeeEventOut,
    EmployeeEventsOut,
    EmployeeKpis,
    EmployeeListOut,
    EmployeeMetaOut,
    EmployeeOut,
    EmployeeRow,
    EmployeeUpdate,
    EmployeeUpdateOut,
    JobGradeIn,
    JobGradeOut,
    JobGradeUpdateIn,
    JobGradesOut,
    MyContactIn,
    MyProfileOut,
    RequestDecisionIn,
    RoleOption,
    ShiftAssignmentOut,
    ShiftAssignmentsOut,
    TransitionIn,
    UpdateRequestIn,
    UpdateRequestOut,
    UpdateRequestsOut,
    UserOption,
)
from ..services.employee_service import (
    EmployeeError,
    EmployeeForbidden,
    EmployeeNotFound,
    EmployeeService,
    EmployeeValidationError,
    SENSITIVE_FIELDS,
)
from ..services.rbac_service import AuthorizationService
from ..services.payroll_service import PayrollError, PayrollService
from ..storage import get_storage, make_key, url_from_key

router = APIRouter(prefix="/api/employees", tags=["employees"])

MODULE = "nhan_su"

# TỰ PHỤC VỤ (tách 10/08/2026) — một ô quyền cho MỌI việc người lao động làm với hồ sơ của CHÍNH
# MÌNH: tự chấm công, xem công/phiếu lương của mình, tự gửi đơn nghỉ / phiếu tăng ca / xin tạm ứng.
# Trước đây nhóm này không gác gì (chỉ cần đăng nhập) nên không có cách nào tắt cho một vai.
# Ba hàng rào cũ GIỮ NGUYÊN (phải có hồ sơ NV nối tài khoản · trong bán kính điểm chấm công · đúng
# khung giờ ca) — chúng chống lạm dụng, còn ô này chống truy cập.
MODULE_TU_PHUC_VU = "self_service"
SelfUser = Annotated[User, Depends(require_permission(MODULE_TU_PHUC_VU, "read"))]

# Ô THAO TÁC của Tự phục vụ (tách 11/08/2026). `SelfUser` (= `read`) chỉ cho XEM công / phiếu /
# đơn của chính mình; mọi đường GHI — chấm công, gửi · sửa · huỷ đơn nghỉ, phiếu tăng ca, xin đi
# muộn, xin tạm ứng, sửa hồ sơ của mình — đòi ô này.
SelfWriter = Annotated[User, Depends(require_permission(MODULE_TU_PHUC_VU, "create"))]

# Hồ sơ HR (CCCD, hợp đồng…) đi qua kho file dùng chung; đọc lại qua /api/files, chỉ người
# có quyền `nhan_su` mới xem được (app/routers/files.py).
_HR_SUBDIR = "hr"

Service = Annotated[EmployeeService, Depends(get_employee_service)]
Payroll = Annotated[PayrollService, Depends(get_payroll_service)]
Authz = Annotated[AuthorizationService, Depends(get_authorization_service)]
Users = Annotated[UserRepository, Depends(get_user_repository)]
Depts = Annotated[DepartmentRepository, Depends(get_department_repository)]
Roles = Annotated[RoleRepository, Depends(get_role_repository)]
Audit = Annotated[AuditLogRepository, Depends(get_audit_repository)]


def _scope_for(authz: AuthorizationService, user: User) -> str:
    """The caller's data scope on nhan_su (own/department/all). Defaults to `own`."""
    return authz.scope_for(user, MODULE) or "own"


def _raise(exc: Exception) -> None:
    if isinstance(exc, EmployeeNotFound):
        raise HTTPException(status_code=404, detail=str(exc))
    if isinstance(exc, EmployeeForbidden):
        raise HTTPException(status_code=403, detail=str(exc))
    if isinstance(exc, EmployeeValidationError):
        raise HTTPException(status_code=400, detail=str(exc))
    raise exc


def _dept_names(depts: DepartmentRepository, ids: set[int]) -> dict[int, str]:
    out: dict[int, str] = {}
    for did in ids:
        d = depts.get_by_id(did)
        if d is not None:
            out[did] = d.name
    return out


def _user_names(users: UserRepository, ids: set[int]) -> dict[int, str]:
    out: dict[int, str] = {}
    for uid in ids:
        u = users.get_by_id(uid)
        if u is not None:
            out[uid] = u.username
    return out


def _role_names(
    users: UserRepository, roles: RoleRepository, ids: set[int]
) -> dict[int, str]:
    """user_id → tên Vai trò (RBAC) của tài khoản, để hiện làm "Chức danh" ở list nhân sự."""
    out: dict[int, str] = {}
    for uid in ids:
        u = users.get_by_id(uid)
        if u is not None and u.role_id is not None:
            r = roles.get_by_id(u.role_id)
            if r is not None:
                out[uid] = r.name
    return out


def _grade_name(employee, svc) -> str | None:
    """Tên bậc tay nghề để hiển thị. Rơi về cột chữ CŨ khi hồ sơ chưa được gán bậc danh mục —
    người cũ vẫn thấy đúng bậc mình đang mang, không bị trống trơn sau khi đổi cơ chế."""
    if employee.job_grade_id is not None:
        g = svc.employees.get_job_grade(employee.job_grade_id)
        if g is not None:
            return g.name
    return employee.job_grade


def _row(
    employee,
    dept_names: dict[int, str],
    user_names: dict[int, str],
    role_names: dict[int, str] | None = None,
    grade_names: dict[int, str] | None = None,
) -> EmployeeRow:
    row = EmployeeRow.model_validate(employee)
    if employee.department_id is not None:
        row.department_name = dept_names.get(employee.department_id)
    if employee.user_id is not None:
        row.account_username = user_names.get(employee.user_id)
        if role_names is not None:
            row.role_name = role_names.get(employee.user_id)
    # Tra sẵn thành dict ở endpoint (danh mục chỉ vài dòng) — không query trong vòng lặp.
    if employee.job_grade_id is not None and grade_names is not None:
        row.job_grade_name = grade_names.get(employee.job_grade_id)
    if row.job_grade_name is None:
        row.job_grade_name = employee.job_grade
    return row


def _full(employee, depts: DepartmentRepository, users: UserRepository,
          svc=None) -> EmployeeOut:
    out = EmployeeOut.model_validate(employee)
    if svc is not None:
        out.job_grade_name = _grade_name(employee, svc)
    if employee.department_id is not None:
        d = depts.get_by_id(employee.department_id)
        out.department_name = d.name if d is not None else None
    if employee.user_id is not None:
        u = users.get_by_id(employee.user_id)
        out.account_username = u.username if u is not None else None
    return out


# Dữ liệu nhạy cảm — DÙNG CHUNG danh sách với service (che khi đọc + gác khi ghi, N5).
_SALARY_FIELDS = SENSITIVE_FIELDS


def _mask_salary(out: EmployeeOut) -> EmployeeOut:
    for f in _SALARY_FIELDS:
        setattr(out, f, None)
    return out


def _dup(employee) -> DuplicateRef | None:
    if employee is None:
        return None
    return DuplicateRef(id=employee.id, code=employee.code, full_name=employee.full_name)


def _can_apply_transition(authz: AuthorizationService, user: User, kind: str) -> bool:
    if kind in {"transfer", "promote"}:
        return authz.can(user, MODULE, "transfer")
    return authz.can(user, MODULE, "manage_status")


# --- list + meta ------------------------------------------------------------


#: Nhãn trạng thái trong file xuất — phải khớp nhãn trên màn, nếu không kế toán đối chiếu là lệch.
_NHAN_TRANG_THAI = {
    "probation": "Thử việc",
    "active": "Chính thức",
    "on_leave": "Nghỉ dài hạn",
    "suspended": "Đình chỉ",
    "resigned": "Đã nghỉ",
}

#: Nhãn cột file xuất — GIỮ ĐÚNG 8 cột đang hiện trên màn. Đổi cột là việc khác, đừng nhét vào đây.
_COT_XUAT = ("Mã", "Họ tên", "Phòng/Tổ", "Chức danh", "Bậc tay nghề", "Trạng thái",
             "Ngày vào", "Tài khoản")

#: Lấy theo mẻ khi xuất. KHÔNG phải trần kết quả — vòng lặp chạy tới khi đủ `total`.
_ME_XUAT = 200


@router.get("/export.xlsx")
def export_employees_xlsx(
    svc: Service,
    authz: Authz,
    users: Users,
    depts: Depts,
    roles: Roles,
    # Ô "Xuất Excel danh sách" (`nhan_su:export`) — trước 11/08/2026 endpoint chỉ đòi `read`, và
    # giao diện cũng KHÔNG hỏi ô nào cả, nên ô đó chưa bao giờ có tác dụng: ai xem được hồ sơ là
    # tải được cả danh sách nhân sự ra file. Xuất file là mang dữ liệu RA KHỎI hệ thống — phải là
    # một quyết định cấp riêng, không đi kèm quyền xem.
    user: Annotated[User, Depends(require_permission(MODULE, "export"))],
    q: str | None = Query(default=None),
    department_id: int | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    has_account: bool | None = Query(default=None),
    sort: str = Query(default="code"),
) -> Response:
    """Xuất danh sách nhân sự ra .xlsx THẬT (chủ chốt 08/08/2026).

    Trước đây giao diện tự nối chuỗi CSV rồi đặt tên nút là "Xuất Excel" — nhãn nói dối, và tệ hơn
    là nó chỉ lấy **200 người đầu** rồi im lặng, ai đứng thứ 201 trở đi biến mất khỏi file.

    Hai ràng buộc BẮT BUỘC, đừng tối giản đi:

    1. **Cùng phạm vi quyền và cùng bộ lọc với màn danh sách.** Dùng lại `_scope_for` và đúng các
       tham số của `list_employees`. Bỏ qua là người có phạm vi `own` tải được cả công ty — rò dữ
       liệu nhân sự, không phải lỗi giao diện.
    2. **Không dùng trần `size` của endpoint danh sách** (`le=200`). Ở đây lặp theo mẻ tới khi đủ
       `total`, nên thêm người không phải sửa lại số nào.
    """
    from io import BytesIO

    from openpyxl import Workbook  # lazy import: thiếu dep chỉ hỏng endpoint này, không sập app
    from openpyxl.styles import Font

    scope = _scope_for(authz, user)
    rows: list = []
    page = 1
    while True:
        batch, total = svc.list_employees(
            scope=scope, actor=user, q=q, department_id=department_id, status=status_filter,
            has_account=has_account, sort=sort, page=page, size=_ME_XUAT,
        )
        rows.extend(batch)
        if len(rows) >= total or not batch:
            break
        page += 1

    dept_ids = {e.department_id for e in rows if e.department_id is not None}
    user_ids = {e.user_id for e in rows if e.user_id is not None}
    names = _dept_names(depts, dept_ids)
    unames = _user_names(users, user_ids)
    rnames = _role_names(users, roles, user_ids)
    gnames = {g.id: g.name for g in svc.list_job_grades()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Nhan su"
    ws.append(list(_COT_XUAT))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for e in rows:
        r = _row(e, names, unames, rnames, gnames)
        ws.append([
            r.code or "",
            r.full_name or "",
            r.department_name or "",
            r.role_name or r.position or "",
            r.job_grade_name or "",
            _NHAN_TRANG_THAI.get(r.status, r.status or ""),
            # Ngày vào ghi dạng chuỗi dd/mm/yyyy: để nguyên kiểu ngày thì Excel mỗi máy hiện một
            # định dạng theo vùng, kế toán đối chiếu là lệch.
            r.hire_date.strftime("%d/%m/%Y") if r.hire_date else "",
            r.account_username or "",
        ])
    for idx, width in enumerate((14, 26, 22, 22, 16, 14, 12, 18), start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="danh-sach-nhan-vien.xlsx"'},
    )


@router.get("", response_model=EmployeeListOut)
def list_employees(
    svc: Service,
    authz: Authz,
    users: Users,
    depts: Depts,
    roles: Roles,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
    q: str | None = Query(default=None),
    department_id: int | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    has_account: bool | None = Query(default=None),
    sort: str = Query(default="code"),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> EmployeeListOut:
    scope = _scope_for(authz, user)
    rows, total = svc.list_employees(
        scope=scope, actor=user, q=q, department_id=department_id, status=status_filter,
        has_account=has_account, sort=sort, page=page, size=size,
    )
    dept_ids = {e.department_id for e in rows if e.department_id is not None}
    user_ids = {e.user_id for e in rows if e.user_id is not None}
    names = _dept_names(depts, dept_ids)
    unames = _user_names(users, user_ids)
    rnames = _role_names(users, roles, user_ids)
    gnames = {g.id: g.name for g in svc.list_job_grades()}

    return EmployeeListOut(
        items=[_row(e, names, unames, rnames, gnames) for e in rows],
        total=total, page=page, size=size,
        kpis=_kpis(svc.list_scoped_all(scope=scope, actor=user)),
    )


def _kpis(book) -> EmployeeKpis:
    from datetime import date, timedelta

    soon = date.today() + timedelta(days=30)
    return EmployeeKpis(
        total=len(book),
        active=sum(1 for e in book if e.status == "active"),
        probation=sum(1 for e in book if e.status == "probation"),
        on_leave=sum(1 for e in book if e.status == "on_leave"),
        resigned=sum(1 for e in book if e.status == "resigned"),
        probation_ending_soon=sum(
            1 for e in book
            if e.status == "probation" and e.probation_end_date is not None
            and date.today() <= e.probation_end_date <= soon
        ),
    )


@router.get("/meta", response_model=EmployeeMetaOut)
def get_meta(
    svc: Service,
    users: Users,
    depts: Depts,
    roles: Roles,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> EmployeeMetaOut:
    """Dropdown data: departments + roles + login accounts not yet linked to any employee."""
    departments = depts.list_all()
    # `fallback_all=False`: chưa ai tick cờ Sản xuất thì trả về RỖNG — cờ phải là sự thật.
    # FE tự xử trường hợp "chưa ai tick" (hiện ô Bậc cho mọi phòng còn hơn giấu mất ô).
    prod_ids = {d.id for d in depts.production_departments(fallback_all=False)}
    dept_opts = [DepartmentOption(id=d.id, name=d.name, la_san_xuat=(d.id in prod_ids))
                 for d in departments]
    linked = {e.user_id for e in svc.list_scoped_all(scope="all", actor=user) if e.user_id is not None}
    unlinked = [
        UserOption(id=u.id, username=u.username, name=u.name or u.username)
        for u in users.list_all()
        if u.id not in linked
    ]
    # Vai trò gắn tài khoản (wizard + tab Tài khoản & Quyền). Role thuộc 1 phòng ban nên gom
    # theo từng phòng; số phòng nhỏ nên vòng lặp này rẻ.
    role_opts = [
        RoleOption(id=r.id, name=r.name, department_id=r.department_id)
        for d in departments
        for r in roles.list_by_department(d.id)
    ]
    return EmployeeMetaOut(
        departments=dept_opts, unlinked_users=unlinked, roles=role_opts
    )


# --- create -----------------------------------------------------------------


@router.post("", response_model=EmployeeCreateOut, status_code=status.HTTP_201_CREATED)
def create_employee(
    body: EmployeeCreate,
    svc: Service,
    payroll_svc: Payroll,
    authz: Authz,
    depts: Depts,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> EmployeeCreateOut:
    data = body.model_dump()
    account = data.pop("account", None)
    initial_salary = data.pop("initial_salary", None)
    department_id = data.pop("department_id")
    status_in = data.pop("status")
    hire_date = data.pop("hire_date")
    if initial_salary is not None:
        if not authz.can(user, MODULE, "edit_salary"):
            raise HTTPException(
                status_code=403,
                detail="Ban khong co quyen khai muc luong ban dau cho nhan vien.",
            )
        effective_from = initial_salary.get("effective_from") or hire_date or date.today()
        if hire_date is not None and effective_from < hire_date:
            raise HTTPException(
                status_code=400,
                detail="Ngay hieu luc luong khong duoc truoc ngay vao lam.",
            )
        initial_salary["effective_from"] = effective_from
    try:
        employee, dup_nid, dup_si = svc.create_employee(
            actor=user, department_id=department_id, status=status_in,
            hire_date=hire_date, fields=data,
            can_edit_salary=authz.can(user, MODULE, "edit_salary"),
        )
        if initial_salary is not None:
            payroll_svc.set_salary(
                employee_id=employee.id,
                actor=user,
                amount_mode="manual",
                **initial_salary,
            )
        account_username = None
        if account and account.get("username"):
            employee, _ = svc.create_account(
                employee_id=employee.id, scope="all", actor=user,
                username=account["username"], password=account["password"],
                name=account.get("name"), role_id=account.get("role_id"),
            )
            account_username = account["username"]
    except EmployeeError as exc:
        _raise(exc)
    except PayrollError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from None
    return EmployeeCreateOut(
        employee=_full(employee, depts, users, svc),
        duplicate_national_id=_dup(dup_nid),
        duplicate_social_insurance=_dup(dup_si),
        account_username=account_username,
    )


# --- self-service "Hồ sơ của tôi" (chỉ cần đăng nhập; KHÔNG cần quyền nhan_su) ------
# Phải khai TRƯỚC route "/{employee_id}" để "me" không bị hiểu là id.

# Field nội bộ HCNS — ẩn khỏi self-view của chính nhân viên.
_MY_HIDDEN = ("note", "payroll_group", "pay_grade_key")


def _my_out(employee, depts: DepartmentRepository, users: UserRepository,
            svc=None) -> EmployeeOut:
    out = _full(employee, depts, users, svc)
    for f in _MY_HIDDEN:
        setattr(out, f, None)
    return out


@router.get("/me", response_model=MyProfileOut)
def my_profile(svc: Service, depts: Depts, users: Users, user: SelfUser) -> MyProfileOut:
    emp = svc.my_employee(user=user)
    if emp is None:
        return MyProfileOut(has_employee=False, employee=None)
    return MyProfileOut(has_employee=True, employee=_my_out(emp, depts, users, svc))


@router.put("/me", response_model=MyProfileOut)
def update_my_profile(body: MyContactIn, svc: Service, depts: Depts, users: Users, user: SelfWriter) -> MyProfileOut:
    try:
        emp = svc.update_my_contact(user=user, fields=body.model_dump(exclude_unset=True))
    except EmployeeError as exc:
        _raise(exc)
    return MyProfileOut(has_employee=True, employee=_my_out(emp, depts, users, svc))


@router.get("/me/events", response_model=EmployeeEventsOut)
def my_events(svc: Service, users: Users, user: SelfUser) -> EmployeeEventsOut:
    items = []
    for ev in svc.my_events(user=user):
        row = EmployeeEventOut.model_validate(ev)
        if ev.actor_user_id is not None:
            u = users.get_by_id(ev.actor_user_id)
            row.actor_name = (u.name or u.username) if u is not None else None
        items.append(row)
    return EmployeeEventsOut(items=items)


@router.get("/me/attachments", response_model=AttachmentsOut)
def my_attachments(svc: Service, user: SelfUser) -> AttachmentsOut:
    return AttachmentsOut(items=[AttachmentOut.model_validate(a) for a in svc.my_attachments(user=user)])


def _req_out(req, emp_names: dict[int, str]) -> UpdateRequestOut:
    out = UpdateRequestOut.model_validate(req)
    out.employee_name = emp_names.get(req.employee_id)
    return out


@router.post("/me/update-requests", response_model=UpdateRequestOut, status_code=201)
def create_my_request(body: UpdateRequestIn, svc: Service, user: SelfWriter) -> UpdateRequestOut:
    try:
        req = svc.create_update_request(user=user, changes=body.changes, reason=body.reason)
    except EmployeeError as exc:
        _raise(exc)
    return UpdateRequestOut.model_validate(req)


@router.get("/me/update-requests", response_model=UpdateRequestsOut)
def my_requests(svc: Service, user: SelfUser) -> UpdateRequestsOut:
    return UpdateRequestsOut(items=[UpdateRequestOut.model_validate(r) for r in svc.my_update_requests(user=user)])


# --- HCNS duyệt yêu cầu cập nhật (quyền chi tiết `approve`) ------------------


@router.get("/update-requests", response_model=UpdateRequestsOut)
def list_requests(svc: Service, users: Users,
                  user: Annotated[User, Depends(require_permission(MODULE, "read"))],
                  status_filter: str | None = Query(default=None, alias="status")) -> UpdateRequestsOut:
    reqs = svc.list_update_requests(status=status_filter)
    emp_names: dict[int, str] = {}
    for eid in {r.employee_id for r in reqs}:
        emp = svc.employees.get_by_id(eid)
        if emp is not None:
            emp_names[eid] = emp.full_name
    return UpdateRequestsOut(items=[_req_out(r, emp_names) for r in reqs])


@router.post("/update-requests/{request_id}/approve", response_model=UpdateRequestOut)
def approve_request(request_id: int, body: RequestDecisionIn, svc: Service, authz: Authz,
                    user: Annotated[User, Depends(require_permission(MODULE, "approve"))]) -> UpdateRequestOut:
    try:
        req = svc.decide_update_request(request_id=request_id, actor=user, approve=True, note=body.note,
                                        scope=_scope_for(authz, user),
                                        can_edit_salary=authz.can(user, MODULE, "edit_salary"))
    except EmployeeError as exc:
        _raise(exc)
    return UpdateRequestOut.model_validate(req)


@router.post("/update-requests/{request_id}/reject", response_model=UpdateRequestOut)
def reject_request(request_id: int, body: RequestDecisionIn, svc: Service, authz: Authz,
                   user: Annotated[User, Depends(require_permission(MODULE, "approve"))]) -> UpdateRequestOut:
    try:
        req = svc.decide_update_request(request_id=request_id, actor=user, approve=False,
                                        note=body.note, scope=_scope_for(authz, user))
    except EmployeeError as exc:
        _raise(exc)
    return UpdateRequestOut.model_validate(req)


# --- detail / edit ----------------------------------------------------------


# --- Danh mục bậc tay nghề (chủ 29/07/2026) ---------------------------------
# Nằm trong module `nhan_su` chứ KHÔNG ở Cấu hình lương: HCNS quản hồ sơ mới là người cần thêm
# bậc (đang tạo hồ sơ mà thiếu bậc thì phải khai được ngay), mà họ thường không có quyền lương.


@router.get("/bac-tay-nghe", response_model=JobGradesOut)
def list_job_grades(
    svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
    active_only: bool = False,
) -> JobGradesOut:
    return JobGradesOut(items=[JobGradeOut.model_validate(g)
                               for g in svc.list_job_grades(active_only=active_only)])


@router.post("/bac-tay-nghe", response_model=JobGradeOut, status_code=201)
def create_job_grade(
    body: JobGradeIn,
    svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "create"))],
) -> JobGradeOut:
    try:
        g = svc.create_job_grade(actor=user, name=body.name, code=body.code,
                                 seq=body.seq, note=body.note)
    except EmployeeError as exc:
        _raise(exc)
    return JobGradeOut.model_validate(g)


@router.put("/bac-tay-nghe/{grade_id}", response_model=JobGradeOut)
def update_job_grade(
    grade_id: int,
    body: JobGradeUpdateIn,
    svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
) -> JobGradeOut:
    try:
        g = svc.update_job_grade(actor=user, grade_id=grade_id,
                                 **body.model_dump(exclude_unset=True))
    except EmployeeError as exc:
        _raise(exc)
    return JobGradeOut.model_validate(g)


@router.delete(
    "/bac-tay-nghe/{grade_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
)
def delete_job_grade(
    grade_id: int,
    svc: Service,
    user: Annotated[User, Depends(require_permission(MODULE, "delete"))],
) -> Response:
    try:
        svc.delete_job_grade(actor=user, grade_id=grade_id)
    except EmployeeError as exc:
        _raise(exc)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/{employee_id}", response_model=EmployeeOut)
def get_employee(
    employee_id: int,
    svc: Service,
    authz: Authz,
    depts: Depts,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> EmployeeOut:
    try:
        employee = svc.get_employee(employee_id=employee_id, scope=_scope_for(authz, user), actor=user)
    except EmployeeError as exc:
        _raise(exc)
    out = _full(employee, depts, users, svc)
    if not authz.can(user, MODULE, "view_salary"):
        _mask_salary(out)
    return out


@router.put("/{employee_id}", response_model=EmployeeUpdateOut)
def update_employee(
    employee_id: int,
    body: EmployeeUpdate,
    svc: Service,
    authz: Authz,
    depts: Depts,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
) -> EmployeeUpdateOut:
    try:
        employee, dup_nid, dup_si = svc.update_employee(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user,
            fields=body.model_dump(),
            can_edit_salary=authz.can(user, MODULE, "edit_salary"),
        )
    except EmployeeError as exc:
        _raise(exc)
    return EmployeeUpdateOut(
        employee=_full(employee, depts, users, svc),
        duplicate_national_id=_dup(dup_nid),
        duplicate_social_insurance=_dup(dup_si),
    )


@router.put("/shift/bulk", response_model=AssignShiftBulkOut)
def assign_default_shift_bulk(
    body: AssignShiftBulkIn,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
) -> AssignShiftBulkOut:
    """Đặt CA NỀN cho nhiều NV trong MỘT request — nút "Đặt ca nền" ở màn Phân ca tháng.

    NV không hợp lệ trả về trong `failed` kèm lý do; các NV còn lại vẫn được ghi."""
    try:
        res = svc.set_default_shift_bulk(
            employee_ids=body.employee_ids, scope=_scope_for(authz, user), actor=user,
            shift_id=body.default_shift_id, effective_from=body.effective_from,
        )
    except EmployeeError as exc:
        _raise(exc)
    return AssignShiftBulkOut(**res)


@router.put("/{employee_id}/shift")
def assign_default_shift(
    employee_id: int,
    body: AssignShiftIn,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
) -> dict:
    """Gán ca mặc định cho NV (an toàn, chỉ đụng default_shift_id) — panel Gán ca ở Chấm công."""
    try:
        emp, assignment = svc.set_default_shift(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user,
            shift_id=body.default_shift_id, effective_from=body.effective_from,
        )
    except EmployeeError as exc:
        _raise(exc)
    return {
        "ok": True,
        "employee_id": emp.id,
        "default_shift_id": emp.default_shift_id,
        "assignment_id": assignment.id,
        "effective_from": assignment.effective_from,
    }


@router.delete("/{employee_id}/shift-history/{assignment_id}", status_code=204)
def delete_shift_assignment(
    employee_id: int,
    assignment_id: int,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
):
    """Gỡ một mốc ca nền gán nhầm (drawer lịch sử ở màn Phân ca tháng)."""
    try:
        svc.delete_shift_assignment(
            employee_id=employee_id, assignment_id=assignment_id,
            scope=_scope_for(authz, user), actor=user,
        )
    except EmployeeError as exc:
        _raise(exc)


@router.get("/{employee_id}/shift-history", response_model=ShiftAssignmentsOut)
def list_shift_history(
    employee_id: int,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> ShiftAssignmentsOut:
    try:
        rows = svc.list_shift_assignments(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user
        )
    except EmployeeError as exc:
        _raise(exc)

    today = date.today()
    items: list[ShiftAssignmentOut] = []
    for index, row in enumerate(rows):
        newer = rows[index - 1] if index > 0 else None
        effective_to = newer.effective_from - timedelta(days=1) if newer is not None else None
        out = ShiftAssignmentOut.model_validate(row)
        out.effective_to = effective_to
        out.is_current = row.effective_from <= today and (
            effective_to is None or effective_to >= today
        )
        items.append(out)
    return ShiftAssignmentsOut(employee_id=employee_id, items=items)


# --- transitions (stage changes) -------------------------------------------


@router.post("/{employee_id}/transitions", response_model=EmployeeOut)
def apply_transition(
    employee_id: int,
    body: TransitionIn,
    svc: Service,
    payroll_svc: Payroll,
    authz: Authz,
    depts: Depts,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> EmployeeOut:
    try:
        if not _can_apply_transition(authz, user, body.kind):
            raise HTTPException(
                status_code=403,
                detail="Ban khong co quyen thuc hien thao tac ho so nay.",
            )
        current = svc.get_employee(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user
        )
        employee = svc.apply_transition(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user,
            kind=body.kind, effective_date=body.effective_date, note=body.note,
            new_department_id=body.new_department_id, new_job_grade=body.new_job_grade,
            new_job_grade_id=body.new_job_grade_id,
            new_position=body.new_position, resign_reason=body.resign_reason,
        )
    except EmployeeError as exc:
        _raise(exc)
    except PayrollError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from None
    return _full(employee, depts, users, svc)


# --- Quá trình công tác + Nhật ký ------------------------------------------


@router.get("/{employee_id}/events", response_model=EmployeeEventsOut)
def list_events(
    employee_id: int,
    svc: Service,
    authz: Authz,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> EmployeeEventsOut:
    try:
        events = svc.list_events(employee_id=employee_id, scope=_scope_for(authz, user), actor=user)
    except EmployeeError as exc:
        _raise(exc)
    items = []
    for ev in events:
        row = EmployeeEventOut.model_validate(ev)
        if ev.actor_user_id is not None:
            u = users.get_by_id(ev.actor_user_id)
            row.actor_name = (u.name or u.username) if u is not None else None
        items.append(row)
    return EmployeeEventsOut(items=items)


@router.get("/{employee_id}/activity", response_model=EmployeeActivityOut)
def list_activity(
    employee_id: int,
    svc: Service,
    authz: Authz,
    users: Users,
    audit: Audit,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> EmployeeActivityOut:
    # Access check (also 404/403 for unknown/out-of-scope) before reading the log.
    try:
        svc.get_employee(employee_id=employee_id, scope=_scope_for(authz, user), actor=user)
    except EmployeeError as exc:
        _raise(exc)
    rows = audit.list_by_target(f"employee:{employee_id}")
    items = []
    for a in rows:
        actor_name = None
        if a.actor_user_id is not None:
            u = users.get_by_id(a.actor_user_id)
            actor_name = (u.name or u.username) if u is not None else None
        items.append(EmployeeActivityRowOut(
            action=a.action, target=a.target, detail=a.detail,
            actor_name=actor_name, created_at=a.created_at,
        ))
    return EmployeeActivityOut(items=items)


# --- attachments ------------------------------------------------------------


@router.get("/{employee_id}/attachments", response_model=AttachmentsOut)
def list_attachments(
    employee_id: int,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "read"))],
) -> AttachmentsOut:
    try:
        atts = svc.list_attachments(employee_id=employee_id, scope=_scope_for(authz, user), actor=user)
    except EmployeeError as exc:
        _raise(exc)
    return AttachmentsOut(items=[AttachmentOut.model_validate(a) for a in atts])


@router.post("/{employee_id}/attachments", response_model=AttachmentOut, status_code=201)
def upload_attachment(
    employee_id: int,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
    file: UploadFile = File(...),
    doc_kind: str = Form(default="khac"),
) -> AttachmentOut:
    scope = _scope_for(authz, user)
    # Access check first so we don't write a file for an inaccessible employee.
    try:
        svc.get_employee(employee_id=employee_id, scope=scope, actor=user)
    except EmployeeError as exc:
        _raise(exc)

    key, safe_name = make_key(_HR_SUBDIR, employee_id, file.filename)
    get_storage().save(key, file.file.read(), file.content_type)
    file_url = url_from_key(key)

    try:
        att = svc.add_attachment(
            employee_id=employee_id, scope=scope, actor=user, doc_kind=doc_kind,
            file_name=safe_name, file_url=file_url, file_type=file.content_type,
        )
    except EmployeeError as exc:
        _raise(exc)
    return AttachmentOut.model_validate(att)


@router.delete("/{employee_id}/attachments/{attachment_id}", status_code=204)
def delete_attachment(
    employee_id: int,
    attachment_id: int,
    svc: Service,
    authz: Authz,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
):
    try:
        svc.delete_attachment(
            employee_id=employee_id, scope=_scope_for(authz, user), actor=user,
            attachment_id=attachment_id,
        )
    except EmployeeError as exc:
        _raise(exc)


# --- account link -----------------------------------------------------------


@router.post("/{employee_id}/account", response_model=EmployeeOut)
def attach_account(
    employee_id: int,
    body: AccountIn,
    svc: Service,
    authz: Authz,
    depts: Depts,
    users: Users,
    user: Annotated[User, Depends(require_permission(MODULE, "update"))],
) -> EmployeeOut:
    """Gắn tài khoản cho hồ sơ: TẠO MỚI (`username`+`password`) — đường chính vì mọi tài
    khoản phải sinh ra từ một hồ sơ; hoặc LIÊN KẾT tài khoản có sẵn (`user_id`) để dọn
    tài khoản mồ côi cũ."""
    scope = _scope_for(authz, user)
    try:
        if body.user_id is not None:
            employee = svc.link_account(
                employee_id=employee_id, scope=scope, actor=user, user_id=body.user_id,
            )
        elif (body.username or "").strip():
            employee, _ = svc.create_account(
                employee_id=employee_id, scope=scope, actor=user,
                username=body.username or "", password=body.password or "",
                role_id=body.role_id,
            )
        else:
            raise HTTPException(
                status_code=400, detail="Cần `username` (tạo mới) hoặc `user_id` (liên kết)."
            )
    except EmployeeError as exc:
        _raise(exc)
    return _full(employee, depts, users, svc)


# GỠ `DELETE /{employee_id}/account` (gỡ liên kết): mọi tài khoản phải thuộc một hồ sơ, nên
# gỡ liên kết = đẻ ra tài khoản mồ côi = vi phạm luật. Muốn chặn một người thì KHÓA tài khoản
# (`PUT /api/users/{id}/active`); người nghỉ việc thì login tự chặn theo trạng thái hồ sơ.
# `POST /{employee_id}/account` (liên kết) GIỮ LẠI — đường dọn các tài khoản mồ côi cũ.
