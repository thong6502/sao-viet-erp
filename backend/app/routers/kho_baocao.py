"""Router — BÁO CÁO KHO (kế toán): sổ nhập-xuất + khóa kỳ (chốt sổ) + export MISA.

docs/spec-bao-cao-kho.md. CHỈ kế toán kho vào (quyền `close_book`): xem báo cáo + export Excel +
khóa/mở kỳ. Đăng ký TRƯỚC `kho.router` trong main.py vì `/api/kho/khoa-so` là path 1 đoạn sẽ bị
`/api/kho/{kho_id}` nuốt nếu khai sau (FastAPI khớp theo thứ tự).
"""
from __future__ import annotations

from datetime import date, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import select
from sqlalchemy.orm import Session

import json

from ..db import get_db
from ..deps import require_permission
from ..models.kho_hang import KhoHang
from ..models.stock_lot import StockLot
from ..models.stock_request import StockRequest
from ..models.stock_voucher import (
    VOUCHER_NHAP,
    VOUCHER_POSTED,
    VOUCHER_XUAT,
    StockVoucher,
    StockVoucherLine,
)
from ..models.user import User
from ..repositories.audit_repo import AuditLogRepository
from ..repositories.don_vi_do_repo import DonViDoRepository
from ..repositories.kho_hang_repo import KhoHangRepository
from ..repositories.kho_khoa_so_repo import KhoKhoaSoRepository
from ..repositories.kho_ky_ton_repo import KhoKyTonRepository
from ..repositories.user_repo import UserRepository
from ..repositories.vat_lieu_kho_repo import VatLieuKhoRepository
from ..services.vat_lieu_kho_service import VatLieuKhoService
from ..schemas.stock import (
    BaoCaoChuyenKhoPage,
    BaoCaoChuyenKhoRow,
    BaoCaoKhoPage,
    BaoCaoKhoRow,
    BaoCaoNXTPage,
    BaoCaoNXTRow,
    KhoaSoKyRow,
    KhoExportLogRow,
    KhoKhoaSoIn,
    KhoKhoaSoRow,
    KyDaTinhRow,
    TinhGiaKyIn,
)

router = APIRouter(prefix="/api/kho", tags=["kho-bao-cao"])
MODULE = "kho"

Db = Annotated[Session, Depends(get_db)]
CloseBookUser = Annotated[User, Depends(require_permission(MODULE, "close_book"))]


# Ghi sổ lưu mốc UTC (`ghi_so_luc`); nhưng NGÀY HẠCH TOÁN + phân kỳ sổ phải theo NGÀY LÀM VIỆC
# giờ VN (+7). Lấy `.date()` thẳng trên UTC sẽ đẩy phiếu ghi sổ lúc 00:00–07:00 VN về hôm trước
# → lệch ngày hạch toán / rớt khỏi tháng khi lọc. Quy về VN trước khi cắt ngày.
VN_TZ = timezone(timedelta(hours=7))


def _ngay_ghi_so_vn(dt) -> date | None:
    """Ngày ghi sổ theo giờ VN. `ghi_so_luc` là UTC (naive trên SQLite, aware trên Postgres)."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(VN_TZ).date()


# --- Truy vấn dòng nhập-xuất (chỉ phiếu ĐÃ GHI SỔ) -----------------------------

def _report_rows(
    db: Session, *, tu: date | None, den: date | None, kho_id: int | None,
    loai: str | None, q: str | None = None,
) -> list[BaoCaoKhoRow]:
    # `q` = tìm số CT / mã hàng / tên hàng (khớp ô Tìm ở màn) → để "lọc gì = xuất nấy".
    ql = (q or "").strip().lower()
    stmt = (
        select(StockVoucher, StockVoucherLine, StockRequest, KhoHang, StockLot)
        .join(StockVoucherLine, StockVoucherLine.voucher_id == StockVoucher.id)
        .join(StockRequest, StockRequest.id == StockVoucher.request_id, isouter=True)
        .join(KhoHang, KhoHang.id == StockVoucher.kho_id, isouter=True)
        .join(StockLot, StockLot.id == StockVoucherLine.lot_id, isouter=True)
        .where(
            StockVoucher.trang_thai == VOUCHER_POSTED,
            # ĐIỀU CHUYỂN có chiều "Chuyển kho" RIÊNG (/bao-cao/chuyen-kho) → KHÔNG lẫn vào Sổ
            # Nhập/Xuất (chuẩn kế toán: nội bộ, không nhập-mua/xuất-bán). 3 chiều tách bạch:
            # Nhập kho = nhập thật · Xuất kho = xuất thật · Chuyển kho = điều chuyển nội bộ.
            StockVoucher.dieu_chuyen.is_(False),
        )
    )
    if loai:
        stmt = stmt.where(StockVoucher.loai == loai)
    if kho_id:
        stmt = stmt.where(StockVoucher.kho_id == kho_id)
    stmt = stmt.order_by(StockVoucher.ghi_so_luc, StockVoucher.id, StockVoucherLine.id)

    results = db.execute(stmt).all()
    # Mã / tên / ĐVT mặt hàng tra từ DANH MỤC GỐC theo (hang_loai, hang_id) — bảng `materials` đã bỏ.
    hang_svc = VatLieuKhoService(VatLieuKhoRepository(db), DonViDoRepository(db))
    hang_map = hang_svc.map_theo_cap(
        list({(ln.hang_loai, ln.hang_id) for _v, ln, _req, _kho, _lot in results})
    )
    # ĐVT trên báo cáo hiện TÊN có dấu (tờ · cái · bản kẽm) thay vì MÃ ascii (to · cai · kem) —
    # `mh.don_vi_gia` là MÃ đơn vị; tra danh mục `don_vi_do` để đổi sang tên hiển thị (khớp danh mục).
    dv_ten = {d.ma: d.ten for d in DonViDoRepository(db).all_active()}

    rows: list[BaoCaoKhoRow] = []
    for v, ln, req, kho, lot in results:
        mh = hang_map.get((ln.hang_loai, ln.hang_id))
        # Lọc theo ngày GHI SỔ ở Python (tránh func.date lệ thuộc dialect + parse tz của SQLite).
        d = _ngay_ghi_so_vn(v.ghi_so_luc)
        if tu and (d is None or d < tu):
            continue
        if den and (d is None or d > den):
            continue
        # SỐ LƯỢNG · ĐƠN GIÁ · THÀNH TIỀN đều theo ĐƠN VỊ GỐC — khớp tồn kho + ĐVT danh mục và
        # khớp cách tính giá vốn phiếu (`cost_of`): NHẬP quy đơn giá về ĐV gốc, XUẤT lấy giá vốn
        # lô (vốn đã theo ĐV gốc) × `sl_goc`. Nhân nhầm `so_luong` là lệch đúng bằng hệ số quy đổi.
        qty = float(ln.sl_goc or 0)
        if v.loai == VOUCHER_NHAP:
            so = float(ln.so_luong or 0)
            price = round(float(ln.don_gia) * so / qty) if (qty and ln.don_gia is not None) else None
        else:
            price = int(lot.don_gia_nhap) if (lot is not None and lot.don_gia_nhap is not None) else None
        row = BaoCaoKhoRow(
            voucher_id=v.id,
            ngay_ghi_so=d,
            ngay_ct=v.ngay,
            so_ct=v.ma,
            loai=v.loai,
            loai_kho=req.loai_kho if req else None,
            ma_hang=getattr(mh, "ma", None),
            ten_hang=getattr(mh, "ten", None),
            dvt=dv_ten.get(getattr(mh, "don_vi_gia", None), getattr(mh, "don_vi_gia", None)),
            so_luong=qty,
            don_gia=price,
            thanh_tien=round(price * qty) if price is not None else None,
            kho_id=v.kho_id,
            kho_ten=getattr(kho, "ten", None),
            han_su_dung=getattr(lot, "hsd", None),  # NHẬP: lô vừa tạo · XUẤT: lô bị xuất
            dieu_chuyen=bool(getattr(v, "dieu_chuyen", False)),
        )
        if ql and not any(
            ql in (val or "").lower() for val in (row.so_ct, row.ma_hang, row.ten_hang)
        ):
            continue
        rows.append(row)
    return rows


@router.get("/bao-cao/dong", response_model=BaoCaoKhoPage)
def bao_cao_dong(
    db: Db,
    _: CloseBookUser,
    tu: date | None = Query(default=None),
    den: date | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    loai: str | None = Query(default=None, pattern="^(NHAP|XUAT)$"),
    q: str | None = Query(default=None),
) -> BaoCaoKhoPage:
    rows = _report_rows(db, tu=tu, den=den, kho_id=kho_id, loai=loai, q=q)
    return BaoCaoKhoPage(items=rows, total=len(rows))


# --- Điều chuyển kho: 1 dòng/mặt hàng (Xuất tại kho → Nhập tại kho) --------------

def _chuyen_kho_rows(
    db: Session, *, tu: date | None, den: date | None, kho_id: int | None, q: str | None = None,
) -> list[BaoCaoChuyenKhoRow]:
    """Dòng điều chuyển ĐÃ GHI SỔ: mỗi dòng phiếu NHẬP đích (đại diện điều chuyển) → gộp kho nguồn
    (Xuất tại kho) + kho đích (Nhập tại kho) trên cùng 1 dòng, giá vốn chốt từ nguồn."""
    ql = (q or "").strip().lower()
    stmt = (
        select(StockVoucher, StockVoucherLine, StockRequest, KhoHang)
        .join(StockVoucherLine, StockVoucherLine.voucher_id == StockVoucher.id)
        .join(StockRequest, StockRequest.id == StockVoucher.request_id, isouter=True)
        .join(KhoHang, KhoHang.id == StockVoucher.kho_id, isouter=True)  # kho ĐÍCH (nhập về)
        .where(
            StockVoucher.trang_thai == VOUCHER_POSTED,
            StockVoucher.dieu_chuyen.is_(True),
            StockVoucher.loai == VOUCHER_NHAP,  # vế NHẬP đích = đại diện cho cả điều chuyển
        )
        .order_by(StockVoucher.ghi_so_luc, StockVoucher.id, StockVoucherLine.id)
    )
    results = db.execute(stmt).all()
    # Tên kho NGUỒN (Xuất tại kho) tra từ `stock_requests.kho_nguon_id` — 1 lượt, tránh N+1.
    kho_repo = KhoHangRepository(db)
    nguon_map = {
        kid: getattr(kho_repo.get(kid), "ten", None)
        for kid in {req.kho_nguon_id for _v, _ln, req, _k in results if req and req.kho_nguon_id}
    }
    hang_svc = VatLieuKhoService(VatLieuKhoRepository(db), DonViDoRepository(db))
    hang_map = hang_svc.map_theo_cap(
        list({(ln.hang_loai, ln.hang_id) for _v, ln, _req, _k in results})
    )
    dv_ten = {d.ma: d.ten for d in DonViDoRepository(db).all_active()}

    rows: list[BaoCaoChuyenKhoRow] = []
    for v, ln, req, kho in results:
        mh = hang_map.get((ln.hang_loai, ln.hang_id))
        d = _ngay_ghi_so_vn(v.ghi_so_luc)
        if tu and (d is None or d < tu):
            continue
        if den and (d is None or d > den):
            continue
        nguon_id = getattr(req, "kho_nguon_id", None) if req else None
        # Lọc theo kho: giữ điều chuyển LIÊN QUAN kho này (xuất khỏi hoặc nhập vào).
        if kho_id and kho_id not in (v.kho_id, nguon_id):
            continue
        # SL · đơn giá vốn theo ĐƠN VỊ GỐC (khớp tồn + cách tính nhập): don_gia quy về đv gốc.
        qty = float(ln.sl_goc or 0)
        so = float(ln.so_luong or 0)
        price = round(float(ln.don_gia) * so / qty) if (qty and ln.don_gia is not None) else None
        row = BaoCaoChuyenKhoRow(
            voucher_id=v.id,
            ngay_ghi_so=d,
            ngay_ct=v.ngay,
            so_ct=v.ma,
            ma_hang=getattr(mh, "ma", None),
            ten_hang=getattr(mh, "ten", None),
            dvt=dv_ten.get(getattr(mh, "don_vi_gia", None), getattr(mh, "don_vi_gia", None)),
            so_luong=qty,
            don_gia_von=price,
            tien_von=round(price * qty) if price is not None else None,
            kho_xuat_ten=nguon_map.get(nguon_id),
            kho_nhap_ten=getattr(kho, "ten", None),
            kho_xuat_id=nguon_id,       # kho nguồn — cho FE tô kỳ đã khóa
            kho_nhap_id=v.kho_id,       # kho đích (phiếu nhập đích ghi sổ ở đây)
            dien_giai=v.ghi_chu or (req.ghi_chu if req else None),
        )
        if ql and not any(
            ql in (val or "").lower() for val in (row.so_ct, row.ma_hang, row.ten_hang)
        ):
            continue
        rows.append(row)
    return rows


@router.get("/bao-cao/chuyen-kho", response_model=BaoCaoChuyenKhoPage)
def bao_cao_chuyen_kho(
    db: Db,
    _: CloseBookUser,
    tu: date | None = Query(default=None),
    den: date | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    q: str | None = Query(default=None),
) -> BaoCaoChuyenKhoPage:
    rows = _chuyen_kho_rows(db, tu=tu, den=den, kho_id=kho_id, q=q)
    return BaoCaoChuyenKhoPage(items=rows, total=len(rows))


# --- Nhập-Xuất-Tồn theo kỳ (bình quân gia quyền cuối kỳ) — kiểu MISA "Tính giá kỳ" ------------
#
# LỚP BÁO CÁO — KHÔNG đụng engine giá vốn (kho vẫn tính đích danh per-lô). Định giá LẠI theo BÌNH
# QUÂN GIA QUYỀN CUỐI KỲ:
#   Đơn giá BQ = (GT đầu kỳ + GT nhập) / (SL đầu kỳ + SL nhập)
#   GT xuất    = BQ × SL xuất       ·   GT cuối = GT đầu + GT nhập − GT xuất   (SL tương tự)
#
# ĐẦU KỲ = CUỐI KỲ TRƯỚC, đọc từ SNAPSHOT `kho_ky_ton` (chốt lúc bấm "Tính giá kỳ") → nối chuỗi
# CHÍNH XÁC từng kỳ. Chưa có snapshot kỳ trước (kỳ đầu tiên) → đầu kỳ dựng từ luỹ kế trước `tu`
# (định giá BQ gộp lịch sử trước — chỉ dùng cho lần bootstrap). SL khớp tồn thật (Σ nhập − Σ xuất);
# GT là bản dựng theo BQ. BAO GỒM điều chuyển (nhập/xuất thật của từng kho).

def _nxt_compute(
    db: Session, *, tu: date, den: date, kho_ids: list[int] | None,
) -> dict[tuple, dict]:
    """Tính N-X-T bình quân cho từng (kho, hàng) trong [tu, den]. Đầu kỳ đọc snapshot kỳ trước;
    fold thêm chuyển động nằm GIỮA snapshot và `tu` (nếu có 'kẽ hở' chưa chốt). Trả dict[key]→số."""
    snap_map = KhoKyTonRepository(db).latest_before_map(tu, kho_ids)

    stmt = (
        select(StockVoucher, StockVoucherLine)
        .join(StockVoucherLine, StockVoucherLine.voucher_id == StockVoucher.id)
        .where(StockVoucher.trang_thai == VOUCHER_POSTED)
    )
    if kho_ids:
        stmt = stmt.where(StockVoucher.kho_id.in_(kho_ids))
    results = db.execute(stmt).all()

    class _Acc:
        __slots__ = ("gap_nhap_sl", "gap_nhap_gt", "gap_xuat_sl",
                     "nhap_sl", "nhap_gt", "xuat_sl")

        def __init__(self):
            self.gap_nhap_sl = self.gap_nhap_gt = self.gap_xuat_sl = 0.0
            self.nhap_sl = self.nhap_gt = self.xuat_sl = 0.0

    # Mọi key cần xét = key có snapshot ĐẦU KỲ ∪ key có chuyển động.
    acc: dict[tuple, _Acc] = {k: _Acc() for k in snap_map}
    for v, ln in results:
        d = _ngay_ghi_so_vn(v.ghi_so_luc)
        if d is None or d > den:
            continue
        key = (v.kho_id, ln.hang_loai, ln.hang_id)
        snap = snap_map.get(key)
        # Chuyển động ĐÃ nằm trong snapshot đầu kỳ (d ≤ ngày chốt snapshot) → bỏ, khỏi đếm 2 lần.
        if snap is not None and d <= snap.den_ngay:
            continue
        a = acc.get(key)
        if a is None:
            a = acc[key] = _Acc()
        sl = float(ln.sl_goc or 0)
        gt = round(float(ln.don_gia or 0) * float(ln.so_luong or 0)) if v.loai == VOUCHER_NHAP else 0
        trong_ky = d >= tu   # d ≤ den đã lọc ở trên; d < tu → "kẽ hở" trước kỳ
        if v.loai == VOUCHER_NHAP:
            if trong_ky:
                a.nhap_sl += sl
                a.nhap_gt += gt
            else:
                a.gap_nhap_sl += sl
                a.gap_nhap_gt += gt
        else:  # XUẤT
            if trong_ky:
                a.xuat_sl += sl
            else:
                a.gap_xuat_sl += sl

    out: dict[tuple, dict] = {}
    for key, a in acc.items():
        snap = snap_map.get(key)
        carry_sl = float(snap.sl_cuoi) if snap else 0.0
        carry_gt = int(snap.gt_cuoi) if snap else 0
        # Fold "kẽ hở" (snapshot → tu) vào đầu kỳ: định giá BQ gộp rồi trừ phần xuất kẽ hở.
        if a.gap_nhap_sl or a.gap_xuat_sl or a.gap_nhap_gt:
            mau_gap = carry_sl + a.gap_nhap_sl
            bq_gap = ((carry_gt + a.gap_nhap_gt) / mau_gap) if mau_gap > 0 else 0.0
            dau_sl = carry_sl + a.gap_nhap_sl - a.gap_xuat_sl
            dau_gt = round(carry_gt + a.gap_nhap_gt - bq_gap * a.gap_xuat_sl)
        else:
            dau_sl, dau_gt = carry_sl, carry_gt
        mau = dau_sl + a.nhap_sl
        bq = ((dau_gt + a.nhap_gt) / mau) if mau > 0 else 0.0
        xuat_gt = round(bq * a.xuat_sl)
        out[key] = dict(
            dau_sl=dau_sl, dau_gt=dau_gt,
            nhap_sl=a.nhap_sl, nhap_gt=a.nhap_gt,
            xuat_sl=a.xuat_sl, xuat_gt=xuat_gt,
            cuoi_sl=dau_sl + a.nhap_sl - a.xuat_sl,
            cuoi_gt=dau_gt + a.nhap_gt - xuat_gt,
            don_gia_bq=round(bq, 4) if mau > 0 else None,
        )
    return out


def _nxt_rows(
    db: Session, *, tu: date, den: date, kho_id: int | None, q: str | None = None,
) -> list[BaoCaoNXTRow]:
    ql = (q or "").strip().lower()
    kho_ids = [kho_id] if kho_id else None
    comp = _nxt_compute(db, tu=tu, den=den, kho_ids=kho_ids)

    hang_svc = VatLieuKhoService(VatLieuKhoRepository(db), DonViDoRepository(db))
    hang_map = hang_svc.map_theo_cap(list({(hl, hi) for _k, hl, hi in comp}))
    dv_ten = {d.ma: d.ten for d in DonViDoRepository(db).all_active()}
    kho_repo = KhoHangRepository(db)
    kho_ten = {kid: getattr(kho_repo.get(kid), "ten", None) for kid in {k[0] for k in comp}}

    rows: list[BaoCaoNXTRow] = []
    for (kid, hloai, hid), c in comp.items():
        # Bỏ dòng RỖNG hoàn toàn (không tồn đầu, không phát sinh).
        if (abs(c["dau_sl"]) < 1e-9 and abs(c["nhap_sl"]) < 1e-9 and abs(c["xuat_sl"]) < 1e-9
                and c["dau_gt"] == 0):
            continue
        mh = hang_map.get((hloai, hid))
        row = BaoCaoNXTRow(
            kho_id=kid, kho_ten=kho_ten.get(kid),
            hang_loai=hloai, hang_id=hid,
            ma_hang=getattr(mh, "ma", None), ten_hang=getattr(mh, "ten", None),
            hang_nhom="Giấy" if hloai == "giay" else "Vật tư",
            dvt=dv_ten.get(getattr(mh, "don_vi_gia", None), getattr(mh, "don_vi_gia", None)),
            dau_sl=c["dau_sl"], dau_gt=c["dau_gt"],
            nhap_sl=c["nhap_sl"], nhap_gt=c["nhap_gt"],
            xuat_sl=c["xuat_sl"], xuat_gt=c["xuat_gt"],
            cuoi_sl=c["cuoi_sl"], cuoi_gt=c["cuoi_gt"],
            don_gia_bq=round(c["don_gia_bq"], 2) if c["don_gia_bq"] is not None else None,
        )
        if ql and not any(ql in (val or "").lower() for val in (row.ma_hang, row.ten_hang)):
            continue
        rows.append(row)
    rows.sort(key=lambda r: ((r.kho_ten or "").lower(), (r.ten_hang or "").lower()))
    return rows


def _scope_kho_ids(db: Session, kho_id: int | None, tu: date, den: date) -> list[int] | None:
    """Kho trong phạm vi 'Tính giá kỳ': 1 kho, hoặc MỌI kho (trừ kho được miễn trong khoảng — nếu
    kỳ toàn kho có kho mở riêng). None = không lọc (mọi kho)."""
    if kho_id is not None:
        return [kho_id]
    mien_tru = set(KhoKhoaSoRepository(db).exempted_khos(tu, den))
    all_ids = [r for r in db.execute(select(KhoHang.id)).scalars()]
    ids = [i for i in all_ids if i not in mien_tru]
    return ids or None


@router.get("/bao-cao/nxt", response_model=BaoCaoNXTPage)
def bao_cao_nxt(
    db: Db,
    _: CloseBookUser,
    tu: date = Query(...),
    den: date = Query(...),
    kho_id: int | None = Query(default=None),
    q: str | None = Query(default=None),
) -> BaoCaoNXTPage:
    """Báo cáo N-X-T theo kỳ (bình quân gia quyền cuối kỳ). Đầu kỳ = snapshot kỳ trước ("Tính giá
    kỳ"). Kỳ chưa tính → hiện TẠM TÍNH (da_tinh=false). Chỉ kế toán kho (`close_book`)."""
    rows = _nxt_rows(db, tu=tu, den=den, kho_id=kho_id, q=q)
    kho_ids = [kho_id] if kho_id else None
    ky_repo = KhoKyTonRepository(db)
    da_tinh = ky_repo.count_for_den(den, kho_ids) > 0
    da_khoa = KhoKhoaSoRepository(db).is_locked(kho_id, den) if kho_id else \
        KhoKhoaSoRepository(db).is_locked(None, den)
    return BaoCaoNXTPage(items=rows, total=len(rows), tu=tu, den=den,
                         da_tinh=da_tinh, da_khoa=da_khoa)


def _chot_ky(
    db: Session, *, tu: date, den: date, ten: str | None, kho_ids: list[int] | None,
    khoa_so_id: int | None,
) -> None:
    """Chốt tồn cuối kỳ vào snapshot `kho_ky_ton` (xoá kỳ cũ rồi ghi mới). KHÔNG commit — caller lo.
    Dùng chung cho 'Tính giá kỳ' (bấm tay) và 'Khóa sổ' (tự chốt)."""
    comp = _nxt_compute(db, tu=tu, den=den, kho_ids=kho_ids)
    ky_repo = KhoKyTonRepository(db)
    ky_repo.delete_for_den(den, kho_ids)
    for (kid, hloai, hid), c in comp.items():
        # Bỏ dòng rỗng hoàn toàn (không tồn cuối, không phát sinh) — khỏi phình snapshot.
        if (abs(c["cuoi_sl"]) < 1e-9 and c["cuoi_gt"] == 0
                and abs(c["nhap_sl"]) < 1e-9 and abs(c["xuat_sl"]) < 1e-9
                and abs(c["dau_sl"]) < 1e-9):
            continue
        ky_repo.upsert(
            kho_id=kid, hang_loai=hloai, hang_id=hid, tu_ngay=tu, den_ngay=den,
            ten_ky=ten, sl_cuoi=c["cuoi_sl"], gt_cuoi=int(c["cuoi_gt"]),
            don_gia_bq=c["don_gia_bq"], khoa_so_id=khoa_so_id,
        )


@router.post("/bao-cao/tinh-gia-ky", response_model=BaoCaoNXTPage)
def tinh_gia_ky(payload: TinhGiaKyIn, db: Db, user: CloseBookUser) -> BaoCaoNXTPage:
    """TÍNH GIÁ KỲ (bình quân) kiểu MISA: chốt tồn cuối kỳ vào snapshot `kho_ky_ton` để kỳ sau đọc
    làm đầu kỳ. Chạy lại được (đè) bao nhiêu lần cũng được. KHÔNG đụng phiếu xuất đích danh. Ghi Lịch sử.

    CHO PHÉP tính cả kỳ ĐÃ KHÓA (luồng chốt 29/08/2026: khóa kỳ trước → vào đây chọn kỳ đã khóa rồi
    tính). An toàn vì kỳ đã khóa thì phiếu đóng băng, tính lại ra đúng số."""
    tu, den = payload.tu, payload.den
    kho_ids = _scope_kho_ids(db, payload.kho_id, tu, den)
    _chot_ky(db, tu=tu, den=den, ten=payload.ten, kho_ids=kho_ids, khoa_so_id=None)
    db.commit()
    _log_tinh_gia(db, user, kho_id=payload.kho_id, tu=tu, den=den, ten=payload.ten)
    rows = _nxt_rows(db, tu=tu, den=den, kho_id=payload.kho_id)
    return BaoCaoNXTPage(items=rows, total=len(rows), tu=tu, den=den, da_tinh=True, da_khoa=False)


@router.get("/bao-cao/ky-da-tinh", response_model=list[KyDaTinhRow])
def get_ky_da_tinh(db: Db, _: CloseBookUser) -> list[KyDaTinhRow]:
    """Danh sách các KỲ ĐÃ TÍNH GIÁ (có snapshot) — cho tab 'Kỳ đã tính'. Mới nhất trước."""
    ky_repo = KhoKyTonRepository(db)
    khoa_repo = KhoKhoaSoRepository(db)
    khos_map = ky_repo.khos_by_period()
    out: list[KyDaTinhRow] = []
    for r in ky_repo.aggregate_periods():
        khos = khos_map.get((r.tu_ngay, r.den_ngay), set())
        da_khoa = any(khoa_repo.is_locked(kid, r.den_ngay) for kid in khos)
        out.append(KyDaTinhRow(
            tu_ngay=r.tu_ngay, den_ngay=r.den_ngay, ten=r.ten,
            so_mat_hang=int(r.so_dong), so_kho=int(r.so_kho),
            tong_gt_cuoi=int(r.tong_gt), tinh_luc=r.tinh_luc, da_khoa=da_khoa,
        ))
    return out


# --- Khóa kỳ (chốt sổ) ---------------------------------------------------------

def _khoa_row(db: Session, row) -> KhoKhoaSoRow:
    kho = KhoHangRepository(db).get(row.kho_id) if row.kho_id else None
    nguoi = UserRepository(db).get_by_id(row.nguoi_khoa_id) if row.nguoi_khoa_id else None
    return KhoKhoaSoRow(
        id=row.id,
        kho_id=row.kho_id,
        kho_ten=getattr(kho, "ten", None),
        tu_ngay=row.tu_ngay,
        den_ngay=row.den_ngay,
        hanh_dong=row.hanh_dong,
        nguoi_khoa_ten=getattr(nguoi, "name", None),
        khoa_luc=row.khoa_luc,
        ten=row.ten,
    )


@router.get("/khoa-so", response_model=list[KhoKhoaSoRow])
def get_khoa_so(db: Db, _: CloseBookUser) -> list[KhoKhoaSoRow]:
    """Lịch sử thao tác khóa/mở kỳ (mới nhất trước) — dùng cho tab Lịch sử + đối chiếu đang khóa."""
    return [_khoa_row(db, r) for r in KhoKhoaSoRepository(db).history()]


@router.get("/khoa-so/ky", response_model=list[KhoaSoKyRow])
def get_khoa_so_ky(db: Db, _: CloseBookUser) -> list[KhoaSoKyRow]:
    """Các KỲ CÒN đang khóa (đã gộp khoảng liền mạch) — cho tab 'Kỳ đã khóa' chọn nhanh + xuất."""
    kho_repo = KhoHangRepository(db)
    repo = KhoKhoaSoRepository(db)
    out: list[KhoaSoKyRow] = []
    for kho_id, tu, den, khoa_luc, ten in repo.locked_periods():
        kho = kho_repo.get(kho_id) if kho_id else None
        # Kỳ TOÀN KHO: liệt kê các kho đã MỞ RIÊNG trong kỳ (miễn trừ) để hiển thị "trừ: …".
        mien_tru: list[str] = []
        if kho_id is None:
            for kid in repo.exempted_khos(tu, den):
                k = kho_repo.get(kid)
                mien_tru.append(getattr(k, "ten", None) or f"Kho #{kid}")
        out.append(KhoaSoKyRow(
            kho_id=kho_id, kho_ten=getattr(kho, "ten", None),
            tu_ngay=tu, den_ngay=den, khoa_luc=khoa_luc, ten=ten, mien_tru=mien_tru,
        ))
    return out


@router.post("/khoa-so", response_model=KhoKhoaSoRow, status_code=status.HTTP_201_CREATED)
def set_khoa_so(payload: KhoKhoaSoIn, db: Db, user: CloseBookUser) -> KhoKhoaSoRow:
    if payload.kho_id is not None and KhoHangRepository(db).get(payload.kho_id) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy kho")
    repo = KhoKhoaSoRepository(db)
    # KHÓA: không cho khóa ĐÈ lên kỳ đã khóa (kỳ mới phải bắt đầu sau ngày đã khóa gần nhất).
    if payload.hanh_dong == "khoa" and repo.overlaps_locked(
        payload.kho_id, payload.tu_ngay, payload.den_ngay
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Khoảng này chồng lấn kỳ đã khóa — chọn khoảng bắt đầu sau ngày đã khóa gần nhất.",
        )
    # KHÓA: TÊN kỳ không được TRÙNG tên một kỳ ĐANG KHÓA khác (so không phân biệt hoa/thường) —
    # chặn nhầm lẫn khi có nhiều kỳ. Bỏ trống thì không kiểm.
    if payload.hanh_dong == "khoa" and (payload.ten or "").strip():
        ten_norm = (payload.ten or "").strip().casefold()
        dang_dung = {(t[4] or "").strip().casefold() for t in repo.locked_periods() if t[4]}
        if ten_norm in dang_dung:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Tên kỳ '{payload.ten.strip()}' đã dùng cho một kỳ đang khóa — đặt tên khác.",
            )
    # MỞ: khóa sổ TUẦN TỰ — chỉ mở được phần ĐUÔI vùng khóa, tính từ NGÀY CUỐI đang khóa (cutoff).
    # → mở được cả vùng [đầu..cutoff] hoặc một đuôi [giữa..cutoff]; KHÔNG mở phần giữa/đầu để hở
    #   kỳ mới hơn, cũng không mở vượt quá vùng khóa.
    if payload.hanh_dong == "mo":
        cutoff = repo.locked_cutoff(payload.kho_id)
        if cutoff is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Không có kỳ nào đang khóa để mở.",
            )
        if payload.den_ngay != cutoff:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Mở sổ phải tính từ NGÀY CUỐI đang khóa — đặt 'Đến ngày' = "
                    f"{cutoff.strftime('%d/%m/%Y')} (chọn 'Từ ngày' tùy ý để mở phần đuôi kỳ khóa)."
                ),
            )
        # KHÔNG cho mở VẮT QUA kẽ hở / ôm luôn kỳ cũ hơn: [tu_ngay, cutoff] phải LIỀN MẠCH đang khóa.
        # → muốn mở kỳ cũ thì phải mở kỳ mới hơn TRƯỚC (mở lần lượt từ đuôi về).
        run_start = repo.locked_run_start(payload.kho_id, cutoff)
        if payload.tu_ngay < run_start:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Khoảng mở vắt qua kỳ khác hoặc ngày chưa khóa — chỉ mở được phần ĐUÔI LIỀN MẠCH "
                    f"của kỳ đang khóa (đặt 'Từ ngày' ≥ {run_start.strftime('%d/%m/%Y')}). Muốn mở kỳ "
                    f"cũ hơn thì phải mở kỳ mới hơn ({run_start.strftime('%d/%m/%Y')}–"
                    f"{cutoff.strftime('%d/%m/%Y')}) TRƯỚC."
                ),
            )
    row = repo.add(
        kho_id=payload.kho_id, tu_ngay=payload.tu_ngay, den_ngay=payload.den_ngay,
        hanh_dong=payload.hanh_dong, nguoi_khoa_id=user.id,
        ten=((payload.ten or "").strip() or None) if payload.hanh_dong == "khoa" else None,
    )
    # KHÓA sổ CHỈ khai + đóng băng kỳ (không tự tính giá — luồng chốt 29/08/2026): tính giá là bước
    # RIÊNG, vào tab N-X-T chọn kỳ đã khóa rồi bấm "Tính giá kỳ". Tách bạch cho đỡ rối.
    # MỞ sổ → BỎ CHỐT giá: xoá snapshot của kỳ vừa mở (den_ngay trong khoảng mở) để khoá lại phải
    # tính lại, không giữ số cũ (khớp luật "phải khoá mới tính" — user 2026-08-29).
    if payload.hanh_dong == "mo":
        kho_ids = [payload.kho_id] if payload.kho_id else None
        KhoKyTonRepository(db).delete_range(payload.tu_ngay, payload.den_ngay, kho_ids)
        db.commit()
    return _khoa_row(db, row)


# --- Export Excel theo mẫu MISA -----------------------------------------------

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Cột đúng thứ tự mẫu MISA (Copy of Nhap_kho.xls 33 cột / Copy of Xuat_kho.xls 51 cột).
_NHAP_HEADERS = [
    "Hiển thị trên sổ", "Loại nhập kho", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số chứng từ (*)", "Mã đối tượng", "Tên đối tượng", "Người giao hàng", "Diễn giải",
    "Nhân viên bán hàng", "Kèm theo", "Loại tiền", "Tỷ giá", "Mã hàng (*)", "Tên hàng",
    "Kho (*)", "Hàng hóa giữ hộ/bán hộ", "TK Nợ (*)", "TK Có (*)", "ĐVT", "Số lượng",
    "Đơn giá", "Thành tiền", "Thành tiền quy đổi", "Số lô", "Hạn sử dụng", "Khoản mục CP",
    "Đơn vị", "Đối tượng THCP", "Công trình", "Đơn đặt hàng", "Hợp đồng bán", "Mã thống kê",
]
_XUAT_HEADERS = [
    "Hiển thị trên sổ", "Loại xuất kho", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số chứng từ (*)", "Mẫu số HĐ", "Ký hiệu HĐ", "Mã đối tượng", "Tên đối tượng",
    "Địa chỉ/Bộ phận", "Tên người nhận/Của", "Lý do xuất/Về việc", "Nhân viên bán hàng",
    "Kèm theo", "Số lệnh điều động", "Ngày lệnh điều động", "Người vận chuyển",
    "Tên người vận chuyển", "Hợp đồng số", "Phương tiện vận chuyển", "Xuất tại kho",
    "Địa chỉ kho xuất", "Nhập tại chi nhánh", "Tên chi nhánh", "MST chi nhánh", "Nhập tại kho",
    "Địa chỉ kho nhập", "Mã hàng (*)", "Tên hàng", "Là hàng khuyến mại", "Kho (*)",
    "Hàng hóa giữ hộ/bán hộ", "TK Nợ (*)", "TK Có (*)", "ĐVT", "Số lượng", "Đơn giá bán",
    "Thành tiền", "Đơn giá vốn", "Tiền vốn", "Số lô", "Hạn sử dụng", "Đối tượng",
    "Khoản mục CP", "Đơn vị", "Đối tượng THCP", "Công trình", "Đơn đặt hàng", "Hợp đồng bán",
    "CP không hợp lý", "Mã thống kê",
]


# Cột SỐ cần format phân cách hàng nghìn (Excel hiện "." theo locale VN). Giá trị vẫn là số →
# MISA import đọc đúng; number_format chỉ đổi cách HIỂN THỊ.
_MONEY_COLS = {"Đơn giá", "Thành tiền", "Đơn giá bán", "Đơn giá vốn", "Tiền vốn", "Thành tiền quy đổi"}
_QTY_COLS = {"Số lượng"}


def _fmt_date(d: date | None) -> str | None:
    return d.strftime("%d/%m/%Y") if d else None


def _map_nhap(r: BaoCaoKhoRow) -> dict:
    return {
        # "Loại nhập kho" để TRỐNG — kế toán tự gõ mã 0/1/2/3 trong Excel/MISA.
        "Ngày hạch toán (*)": _fmt_date(r.ngay_ghi_so),
        "Ngày chứng từ (*)": _fmt_date(r.ngay_ct),
        "Số chứng từ (*)": r.so_ct,
        "Mã hàng (*)": r.ma_hang,
        "Tên hàng": r.ten_hang,
        "Kho (*)": r.kho_ten,
        "ĐVT": r.dvt,
        "Số lượng": r.so_luong,
        "Đơn giá": r.don_gia,
        "Thành tiền": r.thanh_tien,
        "Hạn sử dụng": _fmt_date(r.han_su_dung),
    }


def _map_xuat(r: BaoCaoKhoRow) -> dict:
    return {
        # "Loại xuất kho" để TRỐNG — kế toán tự gõ mã 0/1/2/3 trong Excel/MISA.
        "Ngày hạch toán (*)": _fmt_date(r.ngay_ghi_so),
        "Ngày chứng từ (*)": _fmt_date(r.ngay_ct),
        "Số chứng từ (*)": r.so_ct,
        "Mã hàng (*)": r.ma_hang,
        "Tên hàng": r.ten_hang,
        "Xuất tại kho": r.kho_ten,
        "Kho (*)": r.kho_ten,
        "ĐVT": r.dvt,
        "Số lượng": r.so_luong,
        "Đơn giá vốn": r.don_gia,
        "Tiền vốn": r.thanh_tien,
        "Hạn sử dụng": _fmt_date(r.han_su_dung),
    }


# Cột đúng thứ tự mẫu MISA "Chuyển kho" (Copy of Mau_chuyen_kho.xls — 36 cột).
_CHUYEN_HEADERS = [
    "Hiển thị trên sổ", "Hình thức chuyển kho", "Ngày chứng từ (*)", "Ngày hạch toán (*)",
    "Số chứng từ (*)", "Mẫu số hóa đơn", "Ký hiệu HĐ", "Hợp đồng KT số/Lệnh điều động", "Ngày",
    "Của", "Về việc/Diễn giải", "Đại lý/Đơn vị nhận", "Tên đại lý/Tên đơn vị nhận",
    "Mã số thuế đại lý/đơn vị nhận", "Người vận chuyển", "Tên người VC", "Hợp đồng số",
    "Phương tiện VC", "Mã hàng (*)", "Tên hàng", "Xuất tại kho (*)", "Địa chỉ kho xuất",
    "Nhập tại kho (*)", "Địa chỉ kho nhập", "Hàng hóa giữ hộ/bán hộ", "TK Nợ (*)", "TK Có (*)",
    "ĐVT", "Số lượng", "Đơn giá bán", "Thành tiền", "Đơn giá vốn", "Tiền vốn", "Số lô",
    "Hạn sử dụng", "Mã thống kê",
]


def _map_chuyen(r: BaoCaoChuyenKhoRow) -> dict:
    # CHỈ fill cột có sẵn dữ liệu (Xuất/Nhập tại kho, mặt hàng, SL, giá vốn); còn lại để TRỐNG cho
    # kế toán tự khai trong Excel/MISA ("Hình thức chuyển kho", đối tượng, vận chuyển, TK…).
    return {
        "Ngày chứng từ (*)": _fmt_date(r.ngay_ct),
        "Ngày hạch toán (*)": _fmt_date(r.ngay_ghi_so),
        "Số chứng từ (*)": r.so_ct,
        "Về việc/Diễn giải": r.dien_giai,
        "Mã hàng (*)": r.ma_hang,
        "Tên hàng": r.ten_hang,
        "Xuất tại kho (*)": r.kho_xuat_ten,
        "Nhập tại kho (*)": r.kho_nhap_ten,
        "ĐVT": r.dvt,
        "Số lượng": r.so_luong,
        "Đơn giá vốn": r.don_gia_von,
        "Tiền vốn": r.tien_von,
    }


# Mẫu Excel MISA (đã chuyển .xls→.xlsx, GIỮ màu cột + gợi ý (data-validation) + độ rộng + freeze).
# Header ở dòng 1; data ghi từ dòng 2. Xem backend/app/templates/kho/.
_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates" / "kho"


def _build_xlsx(rows: list[BaoCaoKhoRow], loai: str) -> bytes:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    is_nhap = loai == VOUCHER_NHAP
    headers = _NHAP_HEADERS if is_nhap else _XUAT_HEADERS
    mapper = _map_nhap if is_nhap else _map_xuat
    template = _TEMPLATE_DIR / ("nhap_kho.xlsx" if is_nhap else "xuat_kho.xlsx")

    if template.exists():
        # Dùng mẫu MISA thật (màu + gợi ý cột + độ rộng); header ở dòng 1, data ghi từ dòng 2.
        wb = load_workbook(template)
        ws = wb.active
    else:
        # Dự phòng nếu mẫu chưa deploy: dựng header trơn để export vẫn chạy.
        wb = Workbook()
        ws = wb.active
        ws.title = "Nhập kho" if is_nhap else "Xuất kho"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for i, r in enumerate(rows):
        vals = mapper(r)
        for ci, h in enumerate(headers, start=1):
            v = vals.get(h)
            if v is None:
                continue
            cell = ws.cell(row=2 + i, column=ci, value=v)
            if h in _MONEY_COLS:
                cell.number_format = "#,##0"      # phân cách hàng nghìn (Excel VN hiện ".")
            elif h in _QTY_COLS:
                cell.number_format = "#,##0.###"  # SL có thể lẻ, tối đa 3 số thập phân
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Ghi NHẬT KÝ mỗi lần xuất Excel (vào audit_logs) → hiện ở tab "Lịch sử thao tác" ------------
_ACTION_EXPORT = "kho_export"
_ACTION_TINH_GIA = "kho_tinh_gia"


def _log_tinh_gia(db: Session, user: User, *, kho_id: int | None,
                  tu: date, den: date, ten: str | None) -> None:
    """Ghi 1 dòng nhật ký 'Tính giá kỳ' vào audit_logs → hiện ở tab 'Lịch sử thao tác'."""
    kho_ten = getattr(KhoHangRepository(db).get(kho_id), "ten", None) if kho_id is not None else None
    AuditLogRepository(db).create(
        actor_user_id=getattr(user, "id", None),
        action=_ACTION_TINH_GIA,
        target="Tính giá kỳ",
        detail=json.dumps(
            {
                "pham_vi": f"Tính giá kỳ · {kho_ten or 'Tất cả kho'}",
                "khoang_ngay": f"{_fmt_date(tu)} – {_fmt_date(den)}",
                "ten_ky": ten,
            },
            ensure_ascii=False,
        ),
    )


def _log_export(db: Session, user: User, *, loai_label: str,
                kho_id: int | None, tu: date | None, den: date | None) -> None:
    """Ghi 1 dòng nhật ký 'xuất Excel' (ai · lúc nào · báo cáo gì · kho · khoảng ngày)."""
    kho_ten = None
    if kho_id is not None:
        kho_ten = getattr(KhoHangRepository(db).get(kho_id), "ten", None)
    pham_vi = f"{loai_label} · {kho_ten or 'Tất cả kho'}"
    khoang = (
        f"{_fmt_date(tu) or '…'} – {_fmt_date(den) or '…'}" if (tu or den) else None
    )
    # Tên kỳ: xuất TOÀN BỘ (không lọc ngày) → "Toàn bộ"; khoảng ngày TRÙNG ĐÚNG một kỳ đã khóa
    # (cùng phạm vi kho) → tên kỳ đó; khoảng ngày lẻ (không trùng kỳ) → để trống.
    ten_ky: str | None = None
    if tu is None and den is None:
        ten_ky = "Toàn bộ"
    elif tu is not None and den is not None:
        for k_kho, k_tu, k_den, _luc, k_ten in KhoKhoaSoRepository(db).locked_periods():
            if k_tu == tu and k_den == den and k_kho == kho_id:
                ten_ky = k_ten or None
                break
    AuditLogRepository(db).create(
        actor_user_id=getattr(user, "id", None),
        action=_ACTION_EXPORT,
        target=loai_label,
        detail=json.dumps(
            {"pham_vi": pham_vi, "khoang_ngay": khoang, "ten_ky": ten_ky},
            ensure_ascii=False,
        ),
    )


@router.get("/bao-cao/lich-su-export", response_model=list[KhoExportLogRow])
def get_lich_su_export(db: Db, _: CloseBookUser) -> list[KhoExportLogRow]:
    """Lịch sử thao tác báo cáo kho: XUẤT EXCEL + TÍNH GIÁ KỲ (mới nhất trước) — cho tab 'Lịch sử'."""
    users = UserRepository(db)
    audit = AuditLogRepository(db)
    out: list[KhoExportLogRow] = []
    # (action, hành động discriminator, nhãn mặc định)
    for action, hd, nhan in (
        (_ACTION_EXPORT, "export", "Xuất Excel"),
        (_ACTION_TINH_GIA, "tinh_gia", "Tính giá kỳ"),
    ):
        for r in audit.list_by_action(action, limit=200):
            try:
                d = json.loads(r.detail or "{}")
            except (ValueError, TypeError):
                d = {}
            u = users.get_by_id(r.actor_user_id) if r.actor_user_id else None
            out.append(KhoExportLogRow(
                thoi_diem=r.created_at,
                hanh_dong=hd,
                loai=r.target or nhan,
                pham_vi=d.get("pham_vi") or r.target or "—",
                khoang_ngay=d.get("khoang_ngay"),
                ten_ky=d.get("ten_ky"),
                nguoi_ten=getattr(u, "name", None),
            ))
    out.sort(key=lambda x: x.thoi_diem, reverse=True)
    return out


def _passes_funnel(
    r, *, ct_from, ct_to, sl_from, sl_to, dg_from, dg_to, tt_from, tt_to,
    price_attr: str, total_attr: str,
) -> bool:
    """Lọc funnel theo CỘT giống hệt bảng FE (inDateRange/inNumRange): khoảng BAO GỒM hai đầu, để
    trống = không chặn, giá trị None mà đang có chặn = loại. Nhờ vậy 'xuất Excel' ĐÚNG BẰNG những gì
    màn đang hiển thị, không kéo thừa dòng đã bị lọc cột (Ngày CT · Số lượng · Đơn giá · Thành tiền)."""
    d = r.ngay_ct
    if ct_from is not None and (d is None or d < ct_from):
        return False
    if ct_to is not None and (d is None or d > ct_to):
        return False

    def num_ok(v, lo, hi) -> bool:
        if lo is None and hi is None:
            return True
        if v is None:
            return False
        if lo is not None and v < lo:
            return False
        if hi is not None and v > hi:
            return False
        return True

    return (
        num_ok(r.so_luong, sl_from, sl_to)
        and num_ok(getattr(r, price_attr), dg_from, dg_to)
        and num_ok(getattr(r, total_attr), tt_from, tt_to)
    )


@router.get("/bao-cao/export.xlsx")
def export_bao_cao(
    db: Db,
    user: CloseBookUser,
    loai: str = Query(pattern="^(NHAP|XUAT)$"),
    tu: date | None = Query(default=None),
    den: date | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    q: str | None = Query(default=None),
    ct_from: date | None = Query(default=None),
    ct_to: date | None = Query(default=None),
    sl_from: float | None = Query(default=None),
    sl_to: float | None = Query(default=None),
    dg_from: float | None = Query(default=None),
    dg_to: float | None = Query(default=None),
    tt_from: float | None = Query(default=None),
    tt_to: float | None = Query(default=None),
) -> Response:
    # `_report_rows` đã LOẠI điều chuyển (chuẩn kế toán: nội bộ, không nhập-mua/xuất-bán) — nên file
    # MISA nhập/xuất tự động không dính điều chuyển; điều chuyển có mẫu "Chuyển kho" riêng.
    rows = _report_rows(db, tu=tu, den=den, kho_id=kho_id, loai=loai, q=q)
    # Áp bộ lọc funnel theo cột (nếu FE truyền) → file = đúng bảng đang xem.
    rows = [
        r for r in rows
        if _passes_funnel(r, ct_from=ct_from, ct_to=ct_to, sl_from=sl_from, sl_to=sl_to,
                          dg_from=dg_from, dg_to=dg_to, tt_from=tt_from, tt_to=tt_to,
                          price_attr="don_gia", total_attr="thanh_tien")
    ]
    content = _build_xlsx(rows, loai)
    _log_export(db, user, loai_label="Nhập kho" if loai == VOUCHER_NHAP else "Xuất kho",
                kho_id=kho_id, tu=tu, den=den)
    fname = f"bao-cao-kho-{'nhap' if loai == VOUCHER_NHAP else 'xuat'}.xlsx"
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _build_chuyen_xlsx(rows: list[BaoCaoChuyenKhoRow]) -> bytes:
    """Export điều chuyển theo mẫu MISA 'Chuyển kho' — chỉ fill cột có dữ liệu, còn lại để trống."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    template = _TEMPLATE_DIR / "chuyen_kho.xlsx"
    if template.exists():
        wb = load_workbook(template)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Chuyển kho"
        ws.append(_CHUYEN_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for i, r in enumerate(rows):
        vals = _map_chuyen(r)
        for ci, h in enumerate(_CHUYEN_HEADERS, start=1):
            v = vals.get(h)
            if v is None:
                continue
            cell = ws.cell(row=2 + i, column=ci, value=v)
            if h in _MONEY_COLS:
                cell.number_format = "#,##0"
            elif h in _QTY_COLS:
                cell.number_format = "#,##0.###"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/bao-cao/chuyen-kho/export.xlsx")
def export_chuyen_kho(
    db: Db,
    user: CloseBookUser,
    tu: date | None = Query(default=None),
    den: date | None = Query(default=None),
    kho_id: int | None = Query(default=None),
    q: str | None = Query(default=None),
    ct_from: date | None = Query(default=None),
    ct_to: date | None = Query(default=None),
    sl_from: float | None = Query(default=None),
    sl_to: float | None = Query(default=None),
    dg_from: float | None = Query(default=None),
    dg_to: float | None = Query(default=None),
    tt_from: float | None = Query(default=None),
    tt_to: float | None = Query(default=None),
) -> Response:
    rows = _chuyen_kho_rows(db, tu=tu, den=den, kho_id=kho_id, q=q)
    # Bảng Chuyển kho lọc funnel theo `don_gia_von`/`tien_von` (cột Đơn giá/Thành tiền) — khớp FE.
    rows = [
        r for r in rows
        if _passes_funnel(r, ct_from=ct_from, ct_to=ct_to, sl_from=sl_from, sl_to=sl_to,
                          dg_from=dg_from, dg_to=dg_to, tt_from=tt_from, tt_to=tt_to,
                          price_attr="don_gia_von", total_attr="tien_von")
    ]
    content = _build_chuyen_xlsx(rows)
    _log_export(db, user, loai_label="Chuyển kho", kho_id=kho_id, tu=tu, den=den)
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="bao-cao-kho-chuyen-kho.xlsx"'},
    )
