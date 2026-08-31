"""Xuất / Nhập Excel cho ĐỦ 13 màn Cấu hình danh mục.

Cơ chế dùng chung sống ở `services/catalog_excel.py` + `services/catalog_excel_specs.py`; router
chỉ điều phối (`routers/catalog_base.make_catalog_router`, tham số `excel_spec`). Test đi qua
`TestClient` thật và dựng `.xlsx` bằng `openpyxl` trong bộ nhớ — cùng đường người dùng đi.

VÌ SAO GOM NHIỀU MÀN VÀO MỘT TEST: fixture `client` drop + create lại toàn bộ schema mỗi test
(~3 giây). Kiểm cấu trúc và round-trip cho 13 màn bằng 13 test riêng là hơn một phút chỉ để dựng
đi dựng lại cùng một DB. Các test cấu trúc vì thế lặp TRONG một test; test hành vi thì tách riêng
để khi đỏ còn biết vỡ ở luật nào.

Phủ theo đúng bảng nghiệm thu đã chốt:

* xuất: đúng sheet/cột từng màn · CẢ dòng đã ngừng · đủ công thức · sạch mọi trường lịch sử;
* round-trip: xuất → xem trước → chốt mà không sửa gì ⇒ toàn bộ `khong_doi`, KHÔNG đẻ nhật ký;
* sửa: công thức · bậc tính · thứ tự · dữ liệu con · xoá dòng con · ô trống · thiếu cột/sheet;
* chặn: sai màn · sai phiên bản · trùng mã · sai kiểu · công thức quẩn · tham chiếu ma;
* giao dịch: lỗi ở DÒNG CUỐI vẫn rollback sạch cả file;
* quyền: đọc thì xuất được, nhập cần CẢ `create` lẫn `update`;
* tương thích: file Excel đời cũ (không `_meta`, tên sheet/cột cũ) vẫn nhập được.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.db import SessionLocal
from app.models.audit import AuditLog
from app.models.role import SCOPE_ALL
from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
from app.repositories.user_repo import UserRepository
from app.security import create_access_token, hash_password
from app.services.catalog_excel import PHIEN_BAN, SHEET_GIU, SHEET_META
from app.services.catalog_excel_specs import SPECS

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TO_TEST = "Hành chính nhân sự"


# ======================================================================================
# Tiện ích
# ======================================================================================


def _login(client, username="admin", password="admin123") -> dict[str, str]:
    r = client.post("/api/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _wb_tu(headers: list[str], rows: list[list], *, ten_sheet: str,
           loai: str | None = None, phien_ban: str | None = None) -> bytes:
    """Workbook TỐI THIỂU: sheet `_meta` (chỉ khi khai `loai`) + một sheet chính."""
    wb = Workbook()
    ws = wb.active
    if loai is None:
        ws.title = ten_sheet          # file đời cũ: không có `_meta`
    else:
        ws.title = SHEET_META
        ws.append(["khoa", "gia_tri"])
        ws.append(["loai", loai])
        ws.append(["phien_ban", phien_ban or PHIEN_BAN])
        ws = wb.create_sheet(ten_sheet)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    return _bytes(wb)


def _xuat(client, h, prefix: str) -> Workbook:
    r = client.get(f"{prefix}/mau-excel", headers=h)
    assert r.status_code == 200, r.text
    return load_workbook(BytesIO(r.content))


def _nhap(client, h, prefix: str, noi_dung: bytes, mode: str = "preview"):
    return client.post(
        f"{prefix}/import-excel", params={"mode": mode},
        files={"file": ("import.xlsx", noi_dung, XLSX)}, headers=h,
    )


def _bang(ws) -> tuple[list, list[list]]:
    hang = list(ws.iter_rows(values_only=True))
    if not hang:
        return [], []
    return list(hang[0]), [list(r) for r in hang[1:]
                           if r and not all(o is None for o in r)]


def _chinh(client, h, loai: str) -> tuple[list, list[list]]:
    """`(tiêu đề, dòng)` của SHEET CHÍNH màn `loai`."""
    return _bang(_xuat(client, h, PREFIX[loai])[SPECS[loai].tieu_de[:31]])


def _dong_theo_ma(tieu_de: list, dong: list[list], ma: str) -> list:
    i = tieu_de.index("Mã")
    for d in dong:
        if d[i] == ma:
            return d
    raise AssertionError(f'Không thấy mã "{ma}".')


def _dem_nhat_ky() -> int:
    db = SessionLocal()
    try:
        return db.execute(select(func.count()).select_from(AuditLog)).scalar_one()
    finally:
        db.close()


def _to() -> tuple[int, str, str]:
    """`(id, mã, tên)` của một phòng ban có thật — FK của Công đoạn và Công việc khoán."""
    db = SessionLocal()
    try:
        d = DepartmentRepository(db).get_by_name(TO_TEST)
        return d.id, d.code, d.name
    finally:
        db.close()


# ======================================================================================
# Dữ liệu nền: MỘT dòng cho mỗi màn, đủ công thức + bảng con để round-trip có nghĩa
# ======================================================================================

PREFIX = {
    "kho_hang": "/api/kho",
    "bu_hao": "/api/bu-hao",
    "khuon_be": "/api/khuon-be",
    "loai_san_pham": "/api/loai-san-pham",
    "san_xuat_ly_do": "/api/san-xuat-ly-do",
    "cong_viec_khoan": "/api/cong-viec-khoan",
    "don_vi_do": "/api/don-vi",
    "chung_loai_giay": "/api/vat-lieu-kho/chung-loai-giay",
    "giay": "/api/vat-lieu-kho/giay",
    "vat_tu": "/api/vat-lieu-kho/vat-tu-in-an",
    "thanh_pham": "/api/vat-lieu-kho/thanh-pham",
    "cong_doan": "/api/cong-doan",
    "may_thiet_bi": "/api/may-thiet-bi",
}
assert set(PREFIX) == set(SPECS), "PREFIX và SPECS phải phủ đúng 13 màn như nhau."

MA = {
    "kho_hang": "KHO-T1", "bu_hao": "BH-T1", "khuon_be": "KB-T1", "loai_san_pham": "LSP-T1",
    "san_xuat_ly_do": "LD-T1", "cong_viec_khoan": "KH-T1", "don_vi_do": "dvt1",
    "chung_loai_giay": "CL-T1", "giay": "GI-T1", "vat_tu": "VT-T1", "thanh_pham": "TP-T1",
    "cong_doan": "CD-T1", "may_thiet_bi": "MAY-T1",
}


def _tao(client, h, loai: str, payload: dict) -> dict:
    r = client.post(PREFIX[loai], json=payload, headers=h)
    assert r.status_code == 201, f"{loai}: {r.text}"
    return r.json()


def _dung_nen(client, h) -> dict[str, dict]:
    """Một dòng cho MỖI màn. Thứ tự quan trọng: màn được trỏ tới phải có trước màn trỏ đi."""
    to_id = _to()[0]
    ra: dict[str, dict] = {}
    ra["kho_hang"] = _tao(client, h, "kho_hang", {
        "ma": MA["kho_hang"], "ten": "Kho thử", "vi_tri": "Dãy A", "ghi_chu": "ghi chú kho"})
    ra["bu_hao"] = _tao(client, h, "bu_hao", {
        "ma": MA["bu_hao"], "ten": "Bù hao thử", "ghi_chu": "gc",
        "bac": [{"sl_tu": 0, "sl_den": 1000, "gia_tri": 150, "don_vi": "to"},
                {"sl_tu": 1000, "sl_den": None, "gia_tri": 3, "don_vi": "pct"}]})
    ra["khuon_be"] = _tao(client, h, "khuon_be", {
        "ma": MA["khuon_be"], "ten": "Khuôn thử", "loai": "khuon_be", "so_ke": "K1",
        "tinh_trang": "dang_dung", "ghi_chu": "gc"})
    ra["san_xuat_ly_do"] = _tao(client, h, "san_xuat_ly_do", {
        "ma": MA["san_xuat_ly_do"], "nhom": "loi", "ten": "Lỗi thử",
        "mo_ta": "mô tả", "thu_tu": 7})
    ra["cong_viec_khoan"] = _tao(client, h, "cong_viec_khoan", {
        "ma": MA["cong_viec_khoan"], "ten": "Việc khoán thử", "department_id": to_id,
        "unit": "to", "unit_price": 120, "cong_thuc_luong": "so_to * 2", "note": "gc"})
    ra["don_vi_do"] = _tao(client, h, "don_vi_do", {
        "ma": MA["don_vi_do"], "ten": "Đơn vị thử", "ho": "thanh_pham",
        "ghi_chu": "gc", "hieu_luc_tu": "2026-01-01"})
    ra["chung_loai_giay"] = _tao(client, h, "chung_loai_giay", {
        "ma": MA["chung_loai_giay"], "ten": "Couche thử", "mo_ta": "mô tả"})
    ra["giay"] = _tao(client, h, "giay", {
        "ma": MA["giay"], "ten": "Giấy thử", "gsm": 250,
        "chung_loai_giay_id": ra["chung_loai_giay"]["id"], "caliper_micron": 300,
        "tho": "canh_dai", "don_vi_gia": "kg", "don_gia": 28000, "gia_thi_truong": 30000,
        "kho_tinh_gia": True, "ghi_chu": "gc", "cong_thuc_gia": "khoi_luong * don_gia",
        "cong_thuc_luong": "dinh_luong * dai_nguyen * rong_nguyen * to_nguyen"})
    ra["vat_tu"] = _tao(client, h, "vat_tu", {
        "ma": MA["vat_tu"], "ten": "Mực thử", "don_vi_gia": "kg", "don_gia": 450000,
        "ghi_chu": "gc", "cong_thuc_gia": "so_kg * don_gia", "cong_thuc_luong": "so_to / 1000"})
    ra["thanh_pham"] = _tao(client, h, "thanh_pham", {
        "ma": MA["thanh_pham"], "ten": "Thành phẩm thử", "don_vi_gia": "cai", "ghi_chu": "gc"})
    ra["cong_doan"] = _tao(client, h, "cong_doan", {
        "ma": MA["cong_doan"], "ten": "Công đoạn thử", "ten_hien_thi": "CĐ thử",
        "nhom": "finishing", "don_vi_vao": "to", "don_vi_ra": "con",
        "cong_thuc_gia": "so_to * 100", "kieu_bu_hao": "khong",
        "che_do_tinh": "theo_san_luong", "pricing_basis": "per_other",
        "department_id": to_id, "khoan_ghi_theo": "khong",
        "nhom_may_cho_phep": ["Bế"], "setup_cost": 50000, "setup_time": 15,
        "rate_tiers": [{"from_qty": 0, "rate": 120, "kieu": "moi_dv", "driver": "so_to"},
                       {"from_qty": 5000, "rate": 90, "kieu": "moi_dv", "driver": "so_to"}],
        "size_tiers": [{"den_cm": 50, "don_gia": 1000}, {"den_cm": 80, "don_gia": 1500}],
        "ghi_chu": "gc",
        "dau_viec_dinh_muc": [{
            "piece_rate_id": ra["cong_viec_khoan"]["id"], "nang_suat_nguoi_gio": 500,
            "nang_suat_nguoi_gio_min": 400, "nang_suat_nguoi_gio_max": 600,
            "don_vi_nang_suat": "tờ/giờ", "so_nguoi_toi_thieu": 1,
            "so_nguoi_tieu_chuan": 2, "so_nguoi_toi_da": 3,
            "vat_tu_ids": [ra["vat_tu"]["id"]]}]})
    ra["loai_san_pham"] = _tao(client, h, "loai_san_pham", {
        "ma": MA["loai_san_pham"], "ten": "Hộp thử", "structural_type": "box",
        "box_sub_type": "folding_carton", "has_cover": False,
        "default_stock_class": "couche", "ghi_chu": "gc",
        "routing_template": [ra["cong_doan"]["id"]]})
    ra["may_thiet_bi"] = _tao(client, h, "may_thiet_bi", {
        "ma": MA["may_thiet_bi"], "ten": "Máy thử", "loai_may": "Máy in",
        "hang_san_xuat": "Heidelberg", "model": "SM74", "so_seri": "X1",
        "toc_do": 8000, "toc_do_min": 4000, "toc_do_max": 12000, "don_vi_toc_do": "to",
        "cong_thuc_luong": "so_to", "makeready_time_default": 30, "so_nhan_cong": 3,
        "kho_max_dai": 1020, "kho_max_rong": 720, "kho_min_dai": 300, "kho_min_rong": 200,
        "kho_kem_dai": 1000, "kho_kem_rong": 700, "vung_in_dai": 980, "vung_in_rong": 690,
        "nhip_giay_mm": 10, "le_hong_mm": 8, "duoi_thang_mau_mm": 12, "ghi_chu": "gc",
        "fields_theo_loai": {
            "chuan_bi_khoan": [{"ten": "Canh máy", "phut": 20}, {"ten": "Rửa lô", "phut": 15}],
            "lich_bao_tri": [{"id": "hm-1", "viec": "Tra dầu", "so": 1, "don_vi": "thang",
                              "ngay_bat_dau": "2026-01-05", "dung_phut": 60,
                              "hang_muc": [{"id": "hm-1-1", "ten": "Lô nước"},
                                           {"id": "hm-1-2", "ten": "Ru lô mực"}]}],
            "khoa_la_khong_hieu": {"a": 1}}})
    return ra


# ======================================================================================
# 1 · XUẤT — sheet, cột, chỉ dòng đang dùng, sạch lịch sử
# ======================================================================================


def test_xuat_dung_sheet_va_cot_cho_du_13_man(client, seed_credentials):
    """Mọi màn phải ra `_meta` (ẩn) + sheet chính + đủ sheet con, tiêu đề khớp spec."""
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)

    for loai, spec in SPECS.items():
        wb = _xuat(client, h, PREFIX[loai])

        assert SHEET_META in wb.sheetnames, loai
        assert wb[SHEET_META].sheet_state == "hidden", loai
        meta = {r[0]: r[1] for r in wb[SHEET_META].iter_rows(values_only=True) if r and r[0]}
        assert meta["loai"] == spec.loai
        assert meta["phien_ban"] == PHIEN_BAN

        ten_chinh = spec.tieu_de[:31]
        assert ten_chinh in wb.sheetnames, f"{loai}: thiếu sheet chính {ten_chinh}"
        tieu_de, dong = _bang(wb[ten_chinh])
        assert tieu_de == [c.nhan for c in spec.cot if not c.chi_nhap], loai
        assert dong, f"{loai}: sheet chính rỗng — dữ liệu nền chưa dựng?"

        for sheet in spec.sheets_con:
            assert sheet.ten[:31] in wb.sheetnames, f"{loai}: thiếu sheet con {sheet.ten}"
            cot_du_lieu = [c for c in (*sheet.khoa_phu, *sheet.cot) if not c.chi_nhap]
            mong = ([] if sheet.toan_cuc else [sheet.cot_cha]) \
                + (["Thứ tự"] if sheet.thu_tu else []) \
                + [c.nhan for c in cot_du_lieu]
            assert _bang(wb[sheet.ten[:31]])[0] == mong, f"{loai}/{sheet.ten}"


def test_xuat_khong_kem_bat_ky_truong_lich_su_nao(client, seed_credentials):
    """Nhật ký · "lần trước công thức" · mốc hệ thống · id trần: KHÔNG được lộ ra file.

    Nhập ngược một cột lịch sử là ghi đè quá khứ bằng thứ người ta vừa gõ tay — đúng cái mà cột
    lịch sử sinh ra để chống.
    """
    for loai, spec in SPECS.items():
        for c in spec.cot:
            assert not c.field.endswith(("_truoc", "_sua_luc")), f"{loai}: cột lịch sử {c.field}"
            assert c.field not in ("id", "created_at", "updated_at", "anh_url"), \
                f"{loai}: cột dẫn xuất/hệ thống {c.field}"
        for sheet in spec.sheets_con:
            for c in (*sheet.khoa_phu, *sheet.cot):
                assert not c.field.endswith(("_truoc", "_sua_luc")), f"{loai}/{sheet.ten}"

    # Và trên file thật: không tiêu đề nào mời người ta sửa lịch sử.
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    cam = ("lần trước", "lịch sử", "nhật ký", "sửa lúc", "phiên bản giá")
    for loai in SPECS:
        wb = _xuat(client, h, PREFIX[loai])
        for ten in wb.sheetnames:
            if ten in (SHEET_META, SHEET_GIU):
                continue
            for nhan in _bang(wb[ten])[0]:
                for xau in cam:
                    assert xau not in str(nhan).lower(), f"{loai}/{ten}: cột lịch sử “{nhan}”"


def test_xuat_du_moi_o_cong_thuc_dang_chay(client, seed_credentials):
    """Mọi ô công thức của mọi màn phải có mặt KÈM GIÁ TRỊ — đây chính là thứ người ta mở file ra
    để sửa. Thiếu một cột thì lần nhập sau nó bị hiểu là "giữ nguyên", im lặng, mãi mãi."""
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)

    mong = {
        "giay": {"Công thức giá": "khoi_luong * don_gia",
                 "Công thức lượng": "dinh_luong * dai_nguyen * rong_nguyen * to_nguyen"},
        "vat_tu": {"Công thức giá": "so_kg * don_gia", "Công thức lượng": "so_to / 1000"},
        "cong_doan": {"Công thức giá": "so_to * 100", "Công thức sản lượng": None},
        "cong_viec_khoan": {"Công thức lượng": "so_to * 2"},
        "may_thiet_bi": {"Công thức lượng": "so_to"},
    }
    for loai, cot in mong.items():
        tieu_de, dong = _chinh(client, h, loai)
        d = _dong_theo_ma(tieu_de, dong, MA[loai])
        for nhan, gia_tri in cot.items():
            assert nhan in tieu_de, f"{loai}: thiếu cột “{nhan}”"
            assert d[tieu_de.index(nhan)] == gia_tri, f"{loai}/{nhan}"


def test_xuat_ca_dong_da_ngung_kem_trang_thai_false(client, seed_credentials):
    """Dòng đã NGỪNG vẫn ra file, cột `Trạng thái` = FALSE.

    Bỏ chúng đi thì bộ file không tự nhập lại được sang máy khác (dòng còn hiệu lực ở màn này
    được phép trỏ tới dòng đã ngừng ở màn kia), và ngừng dùng thành đường một chiều.
    """
    h = _login(client, **seed_credentials)
    con = _tao(client, h, "kho_hang", {"ma": "KHO-SONG", "ten": "Kho còn dùng"})
    chet = _tao(client, h, "kho_hang", {"ma": "KHO-NGUNG", "ten": "Kho ngừng dùng"})
    assert client.patch(f"/api/kho/{chet['id']}/active", json={"active": False},
                        headers=h).status_code == 200

    tieu_de, dong = _chinh(client, h, "kho_hang")
    ma = [d[0] for d in dong]
    assert con["ma"] in ma and chet["ma"] in ma
    i = tieu_de.index("Trạng thái")
    assert _dong_theo_ma(tieu_de, dong, con["ma"])[i] is True
    assert _dong_theo_ma(tieu_de, dong, chet["ma"])[i] is False


def test_bat_lai_dong_da_ngung_bang_excel(client, seed_credentials):
    """Đã ngừng thì bật lại được bằng chính file xuất ra — chiều ngược của test trên."""
    h = _login(client, **seed_credentials)
    o = _tao(client, h, "kho_hang", {"ma": "KHO-BAT-LAI", "ten": "Kho bật lại"})
    assert client.patch(f"/api/kho/{o['id']}/active", json={"active": False},
                        headers=h).status_code == 200

    noi_dung = _wb_tu(["Mã", "Tên", "Trạng thái"], [["KHO-BAT-LAI", "Kho bật lại", True]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    kq = _nhap(client, h, "/api/kho", noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq
    assert client.get(f"/api/kho/{o['id']}", headers=h).json()["active"] is True


def test_xuat_dich_fk_thanh_ma_nghiep_vu_khong_phai_id(client, seed_credentials):
    """FK ra file dưới dạng MÃ (kèm cột tên để đối chiếu) — số id chỉ có nghĩa trong đúng một DB."""
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    _, ma_to, ten_to = _to()

    tieu_de, dong = _chinh(client, h, "giay")
    d = _dong_theo_ma(tieu_de, dong, MA["giay"])
    assert d[tieu_de.index("Chủng loại giấy")] == nen["chung_loai_giay"]["ma"]

    tieu_de, dong = _chinh(client, h, "cong_doan")
    d = _dong_theo_ma(tieu_de, dong, MA["cong_doan"])
    assert d[tieu_de.index("Mã tổ phụ trách")] == ma_to
    assert d[tieu_de.index("Tên tổ phụ trách")] == ten_to


def test_xuat_bang_con_ra_sheet_doc_duoc_khong_phai_json(client, seed_credentials):
    """Bậc bù hao · bậc đơn giá · đầu việc · gói bảo trì: mỗi thứ một sheet, một dòng một bậc."""
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)

    tieu_de, dong = _bang(_xuat(client, h, PREFIX["bu_hao"])["Bậc bù hao"])
    assert tieu_de == ["Mã", "Thứ tự", "SL từ", "SL đến", "Giá trị", "Đơn vị"]
    assert [d[2:] for d in dong] == [[0, 1000, 150.0, "to"], [1000, None, 3.0, "pct"]]

    wb = _xuat(client, h, PREFIX["cong_doan"])
    assert [d[2:] for d in _bang(wb["Bậc theo khổ"])[1]] == [[50.0, 1000.0], [80.0, 1500.0]]
    assert [d[2] for d in _bang(wb["Nhóm máy cho phép"])[1]] == ["Bế"]
    assert _bang(wb["Đầu việc định mức"])[1][0][2] == nen["cong_viec_khoan"]["ma"]
    assert [d[2:] for d in _bang(wb["Vật tư đầu việc"])[1]] == [
        [nen["cong_viec_khoan"]["ma"], nen["vat_tu"]["ma"]]]

    wb = _xuat(client, h, PREFIX["may_thiet_bi"])
    assert [d[2:] for d in _bang(wb["Khoản chuẩn bị"])[1]] == [["Canh máy", 20.0],
                                                              ["Rửa lô", 15.0]]
    assert _bang(wb["Gói bảo trì"])[1][0][2] == "hm-1"
    assert [d[2:] for d in _bang(wb["Hạng mục bảo trì"])[1]] == [
        ["hm-1", "hm-1-1", "Lô nước"], ["hm-1", "hm-1-2", "Ru lô mực"]]
    # Khoá JSON không diễn giải được: giữ ở sheet ẨN để round-trip không đánh rơi, nhưng KHÔNG
    # bày ra cho người khai sửa tay.
    assert SHEET_GIU in wb.sheetnames and wb[SHEET_GIU].sheet_state == "hidden"
    assert "khoa_la_khong_hieu" in _bang(wb[SHEET_GIU])[1][0][1]


# ======================================================================================
# 2 · ROUND-TRIP — xuất rồi nhập lại mà không sửa gì ⇒ KHÔNG có gì đổi
# ======================================================================================


def test_round_trip_du_13_man_khong_doi_va_khong_de_nhat_ky(client, seed_credentials):
    """Xuất → xem trước → chốt, không sửa một ô nào ⇒ mọi dòng `khong_doi`, nhật ký đứng yên.

    Test này canh CẢ HAI chiều cùng lúc: xuất sai kiểu (số ra chuỗi, ngày ra số Excel) hay nhập
    sai kiểu đều lộ ra ngay dưới dạng một `cap_nhat` bất ngờ, khỏi cần assert từng ô.
    """
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)

    for loai in SPECS:
        prefix = PREFIX[loai]
        noi_dung = _bytes(_xuat(client, h, prefix))

        xem = _nhap(client, h, prefix, noi_dung).json()
        assert xem["hop_le"], f"{loai}: {xem['loi']}"
        assert xem["da_ghi"] is False, f"{loai}: xem trước KHÔNG được ghi"
        assert (xem["tao_moi"], xem["cap_nhat"]) == (0, 0), f"{loai}: {xem}"
        assert xem["khong_doi"] == xem["tong_dong"] > 0, f"{loai}: {xem}"

        truoc = _dem_nhat_ky()
        chot = _nhap(client, h, prefix, noi_dung, mode="commit").json()
        assert chot["hop_le"] and chot["da_ghi"], f"{loai}: {chot}"
        assert (chot["tao_moi"], chot["cap_nhat"]) == (0, 0), f"{loai}: {chot}"
        assert _dem_nhat_ky() == truoc, f"{loai}: dòng không đổi vẫn đẻ nhật ký"


# ======================================================================================
# 3 · SỬA — công thức, bậc tính, thứ tự, dữ liệu con, ô trống, thiếu cột/sheet
# ======================================================================================


def test_sua_cong_thuc_o_sheet_chinh(client, seed_credentials):
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix, ten = PREFIX["giay"], SPECS["giay"].tieu_de[:31]

    wb = _xuat(client, h, prefix)
    tieu_de = _bang(wb[ten])[0]
    i_ma, i_ct = tieu_de.index("Mã"), tieu_de.index("Công thức giá")
    for hang in wb[ten].iter_rows(min_row=2):
        if hang[i_ma].value == MA["giay"]:
            hang[i_ct].value = "khoi_luong * don_gia * 1.1"

    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"] and (kq["tao_moi"], kq["cap_nhat"]) == (0, 1), kq
    r = client.get(f"{prefix}/{nen['giay']['id']}", headers=h).json()
    assert r["cong_thuc_gia"] == "khoi_luong * don_gia * 1.1"


def test_sua_bac_bu_hao_va_xoa_mot_bac(client, seed_credentials):
    """Sheet con CÓ MẶT ⇒ thay TRỌN tập con: xoá dòng khỏi file là xoá cấu hình con đó."""
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix = PREFIX["bu_hao"]

    wb = _xuat(client, h, prefix)
    ws = wb["Bậc bù hao"]
    ws.cell(row=2, column=5).value = 180          # bậc 1: 150 → 180 tờ
    ws.delete_rows(3)                             # xoá bậc 2

    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq
    bac = client.get(f"{prefix}/{nen['bu_hao']['id']}", headers=h).json()["bac"]
    assert [(b["sl_tu"], b["gia_tri"]) for b in bac] == [(0, 180.0)]


def test_thu_tu_o_sheet_con_quyet_dinh_thu_tu_luu(client, seed_credentials):
    """Cột `Thứ tự` quyết định, KHÔNG phải vị trí dòng — người ta chèn dòng mới ở cuối file."""
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix = PREFIX["bu_hao"]

    wb = _xuat(client, h, prefix)
    ws = wb["Bậc bù hao"]
    ws.cell(row=2, column=2).value = 2
    ws.cell(row=3, column=2).value = 1

    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"], kq
    bac = client.get(f"{prefix}/{nen['bu_hao']['id']}", headers=h).json()["bac"]
    assert [b["sl_tu"] for b in bac] == [1000, 0]


def test_them_dong_con_moi_cho_ma_da_co(client, seed_credentials):
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix = PREFIX["cong_doan"]

    wb = _xuat(client, h, prefix)
    wb["Bậc theo khổ"].append([MA["cong_doan"], 3, 120, 2200])
    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq

    cd = client.get(f"{prefix}/{nen['cong_doan']['id']}", headers=h).json()
    assert [t["den_cm"] for t in cd["size_tiers"]] == [50, 80, 120]


def test_o_trong_xoa_gia_tri_con_cot_vang_mat_thi_giu_nguyen(client, seed_credentials):
    """Hai luật ngược nhau trên cùng một ô trống — phân biệt bằng CỘT CÓ trong tiêu đề hay không."""
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix, ten = PREFIX["kho_hang"], SPECS["kho_hang"].tieu_de[:31]

    # (a) cột "Ghi chú" CÓ mặt, ô để trống ⇒ xoá giá trị.
    wb = _xuat(client, h, prefix)
    tieu_de = _bang(wb[ten])[0]
    for hang in wb[ten].iter_rows(min_row=2):
        if hang[tieu_de.index("Mã")].value == MA["kho_hang"]:
            hang[tieu_de.index("Ghi chú")].value = None
    assert _nhap(client, h, prefix, _bytes(wb), mode="commit").json()["cap_nhat"] == 1
    r = client.get(f"{prefix}/{nen['kho_hang']['id']}", headers=h).json()
    assert r["ghi_chu"] is None and r["vi_tri"] == "Dãy A"

    # (b) file CHỈ có "Mã" + "Tên" ⇒ mọi cột khác giữ nguyên, không bị hiểu là xoá trắng.
    noi_dung = _wb_tu(["Mã", "Tên"], [[MA["kho_hang"], "Kho đổi tên"]],
                      ten_sheet=ten, loai="kho_hang")
    kq = _nhap(client, h, prefix, noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq
    r = client.get(f"{prefix}/{nen['kho_hang']['id']}", headers=h).json()
    assert r["ten"] == "Kho đổi tên" and r["vi_tri"] == "Dãy A"


def test_thieu_sheet_con_thi_giu_nguyen_du_lieu_con(client, seed_credentials):
    """File không có sheet con ⇒ dữ liệu con Ở NGUYÊN.

    Bẫy thật: `CongDoanRepository._sau_gan` thay TRỌN bảng định mức mỗi lần ghi, kể cả khi khoá
    vắng mặt — không gán lại thì nhập một file thiếu sheet cũng xoá sạch định mức mọi công đoạn.
    """
    h = _login(client, **seed_credentials)
    nen = _dung_nen(client, h)
    prefix, ten = PREFIX["cong_doan"], SPECS["cong_doan"].tieu_de[:31]

    noi_dung = _wb_tu(["Mã", "Tên"], [[MA["cong_doan"], "Công đoạn đổi tên"]],
                      ten_sheet=ten, loai="cong_doan")
    kq = _nhap(client, h, prefix, noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq

    cd = client.get(f"{prefix}/{nen['cong_doan']['id']}", headers=h).json()
    assert cd["ten"] == "Công đoạn đổi tên"
    assert len(cd["dau_viec_dinh_muc"]) == 1, "thiếu sheet con KHÔNG được xoá định mức"
    assert cd["dau_viec_dinh_muc"][0]["vat_tu_ids"] == [nen["vat_tu"]["id"]]
    assert len(cd["size_tiers"]) == 2 and cd["nhom_may_cho_phep"] == ["Bế"]


def test_tao_moi_va_cap_nhat_cung_mot_file(client, seed_credentials):
    """UPSERT theo mã; dòng KHÔNG có trong file thì giữ nguyên, không bị coi là đã xoá."""
    h = _login(client, **seed_credentials)
    _tao(client, h, "kho_hang", {"ma": "KHO-CU", "ten": "Kho cũ"})
    _tao(client, h, "kho_hang", {"ma": "KHO-YEN", "ten": "Kho không đụng tới"})

    noi_dung = _wb_tu(["Mã", "Tên", "Vị trí"],
                      [["KHO-CU", "Kho cũ đổi tên", "B2"], ["KHO-MOI", "Kho mới", "C3"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    kq = _nhap(client, h, "/api/kho", noi_dung, mode="commit").json()
    assert (kq["tao_moi"], kq["cap_nhat"], kq["khong_doi"]) == (1, 1, 0), kq

    dong = {d["ma"]: d for d in client.get("/api/kho", headers=h).json()["items"]}
    assert dong["KHO-CU"]["ten"] == "Kho cũ đổi tên"
    assert dong["KHO-MOI"]["vi_tri"] == "C3"
    assert dong["KHO-YEN"]["ten"] == "Kho không đụng tới"


def test_dat_trang_thai_false_de_ngung_dung(client, seed_credentials):
    """Muốn NGỪNG một dòng thì đặt `Trạng thái=FALSE` rồi nhập lại."""
    h = _login(client, **seed_credentials)
    o = _tao(client, h, "kho_hang", {"ma": "KHO-TAT", "ten": "Kho sắp tắt"})
    noi_dung = _wb_tu(["Mã", "Tên", "Trạng thái"], [["KHO-TAT", "Kho sắp tắt", False]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")

    kq = _nhap(client, h, "/api/kho", noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq
    assert client.get(f"/api/kho/{o['id']}", headers=h).json()["active"] is False


# ======================================================================================
# 4 · CHẶN — sai màn, sai phiên bản, trùng mã, sai kiểu, công thức quẩn, tham chiếu ma
# ======================================================================================


def test_file_cua_man_khac_bi_tu_choi_422(client, seed_credentials):
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(["Mã", "Tên"], [["KHO-X", "Kho"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    r = _nhap(client, h, PREFIX["bu_hao"], noi_dung)
    assert r.status_code == 422, r.text
    assert "kho_hang" in r.json()["detail"]


def test_file_phien_ban_moi_hon_bi_tu_choi_422(client, seed_credentials):
    """File của một bản sau: cột có thể đã đổi nghĩa. Đọc bừa thì lỗi không lộ ra ở đây — nó lộ
    ra ở bảng giá sai mấy tuần sau."""
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(["Mã", "Tên"], [["KHO-X", "Kho"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang",
                      phien_ban=str(int(PHIEN_BAN) + 1))
    assert _nhap(client, h, "/api/kho", noi_dung).status_code == 422


def test_file_khong_doc_duoc_tra_422(client, seed_credentials):
    h = _login(client, **seed_credentials)
    assert _nhap(client, h, "/api/kho", b"khong-phai-xlsx").status_code == 422


def test_thieu_cot_ma_tra_422(client, seed_credentials):
    """Không có cột `Mã` thì không biết cập nhật dòng nào — từ chối cả file, đừng đoán."""
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(["Tên"], [["Kho gì đó"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    assert _nhap(client, h, "/api/kho", noi_dung).status_code == 422


def test_trung_ma_trong_cung_file(client, seed_credentials):
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(["Mã", "Tên"], [["KHO-A", "Kho A"], ["KHO-A", "Kho A lần hai"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    kq = _nhap(client, h, "/api/kho", noi_dung, mode="commit").json()
    assert kq["hop_le"] is False and kq["da_ghi"] is False
    assert (kq["loi"][0]["dong"], kq["loi"][0]["cot"]) == (3, "Mã")
    assert client.get("/api/kho", params={"q": "KHO-A"}, headers=h).json()["total"] == 0


def test_sai_kieu_bao_dung_cot_va_dong(client, seed_credentials):
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    noi_dung = _wb_tu(["Mã", "Tên", "Định lượng (gsm)"],
                      [[MA["giay"], "Giấy thử", "không-phải-số"]],
                      ten_sheet=SPECS["giay"].tieu_de[:31], loai="giay")
    kq = _nhap(client, h, PREFIX["giay"], noi_dung).json()
    assert kq["hop_le"] is False
    assert (kq["loi"][0]["dong"], kq["loi"][0]["cot"]) == (2, "Định lượng (gsm)")


def test_tham_chieu_khong_ton_tai(client, seed_credentials):
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    noi_dung = _wb_tu(["Mã", "Tên", "Chủng loại giấy"],
                      [[MA["giay"], "Giấy thử", "KHONG-CO-MA-NAY"]],
                      ten_sheet=SPECS["giay"].tieu_de[:31], loai="giay")
    kq = _nhap(client, h, PREFIX["giay"], noi_dung).json()
    assert kq["hop_le"] is False
    assert "KHONG-CO-MA-NAY" in kq["loi"][0]["ly_do"]
    # Câu lỗi phải chỉ luôn MÀN phải khai trước — thứ tự nhập giữa các màn là ràng buộc thật,
    # người nhập không có cách nào đoán ra từ chữ "không tìm thấy".
    assert "Chủng loại giấy" in kq["loi"][0]["ly_do"]


def test_cap_quy_doi_tro_toi_don_vi_nam_DUOI_trong_cung_file(client, seed_credentials):
    """Cặp quy đổi tra mã ở NGAY TRONG danh mục này — đích nằm dưới nguồn vẫn phải chạy.

    Đây là lý do bộ file xuất ra không tự nhập lại được vào DB trắng trước 30/08/2026: `ram → to`
    mà `to` ở dòng 20 còn `ram` ở dòng 19. Cả file là MỘT giao dịch nên nhập hai lượt cũng vô ích.
    """
    h = _login(client, **seed_credentials)
    prefix = PREFIX["don_vi_do"]

    wb = _xuat(client, h, prefix)
    ws = wb[SPECS["don_vi_do"].tieu_de[:31]]
    tieu_de = [c.value for c in ws[1]]

    def _dong(ma: str, ten: str) -> list:
        o = dict.fromkeys(tieu_de)
        o.update({"Mã": ma, "Tên": ten, "Loại đo": "thanh_pham", "Trạng thái": True})
        return [o[t] for t in tieu_de]

    ws.append(_dong("zz-nguon", "Đơn vị nguồn"))     # nguồn TRƯỚC
    ws.append(_dong("zz-dich", "Đơn vị đích"))       # đích SAU
    wb["Các cặp quy đổi"].append(["zz-nguon", "zz-dich", 12, None])

    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"], kq["loi"]
    assert kq["tao_moi"] == 2, kq

    dong = _chinh(client, h, "don_vi_do")[1]
    assert "zz-dich" in [d[0] for d in dong]
    cap = _bang(_xuat(client, h, prefix)["Các cặp quy đổi"])[1]
    assert ["zz-nguon", "zz-dich", 12.0, None] in cap


def test_ten_lech_ma_bi_chan_chu_khong_ghi_nham(client, seed_credentials):
    """Cột tên `chi_doc` là CHỐT CHẶN: mã do máy cấp nên `PB008` hai máy dễ là hai phòng khác nhau.

    Trước 30/08/2026 cột tên bị bỏ qua hoàn toàn ⇒ dữ liệu vào nhầm phòng, không một dòng cảnh báo.
    """
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    prefix = PREFIX["cong_viec_khoan"]

    wb = _xuat(client, h, prefix)
    ws = wb[SPECS["cong_viec_khoan"].tieu_de[:31]]
    tieu_de = [c.value for c in ws[1]]
    ws.cell(row=2, column=tieu_de.index("Tên tổ") + 1).value = "Tổ Không Có Thật"

    kq = _nhap(client, h, prefix, _bytes(wb), mode="commit").json()
    assert kq["hop_le"] is False and kq["da_ghi"] is False, kq
    assert kq["loi"][0]["cot"] == "Tên tổ"
    assert "Tổ Không Có Thật" in kq["loi"][0]["ly_do"]


def test_ten_khop_ma_thi_van_nhap_binh_thuong(client, seed_credentials):
    """Chốt chặn trên không được cản đường file xuất ra dùng lại y nguyên."""
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    prefix = PREFIX["cong_viec_khoan"]

    kq = _nhap(client, h, prefix, _bytes(_xuat(client, h, prefix)), mode="commit").json()
    assert kq["hop_le"] and kq["khong_doi"] >= 1, kq


def test_cong_thuc_khong_hop_le_bi_service_chan(client, seed_credentials):
    """Công thức sản lượng của bước NGOÀI dòng giấy dùng chính số của bước (`sl_ra`) ⇒ quẩn.

    Service chặn, và vì lỗi nghiệp vụ đi cùng đường với lỗi ô nên nó thành một DÒNG LỖI đọc được
    chứ không phải 500.
    """
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    noi_dung = _wb_tu(
        ["Mã", "Tên", "Đơn vị vào", "Đơn vị ra", "Công thức sản lượng"],
        [[MA["cong_doan"], "Công đoạn thử", "kem", "bai", "sl_ra * 2"]],
        ten_sheet=SPECS["cong_doan"].tieu_de[:31], loai="cong_doan")
    kq = _nhap(client, h, PREFIX["cong_doan"], noi_dung, mode="commit").json()
    assert kq["hop_le"] is False and kq["da_ghi"] is False
    assert "E-CD-VONG-TRON" in kq["loi"][0]["ly_do"], kq


def test_loi_o_dong_cuoi_van_rollback_toan_bo_file(client, seed_credentials):
    """Cả file là MỘT giao dịch — hai dòng đầu hợp lệ vẫn KHÔNG được ghi khi dòng cuối hỏng.

    Nhập nửa vời một bảng cấu hình là để lại một trạng thái chưa ai duyệt, mà không ai biết nó
    đã dừng ở dòng nào.
    """
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(
        ["Mã", "Nhóm", "Tên", "Thứ tự hiện"],
        [["LD-A", "loi", "Lý do A", 1],
         ["LD-B", "loi", "Lý do B", 2],
         ["LD-C", "nhom-khong-ton-tai", "Lý do C", 3]],
        ten_sheet=SPECS["san_xuat_ly_do"].tieu_de[:31], loai="san_xuat_ly_do")

    kq = _nhap(client, h, PREFIX["san_xuat_ly_do"], noi_dung, mode="commit").json()
    assert kq["hop_le"] is False and kq["da_ghi"] is False
    assert (kq["tong_dong"], kq["tao_moi"]) == (3, 2), kq   # đếm được, nhưng KHÔNG ghi
    assert client.get(PREFIX["san_xuat_ly_do"], headers=h).json()["total"] == 0


def test_xem_truoc_khong_bao_gio_ghi(client, seed_credentials):
    h = _login(client, **seed_credentials)
    noi_dung = _wb_tu(["Mã", "Tên"], [["KHO-XT", "Kho xem trước"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")

    kq = _nhap(client, h, "/api/kho", noi_dung).json()
    assert kq["hop_le"] and kq["tao_moi"] == 1 and kq["da_ghi"] is False
    assert client.get("/api/kho", headers=h).json()["total"] == 0

    kq = _nhap(client, h, "/api/kho", noi_dung, mode="commit").json()
    assert kq["da_ghi"] and kq["tao_moi"] == 1
    assert client.get("/api/kho", headers=h).json()["total"] == 1


# ======================================================================================
# 5 · QUYỀN
# ======================================================================================


def _tai_khoan(username: str, **quyen) -> dict[str, str]:
    """Tài khoản mang MỘT vai chỉ có đúng các ô quyền được liệt kê trên `dm_kho_hang`."""
    db = SessionLocal()
    try:
        users, depts, roles = UserRepository(db), DepartmentRepository(db), RoleRepository(db)
        to = depts.get_by_name(TO_TEST)
        vai = roles.get_by_name_and_department(f"vai-{username}", to.id) \
            or roles.create(name=f"vai-{username}", department_id=to.id)
        roles.set_permission(role_id=vai.id, module_key="dm_kho_hang", scope=SCOPE_ALL, **quyen)
        u = users.get_by_username(username) or users.create(
            username=username, name=username, password_hash=hash_password("x"))
        users.set_assignment(u, department_id=to.id, role_id=vai.id, is_active=True)
        return {"Authorization": f"Bearer {create_access_token(str(u.id))}"}
    finally:
        db.close()


def test_quyen_doc_xuat_duoc_nhung_khong_nhap_duoc(client, seed_credentials):
    """XUẤT gác bằng quyền ĐỌC; NHẬP đòi CẢ `create` LẪN `update`.

    Không phải "một trong hai": tới lúc gác thì chưa ai biết file mang dòng mới hay dòng cũ —
    biết được thì đã đọc xong file rồi.
    """
    _login(client, **seed_credentials)          # seeder RBAC chạy khi app khởi động
    hd = _tai_khoan("dm-chi-doc", can_read=True)
    assert client.get("/api/kho/mau-excel", headers=hd).status_code == 200

    noi_dung = _wb_tu(["Mã", "Tên"], [["KHO-Q", "Kho quyền"]],
                      ten_sheet=SPECS["kho_hang"].tieu_de[:31], loai="kho_hang")
    assert _nhap(client, hd, "/api/kho", noi_dung).status_code == 403

    hc = _tai_khoan("dm-chi-tao", can_read=True, can_create=True)
    assert _nhap(client, hc, "/api/kho", noi_dung).status_code == 403

    hu = _tai_khoan("dm-chi-sua", can_read=True, can_update=True)
    assert _nhap(client, hu, "/api/kho", noi_dung).status_code == 403

    hdu = _tai_khoan("dm-du-quyen", can_read=True, can_create=True, can_update=True)
    r = _nhap(client, hdu, "/api/kho", noi_dung, mode="commit")
    assert r.status_code == 200, r.text
    assert r.json()["da_ghi"] is True


# ======================================================================================
# 6 · TƯƠNG THÍCH — file Excel đời cũ của 5 màn từng bật Excel
# ======================================================================================


def test_nhan_file_doi_cu_khong_co_meta_va_ten_cot_cu(client, seed_credentials):
    """File cũ: không `_meta`, sheet tên `cong_doan`, cột "Tổ phụ trách" ghi TÊN tổ, nhóm máy nằm
    trong MỘT ô ngăn bằng dấu phẩy.

    Người dùng còn giữ những file này trên máy. Từ chối chúng là bắt họ khai lại từ đầu chỉ vì hệ
    đổi định dạng.
    """
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    ten_to = _to()[2]

    noi_dung = _wb_tu(["Mã", "Tên", "Nhóm", "Tổ phụ trách", "Nhóm máy cho phép"],
                      [[MA["cong_doan"], "Công đoạn đời cũ", "Gia công sau in",
                        ten_to, "Bế, Máy in"]],
                      ten_sheet="cong_doan")
    kq = _nhap(client, h, PREFIX["cong_doan"], noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq

    cd = client.get(PREFIX["cong_doan"], params={"q": MA["cong_doan"]},
                    headers=h).json()["items"][0]
    assert cd["ten"] == "Công đoạn đời cũ"
    assert cd["nhom"] == "finishing", "nhãn tiếng Việt của dropdown phải dịch được về mã"
    assert cd["nhom_may_cho_phep"] == ["Bế", "Máy in"]


def test_nhan_file_doi_cu_cua_may_voi_cot_json_tho(client, seed_credentials):
    """Cột "Field theo loại (JSON)" đời cũ vẫn đọc được, dù nay đã tách thành ba sheet dễ đọc."""
    h = _login(client, **seed_credentials)
    _dung_nen(client, h)
    noi_dung = _wb_tu(
        ["Mã", "Tên", "Loại máy", "Field theo loại (JSON)"],
        [[MA["may_thiet_bi"], "Máy thử", "Máy in",
          '{"chuan_bi_khoan": [{"ten": "Canh máy", "phut": 25}]}']],
        ten_sheet="may_thiet_bi")
    kq = _nhap(client, h, PREFIX["may_thiet_bi"], noi_dung, mode="commit").json()
    assert kq["hop_le"] and kq["cap_nhat"] == 1, kq

    may = client.get(PREFIX["may_thiet_bi"], params={"q": MA["may_thiet_bi"]},
                     headers=h).json()["items"][0]
    assert may["fields_theo_loai"]["chuan_bi_khoan"] == [{"ten": "Canh máy", "phut": 25}]


# ======================================================================================
# 7 · GUARD — thêm cột cấu hình mà quên khai Excel thì test này ĐỎ
# ======================================================================================


def test_guard_moi_truong_ghi_duoc_deu_di_qua_excel_hoac_duoc_loai_tru():
    """Mọi field `repo.fields` phải có mặt ở Excel, hoặc nằm trong `loai_tru` kèm lý do.

    Đây là cái chốt của cả tính năng: thêm một ô công thức mới ở màn nào đó rồi quên khai spec thì
    nó lặng lẽ rơi khỏi file xuất — người dùng sửa file, nhập lại, và ô đó bị hiểu là "giữ nguyên"
    mãi mãi, không ai thấy gì sai.
    """
    for loai, spec in SPECS.items():
        assert spec.repo_cls is not None, f"{loai}: spec chưa khai `repo_cls` cho guard test"
        cua_cot = {c.field for c in spec.cot if not c.chi_doc}
        cua_sheet = {s.trong_json[0] if s.trong_json else s.field
                     for s in spec.sheets_con if s.field or s.trong_json}
        thieu = set(spec.repo_cls.fields) - cua_cot - cua_sheet - spec.loai_tru
        assert not thieu, (
            f"{loai}: field {sorted(thieu)} ghi được ở repo nhưng KHÔNG có trong Excel. "
            f"Khai cột/sheet ở `catalog_excel_specs.py`, hoặc thêm vào `loai_tru` kèm lý do.")


def test_guard_khong_khai_thua_field_khong_ai_ghi_duoc():
    """Chiều ngược: cột Excel trỏ vào field repo KHÔNG cho ghi thì nhập vào là rơi im lặng."""
    ngoai_le = {"ma"}          # `ma` do nền danh mục quản, không nằm trong `repo.fields`
    for loai, spec in SPECS.items():
        thua = {c.field for c in spec.cot if not c.chi_doc} - set(spec.repo_cls.fields) - ngoai_le
        assert not thua, f"{loai}: cột Excel {sorted(thua)} trỏ vào field repo không cho ghi"
