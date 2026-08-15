"""Lương (module `luong`, Phase 1): params + quy tắc, engine tính lương (prorate công,
%thử việc, chuyên cần, BHXH), khai báo/điều chỉnh lương, tạm ứng, tạo/khóa bảng lương."""
from __future__ import annotations

from datetime import date
from types import SimpleNamespace

from datetime import datetime, timezone

from app.db import SessionLocal
from app.repositories.employee_repo import EmployeeRepository
from app.repositories.payroll_repo import PayrollRepository
from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
from app.repositories.user_repo import UserRepository
from app.security import create_access_token, hash_password
from app.services.payroll_service import PayrollService

ADMIN = {"username": "admin", "password": "admin123"}


def _admin_token(client) -> str:
    return client.post("/api/auth/login", json=ADMIN).json()["access_token"]


def _h(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _dept_id(name: str) -> int:
    db = SessionLocal()
    try:
        return DepartmentRepository(db).get_by_name(name).id
    finally:
        db.close()


def _sales_token() -> str:
    db = SessionLocal()
    try:
        users = UserRepository(db)
        existing = users.get_by_username("sales-luong")
        if existing is not None:
            return create_access_token(str(existing.id))
        kd = DepartmentRepository(db).get_by_name("Kinh doanh")
        role = RoleRepository(db).get_by_name_and_department("NV Sales", kd.id)
        u = users.create(username="sales-luong", name="S", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=kd.id, role_id=role.id, is_active=True)
        return create_access_token(str(u.id))
    finally:
        db.close()


def _make_emp(client, token, *, name, payroll_group=None, pay_grade_key=None,
              gender="male", hire_date="2020-01-01", status=None) -> int:
    body = {"full_name": name, "department_id": _dept_id("Hành chính nhân sự"),
            "hire_date": hire_date, "gender": gender}
    if payroll_group:
        body["payroll_group"] = payroll_group
    if pay_grade_key:
        body["pay_grade_key"] = pay_grade_key
    if status:
        body["status"] = status
    return client.post("/api/employees", json=body, headers=_h(token)).json()["employee"]["id"]


def _sal(**kw):
    """Salary namespace cho engine test — mức nền + BH đều bám `luong_vi_tri` (chủ 2026-07-20:
    lương vị trí = lương cơ bản = mức đóng BH). Field đủ cho mọi getattr của _compute."""
    base = dict(amount_mode="manual", base_amount=None, luong_vi_tri=0, luong_trach_nhiem=0,
                insurance_base=None, allowance=0, chuyen_can=0, phu_cap_ca=0, phu_cap_tham_nien=0,
                insurance_elsewhere=False, union_member=False, source_salary_row_id=None)
    base.update(kw)
    return SimpleNamespace(**base)


# --- params + rules ---------------------------------------------------------


def test_params_defaults_and_rbac(client):
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["standard_cong_default"] == 26 and p["probation_ratio"] == 0.80

    upd = client.put("/api/luong/params", json={"standard_cong_default": 24}, headers=_h(token))
    assert upd.status_code == 200 and upd.json()["standard_cong_default"] == 24
    # khôi phục
    client.put("/api/luong/params", json={"standard_cong_default": 26}, headers=_h(token))

    # NV Sales không có quyền lương
    assert client.get("/api/luong/params", headers=_h(_sales_token())).status_code == 403


def test_rule_crud(client):
    token = _admin_token(client)
    created = client.post("/api/luong/rules", json={
        "payroll_group": "to_in", "pay_grade_key": "tho_1", "monthly_amount": 25_000_000,
        "effective_from": "2026-01-01",
    }, headers=_h(token))
    assert created.status_code == 201
    rid = created.json()["id"]
    listed = client.get("/api/luong/rules", headers=_h(token)).json()["items"]
    assert any(r["id"] == rid and r["monthly_amount"] == 25_000_000 for r in listed)

    upd = client.put(f"/api/luong/rules/{rid}", json={
        "payroll_group": "to_in", "pay_grade_key": "tho_1", "monthly_amount": 26_000_000,
    }, headers=_h(token))
    assert upd.json()["monthly_amount"] == 26_000_000
    assert client.delete(f"/api/luong/rules/{rid}", headers=_h(token)).status_code == 204


# --- engine (unit) ----------------------------------------------------------


def test_payroll_resolves_historical_status_and_department_from_employee_events(client):
    token = _admin_token(client)
    old_dept = _dept_id("Hành chính nhân sự")
    new_dept = _dept_id("Kinh doanh")
    employee_id = _make_emp(
        client, token, name="NV lịch sử lương", hire_date="2026-01-01", status="probation"
    )
    assert client.post(
        f"/api/employees/{employee_id}/transitions",
        json={"kind": "confirm", "effective_date": "2026-03-01"},
        headers=_h(token),
    ).status_code == 200
    assert client.post(
        f"/api/employees/{employee_id}/transitions",
        json={"kind": "transfer", "new_department_id": new_dept, "effective_date": "2026-04-01"},
        headers=_h(token),
    ).status_code == 200

    db = SessionLocal()
    try:
        employees = EmployeeRepository(db)
        svc = PayrollService(PayrollRepository(db), employees, attendance=None)
        employee = employees.get_by_id(employee_id)
        assert svc._employment_context_on(employee, date(2026, 2, 28)) == ("probation", old_dept)
        assert svc._employment_context_on(employee, date(2026, 3, 31)) == ("active", old_dept)
        assert svc._employment_context_on(employee, date(2026, 4, 30)) == ("active", new_dept)
    finally:
        db.close()


def test_compute_engine(client):
    """Prorate công + %thử việc + chuyên cần + BHXH — kiểm trực tiếp _compute."""
    client  # đảm bảo app khởi động (tạo bảng + migration)
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="ut_grp", monthly_amount=10_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        on = date(2026, 6, 1)

        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="ut_grp", pay_grade_key=None)

        # nửa công (13/26) → lương công = 5tr; chưa đủ công → chuyên cần 0.
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=10_000_000), params=params, actual_cong=13,
                         standard_cong=26, on=on)
        assert v["monthly_salary"] == 10_000_000
        assert v["luong_cong"] == 5_000_000
        assert v["chuyen_can"] == 0
        assert v["bhxh"] == round(10_000_000 * 0.105)   # 1.050.000, không prorate

        # đủ công → lương công đủ. Chuyên cần = số khai ở HỒ SƠ NV (chưa khai → 0đ; từ 2026-07-23
        # bỏ mức mặc định công ty, không còn tự rơi về 300k).
        v2 = svc._compute(employee=emp, salary=_sal(luong_vi_tri=10_000_000), params=params, actual_cong=26,
                          standard_cong=26, on=on)
        assert v2["luong_cong"] == 10_000_000 and v2["chuyen_can"] == 0
        assert v2["gross"] == 10_000_000
        # Khai chuyên cần ở hồ sơ NV → mới ra tiền.
        v2b = svc._compute(employee=emp, salary=_sal(luong_vi_tri=10_000_000, chuyen_can=300_000),
                           params=params, actual_cong=26, standard_cong=26, on=on)
        assert v2b["chuyen_can"] == 300_000 and v2b["gross"] == 10_300_000

        # thử việc → ×0.8.
        emp_tv = SimpleNamespace(status="probation", hire_date=date(2026, 5, 1), gender="male",
                                 payroll_group="ut_grp", pay_grade_key=None)
        v3 = svc._compute(employee=emp_tv, salary=_sal(luong_vi_tri=10_000_000), params=params, actual_cong=26,
                          standard_cong=26, on=on)
        assert v3["monthly_salary"] == 10_000_000       # mức gốc (chưa nhân)
        assert v3["luong_cong"] == 8_000_000            # 10tr × 0.80 × 1.0 (công ty dùng 80%)
        assert v3["bhxh"] == 0                          # thử việc KHÔNG đóng BHXH (HĐ thử việc)
    finally:
        db.close()


def test_ot_and_night_pay(client):
    """Tăng ca (hệ số phẳng) cộng vào gross, KHÔNG prorate theo công. Tiền ca đêm KHÔNG còn
    tính theo % đơn giá công (PRD v2 C4) → tổ chưa khai đơn giá ca thì night_pay = 0."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="ot_grp", monthly_amount=26_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()   # standard_hours_per_day=8, ot_multiplier=1.5
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="ot_grp", pay_grade_key=None)
        # std 26 → 1 công = 1.000.000; giờ = 125.000. OT 120' (2h)×1.5 = 375.000.
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params, actual_cong=26,
                         standard_cong=26, ot_minutes=120, night_days=2, on=date(2026, 6, 1))
        assert v["ot_pay"] == 375_000
        assert v["night_pay"] == 0            # chưa khai đơn giá ca của tổ → không ra tiền
        # Chuyên cần = 0 (hồ sơ NV chưa khai; không còn mức mặc định công ty từ 2026-07-23).
        assert v["gross"] == 26_000_000 + 375_000
    finally:
        db.close()


def test_piece_work_dept_skips_ot(client):
    """Tổ khoán (has_piece_work): KHÔNG tính tăng ca theo giờ; tổ thường vẫn có tăng ca."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="pw_grp", monthly_amount=26_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="pw_grp", pay_grade_key=None)
        # Cùng dữ liệu OT 120', chỉ khác cờ tổ khoán.
        v_norm = svc._compute(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params, actual_cong=26,
                              standard_cong=26, ot_minutes=120, on=date(2026, 6, 1))
        v_piece = svc._compute(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params, actual_cong=26,
                               standard_cong=26, ot_minutes=120, has_piece_work=True,
                               on=date(2026, 6, 1))
        assert v_norm["ot_pay"] == 375_000          # tổ thường vẫn tính tăng ca
        assert v_piece["ot_pay"] == 0               # tổ khoán bỏ tăng ca giờ
        # Lương công giữ nguyên, gross tổ khoán KHÔNG có phần tăng ca.
        assert v_piece["gross"] == v_norm["gross"] - 375_000
    finally:
        db.close()


def test_bhxh_cap(client):
    """Pha 4a: BHXH/BHYT áp trần bh_base_cap, BHTN áp trần bhtn_base_cap RIÊNG."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="hi_grp", monthly_amount=60_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()   # bh_base_cap=50.6tr, bhtn_base_cap=106.2tr
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="hi_grp", pay_grade_key=None)
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=60_000_000), params=params, actual_cong=26,
                         standard_cong=26, on=date(2026, 6, 1))
        # base 60tr > trần BHXH/BHYT 50.6tr → phần đó trên 50.6tr; BHTN 60tr < 106.2tr → trên 60tr.
        assert v["bhxh"] == round(50_600_000 * (0.08 + 0.015) + 60_000_000 * 0.01)
    finally:
        db.close()


def test_probation_no_bhxh(client):
    """Pha 4a: NV thử việc KHÔNG đóng BHXH (HĐ thử việc)."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="pb_grp", monthly_amount=12_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="probation", hire_date=date(2026, 5, 1), gender="male",
                              payroll_group="pb_grp", pay_grade_key=None)
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=12_000_000), params=params, actual_cong=26,
                         standard_cong=26, on=date(2026, 6, 1))
        assert v["bhxh"] == 0 and v["insurance_base"] == 0
        assert v["luong_cong"] == round(12_000_000 * 0.80)   # thử việc 80% (mặc định công ty)
    finally:
        db.close()


def test_insurance_elsewhere_only_tnld_bnn(client):
    """BH đóng ở nơi khác: công ty KHÔNG trừ BHXH/BHYT/BHTN của NV (bhxh=0) nhưng GIỮ insurance_base
    (> 0, KHÁC thử việc vốn = 0) để đoàn phí công đoàn VẪN tính. TNLĐ-BNN là chi phí phía công ty,
    không vào bảng lương tháng (chỉ hiển thị ở màn Sửa lương). Param `tnld_bnn_rate` mặc định 0.5%."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.update_params(cong_doan_rate=0.005)
        params = svc.get_params()
        assert float(params.tnld_bnn_rate) == 0.005            # mặc định TNLĐ-BNN 0.5%
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group=None, pay_grade_key=None, dependents_count=0)
        # union_member=True: đoàn viên nên đoàn phí CĐ VẪN trừ dù BH đóng ở nơi khác.
        v = svc._compute(employee=emp,
                         salary=_sal(luong_vi_tri=10_000_000, insurance_elsewhere=True, union_member=True),
                         params=params, actual_cong=26, standard_cong=26, on=date(2026, 6, 1))
        assert v["bhxh"] == 0                                  # KHÔNG trừ BHXH/BHYT/BHTN của NV
        assert v["insurance_base"] == 10_000_000               # GIỮ mức nền (khác thử việc = 0)
        assert v["cong_doan"] == round(10_000_000 * 0.005) and v["cong_doan"] > 0   # đoàn phí CĐ VẪN trừ (đoàn viên)
        # tnld_bnn_rate sửa được (round-trip qua update_params)
        svc.update_params(tnld_bnn_rate=0.007)
        assert float(svc.get_params().tnld_bnn_rate) == 0.007
        svc.update_params(cong_doan_rate=0, tnld_bnn_rate=0.005)   # khôi phục
    finally:
        db.close()


def test_compute_day_cong_late_early_minutes():
    """`compute_day_cong` cho ra SỐ PHÚT đi trễ (quá dung sai) + về sớm — nền phạt tự động."""
    from app.services.attendance_service import compute_day_cong
    # Ca 08:00–17:00 (480–1020), dung sai 5'. Vào 08:20 (500' → trễ 15' quá grace), ra 16:40 (1000' → sớm 20').
    v = compute_day_cong(start_min=480, end_min=1020, is_overnight=False, grace_min=5,
                         first_in_min=500, main_out_min=1000)
    assert v["late_minutes"] == 15 and v["early_minutes"] == 20
    # Vào trong dung sai (08:03) + ra đúng giờ → 0/0.
    v2 = compute_day_cong(start_min=480, end_min=1020, is_overnight=False, grace_min=5,
                          first_in_min=483, main_out_min=1020)
    assert v2["late_minutes"] == 0 and v2["early_minutes"] == 0


def test_late_penalty_amount_tiers(client):
    """Tra bảng phạt cho 1 lần vi phạm: bậc ĐẦU có up_to_minute ≥ phút; ∞ là chốt cuối."""
    from app.services.payroll_service import _late_penalty_amount
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        bks = svc.get_late_penalty_brackets()   # seed mặc định 20/40/100/150k
        assert _late_penalty_amount(0, bks) == 0
        assert _late_penalty_amount(10, bks) == 20_000     # ≤15'
        assert _late_penalty_amount(15, bks) == 20_000
        assert _late_penalty_amount(16, bks) == 40_000     # ≤30'
        assert _late_penalty_amount(45, bks) == 100_000    # ≤60'
        assert _late_penalty_amount(120, bks) == 150_000   # >60' (∞)
    finally:
        db.close()


def test_di_tre_auto_from_attendance(client):
    """END-TO-END: NV đi trễ (không phép) → generate() TỰ điền ô "Đi trễ" theo bảng phạt; HCNS sửa
    tay thì khóa (tính lại không đè); "về tự động" thì tính lại từ chấm công."""
    from datetime import datetime as _dt, timezone as _tz
    from app.repositories.attendance_repo import AttendanceRepository
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Trễ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        shift = arepo.create_shift(name="HC", start_minute=480, end_minute=1020,
                                   is_overnight=False, grace_minutes=5)
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift.id
        db.commit()
        # 2026-06-15 (Thứ Hai): vào 08:30 VN (01:30 UTC → trễ 25') + ra 17:00 VN (10:00 UTC).
        arepo.create_log(employee_id=eid, check_type="in",
                         checked_at=_dt(2026, 6, 15, 1, 30, tzinfo=_tz.utc), within_range=True)
        arepo.create_log(employee_id=eid, check_type="out",
                         checked_at=_dt(2026, 6, 15, 10, 0, tzinfo=_tz.utc), within_range=True)
    finally:
        db.close()
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert line["di_tre"] == 40_000 and line["di_tre_manual"] is False   # 25' → bậc ≤30' = 40.000
    # Sửa tay → khóa; generate lại KHÔNG đè.
    client.put(f"/api/luong/lines/{line['id']}", json={"di_tre": 5_000}, headers=_h(token))
    l2 = next(l for l in client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                                     headers=_h(token)).json()["lines"] if l["employee_id"] == eid)
    assert l2["di_tre"] == 5_000 and l2["di_tre_manual"] is True
    # Về tự động → update_line tính lại từ chấm công NGAY (trả về dòng đã cập nhật).
    l3 = client.put(f"/api/luong/lines/{line['id']}",
                    json={"di_tre_manual": False}, headers=_h(token)).json()
    assert l3["di_tre"] == 40_000 and l3["di_tre_manual"] is False


def test_compute_day_cong_night_minutes():
    """SỐ PHÚT rơi 22h–06h: giờ đêm TRONG ca + giờ TĂNG CA ĐÊM (sau end_ref)."""
    from app.services.attendance_service import compute_day_cong
    # Ca qua đêm 22:00–06:00 (1320→360) làm đủ → 8h đêm, 0 OT đêm.
    v = compute_day_cong(start_min=1320, end_min=360, is_overnight=True, grace_min=5,
                         first_in_min=1320, main_out_min=360)
    assert v["night_minutes"] == 480 and v["ot_night_minutes"] == 0
    # Ca qua đêm 20:00–04:00 → giờ đêm 22:00–04:00 = 6h.
    v2 = compute_day_cong(start_min=1200, end_min=240, is_overnight=True, grace_min=5,
                          first_in_min=1200, main_out_min=240)
    assert v2["night_minutes"] == 360
    # Ca THƯỜNG 14:00–22:00 kết thúc 22h; ra ca chính 22:00 rồi PHIÊN TĂNG CA 22:00–23:00 (phiếu tới
    # 24:00) → giờ đêm TRONG ca = 0, tăng ca đêm 60' (chủ soi).
    v3 = compute_day_cong(start_min=840, end_min=1320, is_overnight=False, grace_min=5,
                          first_in_min=840, main_out_min=1320,
                          ot_in_min=1320, ot_out_min=1380, ot_window=(1320, 1440))
    assert v3["night_minutes"] == 0 and v3["ot_night_minutes"] == 60
    # Ca ngày 08:00–17:00 → không giờ đêm.
    v4 = compute_day_cong(start_min=480, end_min=1020, is_overnight=False, grace_min=5,
                          first_in_min=480, main_out_min=1020)
    assert v4["night_minutes"] == 0 and v4["ot_night_minutes"] == 0


def test_night_premium_engine(client):
    """Engine: (A) giờ đêm trong ca × hệ số → premium; (B) tăng ca đêm Đ98.3 (200%); miễn TNCN."""
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()   # night_pct 0.3 · ot_night_extra_pct 0.2 · ot_multiplier 1.5 · 8h/ngày
        assert float(params.ot_night_extra_pct) == 0.2 and float(params.night_pct) == 0.3
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group=None, pay_grade_key=None, dependents_count=0)
        base = dict(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params,
                    actual_cong=26, standard_cong=26, on=date(2026, 6, 1))   # đơn giá 1tr/công → 125k/giờ
        v0 = svc._compute(**base)
        # (A) 8h đêm × (1.3−1) = 144' → 125k × 144/60 = 300k.
        v = svc._compute(**{**base, "night_premium_minutes": 144})
        assert v["night_premium_pay"] == 300_000
        assert v["pit"] == v0["pit"] and v["pit_taxable"] == v0["pit_taxable"]   # premium MIỄN TNCN
        # (B) tăng ca đêm ngày thường 120' (2h): 125k × 2 × (0.3 + 0.2×1) = 125k (khớp OT đêm 200%).
        vb = svc._compute(**{**base, "ot_night_normal_minutes": 120})
        assert vb["night_premium_pay"] == 125_000
    finally:
        db.close()


def test_night_pay_auto_from_attendance(client):
    """END-TO-END: gán ca qua đêm (hệ số 1.3) + chấm công 1 đêm 22:00–06:00 → generate() tự điền
    'Phụ cấp ca đêm theo giờ' = 30% đơn giá 8 giờ."""
    from datetime import datetime as _dt, timezone as _tz
    from app.repositories.attendance_repo import AttendanceRepository
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Ca đêm", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 26_000_000}, headers=_h(token))
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        shift = arepo.create_shift(name="Ca đêm", start_minute=1320, end_minute=360,
                                   is_overnight=True, grace_minutes=5, night_multiplier=1.3)
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift.id
        db.commit()
        # Ngày công 15/06: vào 22:00 VN 15/06 (15:00 UTC) + ra 06:00 VN 16/06 (23:00 UTC 15/06).
        arepo.create_log(employee_id=eid, check_type="in",
                         checked_at=_dt(2026, 6, 15, 15, 0, tzinfo=_tz.utc), within_range=True)
        arepo.create_log(employee_id=eid, check_type="out",
                         checked_at=_dt(2026, 6, 15, 23, 0, tzinfo=_tz.utc), within_range=True)
    finally:
        db.close()
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert line["night_premium_pay"] == 300_000   # 8h × (1.3−1) = 144' → 125k × 144/60


def _setup_shift_and_punches(client, token, *, name, start_min, end_min, overnight,
                             in_utc, out_utc, ot_in_utc=None, ot_out_utc=None):
    """Tạo NV + mức lương + gán ca + lượt chấm CA CHÍNH (in_utc, out_utc). Nếu truyền ot_*_utc thì
    thêm CẶP CHẤM TĂNG CA (vào/ra tăng ca) — mô hình 2-cặp. Trả employee_id."""
    from app.repositories.attendance_repo import AttendanceRepository
    eid = _make_emp(client, token, name=name, status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 26_000_000}, headers=_h(token))
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        shift = arepo.create_shift(name=name, start_minute=start_min, end_minute=end_min,
                                   is_overnight=overnight, grace_minutes=5)
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift.id
        db.commit()
        arepo.create_log(employee_id=eid, check_type="in", checked_at=in_utc, within_range=True)
        arepo.create_log(employee_id=eid, check_type="out", checked_at=out_utc, within_range=True)
        if ot_in_utc is not None and ot_out_utc is not None:
            arepo.create_log(employee_id=eid, check_type="in", checked_at=ot_in_utc, within_range=True)
            arepo.create_log(employee_id=eid, check_type="out", checked_at=ot_out_utc, within_range=True)
    finally:
        db.close()
    return eid


def _mk_ot(client, token, eid, *, work_date, frm, to):
    """Phiếu tăng ca do người có quyền duyệt tạo hộ ⇒ APPROVED luôn."""
    r = client.post("/api/overtime", json={"employee_id": eid, "work_date": work_date,
                                           "from_minute": frm, "to_minute": to}, headers=_h(token))
    assert r.status_code == 201, r.text
    return r.json()


def _cham_cong_du_dong_bhxh(eid):
    """Chấm công 15 NGÀY THƯỜNG của tháng 6/2026 cho NV — vừa đủ để KHÔNG dính luật 14 ngày.

    Từ 04/08/2026 engine áp QĐ 595 Đ42.4: tháng có ≥14 ngày làm việc mà không làm và không hưởng
    lương thì tháng đó KHÔNG đóng BHXH. Test lương cũ dựng NV không có một lượt chấm nào (0 công)
    nên rơi thẳng vào nhánh đó — phải cho họ đi làm thật thì mới kiểm được đúng thứ đang muốn kiểm.

    Chỉ chọn Thứ Hai–Thứ Sáu (01/06/2026 là Thứ Hai) để không sinh công ngày nghỉ tuần — thứ đó
    kéo theo premium và làm lệch các phép so sánh `gross` viết tay trong test. Vào/ra đúng giờ ca
    nên cũng không sinh phạt đi trễ.
    """
    from datetime import datetime as _dt, timezone as _tz
    from app.repositories.attendance_repo import AttendanceRepository

    ngay_thuong = [1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 15, 16, 17, 18, 19]
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        # ⚠️ `meal_allowance`/`shift_allowance` của model CÓ MẶC ĐỊNH 25.000/50.000 — không khai
        # 0 ở đây thì mỗi ngày công tự cộng 75.000đ và các phép so sánh `gross` viết tay trong
        # test lệch mà không hiểu vì sao. Test này chỉ muốn có CÔNG, không muốn có phụ cấp.
        shift = arepo.create_shift(name=f"HC-{eid}", start_minute=480, end_minute=1020,
                                   is_overnight=False, grace_minutes=5,
                                   meal_allowance=0, shift_allowance=0)
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift.id
        db.commit()
        for d in ngay_thuong:
            # 08:00 VN = 01:00 UTC (vào), 17:00 VN = 10:00 UTC (ra) — khớp ca 480→1020.
            for kind, hour in (("in", 1), ("out", 10)):
                arepo.create_log(employee_id=eid, check_type=kind,
                                 checked_at=_dt(2026, 6, d, hour, 0, tzinfo=_tz.utc),
                                 within_range=True)
    finally:
        db.close()


def _gen_line(client, token, eid):
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                      headers=_h(token)).json()
    return next(l for l in gen["lines"] if l["employee_id"] == eid)


def _line_of(client, token, eid, *, year=2026, month=6):
    """Đọc LẠI dòng lương từ bảng (không generate) — để soi số sau khi sửa/thêm khoản."""
    r = client.get(f"/api/luong/table?year={year}&month={month}", headers=_h(token))
    return next(l for l in r.json()["lines"] if l["employee_id"] == eid)


def _adv(client, token, eid, amount, *, month=6, expect=201):
    """HCNS lập đề nghị tạm ứng hộ NV (kỳ 2026-06)."""
    r = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": month,
        "advance_date": f"2026-{month:02d}-05", "amount": amount}, headers=_h(token))
    assert r.status_code == expect, r.text
    return r


def _emp_luong_10tr(client, token, name):
    eid = _make_emp(client, token, name=name, status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 8_000_000, "luong_trach_nhiem": 2_000_000}, headers=_h(token))
    return eid   # lương tháng = 10tr


def test_tam_ung_khong_con_tran(client):
    """Đã gỡ trần tạm ứng: ứng số lớn (90% lương) và cộng dồn vượt lương đều KHÔNG bị chặn."""
    token = _admin_token(client)
    eid = _emp_luong_10tr(client, token, "NV Tạm ứng lớn")
    _adv(client, token, eid, 9_000_000)          # 90% lương vẫn qua
    _adv(client, token, eid, 5_000_000)          # cộng dồn vượt cả lương vẫn qua
    # NV chưa khai lương cũng ứng được (không còn ràng buộc theo lương)
    eid2 = _make_emp(client, token, name="NV Chưa khai lương", status="active")
    _adv(client, token, eid2, 100_000)


def test_phieu_luong_tach_3_dong_bao_hiem(client):
    """Phiếu lương phải hiện RIÊNG BHXH / BHYT / BHTN (kèm tỷ lệ), và 3 dòng cộng lại ĐÚNG BẰNG
    số bảo hiểm đã đóng băng — để TỔNG TRỪ không bao giờ lệch THỰC NHẬN."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Bảo hiểm", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    _cham_cong_du_dong_bhxh(eid)   # không đi làm ngày nào thì tháng đó không đóng BHXH (Đ42.4)
    line = _gen_line(client, token, eid)

    rows = line["insurance_lines"]
    assert [r["label"].split()[0] for r in rows] == ["BHXH", "BHYT", "BHTN"]
    assert all("%" in r["label"] for r in rows)          # nhãn kèm tỷ lệ
    assert line["bhxh"] > 0
    assert sum(r["amount"] for r in rows) == line["bhxh"]   # BẤT BIẾN: tổng khớp số đã trừ


def test_phieu_luong_bao_hiem_bang_0_van_du_3_dong(client):
    """NV được nơi khác đóng BH → không trừ đồng nào, nhưng phiếu vẫn đủ 3 mục (giá trị 0)."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV BH nơi khác", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000, "insurance_elsewhere": True}, headers=_h(token))
    line = _gen_line(client, token, eid)
    assert line["bhxh"] == 0
    assert [r["amount"] for r in line["insurance_lines"]] == [0, 0, 0]


def test_nv_khong_co_quyen_cau_hinh_luong_van_thay_3_dong(client):
    """CA GỐC CỦA BUG: nhân viên xem phiếu của CHÍNH MÌNH (không có `luong:view_salary`) trước đây
    chỉ thấy 1 dòng gộp vì FE phải đi xin tỷ lệ. Nay backend trả sẵn ⇒ vẫn đủ 3 dòng."""
    from app.db import SessionLocal
    from app.repositories.employee_repo import EmployeeRepository
    from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
    from app.repositories.user_repo import UserRepository
    from app.security import create_access_token, hash_password

    admin = _admin_token(client)
    eid = _make_emp(client, admin, name="NV Xem phiếu", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(admin))
    client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(admin))
    # Từ 12/08/2026 phiếu lương phải được CÔNG BỐ thì NV mới thấy — chốt công → chốt lương →
    # công bố. Thiếu bước này thì `/payslip/me` trả `line: null` và ca đo hỏng trước khi chạm
    # tới thứ nó định kiểm (3 dòng bảo hiểm).
    assert client.post("/api/attendance/period/lock", json={"year": 2026, "month": 6},
                       headers=_h(admin)).status_code == 200
    assert client.post("/api/luong/lock", json={"year": 2026, "month": 6},
                       headers=_h(admin)).status_code == 200
    assert client.post("/api/luong/cong-bo", json={"year": 2026, "month": 6},
                       headers=_h(admin)).status_code == 200

    db = SessionLocal()
    try:
        users = UserRepository(db)
        kd = DepartmentRepository(db).get_by_name("Kinh doanh")
        role = RoleRepository(db).get_by_name_and_department("NV Sales", kd.id)
        u = users.create(username="nv-xem-phieu", name="NV", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=kd.id, role_id=role.id, is_active=True)
        emps = EmployeeRepository(db)
        emps.update(emps.get_by_id(eid), user_id=u.id)
        tok = create_access_token(str(u.id))
    finally:
        db.close()

    # Không có quyền cấu hình lương...
    assert client.get("/api/luong/params", headers=_h(tok)).status_code == 403
    # ...nhưng phiếu lương của chính mình vẫn đủ 3 dòng bảo hiểm.
    slip = client.get("/api/luong/payslip/me", headers=_h(tok))
    assert slip.status_code == 200, slip.text
    rows = slip.json()["line"]["insurance_lines"]
    assert [r["label"].split()[0] for r in rows] == ["BHXH", "BHYT", "BHTN"]
    assert sum(r["amount"] for r in rows) == slip.json()["line"]["bhxh"]


def test_tang_ca_vuot_nua_dem_khong_mat_cong(client):
    """Ca chiều 14:00–22:00, tăng ca tới 03:00 hôm sau (mô hình 2-CẶP: ra ca chính 22:00 → vào
    tăng ca 22:00 → ra tăng ca 03:00). Lượt RA tăng ca rơi sang ngày dương lịch mới vẫn được ghép
    về ngày công 15/06 (bản vá qua nửa đêm). Có phiếu 22:00–03:00: ĐỦ 1 công + 300' tăng ca."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    # Ca chính 14:00→22:00 (07:00→15:00 UTC); tăng ca 22:00→03:00 hôm sau (15:00→20:00 UTC 15/06).
    eid = _setup_shift_and_punches(
        client, token, name="NV TC qua đêm", start_min=840, end_min=1320, overnight=False,
        in_utc=_dt(2026, 6, 15, 7, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 15, 0, tzinfo=_tz.utc),       # ra ca chính 22:00
        ot_in_utc=_dt(2026, 6, 15, 15, 0, 1, tzinfo=_tz.utc),  # vào tăng ca 22:00
        ot_out_utc=_dt(2026, 6, 15, 20, 0, tzinfo=_tz.utc))    # ra tăng ca 03:00 hôm sau
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1320, to=1620)  # 22:00 → 03:00 hôm sau
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0     # công ca chính đủ (ra ca chính đúng giờ tan ca)
    assert line["ot_minutes"] == 300      # phiên TC 22:00 → 03:00 = 5h


def test_ca_dem_tang_ca_qua_gio_tan_ca(client):
    """Ca đêm 22:00–06:00, tăng ca tới 08:00 (2-cặp: ra ca chính 06:00 → vào TC 06:00 → ra TC 08:00).
    Có phiếu 06:00–08:00: 1 công + 120' tăng ca."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    # Ca chính 22:00→06:00 (15:00 UTC 15/06 → 23:00 UTC 15/06); tăng ca 06:00→08:00 (23:00 UTC 15/06
    # → 01:00 UTC 16/06). Trục ngày công 15/06: 06:00=1800, 08:00=1920.
    eid = _setup_shift_and_punches(
        client, token, name="NV ca đêm TC", start_min=1320, end_min=360, overnight=True,
        in_utc=_dt(2026, 6, 15, 15, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 23, 0, tzinfo=_tz.utc),        # ra ca chính 06:00
        ot_in_utc=_dt(2026, 6, 15, 23, 0, 1, tzinfo=_tz.utc),   # vào tăng ca 06:00
        ot_out_utc=_dt(2026, 6, 16, 1, 0, tzinfo=_tz.utc))      # ra tăng ca 08:00
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1800, to=1920)
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0
    assert line["ot_minutes"] == 120      # 06:00 → 08:00


def test_tang_ca_khong_phieu_thi_khong_ra_tien_nhung_giu_du_cong(client):
    """Chốt 23/07: KHÔNG có phiếu (và chỉ chấm 1 lượt ra, không có cặp chấm TC riêng) ⇒ tăng ca KHÔNG
    ra tiền, NHƯNG công ca chính vẫn ĐỦ. Đồng thời minh hoạ chốt 25/07: thiếu cặp chấm TC ⇒ TC = 0."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    eid = _setup_shift_and_punches(
        client, token, name="NV TC không phiếu", start_min=840, end_min=1320, overnight=False,
        in_utc=_dt(2026, 6, 15, 7, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 20, 0, tzinfo=_tz.utc))
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0     # công ca chính GIỮ NGUYÊN (kẹp trần tại giờ tan ca)
    assert line["ot_minutes"] == 0        # không phiếu + không cặp chấm TC → 0


def test_thieu_cap_cham_tang_ca_thi_khong_tinh_du_co_phieu(client):
    """Chốt 25/07/2026: dù CÓ phiếu duyệt, nếu NV chỉ chấm 1 lượt ra (quên chấm ra-ca-chính rồi
    vào-tăng-ca) ⇒ KHÔNG có phiên TC ⇒ tăng ca = 0. Buộc phải chấm đủ 2 cặp; HCNS chỉnh tay nếu quên."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    eid = _setup_shift_and_punches(
        client, token, name="NV quên cặp TC", start_min=840, end_min=1320, overnight=False,
        in_utc=_dt(2026, 6, 15, 7, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 20, 0, tzinfo=_tz.utc))       # 1 lượt ra 03:00, KHÔNG có cặp TC
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1320, to=1620)
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0
    assert line["ot_minutes"] == 0        # có phiếu nhưng thiếu cặp chấm TC → vẫn 0


def test_ve_som_hon_phieu_tra_theo_thuc_te(client):
    """Phiếu là TRẦN, không phải mức trả: phiếu duyệt tới 03:00 nhưng NV ra tăng ca lúc 00:00 ⇒ chỉ
    trả 120' (phiên TC 22:00→00:00)."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    eid = _setup_shift_and_punches(
        client, token, name="NV về sớm hơn phiếu", start_min=840, end_min=1320, overnight=False,
        in_utc=_dt(2026, 6, 15, 7, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 15, 0, tzinfo=_tz.utc),       # ra ca chính 22:00
        ot_in_utc=_dt(2026, 6, 15, 15, 0, 1, tzinfo=_tz.utc),  # vào tăng ca 22:00
        ot_out_utc=_dt(2026, 6, 15, 17, 0, tzinfo=_tz.utc))    # ra tăng ca 00:00
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1320, to=1620)  # duyệt tới 03:00
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0
    assert line["ot_minutes"] == 120      # trả theo giờ THỰC (22:00→00:00), không phải 300'


def test_lam_qua_phieu_thi_chan_tran(client):
    """Làm quá phiếu: phiên TC 22:00→03:00 (5h) nhưng phiếu chỉ tới 00:00 ⇒ kẹp trần 120'."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    eid = _setup_shift_and_punches(
        client, token, name="NV làm quá phiếu", start_min=840, end_min=1320, overnight=False,
        in_utc=_dt(2026, 6, 15, 7, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 15, 0, tzinfo=_tz.utc),       # ra ca chính 22:00
        ot_in_utc=_dt(2026, 6, 15, 15, 0, 1, tzinfo=_tz.utc),  # vào tăng ca 22:00
        ot_out_utc=_dt(2026, 6, 15, 20, 0, tzinfo=_tz.utc))    # ra tăng ca 03:00 (thực 5h)
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1320, to=1440)  # chỉ tới 00:00
    line = _gen_line(client, token, eid)
    assert line["actual_cong"] == 1.0
    assert line["ot_minutes"] == 120      # kẹp trần theo phiếu, không phải 300'


def test_cham_vao_tang_ca_chi_khi_co_phieu(client):
    """Chốt 25/07: sau khi đã RA ca chính, chỉ mở 'chấm VÀO tăng ca' khi có phiếu TC ĐÃ DUYỆT phủ
    giờ hiện tại (ot_mode=True). Không phiếu → chặn với thông điệp nhắc phiếu."""
    from datetime import datetime as _dt, timezone as _tz, date as _date
    from app.repositories.attendance_repo import AttendanceRepository
    from app.repositories.audit_repo import AuditLogRepository
    from app.repositories.overtime_repo import OvertimeRepository
    from app.services.attendance_service import AttendanceService, VN_TZ

    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Gate TC", status="active")
    wd = _date(2026, 6, 15)
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        shift = arepo.create_shift(name="HC gate", start_minute=480, end_minute=1020,  # 08:00–17:00
                                   is_overnight=False, grace_minutes=5)
        shift_id = shift.id
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift_id
        # ca chính: vào 08:00 (01:00 UTC), ra 17:00 (10:00 UTC) ngày 15/06
        arepo.create_log(employee_id=eid, check_type="in",
                         checked_at=_dt(2026, 6, 15, 1, 0, tzinfo=_tz.utc), within_range=True)
        arepo.create_log(employee_id=eid, check_type="out",
                         checked_at=_dt(2026, 6, 15, 10, 0, tzinfo=_tz.utc), within_range=True)
        db.commit()
    finally:
        db.close()

    now = _dt(2026, 6, 15, 18, 30, tzinfo=VN_TZ)   # 18:30, SAU giờ tan ca 17:00

    def _timing():
        d = SessionLocal()
        try:
            svc = AttendanceService(AttendanceRepository(d), EmployeeRepository(d),
                                    AuditLogRepository(d), overtime=OvertimeRepository(d))
            sh = next(s for s in AttendanceRepository(d).list_shifts() if s.id == shift_id)
            return svc._check_timing(eid, sh, wd, now)
        finally:
            d.close()

    # Chưa có phiếu → chặn, thông điệp nhắc phiếu.
    action, reason, ot_mode = _timing()
    assert action == "in" and reason is not None and "phiếu tăng ca" in reason.lower()

    # Có phiếu duyệt 17:30–20:00 (1050–1200) → cho VÀO tăng ca.
    _mk_ot(client, token, eid, work_date="2026-06-15", frm=1050, to=1200)
    action2, reason2, ot_mode2 = _timing()
    assert action2 == "in" and reason2 is None and ot_mode2 is True


def test_day_detail_goi_y_cham_bu_cap_tang_ca(client):
    """`GET /api/attendance/day` gợi ý chấm bù cặp TC khi NV có phiếu duyệt (trong ngày) + mới xong ca
    chính (đúng 1 phiên); null khi đã có phiên TC / không phiếu / phiếu qua nửa đêm."""
    from datetime import datetime as _dt, timezone as _tz
    from app.repositories.attendance_repo import AttendanceRepository
    token = _admin_token(client)

    def _setup(name, *, ot_pair=False, overnight_phieu=False, with_phieu=True):
        eid = _make_emp(client, token, name=name, status="active")
        db = SessionLocal()
        try:
            arepo = AttendanceRepository(db)
            sh = arepo.create_shift(name=name, start_minute=480, end_minute=1020,  # 08:00–17:00
                                    is_overnight=False, grace_minutes=5)
            EmployeeRepository(db).get_by_id(eid).default_shift_id = sh.id
            arepo.create_log(employee_id=eid, check_type="in",
                             checked_at=_dt(2026, 6, 15, 1, 0, tzinfo=_tz.utc), within_range=True)
            arepo.create_log(employee_id=eid, check_type="out",
                             checked_at=_dt(2026, 6, 15, 10, 0, tzinfo=_tz.utc), within_range=True)
            if ot_pair:  # đã có phiên tăng ca (17:30 → 20:30)
                arepo.create_log(employee_id=eid, check_type="in",
                                 checked_at=_dt(2026, 6, 15, 10, 30, tzinfo=_tz.utc), within_range=True)
                arepo.create_log(employee_id=eid, check_type="out",
                                 checked_at=_dt(2026, 6, 15, 13, 30, tzinfo=_tz.utc), within_range=True)
            db.commit()
        finally:
            db.close()
        if with_phieu:
            fr, to = (1320, 1620) if overnight_phieu else (1050, 1200)  # đêm: 22:00→03:00 · ngày: 17:30→20:00
            _mk_ot(client, token, eid, work_date="2026-06-15", frm=fr, to=to)
        return eid

    def _sug(eid):
        r = client.get(f"/api/attendance/day?employee_id={eid}&date=2026-06-15", headers=_h(token))
        assert r.status_code == 200, r.text
        return r.json()["ot_suggestion"]

    s = _sug(_setup("NV Gợi ý"))                                   # có phiếu + 1 phiên ca chính
    assert s is not None and s["from_time"] == "17:30" and s["to_time"] == "20:00"
    assert _sug(_setup("NV Đã TC", ot_pair=True)) is None          # đã có phiên TC
    assert _sug(_setup("NV Không phiếu", with_phieu=False)) is None  # không phiếu
    assert _sug(_setup("NV Phiếu đêm", overnight_phieu=True)) is None  # phiếu qua nửa đêm


def test_ngay_off1x_lam_tra_1x_khong_he_so(client):
    """Ngày nghỉ 'off1x': nghỉ KHÔNG lương; ai đi làm được cộng THÊM 1 công lương chính (1×, uncapped),
    KHÔNG nhân hệ số lễ/nghỉ. is_paid bị ép False."""
    from datetime import datetime as _dt, timezone as _tz
    token = _admin_token(client)
    # NV lương 26tr + ca 8–17 + chấm đúng 1 công ngày 2026-06-15.
    eid = _setup_shift_and_punches(
        client, token, name="NV off1x", start_min=480, end_min=1020, overnight=False,
        in_utc=_dt(2026, 6, 15, 1, 0, tzinfo=_tz.utc),
        out_utc=_dt(2026, 6, 15, 10, 0, tzinfo=_tz.utc))
    # Khai 15/06 là ngày nghỉ off1x — dù gửi is_paid=True vẫn bị ép False.
    sp = client.post("/api/calendar/special-days", json={
        "day": "2026-06-15", "kind": "off1x", "name": "Nghỉ hãng", "is_paid": True}, headers=_h(token))
    assert sp.status_code == 201, sp.text
    assert sp.json()["is_paid"] is False

    line = _gen_line(client, token, eid)
    # off1x loại khỏi BASE → không có lương công thường; chỉ +1× cho ngày ĐÃ LÀM (uncapped), không premium.
    assert line["actual_cong"] == 0 and line["luong_cong"] == 0
    daily = line["monthly_salary"] / line["standard_cong"]
    assert abs(line["ot_pay"] - daily) < 1.0     # đúng 1 công lương chính (×1), KHÔNG ×2/×3


def test_ngay_off1x_khong_lam_thi_khong_luong(client):
    """off1x mà NV KHÔNG đi làm (không chấm công) → 0 công, 0 tiền cho ngày đó (nghỉ không lương)."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV off1x nghỉ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 26_000_000}, headers=_h(token))
    client.post("/api/calendar/special-days", json={
        "day": "2026-06-15", "kind": "off1x", "name": "Nghỉ hãng"}, headers=_h(token))
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(token)).json()
    line = next((l for l in gen["lines"] if l["employee_id"] == eid), None)
    # Không chấm công cả tháng → không có dòng công, hoặc có thì ot_pay = 0 cho ngày off1x.
    if line is not None:
        assert line["ot_pay"] == 0


def test_update_line_keeps_ot_night_pay(client):
    """Pha 4a (bẫy C2): sửa ô tay (vi phạm) KHÔNG được xóa tăng ca/ca đêm khỏi gross."""
    token = _admin_token(client)
    emp_id = _make_emp(client, token, name="NV OT", payroll_group="x", status="active")
    db = SessionLocal()
    try:
        repo = PayrollRepository(db)
        svc = PayrollService(repo, EmployeeRepository(db), attendance=None)
        period = repo.create_period(year=2026, month=6, status="draft", standard_cong=26)
        ln = repo.create_line(period_id=period.id, employee_id=emp_id,
                              luong_cong=10_000_000, chuyen_can=300_000, allowance=0,
                              ot_pay=375_000, night_pay=600_000, bhxh=1_050_000)
        updated = svc.update_line(line_id=ln.id, actor=None, vi_pham=200_000)
        assert float(updated.ot_pay) == 375_000 and float(updated.night_pay) == 600_000
        # gross vẫn còn OT + ca đêm, chỉ trừ thêm vi phạm 200k.
        assert float(updated.gross) == 10_000_000 + 300_000 + 375_000 + 600_000 - 200_000
        assert float(updated.net_pay) == float(updated.gross) - 1_050_000
    finally:
        db.close()


# --- TNCN tự tính (Pha 4b) --------------------------------------------------


def test_params_deduction_2026(client):
    """Giảm trừ gia cảnh mặc định = mức 2026 (NQ 110/2025)."""
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["deduction_self"] == 15_500_000 and p["deduction_dependent"] == 6_200_000


def test_khau_tru_tai_nguon_sua_duoc_tu_man_cau_hinh(client):
    """Hai so cua nhanh "khau tru 10%" phai KHAI DUOC — luat doi muc la sua tren man, khong doi code.

    Truoc 08/08/2026 hai so nay chi nam trong DB, khong man nao pho ra ⇒ doi 10% thanh 8% phai nho
    dev. Test nay giu duong khai do song."""
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["pit_flat_rate"] == 0.10
    assert p["pit_flat_threshold"] == 2_000_000

    upd = client.put(
        "/api/luong/params",
        json={"pit_flat_rate": 0.08, "pit_flat_threshold": 3_000_000},
        headers=_h(token),
    )
    assert upd.status_code == 200, upd.text
    lai = client.get("/api/luong/params", headers=_h(token)).json()
    assert lai["pit_flat_rate"] == 0.08
    assert lai["pit_flat_threshold"] == 3_000_000

    # Ty le phai nam trong 0..1 (0..100%). Go 110% la go nham 1.1 hay 110 — chan o schema.
    assert client.put(
        "/api/luong/params", json={"pit_flat_rate": 1.1}, headers=_h(token)
    ).status_code == 422
    assert client.put(
        "/api/luong/params", json={"pit_flat_rate": -0.1}, headers=_h(token)
    ).status_code == 422
    assert client.put(
        "/api/luong/params", json={"pit_flat_threshold": -1}, headers=_h(token)
    ).status_code == 422

    # 0% la so HOP LE (man hinh chi canh bao mem) — khong duoc chan, nhung phai luu dung.
    assert client.put(
        "/api/luong/params", json={"pit_flat_rate": 0}, headers=_h(token)
    ).status_code == 200
    assert client.get("/api/luong/params", headers=_h(token)).json()["pit_flat_rate"] == 0

    # Tra lai mac dinh cho test khac khong bi anh huong (cung mot DB trong 1 test).
    client.put(
        "/api/luong/params",
        json={"pit_flat_rate": 0.10, "pit_flat_threshold": 2_000_000},
        headers=_h(token),
    )


def test_pit_brackets_seeded_and_editable(client):
    """Biểu thuế TNCN seed 5 bậc 2026 + sửa được."""
    token = _admin_token(client)
    items = client.get("/api/luong/pit-brackets", headers=_h(token)).json()["items"]
    assert len(items) == 5
    assert items[0]["rate"] == 0.05
    assert items[-1]["up_to"] is None and items[-1]["rate"] == 0.35
    bid = items[0]["id"]
    upd = client.put(f"/api/luong/pit-brackets/{bid}",
                     json={"seq": 1, "up_to": 11_000_000, "rate": 0.05}, headers=_h(token))
    assert upd.status_code == 200 and upd.json()["up_to"] == 11_000_000


def test_pit_progressive(client):
    """Lũy tiến từng phần: thu nhập tính thuế 20tr → 1,5tr (10tr×5% + 10tr×10%)."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params, brackets = svc.get_params(), svc.get_pit_brackets()
        # gross 35,5tr, không OT/đêm/BHXH, 0 phụ thuộc → tính thuế = 35,5 − 15,5 = 20tr.
        _chiu, taxable, pit = svc._auto_pit(gross=35_500_000, bhxh=0, ot_pay=0, night_pay=0,
                                     dependents_count=0, params=params, brackets=brackets)
        assert taxable == 20_000_000
        assert pit == 1_500_000
    finally:
        db.close()


def test_pit_exempt_ot_night(client):
    """OT + ca đêm được MIỄN thuế (Luật 109/2025) → giảm thu nhập chịu thuế."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params, brackets = svc.get_params(), svc.get_pit_brackets()
        # gross 35,5tr trong đó 5tr OT + 2tr ca đêm (miễn) → chịu thuế 28,5tr → tính thuế 13tr.
        _chiu, taxable, pit = svc._auto_pit(gross=35_500_000, bhxh=0, ot_pay=5_000_000, night_pay=2_000_000,
                                     dependents_count=0, params=params, brackets=brackets)
        assert taxable == 13_000_000
        assert pit == round(10_000_000 * 0.05 + 3_000_000 * 0.10)   # 800k
    finally:
        db.close()


def test_pit_dependents_reduce_tax(client):
    """Người phụ thuộc giảm thu nhập tính thuế (6,2tr/người)."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params, brackets = svc.get_params(), svc.get_pit_brackets()
        # gross 35,5tr, 2 phụ thuộc (12,4tr) → tính thuế = 20tr − 12,4tr = 7,6tr.
        _chiu, taxable, pit = svc._auto_pit(gross=35_500_000, bhxh=0, ot_pay=0, night_pay=0,
                                     dependents_count=2, params=params, brackets=brackets)
        assert taxable == 7_600_000
        assert pit == round(7_600_000 * 0.05)
    finally:
        db.close()


def test_pit_manual_override_and_reset(client):
    """pit tự tính; HCNS ghi đè tay (pit_manual) → giữ; reset (pit_manual=False) → về auto."""
    token = _admin_token(client)
    emp_id = _make_emp(client, token, name="NV Thuế", payroll_group="x", status="active")
    db = SessionLocal()
    try:
        repo = PayrollRepository(db)
        svc = PayrollService(repo, EmployeeRepository(db), attendance=None)
        period = repo.create_period(year=2026, month=6, status="draft", standard_cong=26)
        ln = repo.create_line(period_id=period.id, employee_id=emp_id,
                              luong_cong=50_000_000, gross=50_000_000, bhxh=0)
        # auto: sửa (chưa manual) → pit tính lại theo gross.
        upd = svc.update_line(line_id=ln.id, actor=None, vi_pham=0)
        auto_pit = float(upd.pit)
        assert auto_pit > 0 and upd.pit_manual is False
        # ghi đè tay.
        upd2 = svc.update_line(line_id=ln.id, actor=None, pit=1_234_000)
        assert float(upd2.pit) == 1_234_000 and upd2.pit_manual is True
        # reset về tự tính.
        upd3 = svc.update_line(line_id=ln.id, actor=None, pit_manual=False)
        assert float(upd3.pit) == auto_pit and upd3.pit_manual is False
    finally:
        db.close()


# --- Chi trả + xuất file + nhật ký (Pha 4c) ---------------------------------


def test_pay_unpay_flow(client):
    """State machine: nháp→chốt→đã chi→hủy chi; chặn pay-khi-chưa-chốt, reopen/generate-khi-đã-chi."""
    token = _admin_token(client)
    _make_emp(client, token, name="NV Chi", payroll_group="x", status="active")
    y, m = 2026, 6
    assert client.post("/api/luong/generate", json={"year": y, "month": m}, headers=_h(token)).status_code == 200
    # pay khi CHƯA chốt → 400
    assert client.post("/api/luong/pay", json={"year": y, "month": m}, headers=_h(token)).status_code == 400
    assert client.post("/api/luong/lock", json={"year": y, "month": m}, headers=_h(token)).status_code == 200
    paid = client.post("/api/luong/pay", json={"year": y, "month": m}, headers=_h(token))
    assert paid.status_code == 200 and paid.json()["status"] == "paid" and paid.json()["paid_at"]
    # reopen / generate khi ĐÃ CHI → chặn
    assert client.post("/api/luong/reopen", json={"year": y, "month": m}, headers=_h(token)).status_code == 400
    assert client.post("/api/luong/generate", json={"year": y, "month": m}, headers=_h(token)).status_code in (400, 409)
    # hủy chi → về locked
    un = client.post("/api/luong/unpay", json={"year": y, "month": m, "note": "nhầm"}, headers=_h(token))
    assert un.status_code == 200 and un.json()["status"] == "locked" and un.json()["paid_at"] is None


def test_export_xlsx_smoke(client):
    """Xuất bảng lương + file chuyển khoản .xlsx trả 200 + đúng content-type Excel."""
    token = _admin_token(client)
    _make_emp(client, token, name="NV Xls", payroll_group="x", status="active")
    y, m = 2026, 6
    client.post("/api/luong/generate", json={"year": y, "month": m}, headers=_h(token))
    r1 = client.get(f"/api/luong/export.xlsx?year={y}&month={m}", headers=_h(token))
    assert r1.status_code == 200 and "spreadsheetml" in r1.headers["content-type"]
    r2 = client.get(f"/api/luong/bank.xlsx?year={y}&month={m}", headers=_h(token))
    assert r2.status_code == 200 and "spreadsheetml" in r2.headers["content-type"]


def test_payroll_audit_logged(client):
    """Thao tác lương ghi nhật ký (tạo bảng / chốt) — hiện ở Nhật ký chung."""
    token = _admin_token(client)
    _make_emp(client, token, name="NV Audit", payroll_group="x", status="active")
    y, m = 2026, 6
    client.post("/api/luong/generate", json={"year": y, "month": m}, headers=_h(token))
    client.post("/api/luong/lock", json={"year": y, "month": m}, headers=_h(token))
    db = SessionLocal()
    try:
        from sqlalchemy import select

        from app.models.audit import AuditLog
        actions = [a.action for a in db.execute(select(AuditLog)).scalars()]
        assert "payroll_generate" in actions and "payroll_lock" in actions
    finally:
        db.close()


# --- lương nhân viên (khai báo + preview) -----------------------------------


def test_salary_declare_and_preview(client):
    token = _admin_token(client)
    eid = _make_emp(client, token, name="Thợ In A")

    # khai lương = lương vị trí + trách nhiệm của NV (mức nền, gõ tay từng ô)
    s = client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-01-01", "luong_vi_tri": 18_000_000,
        "luong_trach_nhiem": 4_000_000, "allowance": 300_000,
    }, headers=_h(token))
    assert s.status_code == 201

    prev = client.get(f"/api/luong/salaries/{eid}/preview", headers=_h(token)).json()
    assert prev["monthly"] == 22_000_000 and prev["source"] == "employee"
    # Mức đóng BH = MỨC NỀN (vị trí 18tr + trách nhiệm 4tr) — chủ chốt 12/08/2026, đảo chốt
    # cũ "chỉ lương vị trí". Preview PHẢI khớp bảng lương, xem `PayrollService.salary_preview`.
    assert prev["insurance_base"] == 22_000_000

    hist = client.get(f"/api/luong/salaries/{eid}", headers=_h(token)).json()
    assert hist["employee_name"] == "Thợ In A" and len(hist["items"]) == 1

    # điều chỉnh hiệu lực sau → preview lấy mức mới
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-07-01", "luong_vi_tri": 24_000_000,
    }, headers=_h(token))
    prev2 = client.get(f"/api/luong/salaries/{eid}/preview", headers=_h(token)).json()
    assert prev2["monthly"] == 24_000_000 and prev2["source"] == "employee"

    hist2 = client.get(f"/api/luong/salaries/{eid}", headers=_h(token)).json()["items"]
    assert len(hist2) == 2
    assert hist2[0]["effective_to"] is None
    assert hist2[1]["effective_to"] == "2026-06-30"

    # NHẬT KÝ điều chỉnh: mỗi lần lưu (kể cả CÙNG ngày hiệu lực) = MỘT bản ghi mới — giữ đủ
    # dấu vết "ai · lúc nào". "Hiện hành" = bản lưu SAU (id lớn hơn) → preview + đầu danh sách
    # lấy 25tr; danh sách nay có 3 dòng (01-01, 07-01@24tr, 07-01@25tr).
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-07-01", "luong_vi_tri": 25_000_000,
    }, headers=_h(token))
    hist3 = client.get(f"/api/luong/salaries/{eid}", headers=_h(token)).json()["items"]
    assert len(hist3) == 3 and hist3[0]["luong_vi_tri"] == 25_000_000
    assert hist3[0]["is_current"] is True
    assert hist3[0]["actor_name"]        # có tên người điều chỉnh (nhật ký "ai sửa")
    prev3 = client.get(f"/api/luong/salaries/{eid}/preview", headers=_h(token)).json()
    assert prev3["monthly"] == 25_000_000


# --- tạm ứng ----------------------------------------------------------------


def test_advance_workflow(client):
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Ứng", payroll_group="van_phong")
    created = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-10", "amount": 2_000_000, "reason": "Ứng",
    }, headers=_h(token))
    assert created.status_code == 201 and created.json()["status"] == "pending"
    aid = created.json()["id"]

    ap = client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token))
    assert ap.status_code == 200 and ap.json()["status"] == "approved"
    # duyệt lại → 400
    assert client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token)).status_code == 400

    listed = client.get("/api/luong/advances?year=2026&month=6&status=approved", headers=_h(token)).json()["items"]
    assert any(a["id"] == aid and a["employee_name"] == "NV Ứng" for a in listed)


def test_my_advance_self_create(client):
    """Nhân viên TỰ lập đề nghị tạm ứng (self-service) → pending → hiện ở list HCNS;
    payload có sẵn dept/bank cho phiếu in."""
    token = _admin_token(client)  # admin có hồ sơ (backfill) → coi như 1 NV tự ứng
    created = client.post("/api/luong/advances/me", json={
        "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-12", "amount": 1_500_000, "reason": "Tự ứng",
    }, headers=_h(token))
    assert created.status_code == 201, created.text
    body = created.json()
    assert body["status"] == "pending"
    import re
    assert re.match(r"^TU-\d{6}-[A-Z0-9]{4}$", body.get("code") or ""), body.get("code")  # TU-YYMMDD-XXXX
    for k in ("department_name", "bank_account", "bank_name", "employee_name"):
        assert k in body  # field cho phiếu in (giá trị có thể None)
    aid = body["id"]
    mine = client.get("/api/luong/advances/me", headers=_h(token)).json()
    assert mine["has_employee"] is True and any(a["id"] == aid for a in mine["items"])
    listed = client.get("/api/luong/advances?year=2026&month=6", headers=_h(token)).json()["items"]
    assert any(a["id"] == aid for a in listed)
    # badge real-time: người có quyền duyệt thấy số 'chờ duyệt' > 0
    summ = client.get("/api/luong/advances/notify-summary", headers=_h(token)).json()
    assert summ["pending_approval_count"] >= 1


def test_my_advance_dot_1_self_create(client):
    """NV TỰ xin lương đợt 1: GET /advances/me trả mức 'Lương trả 1 lần' hiện hành (để FE điền sẵn);
    POST kind='luong_dot_1' → phiếu L1 pending hiện ở list của NV → HCNS duyệt → generate trừ đúng
    dòng đợt-1 (advance_total KHÔNG gồm nó)."""
    import re
    from app.db import SessionLocal
    from app.repositories.employee_repo import EmployeeRepository
    from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
    from app.repositories.user_repo import UserRepository
    from app.security import create_access_token, hash_password

    admin = _admin_token(client)
    eid = _make_emp(client, admin, name="NV Tự xin đợt 1", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000, "luong_dot_1": 3_000_000}, headers=_h(admin))

    db = SessionLocal()
    try:
        users = UserRepository(db)
        kd = DepartmentRepository(db).get_by_name("Kinh doanh")
        role = RoleRepository(db).get_by_name_and_department("NV Sales", kd.id)
        u = users.create(username="nv-tu-xin-dot1", name="NV", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=kd.id, role_id=role.id, is_active=True)
        emps = EmployeeRepository(db)
        emps.update(emps.get_by_id(eid), user_id=u.id)
        tok = create_access_token(str(u.id))
    finally:
        db.close()

    # GET trả mức đợt-1 hiện hành để FE điền sẵn ô số tiền.
    me = client.get("/api/luong/advances/me", headers=_h(tok)).json()
    assert me["has_employee"] is True and me["luong_dot_1"] == 3_000_000

    created = client.post("/api/luong/advances/me", json={
        "period_year": 2026, "period_month": 6, "advance_date": "2026-06-15",
        "amount": 3_000_000, "kind": "luong_dot_1"}, headers=_h(tok))
    assert created.status_code == 201, created.text
    a = created.json()
    assert a["kind"] == "luong_dot_1"
    assert re.match(r"^L1-\d{6}-[A-Z0-9]{4}$", a.get("code") or ""), a.get("code")

    mine = client.get("/api/luong/advances/me", headers=_h(tok)).json()
    assert any(x["id"] == a["id"] and x["kind"] == "luong_dot_1" for x in mine["items"])

    client.post(f"/api/luong/advances/{a['id']}/approve", json={}, headers=_h(admin))
    line = _gen_line(client, admin, eid)
    assert line["luong_dot_1_total"] == 3_000_000
    assert line["advance_total"] == 0


# --- lương trả đợt 1 (phiếu kind=luong_dot_1) -------------------------------


def test_luong_dot_1_phieu_duyet_tru_thuc_nhan(client):
    """Đợt 1 = phiếu kind='luong_dot_1' (điền sẵn từ hồ sơ), DUYỆT xong mới trừ. Phiếu lương tách
    RIÊNG 'Thanh toán lương đợt 1' (luong_dot_1_total) với 'Tạm ứng đã nhận' (advance_total);
    thực nhận (đợt 2) trừ CẢ HAI, KHÔNG gộp chung."""
    import re
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Đợt 1", status="active")
    sal = client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000, "luong_dot_1": 3_000_000}, headers=_h(token))
    # hồ sơ lưu 'mức trả 1 lần'
    assert sal.status_code == 201 and sal.json()["luong_dot_1"] == 3_000_000

    # phiếu đợt 1 (kind=luong_dot_1) — mã tiền tố L1
    d1 = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-15", "amount": 3_000_000, "kind": "luong_dot_1"}, headers=_h(token))
    assert d1.status_code == 201, d1.text
    assert d1.json()["kind"] == "luong_dot_1"
    assert re.match(r"^L1-\d{6}-[A-Z0-9]{4}$", d1.json()["code"] or ""), d1.json()["code"]
    d1id = d1.json()["id"]
    # tạm ứng thường 1tr (kind mặc định = tam_ung)
    tuid = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-05", "amount": 1_000_000}, headers=_h(token)).json()["id"]
    for aid in (d1id, tuid):
        client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token))

    line = _gen_line(client, token, eid)
    assert line["luong_dot_1_total"] == 3_000_000     # đợt 1 tách riêng
    assert line["advance_total"] == 1_000_000          # KHÔNG gộp đợt 1 vào tạm ứng
    # thêm thưởng đủ lớn để thực nhận dương → thấy rõ cả 2 khoản đều bị trừ. Thưởng nay khai qua
    # DANH MỤC (ô "Thưởng khác" đã gỡ 28/07/2026).
    lid = line["id"]
    client.post(f"/api/luong/lines/{lid}/components", headers=_h(token),
                json={"component_id": _comp(client, token, name="Thưởng đợt 1 test"),
                      "amount": 30_000_000})
    upd = _line_of(client, token, eid)
    assert upd["luong_dot_1_total"] == 3_000_000 and upd["advance_total"] == 1_000_000
    exp = (upd["gross"] - upd["bhxh"] - upd["cong_doan"] - upd["pit"]
           - upd["advance_total"] - upd["luong_dot_1_total"])
    assert upd["net_pay"] == exp and upd["net_pay"] > 0


def test_luong_dot_1_chua_duyet_thi_chua_tru(client):
    """Phiếu đợt 1 CHỜ DUYỆT → chưa trừ vào lương (giống tạm ứng pending)."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Đợt 1 chờ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000, "luong_dot_1": 3_000_000}, headers=_h(token))
    client.post("/api/luong/advances", json={          # tạo nhưng KHÔNG duyệt
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-15", "amount": 3_000_000, "kind": "luong_dot_1"}, headers=_h(token))
    line = _gen_line(client, token, eid)
    assert line["luong_dot_1_total"] == 0


def test_dot_1_va_tam_ung_vuot_luong_thuc_nhan_san_0(client):
    """Đợt 1 + tạm ứng > lương → thực nhận (đợt 2) = 0 (sàn), KHÔNG âm, KHÔNG tự đòi lại."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Sàn đợt 1", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 5_000_000}, headers=_h(token))
    d1 = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-15", "amount": 4_000_000, "kind": "luong_dot_1"}, headers=_h(token)).json()["id"]
    tu = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-05", "amount": 4_000_000}, headers=_h(token)).json()["id"]
    for aid in (d1, tu):
        client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token))
    line = _gen_line(client, token, eid)
    assert line["luong_dot_1_total"] == 4_000_000 and line["advance_total"] == 4_000_000
    assert line["net_pay"] == 0


# --- bảng lương tháng: tạo + engine + khóa ----------------------------------


def test_generate_lock_flow(client):
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Bảng", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-01-01", "luong_vi_tri": 10_000_000,
    }, headers=_h(token))
    _cham_cong_du_dong_bhxh(eid)   # không đi làm ngày nào thì tháng đó không đóng BHXH (Đ42.4)
    # tạm ứng đã duyệt 2tr trong kỳ
    aid = client.post("/api/luong/advances", json={
        "employee_id": eid, "period_year": 2026, "period_month": 6,
        "advance_date": "2026-06-05", "amount": 2_000_000,
    }, headers=_h(token)).json()["id"]
    client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token))

    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(token))
    assert gen.status_code == 200
    line = next(l for l in gen.json()["lines"] if l["employee_id"] == eid)
    assert line["monthly_salary"] == 10_000_000
    assert line["bhxh"] == round(10_000_000 * 0.105)
    assert line["advance_total"] == 2_000_000
    lid = line["id"]

    # sửa ô tay: vi phạm 500k + thưởng 5tr → gross/net tính lại. Thưởng đủ lớn để 500k nằm trong
    # trần khấu trừ 30% (Điều 102) → khoản phạt được áp trọn.
    upd = client.put(f"/api/luong/lines/{lid}",
                     json={"vi_pham": 500_000, "dieu_chinh_luong": 5_000_000},
                     headers=_h(token)).json()
    assert upd["vi_pham"] == 500_000 and upd["dieu_chinh_luong"] == 5_000_000
    exp_gross = upd["luong_cong"] + upd["chuyen_can"] + upd["allowance"] + 5_000_000 - 500_000
    assert upd["gross"] == exp_gross
    assert upd["net_pay"] == exp_gross - upd["bhxh"] - upd["pit"] - upd["advance_total"]

    # chốt kỳ → generate lại + sửa dòng đều bị chặn (409)
    lock = client.post("/api/luong/lock", json={"year": 2026, "month": 6}, headers=_h(token))
    assert lock.status_code == 200 and lock.json()["status"] == "locked"
    assert client.post("/api/luong/generate", json={"year": 2026, "month": 6}, headers=_h(token)).status_code == 409
    assert client.put(f"/api/luong/lines/{lid}", json={"vi_pham": 0}, headers=_h(token)).status_code == 409

    # mở lại → sửa được
    assert client.post("/api/luong/reopen", json={"year": 2026, "month": 6}, headers=_h(token)).status_code == 200
    assert client.put(f"/api/luong/lines/{lid}", json={"vi_pham": 0}, headers=_h(token)).status_code == 200


# --- Nhóm ĐỎ: vá lỗi tiền/luật engine lương ---------------------------------


def test_dis_deduction_capped_30pct(client):
    """#2b Điều 102: khấu trừ kỷ luật (vi_pham) ≤ 30% lương sau BHXH+TNCN; vượt thì kẹp."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="d102", monthly_amount=20_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="d102", pay_grade_key=None)
        # phạt khủng 100tr → cột vi_pham LƯU RAW; phần kẹp chỉ ảnh hưởng gross (trần 30%).
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=20_000_000), params=params, actual_cong=26, standard_cong=26,
                         vi_pham=100_000_000, on=date(2026, 6, 1))
        assert v["vi_pham"] == 100_000_000            # RAW (không còn capped)
        income = (v["luong_cong"] + v["chuyen_can"] + v["allowance"] + v["khoan"]
                  + v["ot_pay"] + v["night_pay"] + v["other_bonus"])
        phat_eff = income - v["gross"]                # gross = income − phạt đã kẹp
        base = income - v["bhxh"] - v["pit"]
        assert 0 < phat_eff < 100_000_000
        assert abs(phat_eff - round(0.30 * base)) <= 1
    finally:
        db.close()


def test_bonus_items_taxable(client):
    """Thưởng chi tiết = thu nhập chịu thuế: gross tăng đúng Σ thưởng (dieu_chinh âm giảm); PIT tăng."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="bonus_grp", monthly_amount=30_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="bonus_grp", pay_grade_key=None, dependents_count=0)
        base = svc._compute(employee=emp, salary=_sal(luong_vi_tri=30_000_000), params=params, actual_cong=26, standard_cong=26,
                            on=date(2026, 6, 1))
        withb = svc._compute(employee=emp, salary=_sal(luong_vi_tri=30_000_000), params=params, actual_cong=26, standard_cong=26,
                             thuong_5s=1_000_000, thuong_doanh_so=2_000_000, dieu_chinh_luong=-500_000,
                             on=date(2026, 6, 1))
        assert withb["thuong_5s"] == 1_000_000 and withb["thuong_doanh_so"] == 2_000_000
        assert withb["dieu_chinh_luong"] == -500_000
        assert withb["gross"] == base["gross"] + 2_500_000     # 1tr + 2tr − 0.5tr
        assert withb["pit"] > base["pit"]                      # thưởng chịu thuế → PIT tăng
    finally:
        db.close()


def test_cong_doan_auto(client):
    """Đoàn phí công đoàn = insurance_base × cong_doan_rate — CHỈ đoàn viên (union_member); thử việc = 0;
    không phải đoàn viên = 0 (chủ 2026-07-21: opt-in từng người)."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.update_params(cong_doan_rate=0.005)
        svc.payroll.create_rule(payroll_group="cd_grp", monthly_amount=10_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="cd_grp", pay_grade_key=None, dependents_count=0)
        # Đoàn viên → có trừ đoàn phí.
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=10_000_000, union_member=True),
                         params=params, actual_cong=26, standard_cong=26, on=date(2026, 6, 1))
        assert v["cong_doan"] == round(v["insurance_base"] * 0.005) and v["cong_doan"] > 0
        # KHÔNG phải đoàn viên → 0 dù rate > 0.
        v_no = svc._compute(employee=emp, salary=_sal(luong_vi_tri=10_000_000, union_member=False),
                            params=params, actual_cong=26, standard_cong=26, on=date(2026, 6, 1))
        assert v_no["cong_doan"] == 0
        # Thử việc (dù là đoàn viên) → 0.
        emp_tv = SimpleNamespace(status="probation", hire_date=date(2026, 5, 1), gender="male",
                                 payroll_group="cd_grp", pay_grade_key=None, dependents_count=0)
        v_tv = svc._compute(employee=emp_tv, salary=_sal(luong_vi_tri=10_000_000, union_member=True),
                            params=params, actual_cong=26, standard_cong=26, on=date(2026, 6, 1))
        assert v_tv["cong_doan"] == 0
        svc.update_params(cong_doan_rate=0)   # khôi phục
    finally:
        db.close()


def test_penalties_share_30pct_cap(client):
    """5 khoản phạt gộp CHUNG 1 trần 30%; từng cột lưu RAW (không capped riêng)."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="pen_grp", monthly_amount=20_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="pen_grp", pay_grade_key=None, dependents_count=0)
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=20_000_000), params=params, actual_cong=26, standard_cong=26,
                         vi_pham=30_000_000, di_tre=30_000_000, dt_vuot_troi=30_000_000,
                         phat_bien_ban=30_000_000, phat_5s_dong_phuc=30_000_000, on=date(2026, 6, 1))
        for k in ("vi_pham", "di_tre", "dt_vuot_troi", "phat_bien_ban", "phat_5s_dong_phuc"):
            assert v[k] == 30_000_000                          # RAW
        income = (v["luong_cong"] + v["chuyen_can"] + v["allowance"] + v["khoan"]
                  + v["ot_pay"] + v["night_pay"] + v["other_bonus"])
        phat_eff = income - v["gross"]
        base = income - v["bhxh"] - v["pit"]
        assert abs(phat_eff - round(0.30 * base)) <= 1         # gộp về đúng trần 30%
    finally:
        db.close()


def test_generate_preserves_detail_and_net_cong_doan(client):
    """generate: preserve ô tay chi tiết khi Tính lại; net TRỪ công đoàn."""
    token = _admin_token(client)
    client.put("/api/luong/params", json={"cong_doan_rate": 0.005}, headers=_h(token))
    eid = _make_emp(client, token, name="NV Phiếu", status="active")
    # union_member=True: đoàn viên → có trừ đoàn phí công đoàn (mặc định false thì công đoàn = 0).
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 5_000_000, "union_member": True}, headers=_h(token))
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 8}, headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    # Ô tay còn lại sau 28/07/2026 (thưởng đã chuyển sang danh mục): điều chỉnh lương + đi trễ.
    client.put(f"/api/luong/lines/{line['id']}",
               json={"dieu_chinh_luong": 5_000_000, "di_tre": 200_000}, headers=_h(token))
    gen2 = client.post("/api/luong/generate", json={"year": 2026, "month": 8}, headers=_h(token)).json()
    l2 = next(l for l in gen2["lines"] if l["employee_id"] == eid)
    assert l2["dieu_chinh_luong"] == 5_000_000 and l2["di_tre"] == 200_000     # preserve
    assert l2["cong_doan"] == round(l2["insurance_base"] * 0.005)
    assert l2["net_pay"] == round(max(0.0, l2["gross"] - l2["bhxh"] - l2["cong_doan"]
                                      - l2["pit"] - l2["advance_total"]))
    assert l2["net_pay"] > 0                                    # dương → chứng minh có trừ công đoàn
    client.put("/api/luong/params", json={"cong_doan_rate": 0}, headers=_h(token))   # khôi phục


def test_net_floored_at_zero(client):
    """#2a: tạm ứng vượt lương thực → thực nhận = 0, KHÔNG âm."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Sàn", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    aid = client.post("/api/luong/advances", json={"employee_id": eid, "period_year": 2027,
        "period_month": 3, "advance_date": "2027-03-05", "amount": 5_000_000},
        headers=_h(token)).json()["id"]
    client.post(f"/api/luong/advances/{aid}/approve", json={}, headers=_h(token))
    gen = client.post("/api/luong/generate", json={"year": 2027, "month": 3}, headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert line["net_pay"] == 0   # 0 công + tạm ứng 5tr → sàn 0, không âm


def test_standard_cong_dong_theo_lich(client):
    """Công chuẩn (mẫu số chia lương công) ĐỘNG theo Lịch từng tháng (redesign-hcns Đ3/N4) — KHÔNG
    còn cố định 26: đổi tuần làm việc thì mẫu số của kỳ đổi theo, luôn khớp số ngày làm việc của lịch.
    Nhờ vậy làm ĐỦ tháng = nguyên lương kể cả tháng ngắn (NĐ145/2020 Đ55)."""
    token = _admin_token(client)
    base = client.get("/api/calendar/month", params={"year": 2027, "month": 9},
                      headers=_h(token)).json()["working_days"]
    assert client.post("/api/luong/generate", json={"year": 2027, "month": 9},
                       headers=_h(token)).status_code == 200
    with SessionLocal() as db:
        assert float(PayrollRepository(db).get_period_by_ym(2027, 9).standard_cong) == base

    # Tắt Thứ 7 → công chuẩn tháng GIẢM; mẫu số kỳ lương đi theo LỊCH (không phải tham số chung).
    assert client.put("/api/calendar/config", json={"works_sat": False},
                      headers=_h(token)).status_code == 200
    after = client.get("/api/calendar/month", params={"year": 2027, "month": 9},
                       headers=_h(token)).json()["working_days"]
    assert after < base
    assert client.post("/api/luong/generate", json={"year": 2027, "month": 9},
                       headers=_h(token)).status_code == 200
    with SessionLocal() as db:
        assert float(PayrollRepository(db).get_period_by_ym(2027, 9).standard_cong) == after


def test_luong_cong_capped_at_standard(client):
    """Chặn trần: làm ĐỦ (≥ công chuẩn) → nguyên lương tháng (không trả dư tháng dài); thiếu → prorate."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="cap_grp", monthly_amount=13_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="cap_grp", pay_grade_key=None)
        # 27 công / chuẩn 26 → chặn trần = nguyên 13tr (KHÔNG 27/26 = dư)
        over = svc._compute(employee=emp, salary=_sal(luong_vi_tri=13_000_000), params=params, actual_cong=27,
                            standard_cong=26, on=date(2026, 7, 1))
        assert over["luong_cong"] == 13_000_000
        # 13 công / 26 → nửa lương
        half = svc._compute(employee=emp, salary=_sal(luong_vi_tri=13_000_000), params=params, actual_cong=13,
                            standard_cong=26, on=date(2026, 7, 1))
        assert half["luong_cong"] == 6_500_000
    finally:
        db.close()


def test_special_day_premium(client):
    """#3 Đ98: làm nguyên công ngày lễ = +200% premium (base 100% đã nằm trong lương công);
    OT ngày lễ ×3, OT ngày nghỉ tuần ×2."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        svc.payroll.create_rule(payroll_group="sd", monthly_amount=26_000_000,
                                effective_from=date(2020, 1, 1))
        params = svc.get_params()   # ot 1.5 · ot_rest 2 · ot_hol 3 · hol_wm 3 · rest_wm 2
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group="sd", pay_grade_key=None)
        daily = 26_000_000 / 26   # 1.000.000 ; giờ = 125.000
        # 1 công ngày lễ (nằm trong 26 công) → premium = 1×(3−1)×daily = 2.000.000, không OT.
        v = svc._compute(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params, actual_cong=26, standard_cong=26,
                         holiday_cong=1, on=date(2026, 6, 1))
        assert v["ot_pay"] == round(daily * (3 - 1))   # 2.000.000 premium lễ
        # OT: 60' ngày lễ ×3 + 60' ngày nghỉ tuần ×2 (tổng ot_minutes = 120, không có OT thường).
        v2 = svc._compute(employee=emp, salary=_sal(luong_vi_tri=26_000_000), params=params, actual_cong=26, standard_cong=26,
                          ot_minutes=120, ot_holiday_minutes=60, ot_restday_minutes=60,
                          on=date(2026, 6, 1))
        hourly = daily / 8   # 125.000
        assert v2["ot_pay"] == round(hourly * (1 * 3 + 1 * 2))   # 125k × 5 = 625.000
    finally:
        db.close()


# --- nghỉ phép: ngày phép trả lương VỊ TRÍ + chuyên cần khi có đơn theo giờ ---


def test_ngay_phep_chi_tra_luong_vi_tri(client):
    """Chốt của chủ: ngày nghỉ phép năm CHỈ trả lương vị trí (không lương trách nhiệm).

    `luong_ngay_phep` là số TRONG ĐÓ của `luong_cong` — KHÔNG được cộng lại vào gross."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        on = date(2026, 6, 1)
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group=None, pay_grade_key=None)
        sal = _sal(luong_vi_tri=10_000_000, luong_trach_nhiem=3_000_000)   # nền 13tr, std 26

        def run(actual, leave, **kw):
            return svc._compute(employee=emp, salary=sal, params=params, actual_cong=actual,
                                standard_cong=26, paid_leave_cong=leave, on=on, **kw)

        # HỒI QUY: không có ngày phép → y như trước.
        v0 = run(26, 0)
        assert v0["luong_cong"] == 13_000_000 and v0["luong_ngay_phep"] == 0

        # 24 công làm + 2 công phép: 2 ngày phép mất phần TRÁCH NHIỆM (2 × 3tr/26).
        v = run(26, 2)
        assert v["luong_ngay_phep"] == round(2 * 10_000_000 / 26)
        assert v["luong_cong"] == round(24 * 13_000_000 / 26 + 2 * 10_000_000 / 26)
        assert v0["luong_cong"] - v["luong_cong"] == round(2 * 3_000_000 / 26)

        # ⭐ Làm DÔI công (đi làm lễ/CN): trần đã cắt bớt rồi ⇒ KHÔNG được trừ lần hai.
        v28 = run(28, 2)
        assert v28["luong_cong"] == 13_000_000 and v28["luong_ngay_phep"] == 0

        # Thử việc: đơn giá ngày phép phải mang cùng hệ số 80%.
        vp = run(26, 2, employee_status="probation")
        assert vp["luong_cong"] == round(0.8 * (24 * 13_000_000 / 26 + 2 * 10_000_000 / 26))

        # Hồ sơ CŨ chỉ khai base_amount (không có luong_vi_tri) → ngày phép KHÔNG được ra 0đ.
        legacy = svc._compute(employee=emp, salary=_sal(base_amount=13_000_000), params=params,
                              actual_cong=26, standard_cong=26, paid_leave_cong=2, on=on)
        assert legacy["luong_cong"] == 13_000_000 and legacy["luong_ngay_phep"] > 0

        # `luong_ngay_phep` KHÔNG được cộng vào gross lần nữa.
        assert v["gross"] == round(v["luong_cong"] + v["chuyen_can"] + v["allowance"]
                                   + v["khoan"] + v["ot_pay"] + v["night_pay"]
                                   + v["night_premium_pay"] + v["other_bonus"])
    finally:
        db.close()


def test_chuyen_can_nguyen_khi_co_don_nghi_gio(client):
    """Có đơn nghỉ theo giờ đã duyệt → KHÔNG mất chuyên cần, nhưng TIỀN CÔNG vẫn trừ."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        on = date(2026, 6, 1)
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group=None, pay_grade_key=None)
        sal = _sal(luong_vi_tri=10_000_000, chuyen_can=300_000)

        def run(excused):
            return svc._compute(employee=emp, salary=sal, params=params, actual_cong=25.78,
                                standard_cong=26, excused_cong=excused, on=on)

        co_don, khong_don = run(0.22), run(0)
        assert co_don["chuyen_can"] == 300_000          # nghỉ có phép → nguyên
        assert khong_don["chuyen_can"] < 300_000        # không phép → trừ dần
        # TIỀN CÔNG bằng nhau: đơn chỉ tha chuyên cần, không bù công.
        assert co_don["luong_cong"] == khong_don["luong_cong"]
    finally:
        db.close()


def test_api_dong_luong_phoi_du_field_ngay_phep(client):
    """API phải TRẢ RA `luong_ngay_phep` — thêm cột + tính đúng mà quên khai vào `LineOut`
    thì pydantic nuốt im lặng, phiếu lương không bao giờ hiện được dòng ngày phép."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Phơi Field", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000, "luong_trach_nhiem": 3_000_000}, headers=_h(token))
    line = _gen_line(client, token, eid)
    for f in ("luong_ngay_phep", "paid_leave_cong", "excused_cong"):
        assert f in line, f"API nuốt mất field '{f}' — kiểm tra LineOut trong schemas/payroll.py"


# --- Danh mục khoản thu nhập + cờ chịu thuế TNCN (chủ 27/07/2026) ------------


def _comp(client, token, *, name, kind="thu", taxable=True) -> int:
    r = client.post("/api/luong/components",
                    json={"name": name, "kind": kind, "is_taxable": taxable},
                    headers=_h(token))
    assert r.status_code == 201, r.text
    return r.json()["id"]


def _set_emp_comp(client, token, eid, values: dict[int, float | None], expect=200):
    r = client.put(f"/api/luong/components/employee/{eid}",
                   json={"items": [{"component_id": c, "amount": a} for c, a in values.items()]},
                   headers=_h(token))
    assert r.status_code == expect, r.text
    return r.json() if r.status_code < 400 else None


def test_khoan_mien_thue_khong_vao_thu_nhap_chiu_thue(client):
    """⭐ Ruột của yêu cầu: TÍCH 'chịu thuế' thì tính thuế, BỎ TÍCH thì miễn.

    Trước đây mọi phụ cấp gộp một cục nên bị tính thuế hết — người có trang phục / tiền nhà /
    đi lại / tiền cơm bị thu thuế oan."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Phụ Cấp", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 30_000_000}, headers=_h(token))

    mien = _comp(client, token, name="Trang phục test", taxable=False)
    chiu = _comp(client, token, name="Phụ cấp xăng test", taxable=True)
    # Số phải ĐỦ LỚN mới vượt giảm trừ gia cảnh 15,5tr — nhỏ quá thì thuế bằng 0 cả hai vế và
    # test không chứng minh được gì.
    _set_emp_comp(client, token, eid, {mien: 6_000_000, chiu: 22_000_000})

    line = _gen_line(client, token, eid)
    # Cả 2 khoản đều cộng vào thu nhập; chỉ khoản CHỊU thuế mới vào thu nhập tính thuế.
    assert line["allowance"] == 28_000_000
    assert line["thu_nhap_mien_thue"] == 6_000_000
    assert line["pit"] > 0, "chưa tới ngưỡng chịu thuế thì test vô nghĩa"

    # Bật cờ chịu thuế của khoản đang miễn → thuế phải TĂNG đúng phần đó.
    pit_truoc = line["pit"]
    assert client.put(f"/api/luong/components/{mien}", json={"is_taxable": True},
                      headers=_h(token)).status_code == 200
    line2 = _gen_line(client, token, eid)
    assert line2["thu_nhap_mien_thue"] == 0
    assert line2["pit"] > pit_truoc, "bật cờ chịu thuế mà thuế không tăng"
    assert line2["allowance"] == 28_000_000, "tổng phụ cấp không được đổi khi chỉ đổi cờ thuế"


def test_quy_trinh_2_buoc_khong_de_ra_khoan_moi_o_ho_so(client):
    """⭐ Chốt của chủ: muốn có khoản mới thì TẠO Ở DANH MỤC trước, rồi mới gán cho người.

    Màn hồ sơ nhân sự KHÔNG có đường nào đẻ ra khoản mới — gán id không tồn tại phải bị chặn,
    kèm chỉ đúng chỗ cần đi."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Hai Bước", status="active")

    r = client.put(f"/api/luong/components/employee/{eid}",
                   json={"items": [{"component_id": 999999, "amount": 500_000}]},
                   headers=_h(token))
    assert r.status_code == 400
    assert "Danh mục khoản thu nhập" in r.json()["detail"]

    # Bước 1: tạo ở danh mục. Bước 2: gán. Lúc này mới được.
    cid = _comp(client, token, name="Phụ cấp tiếng Nhật", taxable=True)
    _set_emp_comp(client, token, eid, {cid: 2_000_000})
    got = client.get(f"/api/luong/components/employee/{eid}", headers=_h(token)).json()["items"]
    row = next(x for x in got if x["component_id"] == cid)
    assert row["amount"] == 2_000_000 and row["is_taxable"] is True


def test_co_chiu_thue_khong_sua_duoc_o_tang_nhan_vien(client):
    """Quy tắc chỉ sống ở Tầng 1. Gửi kèm `is_taxable` lúc gán cho NV ⇒ bị bỏ qua."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Bất Biến", status="active")
    cid = _comp(client, token, name="Phụ cấp độc hại v2", taxable=False)

    r = client.put(f"/api/luong/components/employee/{eid}",
                   json={"items": [{"component_id": cid, "amount": 1_000_000,
                                    "is_taxable": True}]},
                   headers=_h(token))
    assert r.status_code == 200, r.text
    row = next(x for x in r.json()["items"] if x["component_id"] == cid)
    assert row["is_taxable"] is False, "cờ chịu thuế bị sửa từ tầng nhân viên"


def test_khoan_da_dung_thi_chi_ngung_dung_khong_xoa(client):
    """Khoản chưa có số liệu ⇒ xoá hẳn. Đã dùng ⇒ chỉ ngưng dùng, phiếu lương kỳ cũ vẫn nguyên."""
    token = _admin_token(client)
    chua_dung = _comp(client, token, name="Khoản chưa dùng", taxable=True)
    r = client.delete(f"/api/luong/components/{chua_dung}", headers=_h(token))
    assert r.status_code == 200 and r.json()["deleted"] is True

    eid = _make_emp(client, token, name="NV Giữ Dấu Vết", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    da_dung = _comp(client, token, name="Khoản đã dùng", taxable=True)
    _set_emp_comp(client, token, eid, {da_dung: 300_000})
    _gen_line(client, token, eid)

    r = client.delete(f"/api/luong/components/{da_dung}", headers=_h(token))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["deleted"] is False and body["deactivated"] is True
    assert body["employee_count"] >= 1 and body["period_count"] >= 1
    assert "KHÔNG THỂ XOÁ" in body["message"] and "NGỪNG SỬ DỤNG" in body["message"]
    # Khoản vẫn còn trong danh mục, chỉ tắt đi.
    items = client.get("/api/luong/components", headers=_h(token)).json()["items"]
    row = next(x for x in items if x["id"] == da_dung)
    assert row["is_active"] is False


def test_doi_co_chiu_thue_khong_sua_so_ky_da_tinh(client):
    """Snapshot: sửa 1 ô trên dòng lương CŨ không được lấy cờ chịu thuế HÔM NAY để tính lại."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Snapshot", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 30_000_000}, headers=_h(token))
    cid = _comp(client, token, name="Tiền cơm snapshot", taxable=False)
    _set_emp_comp(client, token, eid, {cid: 3_000_000})
    line = _gen_line(client, token, eid)
    assert line["thu_nhap_mien_thue"] == 3_000_000
    pit_cu = line["pit"]

    # Đổi cờ ở danh mục → dòng CŨ chưa tính lại thì số không được nhúc nhích.
    client.put(f"/api/luong/components/{cid}", json={"is_taxable": True}, headers=_h(token))
    r = client.put(f"/api/luong/lines/{line['id']}", json={"note": "ghi chú"}, headers=_h(token))
    assert r.status_code == 200, r.text
    assert r.json()["pit"] == pit_cu, "sửa ô ghi chú mà thuế đổi — đang đọc cờ sống thay vì snapshot"


def test_khoan_loai_tru_tru_vao_thuc_nhan(client):
    """Khoản kind='tru' trừ thẳng vào THỰC NHẬN, không gộp vào trần 30% của Điều 102."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Khấu Trừ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    thu = _comp(client, token, name="Thưởng thêm test", kind="thu", taxable=True)
    _set_emp_comp(client, token, eid, {thu: 8_000_000})
    base = _gen_line(client, token, eid)
    assert base["net_pay"] > 0, "cần thực nhận > 0 mới quan sát được khấu trừ"

    cid = _comp(client, token, name="Trừ mua đồng phục", kind="tru", taxable=True)
    _set_emp_comp(client, token, eid, {cid: 500_000})
    after = _gen_line(client, token, eid)
    assert after["net_pay"] == base["net_pay"] - 500_000
    assert after["allowance"] == base["allowance"], "khoản TRỪ không được cộng vào phụ cấp"


def test_chan_tao_khoan_trung_voi_o_da_co(client):
    """⭐ Chặn TRẢ TIỀN HAI LẦN: gõ lại tên khoản mà engine đã tự tính thì phải bị từ chối.

    Chuyên cần / phụ cấp ca / tăng ca / thưởng… đều đã có ô khai riêng. Thêm vào danh mục nữa là
    NV nhận hai lần mà phiếu lương trông vẫn bình thường (hai dòng nằm hai chỗ khác nhau)."""
    token = _admin_token(client)
    for ten in ("Chuyên cần", "chuyen can", "CHUYÊN CẦN", "Phụ cấp ca đêm", "Tăng ca",
                "Lương sản lượng",
                # Tiền ngày nghỉ phép engine TỰ TÍNH từ chấm công (`luong_ngay_phep`, nằm trong
                # `luong_cong`). Khai thêm ở danh mục là trả hai lần — bẫy có thật vì bảng Excel
                # của kế toán có hẳn cột "Phép năm(2026)".
                "Phép năm", "phep nam", "Tiền phép"):
        r = client.post("/api/luong/components",
                        json={"name": ten, "kind": "thu", "is_taxable": True},
                        headers=_h(token))
        assert r.status_code == 400, f"'{ten}' phải bị chặn nhưng lại {r.status_code}"
        assert "HAI LẦN" in r.json()["detail"]

    # Tên khác thật thì vẫn tạo được.
    ok = client.post("/api/luong/components",
                     json={"name": "Phụ cấp độc hại", "kind": "thu", "is_taxable": False},
                     headers=_h(token))
    assert ok.status_code == 201, ok.text

    # ⭐ Ngược lại: 4 khoản thưởng đã GỠ ô tay (28/07/2026) thì PHẢI tạo được — nay đó là đường
    # duy nhất khai chúng, và là chỗ khai được "chịu thuế hay không".
    for ten in ("Thưởng 5S", "Thưởng doanh số", "Thưởng thành tích", "Trả đồng phục"):
        r = client.post("/api/luong/components",
                        json={"name": ten, "kind": "thu", "is_taxable": False},
                        headers=_h(token))
        assert r.status_code == 201, f"'{ten}' phải tạo được nhưng lại {r.status_code}: {r.text}"
    # Và đổi TÊN sang tên bị cấm cũng phải chặn.
    r = client.put(f"/api/luong/components/{ok.json()['id']}",
                   json={"name": "Chuyên cần"}, headers=_h(token))
    assert r.status_code == 400 and "HAI LẦN" in r.json()["detail"]


def test_migration_don_khoan_trung_va_bu_khoan_thieu(client):
    """Sự cố seed 27/07: máy dev seed nhầm bản dở — có 7 khoản TRÙNG cột đã có, lại THIẾU đúng
    4 khoản miễn thuế là lý do sinh ra danh mục. Hai migration phải dọn sạch và bù đủ."""
    from app.db import SessionLocal
    from app.db_migrations import (
        _migrate_drop_duplicate_payroll_components as drop,
        _migrate_seed_missing_payroll_components as topup,
    )
    from sqlalchemy import text
    client  # bảo đảm app đã khởi động (create_all + seed)

    db = SessionLocal()
    try:
        # Dựng lại đúng tình huống hỏng: thêm khoản trùng, xoá khoản miễn thuế.
        db.execute(text(
            "INSERT INTO payroll_components (code, name, kind, is_taxable, in_insurance_base,"
            " sort_order, is_active, created_at) VALUES"
            " ('phu_cap_ca_dem', 'Phụ cấp ca đêm', 'thu', 1, 0, 80, 1, :now)"),
            {"now": datetime.now(timezone.utc)})
        db.execute(text("DELETE FROM payroll_components WHERE code = 'trang_phuc'"))
        db.commit()

        drop(db); topup(db); drop(db); topup(db)      # 2 vòng ⇒ idempotent
        rows = dict(db.execute(text("SELECT code, is_taxable FROM payroll_components")).all())
    finally:
        db.close()

    assert "phu_cap_ca_dem" not in rows, "khoản trùng chưa bị dọn"
    assert "trang_phuc" in rows and not rows["trang_phuc"], "chưa bù lại khoản miễn thuế"
    for code in ("tro_cap_nha_o", "ho_tro_di_lai", "tien_com"):
        assert code in rows and not rows[code], f"{code} phải có và phải MIỄN thuế"


# --- Cấu hình thuế theo TỪNG NGƯỜI (chủ 27/07/2026) -------------------------


def _emp_luong(client, token, *, name, luong=30_000_000, thu_nhap=0, **kw):
    """NV có lương + (tuỳ chọn) một khoản thu nhập CHỊU THUẾ để gross khác 0.

    NV không có chấm công thì `actual_cong = 0` ⇒ lương công = 0 ⇒ gross = 0 ⇒ thuế luôn bằng 0,
    test không chứng minh được gì. Bơm thu nhập qua khoản danh mục là cách nhẹ nhất."""
    eid = _make_emp(client, token, name=name, status="active", **kw)
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": luong}, headers=_h(token))
    if thu_nhap:
        cid = _comp(client, token, name=f"Thu nhập test {name}", taxable=True)
        _set_emp_comp(client, token, eid, {cid: thu_nhap})
    return eid


def _set_pit_mode(client, token, eid, mode):
    """`PUT /api/employees/{id}` đòi đủ trường bắt buộc — gửi thiếu là 422."""
    emp = client.get(f"/api/employees/{eid}", headers=_h(token)).json()
    emp = emp.get("employee", emp)
    body = {"full_name": emp["full_name"], "department_id": emp["department_id"],
            "hire_date": emp["hire_date"], "pit_mode": mode,
            "dependents_count": emp.get("dependents_count", 0)}
    r = client.put(f"/api/employees/{eid}", json=body, headers=_h(token))
    assert r.status_code == 200, r.text


def test_tat_giam_tru_ban_than_thi_thue_tang(client):
    """Người làm 2 nơi chỉ được đăng ký giảm trừ bản thân ở MỘT nơi. Bỏ tích ⇒ mất 15,5tr giảm
    trừ ⇒ thu nhập tính thuế tăng đúng bằng đó."""
    token = _admin_token(client)
    eid = _emp_luong(client, token, name="NV Hai Nơi", luong=40_000_000, thu_nhap=40_000_000)
    base = _gen_line(client, token, eid)
    assert base["pit"] > 0

    p = client.get("/api/luong/params", headers=_h(token)).json()
    r = client.post(f"/api/luong/salaries/{eid}",
                    json={"effective_from": "2026-01-01", "luong_vi_tri": 40_000_000,
                          "apply_self_deduction": False}, headers=_h(token))
    assert r.status_code in (200, 201), r.text
    after = _gen_line(client, token, eid)
    assert after["pit_taxable"] == base["pit_taxable"] + p["deduction_self"]
    assert after["pit"] > base["pit"]


def test_nguoi_phu_thuoc_lay_muc_tu_cau_hinh(client):
    """Giảm trừ người phụ thuộc phải LẤY TỪ CẤU HÌNH, không viết cứng. Đổi mức trong cấu hình
    thì số phải đổi theo — chống hardcode 4,4tr (mức cũ trước 2026)."""
    token = _admin_token(client)
    eid = _emp_luong(client, token, name="NV Có NPT", luong=40_000_000, thu_nhap=40_000_000)
    khong_npt = _gen_line(client, token, eid)["pit_taxable"]

    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["deduction_dependent"] == 6_200_000, "mức 2026 — không phải 4,4tr"
    emp = client.get(f"/api/employees/{eid}", headers=_h(token)).json()
    emp = emp.get("employee", emp)
    assert client.put(f"/api/employees/{eid}", json={
        "full_name": emp["full_name"], "department_id": emp["department_id"],
        "hire_date": emp["hire_date"], "dependents_count": 2}, headers=_h(token)).status_code == 200
    hai_npt = _gen_line(client, token, eid)["pit_taxable"]
    assert khong_npt - hai_npt == 2 * p["deduction_dependent"]


def test_thoi_vu_khau_tru_10_phan_tram(client):
    """HĐ dưới 3 tháng / thời vụ / thực tập: khấu trừ 10% tại nguồn, KHÔNG bảng luỹ tiến, KHÔNG
    giảm trừ gia cảnh. Sheet BLTV của xưởng có 122 dòng nhóm này."""
    token = _admin_token(client)
    eid = _emp_luong(client, token, name="NV Thời Vụ", luong=8_000_000, thu_nhap=8_000_000)
    # Mặc định luỹ tiến: 8tr < giảm trừ 15,5tr ⇒ không phải nộp đồng nào.
    assert _gen_line(client, token, eid)["pit"] == 0

    _set_pit_mode(client, token, eid, "khau_tru_10")
    line = _gen_line(client, token, eid)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    # Thuế = 10% thu nhập CHỊU thuế; và thu nhập tính thuế = chính nó (không giảm trừ gì).
    assert line["pit"] == round(line["thu_nhap_chiu_thue"] * p["pit_flat_rate"])
    assert line["pit_taxable"] == line["thu_nhap_chiu_thue"]
    assert line["pit"] > 0, "8 triệu thời vụ phải bị khấu trừ, luỹ tiến thì không"


def test_duoi_nguong_thi_khong_khau_tru_va_cam_ket_08_mien(client):
    """Dưới ngưỡng/lần trả ⇒ chưa phải khấu trừ. Có cam kết 08/CK-TNCN ⇒ không khấu trừ."""
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["pit_flat_threshold"] == 2_000_000

    nho = _emp_luong(client, token, name="NV Thu Nhập Nhỏ", luong=1_500_000, thu_nhap=1_500_000)
    _set_pit_mode(client, token, nho, "khau_tru_10")
    assert _gen_line(client, token, nho)["pit"] == 0, "dưới ngưỡng mà vẫn khấu trừ"

    ck = _emp_luong(client, token, name="NV Cam Kết 08", luong=8_000_000, thu_nhap=8_000_000)
    _set_pit_mode(client, token, ck, "cam_ket_08")
    assert _gen_line(client, token, ck)["pit"] == 0


def test_luy_tien_van_la_mac_dinh_va_khong_lech_mot_dong(client):
    """⭐ Hồi quy: NV không khai gì ⇒ vẫn luỹ tiến, số y hệt trước khi có 2 nhánh mới."""
    token = _admin_token(client)
    eid = _emp_luong(client, token, name="NV Mặc Định", luong=40_000_000, thu_nhap=40_000_000)
    emp = client.get(f"/api/employees/{eid}", headers=_h(token)).json()
    emp = emp.get("employee", emp)
    assert emp["pit_mode"] == "luy_tien"

    line = _gen_line(client, token, eid)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    # Thu nhập tính thuế = chịu thuế − BHXH − giảm trừ bản thân (không NPT).
    assert line["pit_taxable"] == round(
        line["thu_nhap_chiu_thue"] - line["bhxh"] - p["deduction_self"])


# --- Tầng 3: khoản PHÁT SINH cho riêng một kỳ (chủ 27/07/2026) ---------------


def test_thuong_nong_song_sot_khi_tinh_lai(client):
    """⭐ Chỗ nguy hiểm nhất: bấm "Tính lại" KHÔNG được xoá khoản HCNS thêm tay.

    Hàm ghi snapshot xoá-rồi-ghi-lại; nếu xoá cả dòng `source='line'` thì thưởng nóng bay mất mà
    không một thông báo nào — mất tiền của người lao động."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Thưởng Nóng", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    tu_ho_so = _comp(client, token, name="Tiền cơm thưởng nóng", taxable=False)
    _set_emp_comp(client, token, eid, {tu_ho_so: 700_000})

    line = _gen_line(client, token, eid)
    r = client.post(f"/api/luong/lines/{line['id']}/components",
                    json={"component_id": _comp(client, token, name="Thu nhập khác test",
                                                taxable=True),
                          "amount": 500_000, "note": "Thưởng nóng của Sếp"},
                    headers=_h(token))
    assert r.status_code == 201, r.text
    row_id = r.json()["id"]
    assert r.json()["source"] == "line"

    truoc = client.get(f"/api/luong/lines/{line['id']}/components", headers=_h(token)).json()["items"]
    assert len(truoc) == 2 and {x["source"] for x in truoc} == {"employee", "line"}

    # ⭐ TÍNH LẠI — khoản thêm tay phải còn nguyên, khoản từ hồ sơ được ghi đè.
    _gen_line(client, token, eid)
    sau = client.get(f"/api/luong/lines/{line['id']}/components", headers=_h(token)).json()["items"]
    con_lai = next((x for x in sau if x["id"] == row_id), None)
    assert con_lai is not None, "thưởng nóng bị Tính lại xoá mất"
    assert con_lai["amount"] == 500_000 and con_lai["note"] == "Thưởng nóng của Sếp"
    assert len([x for x in sau if x["source"] == "employee"]) == 1, "khoản từ hồ sơ bị nhân đôi"


def test_thuong_nong_khong_cong_doi_sau_khi_tinh_lai(client):
    """⭐ Thưởng nóng chỉ được cộng MỘT lần, dù đi qua đường nào.

    Bẫy: `generate` từng nối khoản hồ sơ + khoản phát sinh rồi ném chung vào `_compute`, nên
    `allowance` nuốt luôn phần `source='line'`. `update_line` lại cộng phần đó LÊN TRÊN
    `allowance` đã lưu ⇒ thêm thưởng nóng → Tính lại → sửa một ô = trả HAI LẦN.
    Kịch bản dưới đây đi đúng 3 bước đó."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Không Cộng Đôi", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    line = _gen_line(client, token, eid)
    goc = line["gross"]

    cid = _comp(client, token, name="Thưởng nóng cộng đôi", taxable=True)
    client.post(f"/api/luong/lines/{line['id']}/components",
                json={"component_id": cid, "amount": 500_000}, headers=_h(token))
    sau_them = _line_of(client, token, eid)
    assert sau_them["gross"] == goc + 500_000, "thêm khoản: sai ngay từ bước 1"

    # Bước 2 — TÍNH LẠI: `allowance` không được nuốt khoản `source='line'`.
    sau_gen = _gen_line(client, token, eid)
    assert sau_gen["gross"] == goc + 500_000, "Tính lại đã cộng thưởng nóng lần hai"

    # Bước 3 — sửa một ô bất kỳ (đường `update_line`) trên dòng đã Tính lại.
    sau_sua = client.put(f"/api/luong/lines/{line['id']}",
                         json={"note": "sửa vặt"}, headers=_h(token)).json()
    assert sau_sua["gross"] == goc + 500_000, "sửa ô sau khi Tính lại → cộng đôi thưởng nóng"


def test_line_out_tra_ra_khoan_de_phieu_luong_khop_tong(client):
    """⭐ Phiếu lương phải CỘNG RA đúng thực nhận.

    `LineOut` từng không trả `components`, nên khoản `source='line'` cộng vào `gross` mà phiếu
    không có dòng nào ⇒ tổng thu trên phiếu nhỏ hơn thực nhận, NV không đối chiếu được (đúng lớp
    lỗi `luong_ngay_phep` bị quên khai vào `LineOut` trước đây).

    Bất biến kiểm ở đây: Σ thu − Σ trừ == net_pay, dựng CHÍNH những dòng mà phiếu lương render."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Khớp Tổng", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    # Đủ 3 loại: khoản hồ sơ (chịu thuế), khoản phát sinh (miễn thuế), khoản khấu trừ.
    _set_emp_comp(client, token, eid,
                  {_comp(client, token, name="PC hồ sơ khớp tổng", taxable=True): 800_000})
    line = _gen_line(client, token, eid)
    for name, kind, amount in (("Thưởng nóng khớp tổng", "thu", 1_500_000),
                               ("Trừ mua đồng phục khớp tổng", "tru", 250_000)):
        client.post(f"/api/luong/lines/{line['id']}/components", headers=_h(token),
                    json={"component_id": _comp(client, token, name=name, kind=kind,
                                                taxable=False),
                          "amount": amount})
    l = _line_of(client, token, eid)

    comps = l["components"]
    assert len(comps) == 3, f"LineOut phải trả đủ 3 khoản, nhận {comps}"
    assert {c["source"] for c in comps} == {"employee", "line"}

    # Dựng đúng 2 cột của phiếu lương. Khoản hồ sơ ĐÃ nằm trong `allowance` nên không cộng lại;
    # chỉ khoản `source='line'` mới là dòng thu nhập thêm.
    thu = (l["luong_cong"] + l["chuyen_can"] + l["allowance"] + l["khoan"] + l["ot_pay"]
           + l["night_pay"] + l["night_premium_pay"] + l["dieu_chinh_luong"]
           + l["thuong_5s"] + l["thuong_doanh_so"] + l["thuong_thanh_tich"] + l["phep_nam"]
           + l["tra_dong_phuc"] + l["other_bonus"]
           + sum(c["amount"] for c in comps if c["kind"] != "tru" and c["source"] == "line"))
    tru = (l["bhxh"] + l["cong_doan"] + l["pit"] + l["di_tre"] + l["dt_vuot_troi"]
           + l["phat_bien_ban"] + l["phat_5s_dong_phuc"] + l["vi_pham"]
           + l["luong_dot_1_total"] + l["advance_total"]
           + sum(c["amount"] for c in comps if c["kind"] == "tru"))
    assert thu == l["gross"], f"tổng thu {thu} ≠ gross {l['gross']}"
    assert round(thu - tru) == l["net_pay"], f"phiếu ra {thu - tru}, thực nhận {l['net_pay']}"


def test_xuat_excel_cot_thuong_co_khoan_danh_muc(client):
    """File xuất phải khớp bảng lương: cột "Thưởng" gồm khoản phát sinh, không chỉ cột cũ.

    Đường xuất Excel gọi `_lines_out` KHÔNG kèm `svc` ⇒ `components` rỗng ⇒ cột Thưởng ra 0 trong
    khi cột Tổng đã có tiền. Kế toán mở file ra là thấy lệch."""
    from io import BytesIO

    from openpyxl import load_workbook
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Xuất Excel", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    line = _gen_line(client, token, eid)
    client.post(f"/api/luong/lines/{line['id']}/components", headers=_h(token),
                json={"component_id": _comp(client, token, name="Thưởng xuất Excel"),
                      "amount": 1_200_000})

    r = client.get("/api/luong/export.xlsx?year=2026&month=6", headers=_h(token))
    assert r.status_code == 200, r.text
    ws = load_workbook(BytesIO(r.content)).active
    head = [c.value for c in ws[1]]
    i_thuong, i_tong = head.index("Thưởng"), head.index("Tổng")
    row = next(r for r in ws.iter_rows(min_row=2, values_only=True) if r[1] == "NV Xuất Excel")
    assert row[i_thuong] == 1_200_000, f"cột Thưởng ra {row[i_thuong]}, mất khoản danh mục"
    assert row[i_tong] == _line_of(client, token, eid)["gross"]


def _bulk(client, token, cid, **body):
    r = client.post(f"/api/luong/components/{cid}/bulk-assign", json=body, headers=_h(token))
    return r


def _emp_comp_amount(client, token, eid, cid):
    items = client.get(f"/api/luong/components/employee/{eid}", headers=_h(token)).json()["items"]
    row = next((x for x in items if x["component_id"] == cid), None)
    return row["amount"] if row else None


def test_gan_hang_loat_mac_dinh_khong_de_muc_rieng(client):
    """⭐ Mặc định PHẢI an toàn: người đã có mức riêng thì GIỮ NGUYÊN.

    Không gửi cờ `overwrite` (client cũ / quên gửi) cũng không được đè — mức riêng đã khai cho
    từng người không có đường hoàn tác."""
    token = _admin_token(client)
    cid = _comp(client, token, name="PC gán hàng loạt")
    a = _make_emp(client, token, name="NV Bulk Chưa Có", status="active")
    b = _make_emp(client, token, name="NV Bulk Đã Có", status="active")
    _set_emp_comp(client, token, b, {cid: 800_000})

    r = _bulk(client, token, cid, amount=500_000, all_active=True)
    assert r.status_code == 200, r.text
    res = r.json()

    assert _emp_comp_amount(client, token, a, cid) == 500_000, "người chưa có phải được gán"
    assert _emp_comp_amount(client, token, b, cid) == 800_000, "mức riêng bị đè dù không xin đè"
    assert res["skipped_existing"] >= 1 and res["overwritten"] == 0
    assert res["assigned"] >= 1


def test_gan_hang_loat_bat_ghi_de_thi_de_that(client):
    """Bật `overwrite` ⇒ đè thật, và đếm riêng `overwritten` để banner nói đúng."""
    token = _admin_token(client)
    cid = _comp(client, token, name="PC gán đè")
    eid = _make_emp(client, token, name="NV Bulk Bị Đè", status="active")
    _set_emp_comp(client, token, eid, {cid: 800_000})

    r = _bulk(client, token, cid, amount=500_000, employee_ids=[eid], overwrite=True)
    assert r.status_code == 200, r.text
    res = r.json()
    assert _emp_comp_amount(client, token, eid, cid) == 500_000
    assert res["overwritten"] == 1 and res["skipped_existing"] == 0 and res["assigned"] == 0


def test_gan_hang_loat_khong_dung_co_chiu_thue(client):
    """Ghi đè chỉ đổi SỐ TIỀN. Cờ `is_taxable` vẫn bất biến ở Tầng 1."""
    token = _admin_token(client)
    cid = _comp(client, token, name="PC bulk miễn thuế", taxable=False)
    eid = _make_emp(client, token, name="NV Bulk Cờ Thuế", status="active")
    _bulk(client, token, cid, amount=300_000, employee_ids=[eid])
    _bulk(client, token, cid, amount=900_000, employee_ids=[eid], overwrite=True)

    items = client.get(f"/api/luong/components/employee/{eid}", headers=_h(token)).json()["items"]
    row = next(x for x in items if x["component_id"] == cid)
    assert row["amount"] == 900_000 and row["is_taxable"] is False


def test_gan_hang_loat_loai_nguoi_da_nghi_viec(client):
    """"Tất cả" = ĐANG LÀM VIỆC. Rải phụ cấp cho người đã nghỉ là đẻ tiền cho hồ sơ chết."""
    token = _admin_token(client)
    cid = _comp(client, token, name="PC bulk nghỉ việc")
    nghi = _make_emp(client, token, name="NV Bulk Đã Nghỉ", status="resigned")

    r = _bulk(client, token, cid, amount=400_000, all_active=True)
    assert r.status_code == 200, r.text
    assert _emp_comp_amount(client, token, nghi, cid) is None, "người đã nghỉ việc vẫn bị gán"


def test_gan_hang_loat_chan_khoan_da_ngung_ap_dung(client):
    """Khoản đã ngừng áp dụng ⇒ 400, không gán cho ai."""
    token = _admin_token(client)
    cid = _comp(client, token, name="PC bulk sắp tắt")
    eid = _make_emp(client, token, name="NV Bulk Khoản Tắt", status="active")
    assert client.put(f"/api/luong/components/{cid}", json={"is_active": False},
                      headers=_h(token)).status_code == 200

    r = _bulk(client, token, cid, amount=100_000, employee_ids=[eid])
    assert r.status_code == 400, r.text
    assert _emp_comp_amount(client, token, eid, cid) is None


def test_gan_hang_loat_vao_dung_luong(client):
    """⭐ Gán xong chạy lương thì tiền phải vào thật — không dừng ở màn cấu hình."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Bulk Ra Tiền", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    goc = _gen_line(client, token, eid)["allowance"]

    cid = _comp(client, token, name="PC bulk ra tiền")
    _bulk(client, token, cid, amount=750_000, employee_ids=[eid])
    assert _gen_line(client, token, eid)["allowance"] == goc + 750_000


def test_api_khong_con_nhan_o_thuong_cu(client):
    """⭐ Sau 28/07/2026 chỉ còn MỘT đường khai thưởng: danh mục.

    Gửi thẳng 6 cột thưởng cũ vào `PUT /lines/{id}` phải KHÔNG ăn — nếu còn ăn thì vẫn tồn tại
    đường khai thưởng bỏ qua cờ "Chịu thuế", đúng thứ chủ yêu cầu dẹp."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Ô Cũ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    line = _gen_line(client, token, eid)
    goc = line["gross"]

    r = client.put(f"/api/luong/lines/{line['id']}", headers=_h(token), json={
        "thuong_5s": 1_000_000, "thuong_doanh_so": 2_000_000, "thuong_thanh_tich": 3_000_000,
        "phep_nam": 4_000_000, "tra_dong_phuc": 5_000_000, "other_bonus": 6_000_000})
    assert r.status_code == 200, r.text
    sau = r.json()
    for f in ("thuong_5s", "thuong_doanh_so", "thuong_thanh_tich", "phep_nam",
              "tra_dong_phuc", "other_bonus"):
        assert sau[f] == 0, f"{f} vẫn ghi được qua API"
    assert sau["gross"] == goc, "21 triệu lọt vào lương qua cột đã khai tử"


def test_khoan_phat_sinh_khong_lap_sang_ky_sau(client):
    """Thưởng nóng chỉ có ở kỳ khai — kỳ sau phải sạch, không ai phải nhớ vào gỡ."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Không Lặp", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    cid = _comp(client, token, name="Thu nhập khác không lặp", taxable=True)

    line6 = _gen_line(client, token, eid)
    client.post(f"/api/luong/lines/{line6['id']}/components",
                json={"component_id": cid, "amount": 500_000}, headers=_h(token))

    gen7 = client.post("/api/luong/generate", json={"year": 2026, "month": 7},
                       headers=_h(token)).json()
    line7 = next(l for l in gen7["lines"] if l["employee_id"] == eid)
    sau = client.get(f"/api/luong/lines/{line7['id']}/components", headers=_h(token)).json()["items"]
    assert all(x["source"] != "line" for x in sau), "khoản 1 lần lặp sang kỳ sau"


def test_khoan_phat_sinh_vao_dung_tong_va_thue(client):
    """Khoản phát sinh phải cộng vào tổng lương, và chịu/miễn thuế theo đúng cờ ở danh mục."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Tổng Đúng", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    line = _gen_line(client, token, eid)
    goc_gross, goc_mien = line["gross"], line["thu_nhap_mien_thue"]

    mien = _comp(client, token, name="Hỗ trợ xăng chuyến công tác", taxable=False)
    client.post(f"/api/luong/lines/{line['id']}/components",
                json={"component_id": mien, "amount": 1_000_000}, headers=_h(token))
    after = next(l for l in client.get("/api/luong/table?year=2026&month=6",
                                       headers=_h(token)).json()["lines"]
                 if l["employee_id"] == eid)
    assert after["gross"] == goc_gross + 1_000_000
    assert after["thu_nhap_mien_thue"] == goc_mien + 1_000_000


def test_de_so_tien_khoan_tu_ho_so_nhung_KHONG_go_duoc(client):
    """ĐỔI LUẬT 12/08/2026 — trước đó SỬA cũng bị chặn.

    Chủ chốt: *"gán Hỗ trợ chi phí đi lại, nhưng tháng này nó đi nhiều hơn thì sửa thế nào?"*
    Nay SỬA SỐ TIỀN được, cho riêng kỳ này (`da_de_tay`), hồ sơ giữ nguyên.

    Nhưng **GỠ vẫn chặn**, cố ý: gỡ một khoản của hồ sơ ở đây thì lần Tính lại kế tiếp nó mọc lại
    — người dùng tưởng hệ thống nuốt thao tác. Muốn thôi trả khoản đó thì gỡ ở hồ sơ nhân viên."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Chép Hồ Sơ", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    cid = _comp(client, token, name="Phụ cấp từ hồ sơ test", taxable=True)
    _set_emp_comp(client, token, eid, {cid: 300_000})
    line = _gen_line(client, token, eid)

    row = next(x for x in client.get(f"/api/luong/lines/{line['id']}/components",
                                     headers=_h(token)).json()["items"]
               if x["source"] == "employee")
    r = client.put(f"/api/luong/lines/components/{row['id']}",
                   json={"amount": 999_000}, headers=_h(token))
    assert r.status_code == 200, r.text
    assert r.json()["amount"] == 999_000 and r.json()["da_de_tay"] is True

    # GỠ thì vẫn chặn — vế cũ giữ nguyên.
    xoa = client.delete(f"/api/luong/lines/components/{row['id']}", headers=_h(token))
    assert xoa.status_code == 400 and "Lương nhân viên" in xoa.json()["detail"]


def test_ngung_ap_dung_van_tra_luong_va_bao_ai_con_dinh(client):
    """Chốt của chủ: ngừng áp dụng KHÔNG cắt lương ai. Tiền vẫn trả đủ, hệ thống chỉ chỉ ra
    còn ai đang dính để HCNS chủ động gỡ."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Khoản Bị Tắt", status="active")
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    cid = _comp(client, token, name="Phụ cấp sắp bỏ", taxable=True)
    _set_emp_comp(client, token, eid, {cid: 800_000})
    truoc = _gen_line(client, token, eid)["gross"]

    # Tắt khoản (đã gán cho NV ⇒ chỉ ngừng áp dụng)
    r = client.delete(f"/api/luong/components/{cid}", headers=_h(token))
    assert r.status_code == 200 and r.json()["deactivated"] is True
    assert r.json()["employee_count"] == 1

    # Lương VẪN trả đủ.
    assert _gen_line(client, token, eid)["gross"] == truoc, "ngừng áp dụng mà bị cắt lương"

    # Và hệ thống chỉ đúng ai còn dính.
    holders = client.get(f"/api/luong/components/{cid}/holders", headers=_h(token)).json()
    assert [x["employee_id"] for x in holders["items"]] == [eid]

    # Màn hồ sơ thấy cờ để bật cảnh báo đỏ.
    row = next(x for x in client.get(f"/api/luong/components/employee/{eid}",
                                     headers=_h(token)).json()["items"]
               if x["component_id"] == cid)
    assert row["is_active"] is False and row["amount"] == 800_000

    # Không gán MỚI khoản đã tắt cho người khác được.
    eid2 = _make_emp(client, token, name="NV Khác", status="active")
    r2 = client.put(f"/api/luong/components/employee/{eid2}",
                    json={"items": [{"component_id": cid, "amount": 500_000}]},
                    headers=_h(token))
    assert r2.status_code == 400 and "ngừng áp dụng" in r2.json()["detail"]


def test_hai_khoan_mo_thu_nhap_khac_co_san(client):
    """Khoản lặt vặt một lần không phải đẻ danh mục mới — 2 khoản mở phải có sẵn."""
    token = _admin_token(client)
    items = client.get("/api/luong/components", headers=_h(token)).json()["items"]
    by_code = {x["code"]: x for x in items}
    assert "thu_nhap_khac_ct" in by_code and by_code["thu_nhap_khac_ct"]["is_taxable"] is True
    assert "thu_nhap_khac_mt" in by_code and by_code["thu_nhap_khac_mt"]["is_taxable"] is False


# ══════════════════════════════════ ĐỢT A (12/08/2026) — đoàn phí vào thuế


def _nv_luong(client, token, *, ten, vi_tri, trach_nhiem=0, doan_vien=False):
    eid = _make_emp(client, token, name=ten, status="active")
    r = client.post(f"/api/luong/salaries/{eid}",
                    json={"effective_from": "2026-01-01", "luong_vi_tri": vi_tri,
                          "luong_trach_nhiem": trach_nhiem, "union_member": doan_vien},
                    headers=_h(token))
    assert r.status_code == 201, r.text
    return eid


def test_A4_sua_mot_o_va_tinh_lai_ra_CUNG_SO_THUE(client):
    """⭐ Ca quan trọng nhất của đợt A — canh THỨ TỰ trong `update_line`.

    Từ 12/08/2026 thuế TRỪ đoàn phí. `update_line` vốn tính thuế TRƯỚC rồi mới tính đoàn phí — vô
    hại suốt thời gian thuế chưa dùng tới, nhưng nay để nguyên là **thuế ăn số đoàn phí CŨ**.

    ⚠️ Ca đo phải làm ĐOÀN PHÍ ĐỔI GIÁ TRỊ giữa hai lần tính, nếu không thứ tự đúng hay sai đều ra
    cùng số và test XANH GIẢ: dòng vừa `generate` đã mang sẵn đoàn phí đúng, nên `_apply_auto_pit`
    đọc trước hay sau cũng thấy y hệt. Đây đúng là chỗ bản test đầu tiên đã hụt.

    Cách làm: tính lương với tỷ lệ 0,5% → ĐỔI tỷ lệ lên 2% → sờ vào dòng → so với `generate` mới.
    Hai đường phải ra CÙNG đoàn phí và CÙNG thuế."""
    token = _admin_token(client)
    client.put("/api/luong/params", json={"cong_doan_rate": 0.005}, headers=_h(token))
    eid = _nv_luong(client, token, ten="NV Doan Vien A4", vi_tri=60_000_000, doan_vien=True)

    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                      headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert line["cong_doan"] > 0, "ca đo hỏng: NV này phải có đoàn phí"

    # ⚠️ NV thử không có lượt chấm công nào ⇒ lương công = 0 ⇒ gross ≈ 0 ⇒ THUẾ = 0 CẢ HAI ĐƯỜNG,
    # và ca đo thành vô nghĩa (bản đầu tiên của test này đã dính đúng vậy: đột biến dời khối đoàn
    # phí về chỗ cũ vẫn XANH). Bơm một khoản chịu thuế đủ lớn để thuế thật sự khác 0.
    line = client.put(f"/api/luong/lines/{line['id']}",
                      json={"dieu_chinh_luong": 100_000_000}, headers=_h(token)).json()
    assert line["pit"] > 0, "ca đo hỏng: chưa có thuế thì so thuế vô nghĩa"

    # Tỷ lệ đổi ⇒ `update_line` phải tính lại đoàn phí, VÀ thuế phải theo số MỚI đó.
    client.put("/api/luong/params", json={"cong_doan_rate": 0.02}, headers=_h(token))
    sua = client.put(f"/api/luong/lines/{line['id']}", json={"note": "soát lại"},
                     headers=_h(token)).json()
    tinh_lai = next(l for l in client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                                           headers=_h(token)).json()["lines"]
                    if l["employee_id"] == eid)

    assert sua["cong_doan"] == tinh_lai["cong_doan"] > line["cong_doan"]
    assert sua["pit_taxable"] == tinh_lai["pit_taxable"], (
        f'"Sửa 1 ô" ra thu nhập tính thuế {sua["pit_taxable"]}, "Tính lại" ra '
        f'{tinh_lai["pit_taxable"]} — khối đoàn phí đang nằm SAU khối TNCN trong `update_line`'
    )
    assert sua["pit"] == tinh_lai["pit"] > 0


def test_A5_khong_phai_doan_vien_thi_sua_dong_khong_lam_doan_phi_song_lai(client):
    """Lỗi #3 trong `docs/CONG_THUC_TINH_LUONG.md` Phần 14 — nay vá cùng đợt A.

    `update_line` chỉ kiểm `is_probation`, QUÊN cờ đoàn viên. "Tính lại" ra 0đ đúng, nhưng mọi
    thao tác sửa dòng (kể cả thêm/xoá khoản phát sinh) làm đoàn phí SỐNG LẠI và trừ vào thực nhận.

    Từ 12/08/2026 lỗi này còn nặng hơn: đoàn phí ma làm GIẢM THUẾ của người không hề đóng."""
    token = _admin_token(client)
    client.put("/api/luong/params", json={"cong_doan_rate": 0.005}, headers=_h(token))
    eid = _nv_luong(client, token, ten="NV Khong Doan Vien", vi_tri=60_000_000, doan_vien=False)

    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                      headers=_h(token)).json()
    line = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert line["cong_doan"] == 0, "không phải đoàn viên mà Tính lại đã ra tiền"

    sau = client.put(f"/api/luong/lines/{line['id']}", json={"note": "soát lại"},
                     headers=_h(token)).json()
    assert sau["cong_doan"] == 0, "sửa một ô làm đoàn phí sống lại — đúng lỗi #3"
    assert sau["pit"] == line["pit"], "đoàn phí ma còn kéo thuế xuống theo"
    assert sau["net_pay"] == line["net_pay"]


def test_A_doan_phi_bam_muc_nen_va_giam_thue(client):
    """Ba mục của đợt A gặp nhau trên một dòng lương thật, qua đúng đường API.

    Bảng lương công ty (T05/2026) là chuẩn đối chiếu: BH bắt buộc và đoàn phí CÙNG một gốc, và
    đoàn phí nằm trong khối giảm trừ trước thuế."""
    token = _admin_token(client)
    client.put("/api/luong/params", json={"cong_doan_rate": 0.005}, headers=_h(token))
    eid = _nv_luong(client, token, ten="NV Co Trach Nhiem", vi_tri=8_000_000,
                    trach_nhiem=2_000_000, doan_vien=True)

    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                      headers=_h(token)).json()
    ln = next(l for l in gen["lines"] if l["employee_id"] == eid)

    # Mục 3: mức đóng BH = vị trí + trách nhiệm.
    assert ln["insurance_base"] == 10_000_000
    # Mục 4: đoàn phí CÙNG GỐC với BH, không phải thực lĩnh.
    assert ln["cong_doan"] == round(10_000_000 * 0.005)
    assert ln["cong_doan"] != round(float(ln["net_pay"]) * 0.005)
