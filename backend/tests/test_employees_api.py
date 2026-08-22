"""Hồ sơ nhân sự (module `nhan_su`) — lát #1.

CRUD + KPIs, soft duplicate CCCD/BHXH, stage transitions (confirm/resign/reinstate/
transfer/promote) with Quá trình công tác events, account link/create/unlink, attachments,
and the RBAC gate. Admin (Giám đốc) holds full nhan_su by seed; a NV Sales user does not.
"""
from __future__ import annotations

from app.db import SessionLocal
from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
from app.repositories.user_repo import UserRepository
from app.security import create_access_token, hash_password

ADMIN = {"username": "admin", "password": "admin123"}


def _admin_token(client) -> str:
    return client.post("/api/auth/login", json=ADMIN).json()["access_token"]


def _h(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _dept_id(name: str) -> int:
    db = SessionLocal()
    try:
        return DepartmentRepository(db).get_by_name(name).id
    finally:
        db.close()


def _sales_token() -> str:
    db = SessionLocal()
    try:
        users = UserRepository(db)
        existing = users.get_by_username("sales-emp")
        if existing is not None:
            return create_access_token(str(existing.id))
        kd = DepartmentRepository(db).get_by_name("Kinh doanh")
        role = RoleRepository(db).get_by_name_and_department("NV Sales", kd.id)
        u = users.create(username="sales-emp", name="S", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=kd.id, role_id=role.id, is_active=True)
        return create_access_token(str(u.id))
    finally:
        db.close()


def _create(client, token, **over):
    body = {
        "full_name": over.pop("full_name", "Trần Văn A"),
        "department_id": over.pop("department_id", _dept_id("Hành chính nhân sự")),
        "hire_date": over.pop("hire_date", "2024-01-15"),
        # Hồ sơ mặc định là THỬ VIỆC, mà trạng thái đó nay BẮT BUỘC khai ngày hết thử việc.
        # Mốc chọn TRƯỚC ngày bật "tự đánh dấu hết thử việc" (22/08/2026) ⇒ hàm quét bỏ qua ⇒
        # hồ sơ đứng yên ở "probation", hành vi các test cũ không đổi một chữ.
        "probation_end_date": over.pop("probation_end_date", "2025-12-31"),
    }
    body.update(over)
    return client.post("/api/employees", json=body, headers=_h(token))


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _doc_xlsx(data: bytes) -> list[list]:
    """Doc lai file .xlsx thanh danh sach dong — de kiem NOI DUNG chu khong chi kiem status 200."""
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        return [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()


def test_xuat_nhan_su_ra_file_xlsx_that(client):
    """Nut "Xuat Excel" phai ra file .xlsx THAT, khong phai CSV doi ten (chu chot 08/08/2026)."""
    token = _admin_token(client)
    _create(client, token, full_name="Nguyen Thi Xuat", hire_date="2024-03-01")

    resp = client.get("/api/employees/export.xlsx", headers=_h(token))
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(XLSX_MIME)
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK", "phai la file .xlsx (zip), khong phai chuoi CSV"

    rows = _doc_xlsx(resp.content)
    assert rows[0] == ["Mã", "Họ tên", "Phòng/Tổ", "Chức danh", "Bậc tay nghề", "Trạng thái",
                       "Ngày vào", "Tài khoản"]
    ten = [r[1] for r in rows[1:]]
    assert "Nguyen Thi Xuat" in ten
    # Ngay vao phai la chuoi dd/mm/yyyy — de nguyen kieu ngay thi moi may Excel hien mot kieu.
    dong = next(r for r in rows[1:] if r[1] == "Nguyen Thi Xuat")
    assert dong[6] == "01/03/2024"


def test_xuat_nhan_su_khong_bi_cat_o_200_nguoi(client):
    """Ban cu chi lay 200 nguoi dau roi IM LANG. File phai co DU so nguoi nhu o "Tong" tren man."""
    token = _admin_token(client)
    tong = client.get("/api/employees?size=1", headers=_h(token)).json()["total"]

    rows = _doc_xlsx(client.get("/api/employees/export.xlsx", headers=_h(token)).content)
    assert len(rows) - 1 == tong, "so dong trong file phai bang so Tong tren man"


def test_xuat_nhan_su_theo_dung_bo_loc_cua_man(client):
    """File xuat ra phai phan anh DUNG bo loc dang chon, khong phai ca danh sach."""
    token = _admin_token(client)
    _create(client, token, full_name="Loc Theo Phong", department_id=_dept_id("Kinh doanh"))
    kd = _dept_id("Kinh doanh")

    tong_kd = client.get(
        f"/api/employees?size=1&department_id={kd}", headers=_h(token)
    ).json()["total"]
    rows = _doc_xlsx(
        client.get(f"/api/employees/export.xlsx?department_id={kd}", headers=_h(token)).content
    )
    assert len(rows) - 1 == tong_kd
    assert all(r[2] == "Kinh doanh" for r in rows[1:])


def test_xuat_nhan_su_chan_nguoi_khong_co_quyen(client):
    """Ro du lieu nhan su la rui ro lon nhat cua endpoint nay — nguoi khong co quyen phai bi chan."""
    resp = client.get("/api/employees/export.xlsx", headers=_h(_sales_token()))
    assert resp.status_code == 403


# --- self-service "Hồ sơ của tôi" ------------------------------------------


def _vai_tho(client, token) -> int:
    """Vai TRỐNG QUYỀN cho thợ — chỉ có hai ô mặc định (Tự phục vụ + Nội quy).

    Từ 10/08/2026 tự phục vụ là MỘT Ô QUYỀN. Tài khoản KHÔNG GÁN VAI thì không có ô nào, nên
    cũng không mở được "Hồ sơ của tôi" — giống hệt mọi module khác của hệ thống. Vai mới sinh ra
    đã kèm sẵn hai ô đó (xem `RoleRepository.O_MAC_DINH`)."""
    meta = client.get("/api/employees/meta", headers=_h(token)).json()
    dept = next(d["id"] for d in meta["departments"] if d["name"] == "Hành chính nhân sự")
    from app.db import SessionLocal
    from app.repositories.rbac_repo import RoleRepository
    db = SessionLocal()
    try:
        roles = RoleRepository(db)
        vai = roles.get_by_name_and_department("Tho trong quyen", dept) or roles.create(
            name="Tho trong quyen", department_id=dept)
        return vai.id
    finally:
        db.close()


def test_my_profile_self_service(client):
    token = _admin_token(client)
    _create(client, token, full_name="NV Tự", phone="0900", note="ghi chú nội bộ",
            payroll_group="van_phong",
            account={"username": "nvtu", "password": "nvtu12345", "role_id": _vai_tho(client, token)})
    me_tok = client.post("/api/auth/login", json={"username": "nvtu", "password": "nvtu12345"}).json()["access_token"]

    me = client.get("/api/employees/me", headers=_h(me_tok)).json()
    assert me["has_employee"] is True
    assert me["employee"]["full_name"] == "NV Tự" and me["employee"]["phone"] == "0900"
    assert me["employee"]["note"] is None and me["employee"]["payroll_group"] is None  # nội bộ, ẩn

    # tự sửa liên lạc (whitelist): phone đổi được; full_name bị bỏ qua (không whitelist)
    upd = client.put("/api/employees/me", json={"phone": "0911", "full_name": "HACK"}, headers=_h(me_tok)).json()
    assert upd["employee"]["phone"] == "0911" and upd["employee"]["full_name"] == "NV Tự"

    # user không gắn hồ sơ → has_employee false
    assert client.get("/api/employees/me", headers=_h(_sales_token())).json()["has_employee"] is False


def test_profile_update_request_flow(client):
    """NV đề nghị đổi field bảo vệ (số TK) → chỉ field whitelist được ghi nhận; HCNS duyệt
    → áp thật vào hồ sơ; duyệt lại đơn đã xử lý → 400."""
    admin = _admin_token(client)
    _create(client, admin, full_name="NV Yêu Cầu", bank_account="111",
            account={"username": "nvyc", "password": "nvyc12345", "role_id": _vai_tho(client, admin)})
    me_tok = client.post("/api/auth/login", json={"username": "nvyc", "password": "nvyc12345"}).json()["access_token"]

    # đề nghị đổi số TK (được) + phone (không whitelist → bị loại)
    r = client.post("/api/employees/me/update-requests",
                    json={"changes": {"bank_account": "999", "phone": "0900"}, "reason": "Đổi NH"},
                    headers=_h(me_tok))
    assert r.status_code == 201
    rid = r.json()["id"]
    assert r.json()["changes"].get("bank_account") == "999" and "phone" not in r.json()["changes"]

    assert any(x["id"] == rid and x["status"] == "pending"
               for x in client.get("/api/employees/me/update-requests", headers=_h(me_tok)).json()["items"])

    # HCNS thấy + duyệt
    listed = client.get("/api/employees/update-requests?status=pending", headers=_h(admin)).json()["items"]
    assert any(x["id"] == rid and x["employee_name"] == "NV Yêu Cầu" for x in listed)
    ap = client.post(f"/api/employees/update-requests/{rid}/approve", json={}, headers=_h(admin))
    assert ap.status_code == 200 and ap.json()["status"] == "approved"

    # áp thật: số TK đã đổi
    assert client.get("/api/employees/me", headers=_h(me_tok)).json()["employee"]["bank_account"] == "999"
    # duyệt lại đơn đã xử lý → 400
    assert client.post(f"/api/employees/update-requests/{rid}/approve", json={}, headers=_h(admin)).status_code == 400


def test_ho_so_cua_toi_tra_du_o_bhxh_thue_va_quan_ly_truc_tiep(client):
    """Màn "Hồ sơ của tôi" đọc thẳng payload này. Thiếu một field ở schema Out là màn hiện
    "—" trong khi DB có số — nên chốt bằng test: BHXH · cách tính thuế · ngày và nơi cấp CCCD ·
    hạn thử việc · trưởng bộ phận (do route /me tự tra)."""
    admin = _admin_token(client)
    hcns = _dept_id("Hành chính nhân sự")
    r = _create(client, admin, full_name="NV Đủ Ô", department_id=hcns,
                social_insurance_no="7912345678", pit_tax_code="8123456789",
                national_id="079080000123", national_id_date="2021-03-04",
                national_id_place="Cục CS QLHC", probation_end_date="2024-03-15",
                account={"username": "nvdu", "password": "nvdu12345", "role_id": _vai_tho(client, admin)})
    assert r.status_code == 201, r.text

    # Trưởng bộ phận của phòng = tài khoản admin.
    db = SessionLocal()
    try:
        depts = DepartmentRepository(db)
        admin_user = UserRepository(db).get_by_username("admin")
        depts.set_head(depts.get_by_id(hcns), admin_user.id)
        head_name = admin_user.name or admin_user.username
    finally:
        db.close()

    me_tok = client.post("/api/auth/login",
                         json={"username": "nvdu", "password": "nvdu12345"}).json()["access_token"]
    emp = client.get("/api/employees/me", headers=_h(me_tok)).json()["employee"]
    assert emp["social_insurance_no"] == "7912345678"
    assert emp["pit_mode"] == "luy_tien"
    assert emp["national_id_date"] == "2021-03-04" and emp["national_id_place"] == "Cục CS QLHC"
    assert emp["probation_end_date"] == "2024-03-15"
    assert emp["department_head_name"] == head_name


def test_nv_tu_rut_de_nghi_khi_hcns_chua_xu_ly(client):
    """Gõ nhầm thì phải rút lại được — nhưng rút = đổi trạng thái, KHÔNG xoá vết; và chỉ rút
    được đề nghị CỦA MÌNH, còn đang chờ."""
    admin = _admin_token(client)
    _create(client, admin, full_name="NV Rút", bank_account="111",
            account={"username": "nvrut", "password": "nvrut12345", "role_id": _vai_tho(client, admin)})
    _create(client, admin, full_name="NV Khác", account={"username": "nvkhac", "password": "nvkhac1234", "role_id": _vai_tho(client, admin)})
    tok = client.post("/api/auth/login",
                      json={"username": "nvrut", "password": "nvrut12345"}).json()["access_token"]
    tok_khac = client.post("/api/auth/login",
                           json={"username": "nvkhac", "password": "nvkhac1234"}).json()["access_token"]

    rid = client.post("/api/employees/me/update-requests",
                      json={"changes": {"bank_account": "999"}, "reason": "Đổi NH"},
                      headers=_h(tok)).json()["id"]

    # người khác KHÔNG rút hộ được
    assert client.post(f"/api/employees/me/update-requests/{rid}/cancel",
                       headers=_h(tok_khac)).status_code == 404

    ok = client.post(f"/api/employees/me/update-requests/{rid}/cancel", headers=_h(tok))
    assert ok.status_code == 200 and ok.json()["status"] == "cancelled"
    assert ok.json()["decided_at"] is not None and ok.json()["decided_by_name"] == "NV Rút"

    # vẫn còn trong danh sách (có vết), và HCNS không duyệt được nữa
    mine = client.get("/api/employees/me/update-requests", headers=_h(tok)).json()["items"]
    assert any(x["id"] == rid and x["status"] == "cancelled" for x in mine)
    assert client.post(f"/api/employees/update-requests/{rid}/approve", json={},
                       headers=_h(admin)).status_code == 400
    # rút lần hai → 400, và hồ sơ KHÔNG bị áp thay đổi
    assert client.post(f"/api/employees/me/update-requests/{rid}/cancel", headers=_h(tok)).status_code == 400
    assert client.get("/api/employees/me", headers=_h(tok)).json()["employee"]["bank_account"] == "111"


def test_de_nghi_bi_tu_choi_luu_ai_quyet_va_ly_do(client):
    """NV phải thấy ai từ chối, lúc nào, vì sao — không thì màn chỉ có chữ "Từ chối" trơ trọi."""
    admin = _admin_token(client)
    _create(client, admin, full_name="NV Bị Từ Chối", bank_account="111",
            account={"username": "nvtc", "password": "nvtc12345", "role_id": _vai_tho(client, admin)})
    tok = client.post("/api/auth/login",
                      json={"username": "nvtc", "password": "nvtc12345"}).json()["access_token"]
    rid = client.post("/api/employees/me/update-requests",
                      json={"changes": {"bank_account": "999"}}, headers=_h(tok)).json()["id"]

    client.post(f"/api/employees/update-requests/{rid}/reject",
                json={"note": "Sai số tài khoản"}, headers=_h(admin))

    row = next(x for x in client.get("/api/employees/me/update-requests", headers=_h(tok)).json()["items"]
               if x["id"] == rid)
    assert row["status"] == "rejected" and row["decision_note"] == "Sai số tài khoản"
    assert row["decided_at"] is not None and row["decided_by_name"]


def test_de_nghi_cua_toi_cat_trang_o_may_chu_va_dem_theo_trang_thai(client):
    """Đề nghị KHÔNG bị xoá khi rút/từ chối nên danh sách chỉ dài thêm ⇒ cắt trang Ở MÁY CHỦ.

    Hai chốt: (1) `page`/`size` cắt đúng và `total` là tổng THẬT chứ không phải số dòng trả về;
    (2) `dem` đếm trên TOÀN BỘ hồ sơ và KHÔNG đổi theo bộ lọc — badge "N chờ duyệt" cùng số trên
    pill lọc đọc ô này; đếm lại từ `items` của trang đang xem là sai ngay khi qua trang 2."""
    admin = _admin_token(client)
    _create(client, admin, full_name="NV Nhiều Đề Nghị", bank_account="111",
            account={"username": "nvnhieu", "password": "nvnhieu123", "role_id": _vai_tho(client, admin)})
    tok = client.post("/api/auth/login",
                      json={"username": "nvnhieu", "password": "nvnhieu123"}).json()["access_token"]

    ids = [client.post("/api/employees/me/update-requests",
                       json={"changes": {"bank_account": f"90{i}"}}, headers=_h(tok)).json()["id"]
           for i in range(5)]
    client.post(f"/api/employees/me/update-requests/{ids[0]}/cancel", headers=_h(tok))
    client.post(f"/api/employees/update-requests/{ids[1]}/reject", json={"note": "Sai"}, headers=_h(admin))

    trang1 = client.get("/api/employees/me/update-requests?page=1&size=2", headers=_h(tok)).json()
    assert trang1["total"] == 5 and trang1["page"] == 1 and trang1["size"] == 2
    assert [x["id"] for x in trang1["items"]] == sorted(ids, reverse=True)[:2]   # mới nhất trước

    trang3 = client.get("/api/employees/me/update-requests?page=3&size=2", headers=_h(tok)).json()
    assert [x["id"] for x in trang3["items"]] == [ids[0]] and trang3["total"] == 5

    assert trang1["dem"] == {"pending": 3, "cancelled": 1, "rejected": 1}

    loc = client.get("/api/employees/me/update-requests?status=pending&size=10", headers=_h(tok)).json()
    assert loc["total"] == 3 and all(x["status"] == "pending" for x in loc["items"])
    assert loc["dem"]["cancelled"] == 1        # pill "Đã rút" vẫn phải hiện số dù đang lọc pending
    assert client.get("/api/employees/me/update-requests?status=xyz", headers=_h(tok)).status_code == 400


def test_de_nghi_dai_hon_o_ho_so_bi_chan_va_hang_doi_kem_gia_tri_hien_tai(client):
    """Hai chốt của màn HCNS duyệt đề nghị cập nhật:

    1) `changes` là JSON — dài bao nhiêu cũng lưu được, nên phải ĐO NGAY LÚC GỬI theo đúng độ
       dài cột (`bank_account` = String(30)). Không đo thì đề nghị 44 ký tự lọt vào hàng đợi,
       tới lúc HCNS bấm Duyệt mới ghi vào cột thật ⇒ Postgres nổ `value too long` và người
       DUYỆT lãnh lỗi thay người GÕ (SQLite của test không nổ, nên chốt bằng test này).
    2) Hàng đợi phải kèm giá trị ĐANG có: người duyệt cần thấy đổi TỪ GÌ sang gì mới quyết được.
    """
    admin = _admin_token(client)
    _create(client, admin, full_name="NV Dài Chữ", bank_account="111",
            account={"username": "nvdai", "password": "nvdai12345", "role_id": _vai_tho(client, admin)})
    tok = client.post("/api/auth/login",
                      json={"username": "nvdai", "password": "nvdai12345"}).json()["access_token"]

    qua_dai = client.post("/api/employees/me/update-requests",
                          json={"changes": {"bank_account": "9" * 44}}, headers=_h(tok))
    assert qua_dai.status_code == 400
    assert "30 ký tự" in qua_dai.json()["detail"]

    rid = client.post("/api/employees/me/update-requests",
                      json={"changes": {"bank_account": "999"}}, headers=_h(tok)).json()["id"]
    row = next(x for x in client.get("/api/employees/update-requests?status=pending",
                                     headers=_h(admin)).json()["items"] if x["id"] == rid)
    assert row["current"]["bank_account"] == "111"


# --- create + list ----------------------------------------------------------


def test_create_assigns_code_probation_and_hired_event(client):
    token = _admin_token(client)
    resp = _create(client, token, full_name="Nguyễn Thị B")
    assert resp.status_code == 201
    emp = resp.json()["employee"]
    assert emp["code"].startswith("NV")
    assert emp["status"] == "probation"  # mặc định Thử việc
    assert emp["department_name"] == "Hành chính nhân sự"

    events = client.get(f"/api/employees/{emp['id']}/events", headers=_h(token)).json()["items"]
    assert any(e["event_type"] == "hired" for e in events)


def _bac(client, token, code: str) -> int:
    """id của một bậc trong danh mục theo mã (bac_1..bac_5)."""
    items = client.get("/api/employees/bac-tay-nghe", headers=_h(token)).json()["items"]
    return next(g["id"] for g in items if g["code"] == code)


def test_create_with_employee_specific_salary_can_have_different_amounts(client):
    """Bậc thợ KHÔNG quyết định tiền — cùng bậc, 2 NV 2 mức lương vị trí khác nhau (bảng T05:
    cùng bậc 2 mà người 20tr người 10,5tr).

    Từ 29/07/2026 bậc là DANH MỤC (`job_grade_id`) chứ không còn là chữ tự do, nhưng luật trên
    không đổi: chủ chốt "khai bậc thôi, không cần điền tiền"."""
    token = _admin_token(client)
    dept_id = _dept_id("Hành chính nhân sự")
    bac_2 = _bac(client, token, "bac_2")

    first = _create(
        client, token, full_name="NV mức riêng A", department_id=dept_id,
        job_grade_id=bac_2,
        initial_salary={"luong_vi_tri": 8_000_000, "luong_trach_nhiem": 1_000_000,
                        "chuyen_can": 300_000},
    )
    second = _create(
        client, token, full_name="NV mức riêng B", department_id=dept_id,
        job_grade_id=bac_2,
        initial_salary={"luong_vi_tri": 13_000_000, "luong_trach_nhiem": 2_000_000,
                        "chuyen_can": 500_000},
    )
    assert first.status_code == 201 and second.status_code == 201
    assert first.json()["employee"]["job_grade_name"] == "Thợ vững"
    assert second.json()["employee"]["job_grade_id"] == bac_2, "hai người CÙNG một bậc"

    salary_a = client.get(
        f"/api/luong/salaries/{first.json()['employee']['id']}", headers=_h(token)
    ).json()["items"][0]
    salary_b = client.get(
        f"/api/luong/salaries/{second.json()['employee']['id']}", headers=_h(token)
    ).json()["items"][0]
    assert salary_a["amount_mode"] == salary_b["amount_mode"] == "manual"
    assert salary_a["luong_vi_tri"] + salary_a["luong_trach_nhiem"] == 9_000_000
    assert salary_b["luong_vi_tri"] + salary_b["luong_trach_nhiem"] == 15_000_000
    # Mức đóng BH = lương vị trí (khác nhau giữa 2 người cùng bậc).
    assert salary_a["luong_vi_tri"] == 8_000_000 and salary_b["luong_vi_tri"] == 13_000_000


def test_create_employee_with_initial_salary_no_grade_does_not_break(client):
    """Sau khi GỠ bậc: tạo NV + khai lương ban đầu (chỉ lương vị trí, không bậc) vẫn chạy."""
    token = _admin_token(client)
    resp = _create(
        client, token, full_name="NV gõ lương tay", department_id=_dept_id("Kinh doanh"),
        initial_salary={"luong_vi_tri": 8_000_000},
    )
    assert resp.status_code == 201, resp.text
    eid = resp.json()["employee"]["id"]
    prev = client.get(f"/api/luong/salaries/{eid}/preview", headers=_h(token)).json()
    assert prev["monthly"] == 8_000_000 and prev["insurance_base"] == 8_000_000


def test_create_with_initial_salary_requires_payroll_permission_before_creating_employee(client):
    admin = _admin_token(client)
    token = _ns_no_salary_token()
    before = client.get("/api/employees", headers=_h(admin)).json()["total"]

    response = _create(
        client,
        token,
        full_name="NV khong duoc khai luong ban dau",
        initial_salary={"luong_vi_tri": 8_000_000},
    )

    assert response.status_code == 403
    after = client.get("/api/employees", headers=_h(admin)).json()["total"]
    assert after == before


def test_create_rejects_salary_effective_before_hire_date_without_partial_employee(client):
    token = _admin_token(client)
    before = client.get("/api/employees", headers=_h(token)).json()["total"]

    response = _create(
        client,
        token,
        full_name="NV sai ngay hieu luc luong",
        hire_date="2026-02-01",
        initial_salary={
            "effective_from": "2026-01-01",
            "luong_vi_tri": 8_000_000,
        },
    )

    assert response.status_code == 400
    assert "ngay vao lam" in response.json()["detail"].lower()
    after = client.get("/api/employees", headers=_h(token)).json()["total"]
    assert after == before


def test_list_has_kpis_and_status_filter(client):
    token = _admin_token(client)
    # Baseline: mọi tài khoản đều có hồ sơ (`backfill_employee_profiles`) nên DB đã có sẵn hồ sơ
    # nền (vd admin, active). Đo DELTA thay vì số tuyệt đối để không phụ thuộc dữ liệu nền.
    base = client.get("/api/employees", headers=_h(token)).json()["kpis"]
    _create(client, token, full_name="A")   # probation (mặc định)
    active = _create(client, token, full_name="B").json()["employee"]
    client.post(
        f"/api/employees/{active['id']}/transitions",
        json={"kind": "confirm", "effective_date": "2024-03-01"},
        headers=_h(token),
    )

    data = client.get("/api/employees", headers=_h(token)).json()
    assert data["kpis"]["total"] == base["total"] + 2
    assert data["kpis"]["probation"] == base["probation"] + 1
    assert data["kpis"]["active"] == base["active"] + 1

    only_active = client.get("/api/employees?status=active", headers=_h(token)).json()
    assert only_active["total"] == base["active"] + 1
    assert all(it["status"] == "active" for it in only_active["items"])


# --- soft duplicate ---------------------------------------------------------


def test_duplicate_cccd_is_soft_warning(client):
    token = _admin_token(client)
    first = _create(client, token, full_name="X", national_id="079123456789").json()["employee"]
    resp = _create(client, token, full_name="Y", national_id="079123456789")
    assert resp.status_code == 201  # KHÔNG chặn
    dup = resp.json()["duplicate_national_id"]
    assert dup is not None and dup["id"] == first["id"]


# --- edit -------------------------------------------------------------------


def test_update_edits_profile(client):
    token = _admin_token(client)
    emp = _create(client, token).json()["employee"]
    resp = client.put(
        f"/api/employees/{emp['id']}",
        json={"full_name": "Tên Mới", "phone": "0900000000", "dependents_count": 2,
              "probation_end_date": "2025-12-31"},
        headers=_h(token),
    )
    assert resp.status_code == 200
    out = resp.json()["employee"]
    assert out["full_name"] == "Tên Mới"
    assert out["phone"] == "0900000000"
    assert out["dependents_count"] == 2


# --- transitions ------------------------------------------------------------


def test_confirm_then_illegal_transition(client):
    token = _admin_token(client)
    emp = _create(client, token).json()["employee"]
    ok = client.post(
        f"/api/employees/{emp['id']}/transitions",
        json={"kind": "confirm", "effective_date": "2024-03-01"},
        headers=_h(token),
    )
    assert ok.status_code == 200 and ok.json()["status"] == "active"
    # confirm again from `active` is illegal
    bad = client.post(
        f"/api/employees/{emp['id']}/transitions",
        json={"kind": "confirm"},
        headers=_h(token),
    )
    assert bad.status_code == 400


def test_resign_requires_reason_then_locks_edit_until_reinstate(client):
    token = _admin_token(client)
    emp = _create(client, token).json()["employee"]
    eid = emp["id"]

    no_reason = client.post(
        f"/api/employees/{eid}/transitions", json={"kind": "resign"}, headers=_h(token)
    )
    assert no_reason.status_code == 400

    resigned = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "resign", "effective_date": "2024-06-30", "resign_reason": "Tự xin nghỉ"},
        headers=_h(token),
    )
    assert resigned.status_code == 200 and resigned.json()["status"] == "resigned"

    # Resigned hồ sơ is read-only.
    locked = client.put(f"/api/employees/{eid}", json={"full_name": "Z"}, headers=_h(token))
    assert locked.status_code == 400

    # Reinstate reopens it.
    back = client.post(
        f"/api/employees/{eid}/transitions", json={"kind": "reinstate"}, headers=_h(token)
    )
    assert back.status_code == 200 and back.json()["status"] == "active"
    assert client.put(f"/api/employees/{eid}", json={"full_name": "Z"}, headers=_h(token)).status_code == 200


def test_transfer_and_promote_record_events(client):
    token = _admin_token(client)
    hcns = _dept_id("Hành chính nhân sự")
    kd = _dept_id("Kinh doanh")
    emp = _create(client, token, department_id=hcns).json()["employee"]
    eid = emp["id"]

    # transfer to another department
    t = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "transfer", "new_department_id": kd, "effective_date": "2024-04-06"},
        headers=_h(token),
    )
    assert t.status_code == 200 and t.json()["department_id"] == kd
    # transfer to same dept is rejected
    same = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "transfer", "new_department_id": kd},
        headers=_h(token),
    )
    assert same.status_code == 400

    # promote (bậc tay nghề — chọn từ danh mục, không còn gõ chữ tự do)
    bac_3 = _bac(client, token, "bac_3")
    p = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "promote", "new_job_grade_id": bac_3},
        headers=_h(token),
    )
    assert p.status_code == 200 and p.json()["job_grade_name"] == "Thợ thường"

    kinds = {e["event_type"] for e in client.get(f"/api/employees/{eid}/events", headers=_h(token)).json()["items"]}
    assert {"hired", "transferred", "promoted"} <= kinds


def test_transfer_and_promote_change_grade_without_changing_salary(client):
    """Bậc gỡ hẳn → điều chuyển đổi PHÒNG, thăng bậc đổi `job_grade` (free-text) + chức danh;
    lương của NV KHÔNG đổi và KHÔNG sinh mốc lương mới (lương theo NV, không theo bậc/phòng)."""
    token = _admin_token(client)
    hcns = _dept_id("Hành chính nhân sự")
    kd = _dept_id("Kinh doanh")

    employee = _create(
        client, token, full_name="NV giữ nguyên lương khi đổi bậc", department_id=hcns,
        hire_date="2026-01-01", job_grade_id=_bac(client, token, "bac_1"),
        initial_salary={"effective_from": "2026-01-01", "luong_vi_tri": 11_000_000,
                        "luong_trach_nhiem": 2_000_000, "allowance": 700_000},
    ).json()["employee"]

    transferred = client.post(
        f"/api/employees/{employee['id']}/transitions",
        json={"kind": "transfer", "new_department_id": kd,
              "new_job_grade_id": _bac(client, token, "bac_2"),
              "effective_date": "2026-04-01"},
        headers=_h(token),
    )
    assert transferred.status_code == 200, transferred.text
    assert transferred.json()["department_id"] == kd
    assert transferred.json()["job_grade_name"] == "Thợ vững"

    promoted = client.post(
        f"/api/employees/{employee['id']}/transitions",
        json={"kind": "promote", "new_job_grade_id": _bac(client, token, "bac_3"),
              "new_position": "Tổ phó", "effective_date": "2026-07-01"},
        headers=_h(token),
    )
    assert promoted.status_code == 200, promoted.text
    assert promoted.json()["job_grade_name"] == "Thợ thường"

    # Lương KHÔNG đổi, KHÔNG sinh mốc lương mới (vẫn đúng 1 bản ghi ban đầu).
    history = client.get(
        f"/api/luong/salaries/{employee['id']}", headers=_h(token)
    ).json()["items"]
    assert len(history) == 1
    assert history[0]["luong_vi_tri"] == 11_000_000
    assert history[0]["luong_trach_nhiem"] == 2_000_000
    assert history[0]["allowance"] == 700_000


# --- account link -----------------------------------------------------------


def test_create_account_in_wizard_then_login(client):
    token = _admin_token(client)
    resp = _create(
        client, token, full_name="Có Tài Khoản",
        account={"username": "nv_login", "password": "secret1"},
    )
    assert resp.status_code == 201
    payload = resp.json()
    assert payload["account_username"] == "nv_login"

    # the new account can authenticate
    login = client.post("/api/auth/login", json={"username": "nv_login", "password": "secret1"})
    assert login.status_code == 200


def test_create_account_for_existing_employee(client):
    """NV tạo trước, cấp tài khoản sau — đường chính vì không còn tài khoản mồ côi để liên kết."""
    token = _admin_token(client)
    eid = _create(client, token, full_name="Cấp Sau").json()["employee"]["id"]
    resp = client.post(
        f"/api/employees/{eid}/account",
        json={"username": "nv_capsau", "password": "secret1"},
        headers=_h(token),
    )
    assert resp.status_code == 200
    assert resp.json()["account_username"] == "nv_capsau"
    login = client.post(
        "/api/auth/login", json={"username": "nv_capsau", "password": "secret1"}
    )
    assert login.status_code == 200


def test_attach_account_needs_username_or_user_id(client):
    token = _admin_token(client)
    eid = _create(client, token, full_name="Thiếu Tham Số").json()["employee"]["id"]
    resp = client.post(f"/api/employees/{eid}/account", json={}, headers=_h(token))
    assert resp.status_code == 400


def test_unlink_account_endpoint_is_gone(client):
    """Mọi tài khoản phải thuộc một hồ sơ → không còn cửa 'gỡ liên kết' (đẻ tài khoản mồ côi)."""
    token = _admin_token(client)
    payload = _create(
        client, token, full_name="Không Gỡ Được",
        account={"username": "nv_nogo", "password": "secret1"},
    ).json()
    eid = payload["employee"]["id"]
    resp = client.delete(f"/api/employees/{eid}/account", headers=_h(token))
    assert resp.status_code == 405  # method gone; chặn người = KHÓA tài khoản


def test_create_orphan_user_endpoint_is_gone(client):
    """Đường tạo tài khoản duy nhất là qua hồ sơ NV → POST /api/users không còn."""
    token = _admin_token(client)
    resp = client.post(
        "/api/users",
        json={"name": "Mồ Côi", "username": "mocoi", "department_id": _dept_id("Kinh doanh")},
        headers=_h(token),
    )
    assert resp.status_code == 405


def test_resigned_employee_cannot_login_and_reinstate_restores(client):
    """Trạng thái hồ sơ 'Đã nghỉ' ⇒ không đăng nhập được. Đổi trạng thái lại ⇒ vào được."""
    token = _admin_token(client)
    payload = _create(
        client, token, full_name="Sắp Nghỉ",
        account={"username": "nv_nghi", "password": "secret1"},
    ).json()
    eid = payload["employee"]["id"]
    creds = {"username": "nv_nghi", "password": "secret1"}
    assert client.post("/api/auth/login", json=creds).status_code == 200

    # nghỉ việc → login đóng cửa (không ai phải đi khóa tay)
    resp = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "resign", "effective_date": "2026-07-31", "resign_reason": "Chuyển việc"},
        headers=_h(token),
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "resigned"
    assert client.post("/api/auth/login", json=creds).status_code == 401

    # tuyển lại → tự mở, không cần nhớ mở khóa
    back = client.post(
        f"/api/employees/{eid}/transitions",
        json={"kind": "reinstate", "effective_date": "2026-09-01"},
        headers=_h(token),
    )
    assert back.status_code == 200
    assert client.post("/api/auth/login", json=creds).status_code == 200


def test_meta_returns_roles_for_account_form(client):
    token = _admin_token(client)
    meta = client.get("/api/employees/meta", headers=_h(token)).json()
    assert meta["roles"], "meta phải trả vai trò để form tài khoản chọn"
    kd = _dept_id("Kinh doanh")
    assert any(r["name"] == "NV Sales" and r["department_id"] == kd for r in meta["roles"])


# --- attachments ------------------------------------------------------------


def test_attachment_upload_list_delete(client):
    token = _admin_token(client)
    emp = _create(client, token).json()["employee"]
    eid = emp["id"]

    up = client.post(
        f"/api/employees/{eid}/attachments",
        files={"file": ("cccd.pdf", b"%PDF-1.4 fake", "application/pdf")},
        data={"doc_kind": "cccd"},
        headers=_h(token),
    )
    assert up.status_code == 201
    att = up.json()
    assert att["doc_kind"] == "cccd" and att["file_name"] == "cccd.pdf"
    assert att["file_url"].startswith(f"/api/files/hr/{eid}/")

    listed = client.get(f"/api/employees/{eid}/attachments", headers=_h(token)).json()["items"]
    assert len(listed) == 1

    assert client.delete(f"/api/employees/{eid}/attachments/{att['id']}", headers=_h(token)).status_code == 204
    assert client.get(f"/api/employees/{eid}/attachments", headers=_h(token)).json()["items"] == []


# --- RBAC -------------------------------------------------------------------


def test_forbidden_without_permission(client):
    token = _sales_token()
    assert client.get("/api/employees", headers=_h(token)).status_code == 403
    assert _create(client, token).status_code == 403


# --- N5: tách quyền SỬA lương/BHXH khỏi quyền sửa hồ sơ ---------------------


def _ns_no_salary_token() -> str:
    """User có quyền nhan_su read/create/update nhưng KHÔNG view_salary/edit_salary (N5)."""
    db = SessionLocal()
    try:
        users = UserRepository(db)
        existing = users.get_by_username("ns-nosalary")
        if existing is not None:
            return create_access_token(str(existing.id))
        hcns = DepartmentRepository(db).get_by_name("Hành chính nhân sự")
        roles = RoleRepository(db)
        role = roles.get_by_name_and_department("NS Không Lương", hcns.id)
        if role is None:
            role = roles.create(name="NS Không Lương", department_id=hcns.id)
            roles.set_permission(
                role_id=role.id, module_key="nhan_su", scope="all",
                can_read=True, can_create=True, can_update=True,
            )
        u = users.create(username="ns-nosalary", name="NS", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=hcns.id, role_id=role.id, is_active=True)
        return create_access_token(str(u.id))
    finally:
        db.close()


def test_edit_salary_gate_blocks_sensitive_write(client):
    """Người có nhan_su:update nhưng KHÔNG edit_salary → field lương/BHXH bị BỎ QUA khi
    tạo/sửa (field thường vẫn ghi). Admin (có view_salary) đọc lại để kiểm."""
    admin = _admin_token(client)
    tok = _ns_no_salary_token()

    # Tạo NV kèm field nhạy cảm → bị bỏ (không lưu lén)
    created = _create(client, tok, full_name="NV No Salary",
                      bank_account="123456", payroll_group="van_phong", phone="0900")
    assert created.status_code == 201
    eid = created.json()["employee"]["id"]

    seen = client.get(f"/api/employees/{eid}", headers=_h(admin)).json()
    assert seen["phone"] == "0900"          # field thường: ghi được
    assert seen["bank_account"] is None     # nhạy cảm: bị bỏ khi tạo
    assert seen["payroll_group"] is None

    # Admin đặt số TK nền (full_name bắt buộc ở EmployeeUpdate)
    r_admin = client.put(f"/api/employees/{eid}",
                         json={"full_name": "NV No Salary", "bank_account": "111",
                               "probation_end_date": "2025-12-31"}, headers=_h(admin))
    assert r_admin.status_code == 200
    # user không-quyền-lương PUT đổi bank + phone → bank bị bỏ, phone đổi
    upd = client.put(f"/api/employees/{eid}",
                     json={"full_name": "NV No Salary", "bank_account": "999", "phone": "0911",
                           "probation_end_date": "2025-12-31"}, headers=_h(tok))
    assert upd.status_code == 200
    seen2 = client.get(f"/api/employees/{eid}", headers=_h(admin)).json()
    assert seen2["phone"] == "0911"         # field thường: ghi được
    assert seen2["bank_account"] == "111"   # nhạy cảm: KHÔNG đổi


# --- Đ1: hồ sơ là GỐC — đồng bộ danh tính xuống tài khoản gắn kèm ------------


def test_employee_edit_syncs_name_and_department_to_user(client):
    """Đ1/Đ2: sửa tên hồ sơ + điều chuyển phòng → tài khoản gắn kèm đồng bộ tên + phòng
    (phòng = trục data-scope RBAC). Đồng bộ 1 chiều hồ-sơ→tài-khoản."""
    admin = _admin_token(client)
    hcns = _dept_id("Hành chính nhân sự")
    kd = _dept_id("Kinh doanh")
    created = _create(client, admin, full_name="Tên Cũ", department_id=hcns,
                      account={"username": "synced", "password": "synced123"})
    eid = created.json()["employee"]["id"]

    client.put(f"/api/employees/{eid}", json={"full_name": "Tên Mới", "probation_end_date": "2025-12-31"},
               headers=_h(admin))
    client.post(f"/api/employees/{eid}/transitions",
                json={"kind": "transfer", "new_department_id": kd}, headers=_h(admin))

    db = SessionLocal()
    try:
        u = UserRepository(db).get_by_username("synced")
        assert u.name == "Tên Mới"        # tên đồng bộ
        assert u.department_id == kd       # phòng (scope) đồng bộ theo điều chuyển
    finally:
        db.close()


def test_user_with_profile_cannot_self_rename(client):
    """Đ1: tài khoản CÓ hồ sơ → không tự đổi tên hiển thị qua /api/users/me (tên do hồ sơ
    quyết, HCNS cập nhật). Chặn ở backend, không chỉ ẩn ở FE."""
    admin = _admin_token(client)
    _create(client, admin, full_name="Khoa Nguyen",
            account={"username": "hasprofile", "password": "hasprofile1"})
    tok = client.post("/api/auth/login",
                      json={"username": "hasprofile", "password": "hasprofile1"}).json()["access_token"]
    r = client.patch("/api/users/me", json={"name": "Tên Tự Đặt"}, headers=_h(tok))
    assert r.status_code == 400


def test_danh_sach_nhan_vien_chan_size_qua_200(client):
    """Ghim ràng buộc mà FE ĐANG DỰA VÀO: `GET /api/employees` chặn `size > 200`.

    Modal "Gán cho nhân viên" từng gửi `size=500` ⇒ 422, và vì gọi trong `Promise.all` nên danh
    sách treo mãi ở "Đang tải danh sách nhân viên…". Nay FE phân trang theo lô 200
    (`EMP_PAGE` trong `CauHinhLuongTab.tsx`) — hạ trần xuống dưới 200 là làm hỏng chỗ đó."""
    token = _admin_token(client)
    assert client.get("/api/employees?page=1&size=500", headers=_h(token)).status_code == 422
    assert client.get("/api/employees?page=1&size=200", headers=_h(token)).status_code == 200


# --- HẾT THỬ VIỆC: MÁY TỰ ĐÁNH DẤU, HCNS MỚI XÁC NHẬN (chủ chốt 22/08/2026) -------------------
#
# Bốn chốt: ô Ngày hết thử việc là BẮT BUỘC · qua ngày thì máy đổi sang "Hết thử việc" chứ KHÔNG
# tự lên chính thức · tiền không đổi ở bước máy · hồ sơ hết hạn TRƯỚC ngày bật thì máy không đụng.


def _svc_nhan_su(db):
    from app.repositories.audit_repo import AuditLogRepository
    from app.repositories.employee_repo import EmployeeRepository
    from app.repositories.rbac_repo import DepartmentRepository
    from app.repositories.user_repo import UserRepository
    from app.services.employee_service import EmployeeService
    return EmployeeService(EmployeeRepository(db), AuditLogRepository(db),
                           UserRepository(db), DepartmentRepository(db))


def test_thu_viec_bat_buoc_khai_ngay_het_thu_viec(client):
    """Không có ngày thì máy không có mốc nào để so ⇒ chặn ngay từ lúc tạo."""
    token = _admin_token(client)
    r = _create(client, token, full_name="NV Thiếu Ngày", probation_end_date=None)
    assert r.status_code == 400, r.text
    assert "Ngày hết thử việc" in r.json()["detail"]

    # Khai rồi thì tạo được.
    ok = _create(client, token, full_name="NV Đủ Ngày", probation_end_date="2026-10-31")
    assert ok.status_code == 201, ok.text


def test_qua_ngay_het_thu_viec_may_doi_sang_HET_THU_VIEC_chu_khong_len_chinh_thuc(client):
    from datetime import date, timedelta
    from app.db import SessionLocal

    token = _admin_token(client)
    hom_qua = date.today() - timedelta(days=1)
    eid = _create(client, token, full_name="NV Hết Hạn Hôm Qua",
                  probation_end_date=hom_qua.isoformat()).json()["employee"]["id"]

    db = SessionLocal()
    try:
        svc = _svc_nhan_su(db)
        assert svc.tu_danh_dau_het_thu_viec(moc=hom_qua) == 1
    finally:
        db.close()

    emp = client.get(f"/api/employees/{eid}", headers=_h(token)).json()
    # KHÔNG phải "active" — đây là điểm chính của cả tính năng.
    assert emp["status"] == "probation_ended"

    ev = client.get(f"/api/employees/{eid}/events", headers=_h(token)).json()
    moc_moi = [e for e in ev["items"] if e["event_type"] == "probation_ended"]
    assert len(moc_moi) == 1, ev
    # Ngày hiệu lực = ngày đầu tiên KHÔNG còn thử việc.
    assert moc_moi[0]["effective_date"] == (hom_qua + timedelta(days=1)).isoformat()

    # Chạy lại không đẻ thêm mốc (hồ sơ hết là `probation` nên rơi khỏi lượt quét sau).
    db = SessionLocal()
    try:
        assert _svc_nhan_su(db).tu_danh_dau_het_thu_viec(moc=hom_qua) == 0
    finally:
        db.close()


def test_may_KHONG_dung_ho_so_het_han_truoc_ngay_bat_tinh_nang(client):
    """Dữ liệu tồn: hết hạn từ thời chưa có tính năng ⇒ để HCNS tự xử, máy không đụng."""
    from datetime import date, timedelta
    from app.db import SessionLocal

    token = _admin_token(client)
    hom_qua = date.today() - timedelta(days=1)
    eid = _create(client, token, full_name="NV Tồn Cũ",
                  probation_end_date=hom_qua.isoformat()).json()["employee"]["id"]

    db = SessionLocal()
    try:
        # Mốc đặt SAU ngày hết hạn ⇒ hồ sơ này là "người cũ".
        assert _svc_nhan_su(db).tu_danh_dau_het_thu_viec(moc=date.today()) == 0
    finally:
        db.close()
    assert client.get(f"/api/employees/{eid}", headers=_h(token)).json()["status"] == "probation"


def test_chua_toi_han_thi_van_la_thu_viec(client):
    from datetime import date, timedelta
    from app.db import SessionLocal

    token = _admin_token(client)
    mai = date.today() + timedelta(days=1)
    eid = _create(client, token, full_name="NV Còn Hạn",
                  probation_end_date=mai.isoformat()).json()["employee"]["id"]
    db = SessionLocal()
    try:
        assert _svc_nhan_su(db).tu_danh_dau_het_thu_viec(moc=date(2020, 1, 1)) == 0
    finally:
        db.close()
    assert client.get(f"/api/employees/{eid}", headers=_h(token)).json()["status"] == "probation"


def test_HCNS_bam_chuyen_chinh_thuc_tu_trang_thai_het_thu_viec(client):
    """Đường lên chính thức vẫn phải có người bấm — và nó phải mở từ trạng thái mới."""
    from datetime import date, timedelta
    from app.db import SessionLocal

    token = _admin_token(client)
    hom_qua = date.today() - timedelta(days=1)
    eid = _create(client, token, full_name="NV Chờ Xác Nhận",
                  probation_end_date=hom_qua.isoformat()).json()["employee"]["id"]
    db = SessionLocal()
    try:
        _svc_nhan_su(db).tu_danh_dau_het_thu_viec(moc=hom_qua)
    finally:
        db.close()

    r = client.post(f"/api/employees/{eid}/transitions",
                    json={"kind": "confirm", "effective_date": date.today().isoformat()},
                    headers=_h(token))
    assert r.status_code == 200, r.text
    assert client.get(f"/api/employees/{eid}", headers=_h(token)).json()["status"] == "active"
