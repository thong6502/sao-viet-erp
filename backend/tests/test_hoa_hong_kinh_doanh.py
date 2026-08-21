"""Hoa hồng kinh doanh — nối `commission_pct` vào tiền lương (§4.6 + hai quyết định 21/08/2026).

Trước đợt này ô `%` đã khai được từ mg 0128 nhưng **engine không đọc** — chú thích trong model ghi
thẳng *"khai bao nhiêu cũng không đổi một đồng"*. Bộ test này khoá đúng chỗ đó.

Hai quyết định của chủ dự án khác spec, và cả hai đều đổi SỐ TIỀN nên phải có test riêng:
  · MỐC sinh hoa hồng = lúc RA CÔNG NỢ (hoá đơn issued), không phải lúc thu tiền;
  · GỐC tính = TRƯỚC VAT, quy đổi theo tỷ lệ.
"""
from __future__ import annotations

from datetime import date

from app.db import SessionLocal
from app.models.accounting import (
    SALES_INVOICE_CANCELLED,
    SALES_INVOICE_ISSUED,
    SalesInvoice,
)
from app.models.customer import Customer
from app.models.employee import STATUS_ACTIVE, Employee
from app.models.order import STATUS_ORDERED, Order, OrderLine
from app.models.payroll import EmployeeSalary
from app.models.user import User
from app.security import hash_password
from app.services.hoa_hong_service import HoaHongService

KY_TU, KY_DEN = date(2026, 8, 1), date(2026, 8, 31)


def _sales(ten: str, *, pct: float = 0.05) -> tuple[int, int]:
    """NV kinh doanh có tài khoản + mức lương khai `commission_pct`. Trả `(employee_id, user_id)`."""
    db = SessionLocal()
    try:
        u = User(username=f"sale-{ten}", name=ten, password_hash=hash_password("x"))
        db.add(u)
        db.flush()
        e = Employee(code=f"NVKD{abs(hash(ten)) % 9000 + 1000}", full_name=ten,
                     hire_date=date(2020, 1, 1), user_id=u.id, status=STATUS_ACTIVE)
        db.add(e)
        db.flush()
        db.add(EmployeeSalary(employee_id=e.id, effective_from=date(2020, 1, 1),
                              base_amount=10_000_000, commission_pct=pct))
        db.commit()
        return e.id, u.id
    finally:
        db.close()


def _don(ma: str, user_id: int, *, truoc_vat: int, vat_pct: int = 8,
         pct: float | None = 0.05) -> int:
    """Đơn ĐÃ CHỐT của người sales đó. `pct=None` ⇒ đơn không khai hoa hồng."""
    db = SessionLocal()
    try:
        kh = Customer(code=f"KH-{ma}", name=f"Khach {ma}")
        db.add(kh)
        db.flush()
        o = Order(order_no=ma, customer_id=kh.id, status=STATUS_ORDERED,
                  sale_user_id=user_id, commission_pct=pct or 0)
        o.lines.append(OrderLine(description="Hop giay", qty=1, don_vi_tinh="cai",
                                 line_total=truoc_vat, vat_pct_estimate=vat_pct))
        db.add(o)
        db.commit()
        return o.id
    finally:
        db.close()


def _hoa_don(order_id: int, so_tien_co_vat: int, *, ngay: date = date(2026, 8, 15),
             huy: bool = False) -> None:
    db = SessionLocal()
    try:
        o = db.get(Order, order_id)
        db.add(SalesInvoice(
            order_id=order_id, customer_id=o.customer_id,
            invoice_symbol="1C26TAA", invoice_number=f"{order_id}{int(ngay.day):02d}",
            invoice_date=ngay, amount_vnd=so_tien_co_vat,
            customer_name_snapshot="Khach test",
            status=SALES_INVOICE_CANCELLED if huy else SALES_INVOICE_ISSUED,
        ))
        db.commit()
    finally:
        db.close()


def _tinh(employee_id: int, *, tu=KY_TU, den=KY_DEN) -> float:
    db = SessionLocal()
    try:
        return HoaHongService(db).hoa_hong_ky(employee_id, tu_ngay=tu, den_ngay=den)
    finally:
        db.close()


# =============================================================================================
# Công thức
# =============================================================================================
def test_RA_CONG_NO_la_co_hoa_hong(client):
    """⭐ Mốc sinh hoa hồng = hoá đơn bán được ghi nhận (chủ chốt 21/08/2026).

    Spec §4.6 chọn mốc *thu được tiền*; chủ đổi sang mốc công nợ. Đơn 100tr trước VAT + VAT 8%
    ⇒ hoá đơn 108tr; hoa hồng 5% tính trên phần TRƯỚC VAT = 100tr × 5% = 5tr.
    """
    emp, uid = _sales("ra cong no")
    oid = _don("DH-HH-01", uid, truoc_vat=100_000_000, vat_pct=8)
    _hoa_don(oid, 108_000_000)
    assert _tinh(emp) == 5_000_000


def test_GOC_TINH_la_TRUOC_VAT_khong_phai_tien_hoa_don(client):
    """⭐ VAT là tiền thu hộ nhà nước — trả hoa hồng trên đó là trả trên tiền không phải của mình.

    Nếu lấy thẳng tiền hoá đơn thì ra 108tr × 5% = 5.4tr, thừa 400k cho mỗi 100tr doanh thu.
    """
    emp, uid = _sales("truoc vat")
    oid = _don("DH-HH-02", uid, truoc_vat=100_000_000, vat_pct=8)
    _hoa_don(oid, 108_000_000)
    assert _tinh(emp) == 5_000_000
    assert _tinh(emp) != 5_400_000


def test_don_KHONG_VAT_thi_khong_quy_doi_gi(client):
    """Tỷ lệ quy đổi phải là 1.0, không được chia cho 0 hay cắt mất tiền."""
    emp, uid = _sales("khong vat")
    oid = _don("DH-HH-03", uid, truoc_vat=50_000_000, vat_pct=0)
    _hoa_don(oid, 50_000_000)
    assert _tinh(emp) == 2_500_000


def test_HOA_DON_TUNG_PHAN_chi_an_hoa_hong_phan_do(client):
    """⭐ Một đơn xuất nhiều hoá đơn ⇒ mỗi hoá đơn chỉ mở phần hoa hồng của nó.

    Đây là lý do quy đổi theo TỶ LỆ chứ không trừ thẳng VAT: hoá đơn chỉ là một mẩu của đơn.
    """
    emp, uid = _sales("tung phan")
    oid = _don("DH-HH-04", uid, truoc_vat=100_000_000, vat_pct=8)
    _hoa_don(oid, 54_000_000, ngay=date(2026, 8, 10))    # nửa đơn
    assert _tinh(emp) == 2_500_000
    _hoa_don(oid, 54_000_000, ngay=date(2026, 8, 20))    # nửa còn lại
    assert _tinh(emp) == 5_000_000


# =============================================================================================
# Các trường KHÔNG được tính
# =============================================================================================
def test_hoa_don_HUY_khong_tinh(client):
    emp, uid = _sales("huy")
    oid = _don("DH-HH-05", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000, huy=True)
    assert _tinh(emp) == 0


def test_hoa_don_NGOAI_KY_khong_tinh(client):
    """Hoá đơn tháng 7 không rơi vào kỳ tháng 8, và ngược lại."""
    emp, uid = _sales("ngoai ky")
    oid = _don("DH-HH-06", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000, ngay=date(2026, 7, 31))
    assert _tinh(emp) == 0
    assert _tinh(emp, tu=date(2026, 7, 1), den=date(2026, 7, 31)) == 5_000_000


def test_don_cua_NGUOI_KHAC_khong_tinh(client):
    """⭐ Hoa hồng bám `orders.sale_user_id`. Sai chỗ này là trả tiền cho nhầm người."""
    a, uid_a = _sales("nguoi A")
    b, _uid_b = _sales("nguoi B")
    oid = _don("DH-HH-07", uid_a, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    assert _tinh(a) == 5_000_000
    assert _tinh(b) == 0


def test_don_KHONG_KHAI_hoa_hong_thi_bang_0(client):
    """`orders.commission_pct = 0` ⇒ không có hoa hồng, dù người sales có mức % riêng."""
    emp, uid = _sales("don khong khai")
    oid = _don("DH-HH-08", uid, truoc_vat=100_000_000, pct=None)
    _hoa_don(oid, 108_000_000)
    assert _tinh(emp) == 0


def test_pct_AM_khong_duoc_TRU_vao_hoa_hong_don_khac(client):
    """% âm (gõ nhầm dấu) chỉ được BỎ QUA đơn đó, không được ăn lẹm sang đơn lành.

    Không có hàng rào `pct <= 0` thì đơn âm trừ thẳng vào tổng: người ta làm 2 đơn mà nhận ít hơn
    làm 1 đơn, và trên phiếu lương không có gì để lần ra vì sao.
    """
    emp, uid = _sales("pct am")
    lanh = _don("DH-HH-11", uid, truoc_vat=100_000_000)
    _hoa_don(lanh, 108_000_000, ngay=date(2026, 8, 11))
    am = _don("DH-HH-12", uid, truoc_vat=100_000_000, pct=-0.05)
    _hoa_don(am, 108_000_000, ngay=date(2026, 8, 12))
    assert _tinh(emp) == 5_000_000, "đơn % âm đã ăn lẹm vào hoa hồng của đơn lành"


def test_nhan_vien_CHUA_CO_TAI_KHOAN_thi_bang_0(client):
    """Đơn ghi USER, bảng lương chạy theo EMPLOYEE — không có tài khoản thì không thể là sales."""
    db = SessionLocal()
    try:
        e = Employee(code="NVKD-NOUSER", full_name="Khong co tai khoan",
                     hire_date=date(2020, 1, 1))
        db.add(e)
        db.commit()
        eid = e.id
    finally:
        db.close()
    assert _tinh(eid) == 0


def test_don_khong_co_dong_hang_khong_lam_vo(client):
    """Đơn chưa có dòng nào ⇒ tỷ lệ quy đổi lùi về 1.0, không chia cho 0."""
    db = SessionLocal()
    try:
        u = User(username="sale-rong", name="Sale rong", password_hash=hash_password("x"))
        db.add(u)
        db.flush()
        e = Employee(code="NVKD-RONG", full_name="Sale rong", hire_date=date(2020, 1, 1),
                     user_id=u.id)
        db.add(e)
        db.flush()
        kh = Customer(code="KH-RONG", name="Khach rong")
        db.add(kh)
        db.flush()
        o = Order(order_no="DH-HH-09", customer_id=kh.id, status=STATUS_ORDERED,
                  sale_user_id=u.id, commission_pct=0.05)
        db.add(o)
        db.commit()
        eid, oid = e.id, o.id
    finally:
        db.close()
    _hoa_don(oid, 10_000_000)
    assert _tinh(eid) == 500_000


# =============================================================================================
# Chụp ảnh % lúc chốt đơn
# =============================================================================================
def test_CHOT_DON_chup_pct_cua_sale_vao_don(client):
    """⭐ Chụp lúc chốt, không đọc-sống: đổi % tháng sau không được sửa ngược đơn cũ."""
    from app.services.order_service import _pct_hoa_hong_cua_sale

    _emp, uid = _sales("chup anh", pct=0.07)
    db = SessionLocal()
    try:
        assert _pct_hoa_hong_cua_sale(db, uid) == 0.07
        assert _pct_hoa_hong_cua_sale(db, None) == 0.0
        assert _pct_hoa_hong_cua_sale(db, 999999) == 0.0
    finally:
        db.close()


def test_doi_pct_cua_NGUOI_khong_lam_doi_don_DA_CHOT(client):
    """Đơn đã chốt giữ số cũ — tiền đã hứa thì không sửa ngược."""
    emp, uid = _sales("doi pct", pct=0.05)
    oid = _don("DH-HH-10", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    assert _tinh(emp) == 5_000_000

    db = SessionLocal()
    try:
        db.add(EmployeeSalary(employee_id=emp, effective_from=date(2026, 9, 1),
                              base_amount=10_000_000, commission_pct=0.20))
        db.commit()
    finally:
        db.close()
    assert _tinh(emp) == 5_000_000, "đổi % của người làm đổi luôn hoa hồng đơn đã chốt"


# =============================================================================================
# Nối vào BẢNG LƯƠNG — đây mới là chỗ tiền tới tay người ta
# =============================================================================================
ADMIN = {"username": "admin", "password": "admin123"}


def _h(client) -> dict[str, str]:
    tok = client.post("/api/auth/login", json=ADMIN).json()["access_token"]
    return {"Authorization": f"Bearer {tok}"}


def _tinh_luong(client, eid: int, *, year=2026, month=8) -> dict:
    """Bấm "Tính lại" cả kỳ rồi lấy dòng của NV này."""
    r = client.post("/api/luong/generate", json={"year": year, "month": month}, headers=_h(client))
    assert r.status_code in (200, 201), r.text
    return next(l for l in r.json()["lines"] if l["employee_id"] == eid)


def _dong_hh(line: dict) -> list[dict]:
    return [c for c in line["components"] if c["code"] == "hoa_hong_kd"]


def test_hoa_hong_LEN_PHIEU_LUONG_va_cong_vao_gross(client):
    """⭐ Cả tính năng chỉ có nghĩa ở đây: % khai xong phải RA TIỀN trên phiếu lương.

    Trước đợt này ô `%` khai được từ mg 0128 mà engine không đọc — khai bao nhiêu cũng bằng 0.
    """
    emp, uid = _sales("len phieu")
    truoc = _tinh_luong(client, emp)                  # chưa có hoá đơn nào
    assert _dong_hh(truoc) == []

    oid = _don("DH-LUONG-01", uid, truoc_vat=100_000_000, vat_pct=8)
    _hoa_don(oid, 108_000_000)
    sau = _tinh_luong(client, emp)

    assert len(_dong_hh(sau)) == 1
    assert _dong_hh(sau)[0]["amount"] == 5_000_000
    assert sau["gross"] == truoc["gross"] + 5_000_000, (
        f"hoa hồng có dòng nhưng gross đi từ {truoc['gross']} sang {sau['gross']}")


def test_KHONG_de_dong_0_dong_cho_nguoi_khong_lam_kinh_doanh(client):
    """Cả trăm người không làm kinh doanh — thêm dòng "Hoa hồng 0đ" cho từng người là rác phiếu."""
    emp, _uid = _sales("khong kinh doanh")
    assert _dong_hh(_tinh_luong(client, emp)) == []


def test_TINH_LAI_khong_cong_doi(client):
    """⭐ Bẫy chết người của khoản hệ tự tính: mỗi lần bấm "Tính lại" lại đẻ thêm một dòng."""
    emp, uid = _sales("tinh lai")
    oid = _don("DH-LUONG-02", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)

    l1 = _tinh_luong(client, emp)
    l2 = _tinh_luong(client, emp)
    l3 = _tinh_luong(client, emp)
    assert len(_dong_hh(l3)) == 1, "tính lại 3 lần ra 3 dòng hoa hồng"
    assert l3["gross"] == l1["gross"] == l2["gross"]


def test_HOA_DON_MOI_thi_so_moi_THAY_so_cu(client):
    """Xuất thêm hoá đơn giữa kỳ ⇒ tính lại phải ra số MỚI: không giữ số cũ, cũng không cộng dồn."""
    emp, uid = _sales("hoa don moi")
    oid = _don("DH-LUONG-03", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 54_000_000, ngay=date(2026, 8, 10))
    assert _dong_hh(_tinh_luong(client, emp))[0]["amount"] == 2_500_000

    _hoa_don(oid, 54_000_000, ngay=date(2026, 8, 20))
    dong = _dong_hh(_tinh_luong(client, emp))
    assert len(dong) == 1
    assert dong[0]["amount"] == 5_000_000, "phải là 5tr (số mới), không phải 2,5tr hay 7,5tr"


def test_tinh_lai_hoa_hong_KHONG_XOA_thuong_nong_them_tay(client):
    """⭐ Hoa hồng xoá-rồi-ghi-lại mỗi lần tính. Xoá lố sang nguồn `line` là mất thưởng nóng HCNS
    đã nhập — mất tiền của người lao động mà không một thông báo nào."""
    emp, uid = _sales("giu thuong nong")
    oid = _don("DH-LUONG-04", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    line = _tinh_luong(client, emp)

    comps = client.get("/api/luong/components", headers=_h(client)).json()["items"]
    khac = next(c for c in comps if c["code"] == "thu_nhap_khac_ct")
    r = client.post(f"/api/luong/lines/{line['id']}/components",
                    json={"component_id": khac["id"], "amount": 300_000}, headers=_h(client))
    assert r.status_code == 201, r.text

    sau = _tinh_luong(client, emp)                     # bấm "Tính lại"
    ma = {c["code"]: c["amount"] for c in sau["components"]}
    assert ma.get("thu_nhap_khac_ct") == 300_000, "tính lại hoa hồng đã nuốt mất thưởng nóng"
    assert ma.get("hoa_hong_kd") == 5_000_000


def test_hoa_hong_CHIU_THUE_TNCN(client):
    """Hoa hồng là thu nhập chịu thuế — cờ lấy từ DANH MỤC chứ không đóng đinh trong engine."""
    emp, uid = _sales("chiu thue")
    truoc = _tinh_luong(client, emp)
    oid = _don("DH-LUONG-05", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    sau = _tinh_luong(client, emp)

    assert _dong_hh(sau)[0]["is_taxable"] is True
    assert sau["thu_nhap_chiu_thue"] == truoc["thu_nhap_chiu_thue"] + 5_000_000


def test_khoan_hoa_hong_la_nguon_AUTO_khong_phai_tay_go(client):
    """Nguồn `auto` là thứ phân biệt "hệ tự tính" với "HCNS gõ" — giao diện dựa vào đó để KHOÁ ô."""
    emp, uid = _sales("nguon auto")
    oid = _don("DH-LUONG-06", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    assert _dong_hh(_tinh_luong(client, emp))[0]["source"] == "auto"


def test_KHONG_CHO_GO_TAY_dong_hoa_hong(client):
    """⭐ Sửa/gỡ tay dòng hoa hồng phải bị CHẶN, kèm câu chỉ đúng chỗ sửa.

    Không chặn thì: sửa số → "Tính lại" xoá sạch âm thầm; gỡ dòng → tính lại là mọc lại. Cả hai
    đều để HCNS tưởng mình vừa làm được việc gì đó.
    """
    emp, uid = _sales("go tay")
    oid = _don("DH-LUONG-07", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    line = _tinh_luong(client, emp)
    row_id = _dong_hh(line)[0]["id"]

    r = client.put(f"/api/luong/lines/components/{row_id}", json={"amount": 99_000_000},
                   headers=_h(client))
    assert r.status_code == 400, r.text
    assert "tự tính" in r.json()["detail"]

    r = client.delete(f"/api/luong/lines/components/{row_id}", headers=_h(client))
    assert r.status_code == 400, r.text
    assert "Lương nhân viên" not in r.json()["detail"], "chỉ sai chỗ — hoa hồng không nằm ở đó"

    assert _dong_hh(_tinh_luong(client, emp))[0]["amount"] == 5_000_000


def test_FILE_XUAT_khong_duoc_NUOT_hoa_hong(client):
    """⭐ Hoa hồng nằm trong cột "Tổng" thì phải có MỘT cột giải thích được nó.

    Cột "Thưởng" của bảng + file xuất trước đây chỉ cộng `source='line'`, nên khoản `auto` lọt
    vào Tổng mà không cột nào nói ra — kế toán dò lệch mãi không ra, đúng chuyện đã xảy ra với
    "Cơm ca"/"Phụ cấp ca" ngày 03/08/2026.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    def _xuat(ten_nv: str) -> tuple[list, tuple]:
        r = client.get("/api/luong/export.xlsx?year=2026&month=8", headers=_h(client))
        assert r.status_code == 200, r.text
        ws = load_workbook(BytesIO(r.content)).active
        head = [c.value for c in ws[1]]
        return head, next(x for x in ws.iter_rows(min_row=2, values_only=True) if x[1] == ten_nv)

    emp, uid = _sales("file xuat")
    _tinh_luong(client, emp)
    head, truoc = _xuat("file xuat")

    oid = _don("DH-LUONG-08", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)
    _tinh_luong(client, emp)
    _, sau = _xuat("file xuat")

    i_thuong, i_tong = head.index("Thưởng"), head.index("Tổng")
    assert sau[i_tong] - truoc[i_tong] == 5_000_000, "Tổng không nhận hoa hồng"
    assert sau[i_thuong] - truoc[i_thuong] == 5_000_000, (
        "Tổng có thêm 5tr mà không cột khoản nào tăng — file xuất cộng lại không khớp")


def test_NV_DA_NGHI_VIEC_van_duoc_tra_hoa_hong_don_cu(client):
    """⭐ Nghỉ việc rồi hoá đơn mới ra ⇒ vẫn còn tiền phải trả, KHÔNG được quỵt.

    Vòng `generate` bỏ qua NV nghỉ việc "không có công/khoán/dòng lương trong kỳ". NV kinh doanh
    nghỉ tháng trước, tháng này kế toán mới xuất hoá đơn của đơn họ đã chốt — không có công, không
    có khoán, chưa có dòng lương nào ⇒ trước bản vá là rơi khỏi bảng lương, tiền bốc hơi không một
    dòng cảnh báo.
    """
    from app.models.employee import STATUS_RESIGNED

    emp, uid = _sales("da nghi viec")
    oid = _don("DH-LUONG-09", uid, truoc_vat=100_000_000)
    _hoa_don(oid, 108_000_000)

    db = SessionLocal()
    try:
        e = db.get(Employee, emp)
        e.status = STATUS_RESIGNED
        e.resign_date = date(2026, 7, 31)
        db.commit()
    finally:
        db.close()

    r = client.post("/api/luong/generate", json={"year": 2026, "month": 8}, headers=_h(client))
    assert r.status_code in (200, 201), r.text
    dong = next((l for l in r.json()["lines"] if l["employee_id"] == emp), None)
    assert dong is not None, "NV nghỉ việc còn hoa hồng mà không có dòng lương nào"
    assert _dong_hh(dong)[0]["amount"] == 5_000_000


def test_KHONG_CO_DUONG_NAO_de_sale_tu_dat_pct_cho_don(client):
    """⭐ Ai đặt % hoa hồng — chốt bằng SỐ, không bằng lời (chủ 21/08/2026).

    "% của người kinh doanh nó bán được đơn rồi thì nó theo chỗ mình setup trước đó chứ, làm sao
    nó có quyền chỉnh sửa % hoa hồng được." Cho sale gõ % trên chính đơn mình bán là để người ta
    TỰ VIẾT PHIẾU LƯƠNG CỦA MÌNH.

    Test này ĐỎ ngay khi ai đó phơi `commission_pct` ra API đơn hàng.
    """
    from app.schemas.order import OrderDetailOut, OrderUpdate

    assert "commission_pct" not in OrderUpdate.model_fields, (
        "đã mở đường ghi % hoa hồng qua API sửa đơn — sale tự đặt được hoa hồng của mình")
    assert "commission_pct" not in OrderDetailOut.model_fields, (
        "% hoa hồng của người khác lộ ra trên đơn — ai xem được đơn là biết mức hoa hồng")


def test_pct_lay_dung_muc_DA_SETUP_o_ho_so_luong(client):
    """% của đơn = mức khai ở HỒ SƠ LƯƠNG tại thời điểm chốt, không phải số ai đó gõ vào đơn."""
    from app.models.user import User as _U

    emp, uid = _sales("theo ho so", pct=0.07)
    db = SessionLocal()
    try:
        from app.services.order_service import _pct_hoa_hong_cua_sale
        assert _pct_hoa_hong_cua_sale(db, uid) == 0.07

        # Chưa khai gì ⇒ 0, tức đơn của người đó không có hoa hồng (không đoán, không lấy mặc định)
        u = _U(username="sale-chua-khai", name="Chua khai", password_hash=hash_password("x"))
        db.add(u)
        db.flush()
        db.add(Employee(code="NVKD-CHUAKHAI", full_name="Chua khai", hire_date=date(2020, 1, 1),
                        user_id=u.id, status=STATUS_ACTIVE))
        db.commit()
        assert _pct_hoa_hong_cua_sale(db, u.id) == 0.0
    finally:
        db.close()


def test_chup_pct_KHONG_duoc_lay_muc_CUA_TUONG_LAI(client):
    """⭐ Chốt đơn HÔM NAY phải ăn mức ĐANG hiệu lực hôm nay, không phải mức tháng sau.

    Nhân sự khai trước "từ 01/12 lên 20%". Đơn chốt hôm nay mà chụp nhầm 20% là công ty trả
    theo lời hứa CHƯA tới hạn — mỗi 100tr doanh thu chi dư 15tr.

    Cả engine lương tra mức bằng `effective_from <= ngày` (`latest_salaries_map`); hàm chụp này
    phải theo đúng nếp đó.
    """
    from app.services.order_service import _pct_hoa_hong_cua_sale

    emp, uid = _sales("muc tuong lai", pct=0.05)
    db = SessionLocal()
    try:
        db.add(EmployeeSalary(employee_id=emp, effective_from=date(2099, 12, 1),
                              base_amount=10_000_000, commission_pct=0.20))
        db.commit()
        assert _pct_hoa_hong_cua_sale(db, uid) == 0.05, (
            "chụp nhầm mức của tương lai — đơn chốt hôm nay ăn % chưa tới hạn")
    finally:
        db.close()
