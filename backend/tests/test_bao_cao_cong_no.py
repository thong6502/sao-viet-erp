"""Sổ TỔNG HỢP CÔNG NỢ theo kỳ (TK 131 / 331) — docs/prd-bao-cao-cong-no.md §5.1.

ĐỪNG nhầm với `test_payables_api.py` / `test_sales_invoices_api.py`: hai file kia canh màn Công
nợ — ảnh chụp TẠI HÔM NAY để đi đòi nợ. File này canh SỔ THEO KỲ: đầu kỳ · phát sinh · cuối kỳ,
khoảng ngày tự chọn, dùng để đối chiếu với MISA.

Bốn luật xương sống:

1. ⭐ **CÂN SỔ.** `đầu kỳ + PS Nợ − PS Có = cuối kỳ`, đúng cho TỪNG DÒNG và cho dòng TỔNG. Đây là
   test quan trọng nhất cả file — một sổ không cân thì mọi con số trên đó đều vô nghĩa, và kế
   toán sẽ phát hiện ra đúng lúc ngồi đối chiếu với NCC chứ không sớm hơn.
2. **Chứng từ rơi đúng ngăn theo NGÀY.** Trước kỳ → đầu kỳ; trong kỳ → phát sinh; sau kỳ → biến
   mất hẳn. Kéo `tu_ngay` lùi lại thì tiền chuyển từ cột đầu kỳ sang cột phát sinh, còn CUỐI KỲ
   không đổi — cuối kỳ là sự thật tại `den_ngay`, không phụ thuộc mình bắt đầu nhìn từ đâu.
3. **Số ÂM không bao giờ lọt ra.** Âm thì nhảy sang cột bên kia. Nhờ luật này mà "mình cọc trước
   NCC" hiện đúng ở cột Nợ của TK 331, và "khách ứng trước" hiện ở cột Có của TK 131.
4. **Đầu kỳ dựng THUẦN từ chứng từ.** Không có chỗ khai tay nợ cũ (màn nhập từ file MISA đã bỏ
   04/09/2026), nên nợ phát sinh trước khi dùng hệ này thì báo cáo không thấy.

Ngày lấy từ seam `_business_today` của service, KHÔNG cắm ngày cứng — cắm cứng là hẹn giờ cho
test tự đỏ vài tháng sau (bài học từ 5 test đỏ hôm 02/09/2026).
"""
from __future__ import annotations

from datetime import timedelta

from app.services.accounting_service import _business_today as _hom_nay
from tests.test_payables_api import (
    _da_mua,
    _dong_dau_tien,
    _don,
    _ghi_dot,
    _headers,
    _khai_coc,
    _phieu_chi,
    _supplier,
)
from tests.test_sales_invoices_api import _add_deposit, _invoice_payload, _sales_order


def _ngay(lui: int) -> str:
    """ISO của ngày cách hôm nay `lui` ngày về TRƯỚC."""
    return (_hom_nay() - timedelta(days=lui)).isoformat()


def _bao_cao(client, headers, duong: str, *, tu: str, den: str, expect: int = 200) -> dict:
    r = client.get(
        f"/api/accounting/reports/{duong}",
        params={"tu_ngay": tu, "den_ngay": den},
        headers=headers,
    )
    assert r.status_code == expect, r.text
    return r.json()


def _dong(bao_cao: dict, ten_chua: str) -> dict:
    """Lấy đúng một dòng theo mẩu tên. Nổ rõ ràng nếu không có / có nhiều hơn một."""
    khop = [d for d in bao_cao["items"] if ten_chua.lower() in (d["ten"] or "").lower()]
    assert len(khop) == 1, f"tìm {ten_chua!r}: {[d['ten'] for d in bao_cao['items']]}"
    return khop[0]


def _kiem_can_so(bao_cao: dict) -> None:
    """⭐ Bất biến sống còn: đầu kỳ + PS Nợ − PS Có = cuối kỳ, từng dòng VÀ dòng tổng.

    Cộng theo `net` (Nợ − Có) vì một dòng chỉ có tiền ở MỘT trong hai cột mỗi cụm.
    """
    for d in bao_cao["items"]:
        dau = d["dau_no"] - d["dau_co"]
        cuoi = d["cuoi_no"] - d["cuoi_co"]
        assert dau + d["ps_no"] - d["ps_co"] == cuoi, f"lệch cân sổ ở dòng {d['ten']!r}: {d}"
        # Không cột nào âm, và không dòng nào có tiền ở CẢ HAI cột cùng cụm.
        for k in ("dau_no", "dau_co", "ps_no", "ps_co", "cuoi_no", "cuoi_co"):
            assert d[k] >= 0, f"cột {k} âm ở dòng {d['ten']!r}"
        assert not (d["dau_no"] and d["dau_co"]), d
        assert not (d["cuoi_no"] and d["cuoi_co"]), d

    t = bao_cao["tong"]
    assert t["so_dong"] == len(bao_cao["items"])
    for k in ("dau_no", "dau_co", "ps_no", "ps_co", "cuoi_no", "cuoi_co"):
        assert t[k] == sum(d[k] for d in bao_cao["items"]), f"tổng cột {k} không khớp"


# ══ TK 331 — PHẢI TRẢ ═══════════════════════════════════════════════════════════════════════


def test_hang_ve_trong_ky_thi_vao_ps_co_va_can_so(client):
    """Hàng về = nợ NCC tăng = PS **Có** (ghi theo NGÀY HÀNG VỀ, chốt 03/09/2026)."""
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC So Ky A")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(10),
    )  # 400 × 2.200 = 880.000

    bc = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    d = _dong(bc, "NCC So Ky A")
    assert (d["dau_no"], d["dau_co"]) == (0, 0), "trước kỳ chưa có gì"
    assert (d["ps_no"], d["ps_co"]) == (0, 880_000)
    assert (d["cuoi_no"], d["cuoi_co"]) == (0, 880_000), "dư Có = còn nợ NCC"
    assert d["tk"] == "331"
    _kiem_can_so(bc)


def test_hang_ve_truoc_ky_thi_nam_o_dau_ky_khong_phai_phat_sinh(client):
    """Kéo `tu_ngay` tới sau ngày hàng về ⇒ tiền chuyển sang cột ĐẦU KỲ, cuối kỳ KHÔNG đổi."""
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC So Ky B")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(20),
    )

    rong = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    hep = _bao_cao(client, headers, "payables", tu=_ngay(10), den=_ngay(0))
    a, b = _dong(rong, "NCC So Ky B"), _dong(hep, "NCC So Ky B")

    assert (a["dau_co"], a["ps_co"]) == (0, 880_000), "kỳ rộng: nằm ở phát sinh"
    assert (b["dau_co"], b["ps_co"]) == (880_000, 0), "kỳ hẹp: nằm ở đầu kỳ"
    assert a["cuoi_co"] == b["cuoi_co"] == 880_000, "CUỐI KỲ không phụ thuộc chỗ bắt đầu nhìn"
    _kiem_can_so(rong)
    _kiem_can_so(hep)


def test_chung_tu_sau_ky_khong_xuat_hien(client):
    """Hàng về hôm nay mà báo cáo kỳ cũ ⇒ NCC này không có dòng nào."""
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC So Ky C")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(0),
    )

    bc = _bao_cao(client, headers, "payables", tu=_ngay(60), den=_ngay(30))
    assert not [d for d in bc["items"] if "NCC So Ky C" in (d["ten"] or "")]
    _kiem_can_so(bc)


def test_coc_truoc_khi_hang_ve_thi_331_du_no(client):
    """Ứng trước cho NCC mà hàng chưa về ⇒ dư **Nợ** của 331 (mình đang là chủ nợ).

    Chính nhánh này làm nên cột "dư cuối kỳ bên Nợ" 1,7 tỷ trong bản xuất MISA. Số âm không được
    phép lọt ra ngoài — nó phải nhảy sang cột bên kia.
    """
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC Coc Truoc")
    don = _don(client, headers, ncc["id"], coc=500_000)
    _phieu_chi(client, headers, don["id"], 500_000, stage="advance")

    bc = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    d = _dong(bc, "NCC Coc Truoc")
    assert (d["ps_no"], d["ps_co"]) == (500_000, 0), "chi tiền = PS Nợ"
    assert (d["cuoi_no"], d["cuoi_co"]) == (500_000, 0), "dư NỢ: NCC đang giữ tiền của mình"
    _kiem_can_so(bc)


def test_hang_ve_roi_tra_du_thi_cuoi_ky_ve_0(client):
    """Vòng đời trọn vẹn: hàng về 880k, trả 880k ⇒ cuối kỳ 0/0 nhưng dòng VẪN hiện.

    Dòng phải còn: sổ tổng hợp là để đối chiếu, mà "đã phát sinh rồi tất toán" khác hẳn "không
    giao dịch gì" — ẩn đi là kế toán không tìm thấy NCC mình vừa trả tiền xong.
    """
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC Tra Du")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    dot = _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(5),
    )
    _phieu_chi(
        client, headers, don["id"], 880_000,
        stage="final", delivery_id=dot["deliveries"][0]["id"],
    )

    bc = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    d = _dong(bc, "NCC Tra Du")
    assert (d["ps_no"], d["ps_co"]) == (880_000, 880_000)
    assert (d["cuoi_no"], d["cuoi_co"]) == (0, 0)
    _kiem_can_so(bc)


# ══ TK 131 — PHẢI THU ═══════════════════════════════════════════════════════════════════════


def test_hoa_don_ban_vao_ps_no_va_can_so(client):
    """Xuất hoá đơn = khách nợ mình = PS **Nợ** của 131."""
    headers = _headers(client)
    order_id, _ = _sales_order(suffix="BC1")
    r = client.post(
        "/api/accounting/sales-invoices",
        json=_invoice_payload(order_id, number="70000001", amount=400_000),
        headers=headers,
    )
    assert r.status_code == 201, r.text

    bc = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))
    d = _dong(bc, "Customer invoice test BC1")
    assert (d["ps_no"], d["ps_co"]) == (400_000, 0)
    assert (d["cuoi_no"], d["cuoi_co"]) == (400_000, 0)
    assert d["tk"] == "131"
    _kiem_can_so(bc)


def test_coc_don_hang_chua_can_tru_lam_131_du_co(client):
    """Khách ứng trước mà chưa có hoá đơn nào ⇒ dư **Có** của 131 — mình đang giữ tiền của họ.

    Đây chính là 4,5 tỷ dư Có trong bản xuất MISA. Màn Công nợ phải thu hiện tại KHÔNG thấy được
    con số này (nó chỉ đếm nợ còn lại ≥ 0), nên đây là thứ sổ tổng hợp làm được mà màn kia không.
    """
    headers = _headers(client)
    order_id, _ = _sales_order(suffix="BC2")
    _add_deposit(order_id, 250_000)

    bc = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))
    d = _dong(bc, "Customer invoice test BC2")
    assert (d["ps_no"], d["ps_co"]) == (0, 250_000)
    assert (d["cuoi_no"], d["cuoi_co"]) == (0, 250_000), "dư CÓ: khách ứng trước"
    _kiem_can_so(bc)


def test_tong_cong_theo_cot_khong_triet_tieu(client):
    """Một khách dư Nợ và một khách dư Có phải hiện ĐỦ HAI cột ở dòng tổng, không bù trừ nhau.

    Bản xuất MISA cũng thế: cuối kỳ phải thu của họ có Nợ 36,1 tỷ **và** Có 4,5 tỷ cùng lúc.
    """
    headers = _headers(client)
    no_id, _ = _sales_order(suffix="BC3")
    co_id, _ = _sales_order(suffix="BC4")
    assert client.post(
        "/api/accounting/sales-invoices",
        json=_invoice_payload(no_id, number="70000002", amount=300_000),
        headers=headers,
    ).status_code == 201
    _add_deposit(co_id, 300_000)

    bc = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))
    assert _dong(bc, "test BC3")["cuoi_no"] == 300_000
    assert _dong(bc, "test BC4")["cuoi_co"] == 300_000
    assert bc["tong"]["cuoi_no"] >= 300_000 and bc["tong"]["cuoi_co"] >= 300_000
    _kiem_can_so(bc)


# ══ CỬA VÀO ═════════════════════════════════════════════════════════════════════════════════


def test_tu_ngay_sau_den_ngay_bi_chan(client):
    """Khoảng ngày ngược ⇒ 422. Không chặn thì báo cáo trả bảng RỖNG một cách bí ẩn: mọi chứng từ
    rơi vào nhánh 'ngoài kỳ' mà không ai hiểu vì sao."""
    headers = _headers(client)
    for duong in ("payables", "receivables"):
        _bao_cao(client, headers, duong, tu=_ngay(0), den=_ngay(30), expect=422)


def test_nhan_cot_khac_nhau_giua_hai_bao_cao(client):
    """Giao diện dùng CHUNG một component nên nhãn phải do server nói."""
    headers = _headers(client)
    tra = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    thu = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))
    assert (tra["nhan_ma"], tra["tk"]) == ("Mã nhà cung cấp", "331")
    assert (thu["nhan_ma"], thu["tk"]) == ("Mã khách hàng", "131")
    assert tra["tieu_de"] == "TỔNG HỢP CÔNG NỢ PHẢI TRẢ"
    assert thu["tieu_de"] == "TỔNG HỢP CÔNG NỢ PHẢI THU"


# ══ XUẤT EXCEL — KHOÁ KHUÔN MISA (§5.5) ═════════════════════════════════════════════════════
#
# Mấy con số dưới đây bóc từ chính bản xuất của họ (`TONG_HOP_CONG_NO_PHAI_THU.xlsx`, đọc
# 03/09/2026). Đổi bất kỳ cái nào là file lệch khuôn và mất luôn lý do tồn tại: kế toán phải dán
# thẳng vào bộ hồ sơ đang dùng, không sửa tay.


def _mo_file(client, headers, duong: str, *, tu: str, den: str):
    from io import BytesIO

    from openpyxl import load_workbook

    r = client.get(
        f"/api/accounting/reports/{duong}.xlsx",
        params={"tu_ngay": tu, "den_ngay": den},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    assert "attachment" in r.headers["content-disposition"]
    return load_workbook(BytesIO(r.content)).active


def test_file_xuat_dung_khuon_misa(client):
    headers = _headers(client)
    ws = _mo_file(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))

    assert ws["A1"].value == "TỔNG HỢP CÔNG NỢ PHẢI THU"
    assert ws["A2"].value.startswith("Tài khoản: 131; Loại tiền: Tổng hợp; Từ ngày ")
    assert sorted(str(m) for m in ws.merged_cells.ranges) == [
        "A1:I1", "A2:I2", "A3:A4", "B3:B4", "C3:C4", "D3:E3", "F3:G3", "H3:I3",
    ]
    assert (ws["A3"].value, ws["B3"].value, ws["C3"].value) == (
        "Mã khách hàng", "Tên khách hàng", "TK công nợ",
    )
    assert (ws["D3"].value, ws["F3"].value, ws["H3"].value) == (
        "Số dư đầu kỳ", "Số phát sinh", "Số dư cuối kỳ",
    )
    assert [ws[f"{c}4"].value for c in "DEFGHI"] == ["Nợ", "Có"] * 3

    # Phông chữ: hai dòng đầu Times New Roman, phần bảng Microsoft Sans Serif 8.
    assert (ws["A1"].font.name, ws["A1"].font.size, ws["A1"].font.bold) == (
        "Times New Roman", 14, True,
    )
    assert (ws["A2"].font.name, ws["A2"].font.size, ws["A2"].font.bold) == (
        "Times New Roman", 11, True,
    )
    assert (ws["A3"].font.name, ws["A3"].font.size) == ("Microsoft Sans Serif", 8)

    # Cột A-C theo mẫu MISA; D-I (6 cột tiền) đặt 17.1 để số tiền lớn không bị tràn.
    assert {k: round(v.width, 1) for k, v in ws.column_dimensions.items() if v.width} == {
        "A": 17.1, "B": 30.0, "C": 14.3, "D": 17.1, "E": 17.1, "F": 17.1, "G": 17.1, "H": 17.1, "I": 17.1,
    }


def test_file_xuat_co_dong_chan_va_dinh_dang_so(client):
    """Dòng chân `Số dòng = N` + tổng 6 cột, và ô tiền mang đúng định dạng số của MISA."""
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC Xuat File")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(3),
    )

    ws = _mo_file(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    chan = ws.max_row
    assert str(ws[f"A{chan}"].value).startswith("Số dòng = ")
    so_dong = int(str(ws[f"A{chan}"].value).split("=")[1])
    assert so_dong == chan - 5, "dòng chân phải đếm đúng số dòng dữ liệu (bắt đầu từ hàng 5)"

    # Tổng cột "PS Có" (cột G) = cộng các dòng dữ liệu ở chính cột đó.
    assert ws[f"G{chan}"].value == sum(ws[f"G{h}"].value or 0 for h in range(5, chan))
    assert ws[f"G{chan}"].value >= 880_000

    assert ws["D5"].number_format == "#,##0_);[Red](#,##0)"
    assert ws["A5"].alignment.horizontal == "left"
    assert ws["D5"].alignment.horizontal == "right"


def test_ten_file_co_ky_de_khong_de_len_nhau(client):
    """Tải hai kỳ khác nhau phải ra hai tên file khác nhau — `bao-cao (1).xlsx` là vô dụng."""
    headers = _headers(client)
    r1 = client.get(
        "/api/accounting/reports/payables.xlsx",
        params={"tu_ngay": _ngay(30), "den_ngay": _ngay(0)}, headers=headers,
    )
    r2 = client.get(
        "/api/accounting/reports/payables.xlsx",
        params={"tu_ngay": _ngay(60), "den_ngay": _ngay(31)}, headers=headers,
    )
    assert r1.headers["content-disposition"] != r2.headers["content-disposition"]
    assert "tong-hop-cong-no-phai-tra-" in r1.headers["content-disposition"]


# ══ PHÂN TUỔI NỢ TẠI "ĐẾN NGÀY" (§5.3, chủ chốt 04/09/2026) ═════════════════════════════════
#
# Đây là chỗ báo cáo khác hẳn màn Công nợ: màn kia neo vào HÔM NAY, báo cáo neo vào ĐẾN NGÀY của
# kỳ. Chọn kỳ đến 31/08 phải ra tuổi nợ đúng như tại 31/08, và in lại tháng sau vẫn ra con số đó.


def _ro(bc: dict, khoa: str) -> int:
    return next(b["amount"] for b in bc["aging"] if b["key"] == khoa)


def test_ro_tuoi_tinh_tai_den_ngay_khong_phai_hom_nay(client):
    """⭐ Cùng một đợt giao, hai kỳ khác nhau ⇒ rơi HAI rổ khác nhau.

    Đợt quá hạn 40 ngày tính tới hôm nay, nhưng mới quá hạn 5 ngày tính tới 35 ngày trước. Nếu
    rổ neo vào hôm nay thì hai lần xem ra cùng một rổ — và báo cáo kỳ cũ in lại sẽ sai.
    """
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC Ro Theo Moc")
    # `credit_days=0` ⇒ hạn trả đúng bằng ngày giao, khỏi phải cộng trừ.
    client.put(
        f"/api/suppliers/{ncc['id']}",
        json={**{k: ncc[k] for k in
                 ("name", "tax_code", "phone", "email", "address", "contact_name", "supplier_group")},
              "credit_days": 0,
              "items": [{"item_name": "Giấy Duplex", "unit": "tờ", "unit_price": 2200, "vat_percent": 0}]},
        headers=headers,
    )
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(40),
    )

    hom_nay = _bao_cao(client, headers, "payables", tu=_ngay(90), den=_ngay(0))
    ky_cu = _bao_cao(client, headers, "payables", tu=_ngay(90), den=_ngay(35))

    assert _ro(hom_nay, "d31_60") == 880_000, "tính tới hôm nay: trễ 40 ngày"
    assert _ro(ky_cu, "d1_7") == 880_000, "tính tới 35 ngày trước: mới trễ 5 ngày"
    assert _ro(ky_cu, "d31_60") == 0


def test_tien_thu_sau_ky_khong_lam_nhe_tuoi_no_cua_ky_cu(client):
    """⭐ Khách trả tiền SAU kỳ thì trong kỳ vẫn còn nợ nguyên.

    Cạm bẫy: mượn `remaining_amount` của `_receivable_rows` (số "còn lại tính tới HÔM NAY") thì
    tiền trả sau kỳ đã bị trừ mất, và tuổi nợ của kỳ cũ nhẹ đi — in lại kỳ tháng 7 vào tháng 9 ra
    số khác lần in tháng 8.
    """
    headers = _headers(client)
    order_id, _ = _sales_order(suffix="RO1", term_days=0)
    r = client.post(
        "/api/accounting/sales-invoices",
        json={**_invoice_payload(order_id, number="70000009", amount=500_000),
              "invoice_date": _ngay(20)},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    _add_deposit(order_id, 500_000)                   # cọc ghi HÔM NAY, tức SAU kỳ dưới đây

    ky_cu = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(10))
    assert _ro(ky_cu, "d8_15") == 500_000, "tại mốc đó khách vẫn nợ nguyên, trễ 10 ngày"

    bay_gio = _bao_cao(client, headers, "receivables", tu=_ngay(30), den=_ngay(0))
    assert sum(b["amount"] for b in bay_gio["aging"]) == 0, "trả rồi thì hết nợ"


def test_loc_ro_giu_nguyen_so_tien_cua_dong(client):
    """Lọc chọn DÒNG, KHÔNG cắt SỐ — 9 cột tiền của đối tượng vẫn hiện đủ."""
    headers = _headers(client)
    ncc = _supplier(client, headers, name="NCC Loc Ro")
    don = _don(client, headers, ncc["id"])
    _da_mua(client, headers, don["id"])
    _ghi_dot(
        client, headers, don["id"],
        lines=[{"purchase_request_line_id": _dong_dau_tien(don), "quantity": 400}],
        ngay=_ngay(2),
    )

    day_du = _bao_cao(client, headers, "payables", tu=_ngay(30), den=_ngay(0))
    goc = _dong(day_du, "NCC Loc Ro")

    r = client.get(
        "/api/accounting/reports/payables",
        params={"tu_ngay": _ngay(30), "den_ngay": _ngay(0), "aging_bucket": "chua_toi_han"},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    loc = r.json()
    sau = _dong(loc, "NCC Loc Ro")
    assert sau["cuoi_co"] == goc["cuoi_co"], "lọc rổ không được cắt số tiền của dòng"
    assert len(loc["items"]) <= len(day_du["items"])
    # Dải rổ giữ NGUYÊN toàn màn: bấm để soi, không phải để đổi bức tranh.
    assert loc["aging"] == day_du["aging"]


def test_khoa_ro_la_thi_bo_qua_khong_no_422(client):
    """Cửa lọc không phải chỗ ném lỗi vào mặt người đang xem sổ."""
    headers = _headers(client)
    r = client.get(
        "/api/accounting/reports/payables",
        params={"tu_ngay": _ngay(30), "den_ngay": _ngay(0), "aging_bucket": "khoa_bay_ra"},
        headers=headers,
    )
    assert r.status_code == 200, r.text
