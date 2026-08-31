"""Báo cáo KCS — tổng hợp theo filter/scope + xuất Excel hai sheet (§5.7, §6.2, §9 mục 10, mg 0250).

Soi tầng service mới `services/san_xuat/kcs_bao_cao.py` (KHÔNG đụng `kcs.py`):
  · mục 2: tỷ lệ đạt = tổng Đạt / tổng nhận (SUM/SUM), KHÔNG phải trung bình tỷ lệ từng phiếu;
  · mục 3: nhóm lỗi/công đoạn/tổ xếp theo TỔNG số lượng, không phải đếm dòng;
  · mục 3.4: "tổ hiệu lực" — batch `routing` lấy `cong_viec.department_id` (KHÔNG có
    `kcs_department_id`, cột này NULL); batch `dot_xuat` lấy `kcs.kcs_department_id`;
  · mục 4: hồ sơ `to_chiu_id IS NULL` KHÔNG vào bảng xếp hạng "tổ", nhưng VẪN cộng vào KPI tổng
    lỗi và bảng xếp hạng nhóm lỗi;
  · §9 mục 10: JSON (`bao_cao_kcs`) và Excel (`xuat_excel_kcs`) đọc CHUNG `_hang_kcs_theo_scope`
    — cùng filter phải trả cùng tổng (test 8);
  · RBAC: `GET /kcs/bao-cao` gác `read`, `GET /kcs/bao-cao/export.xlsx` gác `export` RIÊNG (test 7,
    đi qua HTTP thật vì đây là chỗ RBAC thật sự áp — service không tự gác quyền).

Tái dùng dàn cảnh + helper của test KCS (`_batch`, `_cv_kcs`, `_cv_production`, `_to_kiem`, `_ly_do`,
`_anh`) và `_authz`/`_FakeAuthz` của test board (ép scope không phụ thuộc tên role seed).
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.db import SessionLocal
from app.models.cong_doan import CongDoan
from app.models.department import Department
from app.models.lsx import Lsx
from app.models.role import SCOPE_ALL, SCOPE_OWN
from app.models.san_xuat_kcs import KCS_LOAI_DOT_XUAT, KCS_LOAI_ROUTING, SanXuatKcsBatch
from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
from app.repositories.user_repo import UserRepository
from app.security import create_access_token, hash_password
from app.services.san_xuat import kcs, kcs_bao_cao

# _authz (scope THẬT theo role admin) + _FakeAuthz (ép cứng scope, không phụ thuộc tên role seed).
from tests.test_san_xuat_board import _FakeAuthz, _authz

# Fixtures + helper dàn cảnh dùng chung — TỪ test KCS (theo đúng mẫu test_san_xuat_kho.py).
from tests.test_san_xuat_kcs import (  # noqa: F401
    _T0,
    _T1,
    _anh,
    _batch,
    _cv_kcs,
    _cv_production,
    _ly_do,
    _to_kiem,
    admin,
    customer,
    db,
    lsx_svc,
    orders,
)

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Checklist mẫu 2 tiêu chí, cả hai KHÔNG bắt buộc — để test gửi/không-gửi checklist đều hợp lệ,
# không vướng luật `_validate_checklist_bat_buoc` (mục 3.7).
_TIEU_CHI_2 = [
    {"tieu_chi_id": 1, "ma": "TC-01", "ten": "Không lệch màu", "huong_dan": None,
     "bat_buoc": False, "nguon": "danh_muc", "thu_tu": 1},
    {"tieu_chi_id": 2, "ma": "TC-02", "ten": "Không rách giấy", "huong_dan": None,
     "bat_buoc": False, "nguon": "danh_muc", "thu_tu": 2},
]


def _batch_voi_checklist(db, orders, lsx_svc, admin, customer, *, ma, checklist_ket_qua,
                          nhan=100, dat=90, khong_dat=10):
    """Batch routing CÓ gắn snapshot `cv.kcs_tieu_chi_json` TRƯỚC khi ghi — khác `_batch` (không
    gắn gì) — dùng cho test Sheet 2 (mục 3.7)."""
    to, cv = _cv_kcs(db, orders, lsx_svc, admin, customer, ma=ma)
    cv.kcs_tieu_chi_json = _TIEU_CHI_2
    db.commit()
    res = kcs.tao_batch_kcs(
        db, user=admin, cong_viec_id=cv.id, bat_dau=_T0, ket_thuc=_T1,
        so_luong_nhan=nhan, so_luong_dat=dat, so_luong_khong_dat=khong_dat,
        checklist_ket_qua=checklist_ket_qua,
    )
    return to, cv, res


# --- Mục 2: tỷ lệ đạt = tổng/tổng --------------------------------------------------------------
def test_ty_le_dat_la_tong_tren_tong_khong_phai_trung_binh(db, orders, lsx_svc, admin, customer):
    _batch(db, orders, lsx_svc, admin, customer, nhan=100, dat=90, khong_dat=10, ma="KCS-TY-A")
    _batch(db, orders, lsx_svc, admin, customer, nhan=10, dat=1, khong_dat=9, ma="KCS-TY-B")

    out = kcs_bao_cao.bao_cao_kcs(db, admin, _authz(db))

    assert out["tong_nhan"] == 110
    assert out["tong_dat"] == 91
    assert abs(out["ty_le_dat"] - (91 / 110)) < 1e-9   # tổng/tổng ≈ 0.827
    assert abs(out["ty_le_dat"] - 0.5) > 0.1            # KHÔNG phải trung bình (90%+10%)/2 = 50%


# --- Mục 3: xếp hạng theo TỔNG số lượng, không phải đếm dòng ----------------------------------
def test_xep_hang_nhom_loi_cong_doan_to_theo_tong_so_luong(db, orders, lsx_svc, admin, customer):
    _to, _cv, res = _batch(db, orders, lsx_svc, admin, customer, nhan=100, dat=46, khong_dat=54,
                            ma="KCS-XH")
    ld_x = _ly_do(db, ma="LOI-XH-X", ten="Nhóm X")
    ld_y = _ly_do(db, ma="LOI-XH-Y", ten="Nhóm Y")
    kcs.ghi_loi(db, user=admin, kcs_batch_id=res["kcs_batch_id"], nhom_loi_id=ld_x.id,
                so_luong=50, anh=_anh())
    for _ in range(3):
        kcs.ghi_loi(db, user=admin, kcs_batch_id=res["kcs_batch_id"], nhom_loi_id=ld_y.id,
                     so_luong=1, anh=_anh())

    out = kcs_bao_cao.bao_cao_kcs(db, admin, _authz(db))

    assert out["nhom_loi"][0]["nhom_loi_id"] == ld_x.id   # tổng 50 > tổng 3, KHÔNG phải Y (3 dòng)
    assert out["nhom_loi"][0]["tong_so_luong"] == 50
    ten_by_id = {r["nhom_loi_id"]: r["ten"] for r in out["nhom_loi"]}
    assert ten_by_id[ld_x.id] == "Nhóm X" and ten_by_id[ld_y.id] == "Nhóm Y"


# --- Mục 4: hồ sơ không xác định trách nhiệm KHÔNG vào bảng xếp hạng "tổ" (nhưng vẫn vào KPI +
# xếp hạng nhóm lỗi) --------------------------------------------------------------------------
def test_loi_khong_to_chiu_khong_len_bang_xep_hang_to(db, orders, lsx_svc, admin, customer):
    _to, _cv, res = _batch(db, orders, lsx_svc, admin, customer, nhan=100, dat=90, khong_dat=10,
                            ma="KCS-M4")
    ld = _ly_do(db, ma="LOI-M4", ten="Lỗi chung")
    to_chiu = Department(name="Tổ Chịu M4", code="KCS-M4-TC", la_san_xuat=True)
    db.add(to_chiu)
    db.commit()
    kcs.ghi_loi(db, user=admin, kcs_batch_id=res["kcs_batch_id"], nhom_loi_id=ld.id,
                to_chiu_id=to_chiu.id, so_luong=6, anh=_anh())
    kcs.ghi_loi(db, user=admin, kcs_batch_id=res["kcs_batch_id"], nhom_loi_id=ld.id,
                to_chiu_id=None, so_luong=4, anh=_anh())

    out = kcs_bao_cao.bao_cao_kcs(db, admin, _authz(db))

    assert out["tong_loi"] == 10                    # KPI lấy từ batch, KHÔNG mất dòng to_chiu=None
    to_ids = {r["to_id"] for r in out["to"]}
    assert None not in to_ids and to_chiu.id in to_ids
    to_row = next(r for r in out["to"] if r["to_id"] == to_chiu.id)
    assert to_row["tong_so_luong"] == 6              # CHỈ dòng có to_chiu được cộng vào bảng "tổ"
    nhom_row = next(r for r in out["nhom_loi"] if r["nhom_loi_id"] == ld.id)
    assert nhom_row["tong_so_luong"] == 10           # CẢ HAI dòng (6+4) vẫn cộng vào nhóm lỗi


# --- Mục 3.4 (Ruling 2): batch routing lọc theo tổ ĐANG CHẠY, không phải kcs_department_id -----
def test_batch_routing_loc_theo_kcs_department_id_dung_to_dang_chay(db, orders, lsx_svc, admin,
                                                                      customer):
    to_x, _cv_x, res_x = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-RT-X")
    kb = db.get(SanXuatKcsBatch, res_x["kcs_batch_id"])
    assert kb.kcs_department_id is None and kb.loai == KCS_LOAI_ROUTING   # cột NULL — đúng mục 3.4

    to_y = Department(name="Tổ Khác RT Y", code="KCS-RT-Y", la_san_xuat=True)
    db.add(to_y)
    db.commit()

    authz = _authz(db)
    out_x = kcs_bao_cao.bao_cao_kcs(db, admin, authz, kcs_department_id=to_x.id)
    assert out_x["tong_luot"] == 1                    # resolve qua cv.department_id, THẤY được
    out_y = kcs_bao_cao.bao_cao_kcs(db, admin, authz, kcs_department_id=to_y.id)
    assert out_y["tong_luot"] == 0                     # tổ khác — KHÔNG thấy


# --- §4.1: scope `own` chỉ thấy báo cáo tổ mình -------------------------------------------------
def test_scope_own_chi_thay_bao_cao_to_minh(db, orders, lsx_svc, admin, customer):
    to_x, _cv_x, _res_x = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-SC-X")
    _to_y, _cv_y, _res_y = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-SC-Y")

    user_x = SimpleNamespace(id=admin.id, department_id=to_x.id)
    out = kcs_bao_cao.bao_cao_kcs(db, user_x, _FakeAuthz(SCOPE_OWN))

    assert out["tong_luot"] == 1                        # chỉ tổ X, KHÔNG cộng tổ Y


# --- Filter riêng lẻ: loai / cong_doan_id / tu-den / tu_khoa -----------------------------------
def test_filter_loai_tu_den_cong_doan_id_tu_khoa_thu_hep_dung(db, orders, lsx_svc, admin, customer):
    # Routing: công đoạn "In-Test-FL", ngày _T0 (2026-08-20 giờ VN).
    _to_a, cv_a, res_a = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-FL-A")
    cv_a.ten_cong_doan = "In-Test-FL"
    db.commit()
    cd_in = CongDoan(ma="CD-FL-IN", ten="In-Test-FL", nhom="test")
    db.add(cd_in)
    db.flush()
    db.commit()

    # Đột xuất: công đoạn khác tên, tổ khác, LSX khác, ngày khác (+5 ngày).
    _to_prod, cv_b = _cv_production(db, orders, lsx_svc, admin, customer, ma="KCS-FL-B")
    cv_b.ten_cong_doan = "Be-Test-FL"
    db.commit()
    to_kiem, u_kiem = _to_kiem(db, ma="KCS-FL-KIEM")
    ngay_dx = _T0 + timedelta(days=5)
    kcs.tao_kiem_dot_xuat(
        db, user=u_kiem, cong_viec_id=cv_b.id, kcs_department_id=to_kiem.id,
        bat_dau=ngay_dx, ket_thuc=ngay_dx + timedelta(hours=1),
        so_luong_nhan=20, so_luong_dat=20, so_luong_khong_dat=0, don_vi="cái",
    )

    authz = _authz(db)

    out_dx = kcs_bao_cao.bao_cao_kcs(db, admin, authz, loai=KCS_LOAI_DOT_XUAT)
    assert out_dx["tong_luot"] == 1
    out_rt = kcs_bao_cao.bao_cao_kcs(db, admin, authz, loai=KCS_LOAI_ROUTING)
    assert out_rt["tong_luot"] == 1

    out_cd = kcs_bao_cao.bao_cao_kcs(db, admin, authz, cong_doan_id=cd_in.id)
    assert out_cd["tong_luot"] == 1
    assert out_cd["tong_luot"] == out_rt["tong_luot"]   # cùng đúng 1 batch routing công đoạn "In"

    out_tu = kcs_bao_cao.bao_cao_kcs(db, admin, authz, tu=date(2026, 8, 20), den=date(2026, 8, 20))
    assert out_tu["tong_luot"] == 1   # chỉ ngày của routing, không lấy batch đột xuất +5 ngày

    lsx_a = db.get(Lsx, cv_a.lsx_id)
    assert lsx_a is not None
    out_kw = kcs_bao_cao.bao_cao_kcs(db, admin, authz, tu_khoa=lsx_a.ma)
    assert out_kw["tong_luot"] == 1


# --- RBAC: export gác quyền RIÊNG với read (§4.4, §4.1) — qua HTTP thật --------------------------
def test_export_yeu_cau_quyen_export_rieng_voi_read(client):
    login = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert login.status_code == 200, login.text
    headers_admin = {"Authorization": f"Bearer {login.json()['access_token']}"}

    db2 = SessionLocal()
    try:
        dept = DepartmentRepository(db2).get_by_name("Sản xuất")
        assert dept is not None
        role = RoleRepository(db2).create(name="KCS Bao Cao Chi Doc", department_id=dept.id)
        RoleRepository(db2).set_permission(
            role_id=role.id, module_key="san_xuat", can_read=True, scope=SCOPE_ALL,
        )
        u = UserRepository(db2).create(
            username="kcs-bc-doc", name="NV Đọc Báo Cáo KCS", password_hash=hash_password("x"),
        )
        UserRepository(db2).set_assignment(u, department_id=dept.id, role_id=role.id, is_active=True)
        headers_doc = {"Authorization": f"Bearer {create_access_token(str(u.id))}"}
    finally:
        db2.close()

    r_json = client.get("/api/san-xuat/kcs/bao-cao", headers=headers_doc)
    assert r_json.status_code == 200, r_json.text   # read đủ cho JSON

    r_xlsx_doc = client.get("/api/san-xuat/kcs/bao-cao/export.xlsx", headers=headers_doc)
    assert r_xlsx_doc.status_code == 403, r_xlsx_doc.text   # KHÔNG có can_export → chặn

    r_xlsx_admin = client.get("/api/san-xuat/kcs/bao-cao/export.xlsx", headers=headers_admin)
    assert r_xlsx_admin.status_code == 200, r_xlsx_admin.text
    assert r_xlsx_admin.headers["content-type"] == _XLSX_MEDIA


# --- §9 mục 10: Excel áp đúng scope như dashboard (cùng filter → cùng tổng) ---------------------
def test_export_ap_dung_scope_giong_dashboard(db, orders, lsx_svc, admin, customer):
    to_x, _cv_x, _res_x = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-SC8-X")
    _to_y, _cv_y, _res_y = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-SC8-Y")

    user_x = SimpleNamespace(id=admin.id, department_id=to_x.id)
    authz_own = _FakeAuthz(SCOPE_OWN)

    out = kcs_bao_cao.bao_cao_kcs(db, user_x, authz_own)
    content, _fn = kcs_bao_cao.xuat_excel_kcs(db, user_x, authz_own)
    wb = load_workbook(BytesIO(content))
    ws1 = wb["Kết quả KCS"]

    assert ws1.max_row - 1 == out["tong_luot"] == 1


# --- Excel: tên sheet + header đúng -------------------------------------------------------------
def test_excel_ten_sheet_va_header_dung(db, orders, lsx_svc, admin, customer):
    _batch(db, orders, lsx_svc, admin, customer, ma="KCS-XL-HD")

    content, _fn = kcs_bao_cao.xuat_excel_kcs(db, admin, _authz(db))
    wb = load_workbook(BytesIO(content))

    assert wb.sheetnames == ["Kết quả KCS", "Chi tiết checklist"]
    header1 = [c.value for c in wb["Kết quả KCS"][1]]
    assert header1 == [
        "Mã kết quả", "Thời điểm", "Loại", "Tổ KCS", "Mã LSX", "Công đoạn", "Số nhận",
        "Số đạt", "Số không đạt", "Đơn vị", "Kết luận", "Ghi chú", "Người ghi",
        "Số lỗi ghi nhận", "Nhóm lỗi", "Tổ chịu trách nhiệm", "URL ảnh",
    ]
    header2 = [c.value for c in wb["Chi tiết checklist"][1]]
    assert header2 == [
        "Mã kết quả", "Thời điểm", "Mã tiêu chí", "Tên tiêu chí", "Bắt buộc", "Đạt", "Ghi chú",
    ]


# --- Excel: số dòng khớp dữ liệu đã lọc, và batch thiếu 1 trong 2 field JSON thì Sheet 2 = 0 dòng
def test_excel_so_dong_khop_du_lieu(db, orders, lsx_svc, admin, customer):
    _to_a, _cv_a, res_a = _batch_voi_checklist(
        db, orders, lsx_svc, admin, customer, ma="KCS-XL-CNT-A",
        checklist_ket_qua=[
            {"thu_tu": 1, "dat": True, "ghi_chu": None},
            {"thu_tu": 2, "dat": False, "ghi_chu": "lem mực"},
        ],
    )
    _batch(db, orders, lsx_svc, admin, customer, ma="KCS-XL-CNT-B")   # KHÔNG gửi checklist

    authz = _authz(db)
    out = kcs_bao_cao.bao_cao_kcs(db, admin, authz)
    content, _fn = kcs_bao_cao.xuat_excel_kcs(db, admin, authz)
    wb = load_workbook(BytesIO(content))

    ws1 = wb["Kết quả KCS"]
    assert ws1.max_row - 1 == out["tong_luot"] == 2

    ws2 = wb["Chi tiết checklist"]
    assert ws2.max_row - 1 == 2   # chỉ batch A góp 2 dòng (2 tiêu chí); B góp 0 dòng


def test_excel_url_anh_xuat_hien_dung_cot(db, orders, lsx_svc, admin, customer):
    _to, _cv, res = _batch(db, orders, lsx_svc, admin, customer, ma="KCS-XL-IMG")
    ld = _ly_do(db, ma="LOI-XL-IMG", ten="Lỗi có ảnh")
    anh_2 = [
        {"file_name": "a.jpg", "file_url": "https://x/a.jpg", "file_type": "image/jpeg"},
        {"file_name": "b.jpg", "file_url": "https://x/b.jpg", "file_type": "image/jpeg"},
    ]
    kcs.ghi_loi(db, user=admin, kcs_batch_id=res["kcs_batch_id"], nhom_loi_id=ld.id,
                so_luong=5, anh=anh_2)

    content, _fn = kcs_bao_cao.xuat_excel_kcs(db, admin, _authz(db))
    wb = load_workbook(BytesIO(content))
    ws1 = wb["Kết quả KCS"]

    dong = next(r for r in ws1.iter_rows(min_row=2) if r[0].value == res["kcs_batch_id"])
    url_cell = dong[16].value   # cột 17 "URL ảnh" (index 0-based 16)
    assert "https://x/a.jpg" in url_cell and "https://x/b.jpg" in url_cell


def test_checklist_thieu_snapshot_khong_vo_khong_ra_dong_rac(db, orders, lsx_svc, admin, customer):
    to, cv = _cv_kcs(db, orders, lsx_svc, admin, customer, ma="KCS-XL-MISS")
    # cv.kcs_tieu_chi_json CHƯA gắn (None) — việc cũ trước module này — nhưng batch VẪN gửi
    # checklist_ket_qua (form không biết cv thiếu snapshot).
    assert cv.kcs_tieu_chi_json is None
    res = kcs.tao_batch_kcs(
        db, user=admin, cong_viec_id=cv.id, bat_dau=_T0, ket_thuc=_T1,
        so_luong_nhan=50, so_luong_dat=50, so_luong_khong_dat=0,
        checklist_ket_qua=[{"thu_tu": 1, "dat": True, "ghi_chu": None}],
    )
    kb = db.get(SanXuatKcsBatch, res["kcs_batch_id"])
    assert kb.checklist_json   # batch CÓ checklist_json, chỉ thiếu snapshot ở cv

    content, _fn = kcs_bao_cao.xuat_excel_kcs(db, admin, _authz(db))
    wb = load_workbook(BytesIO(content))
    ws2 = wb["Chi tiết checklist"]

    assert ws2.max_row == 1   # chỉ header — không văng lỗi, không dòng rác
