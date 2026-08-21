"""Cấu hình lương (docs/prd-cau-hinh-luong.md + chốt của chủ 2026-07-20):

- Thành phần lương theo TỔ (lương bậc · KPI · chuyên cần · khoán · tăng ca), ghi đè 2 cấp NV → tổ.
- 4 khoản phụ cấp (ca · trách nhiệm · thâm niên · khác) KHAI TAY theo từng NV, cộng phẳng.
- Chuyên cần TRỪ DẦN · bậc là KHUNG + cảnh báo mềm · emp_count · quyền ghi thang bậc dưới
  /api/luong · ghi chú chính sách của tổ — và vá lệch "Sửa 1 ô" vs "Tính lại".
"""
from __future__ import annotations

from datetime import date
from types import SimpleNamespace

from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db import SessionLocal
from app.db_migrations import (
    _migrate_ca_phu_cap_com_ca_dem,
    _migrate_ca_rename_shift_allowance_go_night_shift,
    _migrate_cau_hinh_luong,
    _migrate_luong_bo_bac_luong,
    _migrate_luong_phu_cap_com_ca_dem,
    _migrate_luong_phu_cap_khai_tay,
    _migrate_luong_v2_khung_bac,
    _migrate_payroll_line_allowance_split,
)
from app.repositories.employee_repo import EmployeeRepository
from app.repositories.payroll_repo import PayrollRepository
from app.repositories.rbac_repo import DepartmentRepository, RoleRepository
from app.repositories.user_repo import UserRepository
from app.security import create_access_token, hash_password
from app.services.payroll_service import PayrollService

ADMIN = {"username": "admin", "password": "admin123"}


def _admin_token(client) -> str:
    return client.post("/api/auth/login", json=ADMIN).json()["access_token"]


def _h(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _dept_id(name: str) -> int:
    db = SessionLocal()
    try:
        return DepartmentRepository(db).get_by_name(name).id
    finally:
        db.close()


def _sales_token() -> str:
    """Tài khoản KHÔNG có quyền module `luong` — để kiểm RBAC của endpoint mới."""
    db = SessionLocal()
    try:
        users = UserRepository(db)
        existing = users.get_by_username("sales-cauhinh")
        if existing is not None:
            return create_access_token(str(existing.id))
        kd = DepartmentRepository(db).get_by_name("Kinh doanh")
        role = RoleRepository(db).get_by_name_and_department("NV Sales", kd.id)
        u = users.create(username="sales-cauhinh", name="S", password_hash=hash_password("x"))
        users.set_assignment(u, department_id=kd.id, role_id=role.id, is_active=True)
        return create_access_token(str(u.id))
    finally:
        db.close()


def _luong_config_token(username: str, *, can_read: bool = False,
                        can_view_salary: bool = False, can_update: bool = False) -> str:
    """Tạo vai trò giới hạn để chứng minh `luong:read` không còn làm lộ cấu hình."""
    db = SessionLocal()
    try:
        users = UserRepository(db)
        depts = DepartmentRepository(db)
        roles = RoleRepository(db)
        kd = depts.get_by_name("Kinh doanh")
        role_name = f"role-{username}"
        role = roles.get_by_name_and_department(role_name, kd.id)
        if role is None:
            role = roles.create(name=role_name, department_id=kd.id)
        roles.set_permission(
            role_id=role.id, module_key="luong", scope="all",
            can_read=can_read, can_view_salary=can_view_salary, can_update=can_update,
        )
        # Hai endpoint tương thích cũ từng chỉ gác theo Phòng ban/Nhân sự.
        roles.set_permission(
            role_id=role.id, module_key="phong_ban", scope="all", can_read=True,
        )
        roles.set_permission(
            role_id=role.id, module_key="nhan_su", scope="all", can_read=True,
        )
        user = users.get_by_username(username)
        if user is None:
            user = users.create(username=username, name=username, password_hash=hash_password("x"))
        users.set_assignment(user, department_id=kd.id, role_id=role.id, is_active=True)
        return create_access_token(str(user.id))
    finally:
        db.close()


def _make_emp(client, token, *, name, status="active") -> int:
    body = {"full_name": name, "department_id": _dept_id("Hành chính nhân sự"),
            "hire_date": "2020-01-01", "gender": "male", "status": status}
    return client.post("/api/employees", json=body, headers=_h(token)).json()["employee"]["id"]


def _cfg_svc(db, dept_name):
    """Service đủ bộ (kèm DepartmentRepository) + 1 bộ phận trắng để khai cấu hình lương."""
    depts = DepartmentRepository(db)
    svc = PayrollService(PayrollRepository(db), EmployeeRepository(db),
                         attendance=None, departments=depts)
    return svc, depts.create(name=dept_name)


def _emp_ns(dept_id, **kw):
    base = dict(status="active", hire_date=date(2020, 1, 1), gender="male",
                payroll_group=None, pay_grade_key=None, dependents_count=0,
                department_id=dept_id)
    base.update(kw)
    return SimpleNamespace(**base)


def _salary_ns(**kw):
    base = dict(amount_mode="manual", base_amount=9_000_000, luong_vi_tri=0, luong_trach_nhiem=0,
                allowance=0, insurance_base=None, chuyen_can=0, source_salary_row_id=None,
                phu_cap_ca=0, phu_cap_tham_nien=0)
    base.update(kw)
    return SimpleNamespace(**base)


def test_params_employer_insurance_rates(client):
    """3 tỷ lệ phía NSDLĐ: mặc định 17.5/3/1, sửa được, KHÔNG trừ vào lương NV."""
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert p["bhxh_rate_er"] == 0.175 and p["bhyt_rate_er"] == 0.03 and p["bhtn_rate_er"] == 0.01
    upd = client.put("/api/luong/params", json={"bhxh_rate_er": 0.18}, headers=_h(token)).json()
    assert upd["bhxh_rate_er"] == 0.18
    client.put("/api/luong/params", json={"bhxh_rate_er": 0.175}, headers=_h(token))

    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        v = svc._compute(employee=_emp_ns(None), salary=_salary_ns(luong_vi_tri=10_000_000),
                         params=svc.get_params(), actual_cong=26, standard_cong=26,
                         on=date(2026, 6, 1))
        assert v["bhxh"] == round(10_000_000 * 0.105)   # NV vẫn chỉ đóng 10.5% (trên lương vị trí)
    finally:
        db.close()


# --- Ghi đè 2 cấp: NV → tổ (chuyên cần) -------------------------------------


def test_chuyen_can_tien_chi_khai_o_ho_so_nv(client):
    """Chuyên cần (chủ chốt 2026-07-23): TIỀN chỉ khai ở HỒ SƠ NV; tổ chỉ còn CÔNG TẮC.
    Bỏ mức tiền cấp tổ + bỏ mức mặc định công ty — trước đây bật ở tổ mà bỏ trống ô tiền thì
    màn hình báo 0đ nhưng engine vẫn trả 300k (số trên màn lệch tiền thật)."""
    client
    db = SessionLocal()
    try:
        svc, dept = _cfg_svc(db, "Tổ 3 cấp")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)

        def cc(chuyen_can_nv=0):
            return svc._compute(employee=emp, salary=_salary_ns(chuyen_can=chuyen_can_nv),
                                params=params, actual_cong=26, standard_cong=26,
                                on=on)["chuyen_can"]

        assert cc() == 0                                    # chưa khai ở hồ sơ → 0đ (KHÔNG rơi về 300k)
        assert cc(chuyen_can_nv=600_000) == 600_000         # khai ở hồ sơ NV → dùng số đó

        # Số tiền khai ở TỔ (dữ liệu cũ) KHÔNG còn tác dụng — tổ chỉ bật/tắt.
        svc.set_dept_components(department_id=dept.id, items=[
            {"component_key": "chuyen_can", "is_enabled": True, "value": 450_000}])
        assert cc() == 0
        assert cc(chuyen_can_nv=600_000) == 600_000

        svc.set_dept_components(department_id=dept.id, items=[
            {"component_key": "chuyen_can", "is_enabled": False, "value": 450_000}])
        assert cc() == 0 and cc(chuyen_can_nv=600_000) == 0  # tổ TẮT → NV khai rồi cũng mất
    finally:
        db.close()


def test_manual_allowances_add_flat(client):
    """Phụ cấp thâm niên · khác KHAI TAY theo từng NV, một số cố định — engine cộng PHẲNG (không
    prorate theo công, không vào gốc tính tăng ca).

    ⚠️ `phu_cap_ca` KHÔNG còn ra tiền từ 03/08/2026: phụ cấp cơm/ca nay tính theo CA THỰC LÀM
    (`work_shifts.meal_allowance` / `.shift_allowance`). Đường per-người phải tắt CÙNG LƯỢT với
    việc bật đường theo ca — để cả hai cùng chạy là TRẢ HAI LẦN."""
    client
    db = SessionLocal()
    try:
        svc, dept = _cfg_svc(db, "Tổ khai tay")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        # Mức nền = vị trí 20tr + trách nhiệm 6tr = 26tr; phụ cấp khai tay: khác 400k, ca 1.2M, TN 900k.
        sal = _salary_ns(base_amount=None, luong_vi_tri=20_000_000, luong_trach_nhiem=6_000_000,
                         allowance=400_000, phu_cap_ca=1_200_000, phu_cap_tham_nien=900_000)
        kw = dict(employee=emp, salary=sal, params=params, standard_cong=26, on=on)

        full = svc._compute(actual_cong=26, **kw)
        assert full["monthly_salary"] == 26_000_000                  # vị trí + trách nhiệm
        assert full["allowance"] == 400_000 + 900_000                # khác + thâm niên
        assert full["phu_cap_tham_nien"] == 900_000
        # ⭐ Số `phu_cap_ca` cũ của hồ sơ KHÔNG được chảy vào lương nữa — đây là chốt chống trả
        # hai lần khi đường "phụ cấp theo ca" đã bật.
        assert full["night_pay"] == 0
        # Chuyên cần = 0 vì hồ sơ NV chưa khai (từ 2026-07-23 không còn mức mặc định công ty).
        assert full["gross"] == (26_000_000 + 1_300_000)

        # Nửa công: lương công prorate, phụ cấp khai tay giữ NGUYÊN số (cộng phẳng).
        half = svc._compute(actual_cong=13, **kw)
        assert half["luong_cong"] == 13_000_000
        assert half["allowance"] == 1_300_000 and half["night_pay"] == 0

        # Tăng ca bám LƯƠNG VỊ TRÍ (20tr), KHÔNG bám mức nền 26tr — chủ chốt 12/08/2026.
        # Phụ cấp khai tay vẫn không làm tiền tăng ca nhảy (vế cũ, giữ nguyên).
        base_only = _salary_ns(base_amount=None, luong_vi_tri=20_000_000, luong_trach_nhiem=6_000_000)
        ot_no_pc = svc._compute(actual_cong=26, ot_minutes=120, **dict(kw, salary=base_only))
        ot_pc = svc._compute(actual_cong=26, ot_minutes=120, **kw)
        assert ot_pc["ot_pay"] == ot_no_pc["ot_pay"] == 288_462   # 20tr/26/8 × 2h × 1,5
        # Chấm công có ca đêm nhưng KHÔNG khai phụ cấp ca → 0đ (hệ thống không tự tính).
        no_ca = svc._compute(actual_cong=26, night_days=5, **dict(kw, salary=base_only))
        assert no_ca["night_pay"] == 0 and no_ca["night_days"] == 5
    finally:
        db.close()


def test_manual_allowances_roundtrip_through_api(client):
    """Phụ cấp khai ở màn Lương nhân viên → lưu, preview đọc lại, ra đúng tiền trên bảng lương
    (phụ cấp ca đi vào `night_pay`/`ca_pay`). Mức đóng BH = lương vị trí."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Phụ cấp tay")
    res = client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-01-01", "luong_vi_tri": 8_000_000,
        "luong_trach_nhiem": 2_000_000, "allowance": 300_000, "phu_cap_ca": 1_500_000,
        "phu_cap_tham_nien": 600_000}, headers=_h(token))
    assert res.status_code == 201, res.text
    body = res.json()
    assert body["phu_cap_ca"] == 1_500_000 and body["phu_cap_tham_nien"] == 600_000

    prev = client.get(f"/api/luong/salaries/{eid}/preview", headers=_h(token)).json()
    assert prev["phu_cap_ca"] == 1_500_000 and prev["phu_cap_tham_nien"] == 600_000
    assert prev["insurance_base"] == 10_000_000      # vị trí 8tr + trách nhiệm 2tr

    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 11},
                      headers=_h(token)).json()
    ln = next(l for l in gen["lines"] if l["employee_id"] == eid)
    # ⭐ Hồ sơ VẪN GIỮ số `phu_cap_ca` (kiểm ở trên) nhưng nó KHÔNG còn ra tiền từ 03/08/2026 —
    # phụ cấp cơm/ca tính theo CA THỰC LÀM. Giữ cột để không mất lịch sử, tắt đường tiền để không
    # trả hai lần.
    assert ln["night_pay"] == 0 and ln["ca_pay"] == 0
    assert ln["allowance"] == 300_000 + 600_000
    assert ln["phu_cap_khac"] == 300_000
    assert ln["insurance_base"] == 10_000_000        # vị trí 8tr + trách nhiệm 2tr
    # Sửa số → kỳ draft đổi theo (khai lại bản hiệu lực mới rồi Tính lại).
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-11-01", "luong_vi_tri": 8_000_000,
        "luong_trach_nhiem": 2_000_000, "allowance": 300_000, "phu_cap_ca": 0,
        "phu_cap_tham_nien": 600_000}, headers=_h(token))
    gen2 = client.post("/api/luong/generate", json={"year": 2026, "month": 11},
                       headers=_h(token)).json()
    l2 = next(l for l in gen2["lines"] if l["employee_id"] == eid)
    # ⭐ Đổi `phu_cap_ca` từ 1.500.000 về 0 mà GROSS KHÔNG NHÚC NHÍCH — bằng chứng ô này đã thật
    # sự ngắt khỏi đường tiền, không phải chỉ hiện 0 trên một cột.
    assert l2["night_pay"] == 0 and l2["gross"] == ln["gross"]


def test_removed_shift_rate_and_allowance_type_endpoints_are_gone(client):
    """Đơn giá ca + danh mục phụ cấp công ty + THANG BẬC LƯƠNG + ghi chú chính sách đã GỠ HẲN
    — không còn route nào."""
    token = _admin_token(client)
    dept_id = _dept_id("Kinh doanh")
    # GET các route đã gỡ → 404.
    for url in (f"/api/luong/shift-rates/{dept_id}", "/api/luong/allowance-types",
                f"/api/luong/salary-rows/{dept_id}", f"/api/luong/salary-policy-note/{dept_id}",
                f"/api/departments/{dept_id}/salary-rows",
                f"/api/employees/meta/salary-rows/{dept_id}"):
        assert client.get(url, headers=_h(token)).status_code == 404, url
    assert client.put(f"/api/luong/shift-rates/{dept_id}", json={"items": []},
                      headers=_h(token)).status_code == 404
    assert client.post(f"/api/luong/salary-rows/{dept_id}", json={"label": "X"},
                       headers=_h(token)).status_code == 404
    # Component key đã bỏ (phụ cấp khai tay + luong_bac) → validate chặn ngay ở schema.
    for key in ("phu_cap_tham_nien", "phu_cap_trach_nhiem", "phu_cap_ca_dem", "luong_bac"):
        assert client.put(f"/api/luong/dept-components/{dept_id}", json={"items": [
            {"component_key": key, "is_enabled": True, "value": 1}]},
            headers=_h(token)).status_code == 422, key


# --- KPI ---------------------------------------------------------------------


# Hai test KPI (thưởng năng suất) ĐÃ XOÁ 29/07/2026 cùng tính năng — chủ: "xưởng không dùng tới,
# xóa backend luôn, đang phát triển mà chưa chạy thật đâu". Việc gỡ được canh bằng
# `test_go_kpi_migration.py` (hãm không cho drop cột khi còn tiền) + hồi quy: `kpi_bonus` vốn luôn
# bằng 0 nên bỏ nó khỏi công thức KHÔNG được đổi một đồng nào ở các test lương còn lại.


# --- vá lệch "Sửa 1 ô" vs "Tính lại" ----------------------------------------


def test_update_line_matches_generate_after_edit(client):
    """Sửa 1 ô và Tính lại phải ra CÙNG gross/net/công đoàn (chung `_capped_penalty`)."""
    token = _admin_token(client)
    client.put("/api/luong/params", json={"cong_doan_rate": 0.005}, headers=_h(token))
    try:
        eid = _make_emp(client, token, name="NV Lệch")
        # union_member=True: đoàn viên → có trừ đoàn phí (mặc định false thì cong_doan = 0).
        client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                    "luong_vi_tri": 20_000_000, "union_member": True}, headers=_h(token))
        gen = client.post("/api/luong/generate", json={"year": 2026, "month": 10},
                          headers=_h(token)).json()
        line = next(l for l in gen["lines"] if l["employee_id"] == eid)
        # phạt khủng → chạm trần 30% Đ102 ở CẢ hai đường tính.
        edited = client.put(f"/api/luong/lines/{line['id']}",
                            json={"vi_pham": 30_000_000, "di_tre": 500_000},
                            headers=_h(token)).json()
        gen2 = client.post("/api/luong/generate", json={"year": 2026, "month": 10},
                           headers=_h(token)).json()
        l2 = next(l for l in gen2["lines"] if l["employee_id"] == eid)
        for k in ("gross", "net_pay", "pit", "cong_doan", "bhxh"):
            assert edited[k] == l2[k], f"lệch ở {k}: sửa 1 ô {edited[k]} vs tính lại {l2[k]}"
        assert edited["cong_doan"] == round(edited["insurance_base"] * 0.005) > 0
    finally:
        client.put("/api/luong/params", json={"cong_doan_rate": 0}, headers=_h(token))


def _nv_co_ca_lam(client, token, *, name, ca_ten, com, ca, ngay=15):
    """NV làm ĐÚNG MỘT ngày ca có mức cơm + phụ cấp ca → dòng lương chắc chắn có 2 khoản đó.

    Test canh `test_update_line_matches_generate_after_edit` ở trên dùng NV KHÔNG làm ca nào, nên
    `meal_allowance_pay`/`shift_allowance_pay` = 0 ở cả hai vế và nó không thể bắt được lệch của
    hai cột này. Phải có ca thật mới có răng — đó là lý do helper này tồn tại.
    """
    from datetime import datetime as _dt, timezone as _tz
    from app.repositories.attendance_repo import AttendanceRepository

    eid = _make_emp(client, token, name=name)
    client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                "luong_vi_tri": 10_000_000}, headers=_h(token))
    db = SessionLocal()
    try:
        arepo = AttendanceRepository(db)
        shift = arepo.create_shift(name=ca_ten, start_minute=480, end_minute=1020,
                                   is_overnight=False, grace_minutes=5,
                                   meal_allowance=com, shift_allowance=ca)
        EmployeeRepository(db).get_by_id(eid).default_shift_id = shift.id
        db.commit()
        # 2026-06-15 là Thứ Hai. Vào 08:00 VN (01:00 UTC) đúng giờ, ra 17:00 VN (10:00 UTC) → đủ
        # 1 công, không đi trễ (nếu trễ thì `generate` tự điền ô phạt và làm nhiễu phép so sánh).
        for kind, hour in (("in", 1), ("out", 10)):
            arepo.create_log(employee_id=eid, check_type=kind,
                             checked_at=_dt(2026, 6, ngay, hour, 0, tzinfo=_tz.utc),
                             within_range=True)
    finally:
        db.close()
    return eid


def _tinh_lai(client, token, eid):
    """Chạy lại `generate` rồi lấy dòng của NV — đây là đường TÍNH LẠI."""
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 6},
                      headers=_h(token)).json()
    return next(l for l in gen["lines"] if l["employee_id"] == eid)


def _doc_dong(client, token, eid):
    """ĐỌC dòng đang lưu, KHÔNG tính lại.

    Quan trọng: `generate` tính lại từ đầu nên nó XOÁ DẤU VẾT của bệnh "sửa 1 ô làm hụt tiền" —
    dùng nó để kiểm tra sau khi sửa là tự chữa lành, test xanh mà bệnh còn nguyên.
    """
    tbl = client.get("/api/luong/table", params={"year": 2026, "month": 6},
                     headers=_h(token)).json()
    return next(l for l in tbl["lines"] if l["employee_id"] == eid)


def test_sua_1_o_khong_duoc_lam_bay_com_ca_va_phu_cap_ca(client):
    """⭐ NV CÓ ca làm: sửa một ô rồi Tính lại phải ra CÙNG số.

    Bệnh thật (03/08/2026): `_compute` cộng `meal_allowance_pay + shift_allowance_pay` vào gross
    còn `update_line` thì không ⇒ chạm vào dòng một cái là hai khoản bay mất, trong khi phiếu
    lương vẫn in đủ hai dòng nên cộng lại không ra tổng.
    """
    token = _admin_token(client)
    eid = _nv_co_ca_lam(client, token, name="NV Có Ca", ca_ten="HC cơm", com=30_000, ca=50_000)
    line = _tinh_lai(client, token, eid)
    # Chốt răng: hỏng dữ liệu dựng thì test này thành vô nghĩa mà vẫn xanh.
    assert line["meal_allowance_pay"] == 30_000 and line["shift_allowance_pay"] == 50_000

    edited = client.put(f"/api/luong/lines/{line['id']}",
                        json={"vi_pham": 100_000}, headers=_h(token)).json()
    l2 = _tinh_lai(client, token, eid)
    for k in ("gross", "net_pay", "thu_nhap_chiu_thue", "thu_nhap_mien_thue", "pit"):
        assert edited[k] == l2[k], f"lệch ở {k}: sửa 1 ô {edited[k]} vs tính lại {l2[k]}"


def test_them_khoan_phat_sinh_khong_duoc_lam_tut_gross(client):
    """Đường người dùng hay đi nhất: thêm một khoản thưởng nóng. `add_line_component` gọi
    `_recompute_line` → `update_line`, nên nó dính CÙNG bệnh với test trên."""
    token = _admin_token(client)
    eid = _nv_co_ca_lam(client, token, name="NV Thưởng Nóng", ca_ten="HC thưởng",
                        com=25_000, ca=40_000)
    line = _tinh_lai(client, token, eid)
    truoc = line["gross"]

    comps = client.get("/api/luong/components", headers=_h(token)).json()["items"]
    khoan_thu = next(c for c in comps if c["kind"] != "tru")
    res = client.post(f"/api/luong/lines/{line['id']}/components",
                      json={"component_id": khoan_thu["id"], "amount": 200_000},
                      headers=_h(token))
    assert res.status_code == 201, res.text

    # ĐỌC, không tính lại — xem `_doc_dong`.
    sau = _doc_dong(client, token, eid)["gross"]
    assert sau == truoc + 200_000, (
        f"thêm 200.000đ mà gross đi từ {truoc} sang {sau} — cơm ca + phụ cấp ca bị nuốt mất")


def test_xuat_excel_co_cot_com_ca_va_phu_cap_ca(client):
    """File xuất liệt kê từng khoản rồi tới cột "Tổng" — thiếu cột nào là kế toán dò mãi không ra
    chênh ở đâu. Hai khoản theo ca từng bị bỏ quên đúng kiểu đó."""
    from io import BytesIO
    from openpyxl import load_workbook

    token = _admin_token(client)
    eid = _nv_co_ca_lam(client, token, name="NV Xuất Ca", ca_ten="HC xuất",
                        com=35_000, ca=45_000, ngay=16)
    _tinh_lai(client, token, eid)

    r = client.get("/api/luong/export.xlsx?year=2026&month=6", headers=_h(token))
    assert r.status_code == 200, r.text
    ws = load_workbook(BytesIO(r.content)).active
    head = [c.value for c in ws[1]]
    row = next(x for x in ws.iter_rows(min_row=2, values_only=True) if x[1] == "NV Xuất Ca")
    assert row[head.index("Cơm ca")] == 35_000
    assert row[head.index("Phụ cấp ca")] == 45_000


# --- Tab 2 (đọc/ghi thành phần) + Tab 1 (điều kiện thăng bậc) ---------------


def test_dept_components_default_view_and_rbac(client):
    """GET trả đủ 5 khoản CÒN khai theo tổ; gác quyền xem cấu hình."""
    token = _admin_token(client)
    dept_id = _dept_id("Kinh doanh")
    res = client.get(f"/api/luong/dept-components/{dept_id}", headers=_h(token))
    assert res.status_code == 200
    items = {c["component_key"]: c for c in res.json()["items"]}
    # Bậc lương gỡ hẳn (luong_bac) + 3 phụ cấp chuyển sang KHAI TAY theo NV + KPI xoá 29/07/2026
    # (chủ: xưởng không chấm KPI) → còn 3 khoản.
    assert set(items) == {"chuyen_can", "luong_khoan", "tang_ca"}
    assert items["chuyen_can"]["is_set"] is False
    assert items["luong_khoan"]["is_enabled"] is False   # soi cờ departments.has_piece_work
    assert client.get(f"/api/luong/dept-components/{dept_id}",
                      headers=_h(_sales_token())).status_code == 403
    assert client.put(f"/api/luong/dept-components/{dept_id}",
                      json={"items": [{"component_key": "chuyen_can", "is_enabled": True}]},
                      headers=_h(_sales_token())).status_code == 403
    assert client.put("/api/luong/dept-components/999999",
                      json={"items": [{"component_key": "chuyen_can", "is_enabled": True}]},
                      headers=_h(token)).status_code == 404


def test_salary_config_is_hidden_from_luong_read_only_but_self_service_stays_open(client):
    """Quyền Xem module không được làm lộ cấu hình; cờ nhạy cảm hoặc update mới đọc được.

    Khóa cả ba đường: `/api/luong`, endpoint Phòng ban cũ và endpoint meta Nhân sự.
    Hai endpoint ``/me`` vẫn chỉ dựa vào tài khoản/hồ sơ của chính người đăng nhập.
    """
    dept_id = _dept_id("Kinh doanh")
    read_only = _luong_config_token("luong-read-no-config", can_read=True)
    protected_paths = [
        "/api/luong/params",
        "/api/luong/rules",
        "/api/luong/pit-brackets",
        f"/api/luong/dept-components/{dept_id}",
    ]
    for path in protected_paths:
        response = client.get(path, headers=_h(read_only))
        assert response.status_code == 403, (path, response.text)

    # Tự phục vụ không cần luong:read/view_salary; chưa gắn hồ sơ thì trả payload rỗng an toàn.
    no_luong_permission = _sales_token()
    assert client.get("/api/luong/payslip/me", headers=_h(no_luong_permission)).status_code == 200
    assert client.get("/api/luong/advances/me", headers=_h(no_luong_permission)).status_code == 200

    config_viewer = _luong_config_token(
        "luong-config-viewer", can_view_salary=True,
    )
    for path in protected_paths:
        response = client.get(path, headers=_h(config_viewer))
        assert response.status_code == 200, (path, response.text)

    # Tương thích vai trò cũ: người có quyền Sửa luôn được đọc cấu hình.
    updater = _luong_config_token("luong-config-updater", can_update=True)
    assert client.get("/api/luong/params", headers=_h(updater)).status_code == 200


# --- bảng phạt đi trễ / về sớm (mirror biểu TNCN) ---------------------------


def test_late_penalty_brackets_seeded_and_editable(client):
    """Bảng phạt trễ/sớm auto-seed 4 bậc mặc định (20k/40k/100k/150k) + đọc/sửa/thêm/xóa được."""
    token = _admin_token(client)
    items = client.get("/api/luong/late-penalty-brackets", headers=_h(token)).json()["items"]
    assert len(items) == 4
    assert [i["up_to_minute"] for i in items] == [15, 30, 60, None]   # phút; bậc cuối = ∞
    assert [i["amount"] for i in items] == [20000, 40000, 100000, 150000]

    # Sửa 1 bậc.
    bid = items[0]["id"]
    upd = client.put(f"/api/luong/late-penalty-brackets/{bid}",
                     json={"seq": 1, "up_to_minute": 10, "amount": 25000}, headers=_h(token))
    assert upd.status_code == 200 and upd.json()["up_to_minute"] == 10 and upd.json()["amount"] == 25000

    # Thêm rồi xóa 1 bậc (CRUD đủ vòng).
    created = client.post("/api/luong/late-penalty-brackets",
                          json={"seq": 5, "up_to_minute": 120, "amount": 200000}, headers=_h(token))
    assert created.status_code == 201
    assert client.delete(f"/api/luong/late-penalty-brackets/{created.json()['id']}",
                         headers=_h(token)).status_code == 204
    # Trả bậc 1 về mặc định để không nhiễu test khác.
    client.put(f"/api/luong/late-penalty-brackets/{bid}",
               json={"seq": 1, "up_to_minute": 15, "amount": 20000}, headers=_h(token))


def test_late_penalty_brackets_rbac(client):
    """Gác quyền y hệt biểu TNCN: `luong:read` đơn lẻ KHÔNG xem được cấu hình nhạy cảm;
    view_salary/update mới đọc; ghi cần `luong:update`. Không có module luong → 403."""
    # Không có quyền module luong → 403 cả đọc lẫn ghi.
    sales = _sales_token()
    assert client.get("/api/luong/late-penalty-brackets", headers=_h(sales)).status_code == 403
    assert client.post("/api/luong/late-penalty-brackets",
                       json={"seq": 1, "up_to_minute": 5, "amount": 1000},
                       headers=_h(sales)).status_code == 403

    # luong:read đơn lẻ KHÔNG lộ cấu hình (giống pit-brackets).
    read_only = _luong_config_token("late-penalty-read", can_read=True)
    assert client.get("/api/luong/late-penalty-brackets", headers=_h(read_only)).status_code == 403

    # view_salary đọc được nhưng KHÔNG ghi được.
    viewer = _luong_config_token("late-penalty-viewer", can_view_salary=True)
    assert client.get("/api/luong/late-penalty-brackets", headers=_h(viewer)).status_code == 200
    assert client.post("/api/luong/late-penalty-brackets",
                       json={"seq": 9, "up_to_minute": 5, "amount": 1000},
                       headers=_h(viewer)).status_code == 403

    # update ghi được (rồi dọn dấu vết).
    updater = _luong_config_token("late-penalty-updater", can_update=True)
    created = client.post("/api/luong/late-penalty-brackets",
                          json={"seq": 9, "up_to_minute": 5, "amount": 1000}, headers=_h(updater))
    assert created.status_code == 201
    client.delete(f"/api/luong/late-penalty-brackets/{created.json()['id']}", headers=_h(updater))


def test_luong_khoan_component_mirrors_department_flag(client):
    """`luong_khoan` chỉ phơi lại cờ `departments.has_piece_work` — không dựng nguồn thứ 2."""
    token = _admin_token(client)
    dept_id = _dept_id("Kinh doanh")
    client.put(f"/api/luong/dept-components/{dept_id}", json={
        "items": [{"component_key": "luong_khoan", "is_enabled": True}]}, headers=_h(token))
    db = SessionLocal()
    try:
        assert DepartmentRepository(db).get_by_id(dept_id).has_piece_work is True
    finally:
        db.close()
    client.put(f"/api/luong/dept-components/{dept_id}", json={
        "items": [{"component_key": "luong_khoan", "is_enabled": False}]}, headers=_h(token))
    db = SessionLocal()
    try:
        assert DepartmentRepository(db).get_by_id(dept_id).has_piece_work is False
    finally:
        db.close()


def test_khoan_va_tang_ca_DOC_LAP(client):
    """⚠️ ĐẢO 17/08/2026 — Khoán ⟷ Tăng ca KHÔNG còn loại trừ nhau, bật CẢ HAI được.

    Chủ chốt: *"Tổ khoán vẫn có tăng ca"*, đảo luật loại trừ ngày 22/07/2026. Engine cũng đã gỡ
    vế `has_piece_work` khỏi `ot_pay`. Test này canh không ai dựng lại luật loại trừ."""
    token = _admin_token(client)
    dept_id = _dept_id("Kinh doanh")

    def _state():
        items = {c["component_key"]: c for c in client.get(
            f"/api/luong/dept-components/{dept_id}", headers=_h(token)).json()["items"]}
        return items["luong_khoan"]["is_enabled"], items["tang_ca"]["is_enabled"]

    # Bật CẢ hai → GIỮ NGUYÊN cả hai, backend KHÔNG ép tắt cái nào.
    client.put(f"/api/luong/dept-components/{dept_id}", json={"items": [
        {"component_key": "luong_khoan", "is_enabled": True},
        {"component_key": "tang_ca", "is_enabled": True},
    ]}, headers=_h(token))
    assert _state() == (True, True)

    # Tắt khoán + bật tăng ca → vẫn đúng như khai.
    client.put(f"/api/luong/dept-components/{dept_id}", json={"items": [
        {"component_key": "luong_khoan", "is_enabled": False},
        {"component_key": "tang_ca", "is_enabled": True},
    ]}, headers=_h(token))
    assert _state() == (False, True)


# --- migration 0076 (DB CŨ → có cột mới, idempotent) -------------------------


def _legacy_db():
    """DB hình dạng TRƯỚC migration 0076 (thiếu promotion_condition / *_er / kpi_*)."""
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False},
                        poolclass=StaticPool)
    db = sessionmaker(bind=eng)()
    db.execute(text(
        "CREATE TABLE department_salary_rows ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT, department_id INTEGER NOT NULL,"
        " label VARCHAR(120) NOT NULL, luong_vi_tri NUMERIC(14,2) NOT NULL DEFAULT 0)"
    ))
    db.execute(text(
        "CREATE TABLE payroll_params ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " bhxh_rate NUMERIC(6,4) NOT NULL DEFAULT 0.08,"
        " chuyen_can_default NUMERIC(14,2) NOT NULL DEFAULT 300000)"
    ))
    db.execute(text("INSERT INTO payroll_params (bhxh_rate) VALUES (0.08)"))
    db.execute(text(
        "CREATE TABLE payroll_lines ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT, period_id INTEGER NOT NULL,"
        " employee_id INTEGER NOT NULL, gross NUMERIC(14,2) NOT NULL DEFAULT 0)"
    ))
    db.execute(text("INSERT INTO payroll_lines (period_id, employee_id, gross) VALUES (1,1,5000000)"))
    db.commit()
    return db


def test_migration_0076_adds_columns_with_defaults():
    db = _legacy_db()
    _migrate_cau_hinh_luong(db)
    insp = inspect(db.get_bind())
    assert "promotion_condition" in {c["name"] for c in insp.get_columns("department_salary_rows")}
    pcols = {c["name"] for c in insp.get_columns("payroll_params")}
    assert {"bhxh_rate_er", "bhyt_rate_er", "bhtn_rate_er"} <= pcols
    lcols = {c["name"] for c in insp.get_columns("payroll_lines")}
    assert {"kpi_percent", "kpi_bonus"} <= lcols
    # hàng CŨ nhận mặc định, không NULL (NOT NULL DEFAULT)
    row = db.execute(text("SELECT bhxh_rate_er, bhyt_rate_er, bhtn_rate_er FROM payroll_params")).first()
    assert [float(x) for x in row] == [0.175, 0.03, 0.01]
    ln = db.execute(text("SELECT kpi_percent, kpi_bonus FROM payroll_lines")).first()
    assert [float(x) for x in ln] == [0.0, 0.0]


def test_migration_0076_idempotent_and_noop_on_missing_tables():
    db = _legacy_db()
    _migrate_cau_hinh_luong(db)
    _migrate_cau_hinh_luong(db)     # chạy lại không lỗi
    assert "kpi_bonus" in {c["name"] for c in inspect(db.get_bind()).get_columns("payroll_lines")}
    # DB trắng (chưa có bảng nào) → no-op, không ném lỗi
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_cau_hinh_luong(empty)


# =============================================================================
# PRD v2 — bậc là KHUNG (C1) · mức riêng của NV (C2) · chuyên cần trừ dần (C3)
# =============================================================================


# --- C3: chuyên cần TRỪ DẦN, kiểm bằng SỐ THẬT của bảng lương T05.2026 -------


def test_chuyen_can_tru_dan_bang_so_that(client):
    """Nghiệm thu §12.2: mức 300.000 → 26 công 300.000 · 25,5 → 225.000 · 25 → 150.000 ·
    23 → 0. Mức 500.000 → 25 công = 250.000."""
    client
    db = SessionLocal()
    try:
        svc, dept = _cfg_svc(db, "Tổ chuyên cần")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 5, 1)

        def cc(cong, muc):
            return svc._compute(employee=emp, salary=_salary_ns(chuyen_can=muc), params=params,
                                actual_cong=cong, standard_cong=26, on=on)["chuyen_can"]

        assert cc(26, 300_000) == 300_000      # đủ công
        assert cc(25.5, 300_000) == 225_000    # nghỉ 0,5 ngày → −25%
        assert cc(25, 300_000) == 150_000      # nghỉ 1 ngày → −50%
        assert cc(24, 300_000) == 0            # nghỉ 2 ngày → mất hết
        assert cc(23, 300_000) == 0
        assert cc(25, 500_000) == 250_000      # quản lý cấp cao
        assert cc(27, 300_000) == 300_000      # làm dư công vẫn 100% (không nhân thêm)
    finally:
        db.close()


def test_allowance_split_visible_without_changing_totals(client):
    """B2: phiếu lương tách "Phụ cấp thâm niên" thành DÒNG RIÊNG khỏi "Phụ cấp khác" — mà TỔNG
    THU NHẬP y nguyên (2 dòng cộng lại đúng bằng `allowance`). Trách nhiệm KHÔNG ở đây (nó là
    `luong_trach_nhiem` trong mức nền)."""
    token = _admin_token(client)
    eid = _make_emp(client, token, name="NV Tách phụ cấp")
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-01-01", "luong_vi_tri": 8_000_000,
        "luong_trach_nhiem": 2_000_000, "allowance": 700_000}, headers=_h(token))

    # (1) Mới chỉ khai "phụ cấp khác" → thâm niên = 0.
    gen = client.post("/api/luong/generate", json={"year": 2026, "month": 12},
                      headers=_h(token)).json()
    before = next(l for l in gen["lines"] if l["employee_id"] == eid)
    assert before["allowance"] == 700_000 and before["phu_cap_khac"] == 700_000
    assert before["phu_cap_tham_nien"] == 0

    # (2) Khai TAY thêm thâm niên 1.950.000 cho CHÍNH NV này.
    client.post(f"/api/luong/salaries/{eid}", json={
        "effective_from": "2026-02-01", "luong_vi_tri": 8_000_000,
        "luong_trach_nhiem": 2_000_000, "allowance": 700_000,
        "phu_cap_tham_nien": 1_950_000}, headers=_h(token))
    gen2 = client.post("/api/luong/generate", json={"year": 2026, "month": 12},
                       headers=_h(token)).json()
    ln = next(l for l in gen2["lines"] if l["employee_id"] == eid)
    assert ln["phu_cap_tham_nien"] == 1_950_000
    assert ln["phu_cap_khac"] == 700_000
    # 2 dòng cộng lại = TỔNG phụ cấp; tổng thu nhập chỉ tính MỘT lần.
    assert (ln["phu_cap_khac"] + ln["phu_cap_tham_nien"]
            == ln["allowance"] == 700_000 + 1_950_000)
    assert ln["gross"] == before["gross"] + 1_950_000
    assert ln["net_pay"] == round(max(0.0, ln["gross"] - ln["bhxh"] - ln["cong_doan"]
                                      - ln["pit"] - ln["advance_total"]))

    # (3) Tính lại lần nữa: số không nhảy (idempotent, không cộng dồn).
    gen3 = client.post("/api/luong/generate", json={"year": 2026, "month": 12},
                       headers=_h(token)).json()
    l3 = next(l for l in gen3["lines"] if l["employee_id"] == eid)
    for k in ("allowance", "phu_cap_tham_nien", "gross", "net_pay"):
        assert l3[k] == ln[k], f"lệch ở {k}"


def test_migration_0089_adds_allowance_split_columns():
    """Cột tách phụ cấp: idempotent, dòng lương CŨ nhận 0 → `allowance` vẫn là tổng đúng."""
    db = _legacy_db()      # DB cũ có payroll_lines 1 dòng gross 5tr
    db.execute(text("ALTER TABLE payroll_lines ADD COLUMN allowance NUMERIC(14,2) NOT NULL DEFAULT 0"))
    db.execute(text("UPDATE payroll_lines SET allowance = 1200000"))
    db.commit()
    _migrate_payroll_line_allowance_split(db)
    _migrate_payroll_line_allowance_split(db)      # chạy lại không lỗi
    cols = {c["name"] for c in inspect(db.get_bind()).get_columns("payroll_lines")}
    assert {"phu_cap_trach_nhiem", "phu_cap_tham_nien"} <= cols
    row = db.execute(text(
        "SELECT allowance, phu_cap_trach_nhiem, phu_cap_tham_nien FROM payroll_lines")).first()
    assert [float(x) for x in row] == [1_200_000, 0, 0]   # tổng phụ cấp KHÔNG đổi
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_payroll_line_allowance_split(empty)   # DB trắng → no-op


def test_migration_0090_adds_manual_allowance_cols_and_drops_old_tables():
    """3 cột phụ cấp khai tay + gỡ 3 bảng của cách tính cũ. Idempotent, DB trắng no-op."""
    db = _legacy_db_v1()          # có employee_salaries kiểu cũ
    db.execute(text("CREATE TABLE department_shift_rates (id INTEGER PRIMARY KEY, department_id INTEGER)"))
    db.execute(text("CREATE TABLE payroll_line_shifts (id INTEGER PRIMARY KEY, payroll_line_id INTEGER)"))
    db.execute(text("CREATE TABLE allowance_types (id INTEGER PRIMARY KEY, key VARCHAR(40))"))
    db.execute(text(
        "CREATE TABLE department_salary_components ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT, department_id INTEGER NOT NULL,"
        " component_key VARCHAR(32) NOT NULL)"))
    for key in ("chuyen_can", "phu_cap_ca_dem", "phu_cap_trach_nhiem", "phu_cap_tham_nien"):
        db.execute(text("INSERT INTO department_salary_components (department_id, component_key)"
                        f" VALUES (1, '{key}')"))
    db.commit()

    _migrate_luong_phu_cap_khai_tay(db)
    _migrate_luong_phu_cap_khai_tay(db)      # chạy lại không lỗi
    insp = inspect(db.get_bind())
    cols = {c["name"] for c in insp.get_columns("employee_salaries")}
    assert {"phu_cap_ca", "phu_cap_trach_nhiem", "phu_cap_tham_nien"} <= cols
    tables = set(insp.get_table_names())
    assert not ({"department_shift_rates", "payroll_line_shifts", "allowance_types"} & tables)
    left = [r[0] for r in db.execute(text(
        "SELECT component_key FROM department_salary_components")).all()]
    assert left == ["chuyen_can"]            # 3 khoản đã chuyển về cấp NV bị dọn
    # Hàng cũ nhận 0, không NULL → không ai tự dưng có/mất phụ cấp.
    row = db.execute(text(
        "SELECT phu_cap_ca, phu_cap_trach_nhiem, phu_cap_tham_nien FROM employee_salaries")).first()
    assert [float(x) for x in row] == [0, 0, 0]
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_luong_phu_cap_khai_tay(empty)   # DB trắng → no-op


# --- migration 0088: GIỮ NGUYÊN lương của người cũ --------------------------


def _legacy_db_v1():
    """DB hình dạng TRƯỚC migration 0088: `employee_salaries` còn trỏ dòng bậc để lấy TIỀN."""
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False},
                        poolclass=StaticPool)
    db = sessionmaker(bind=eng)()
    db.execute(text(
        "CREATE TABLE department_salary_rows ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT, department_id INTEGER NOT NULL,"
        " label VARCHAR(120) NOT NULL,"
        " luong_vi_tri NUMERIC(14,2) NOT NULL DEFAULT 0,"
        " luong_trach_nhiem NUMERIC(14,2) NOT NULL DEFAULT 0)"
    ))
    db.execute(text(
        "INSERT INTO department_salary_rows (department_id, label, luong_vi_tri, luong_trach_nhiem)"
        " VALUES (1, 'Thợ bậc 2', 6000000, 2000000)"
    ))
    db.execute(text(
        "CREATE TABLE employee_salaries ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER NOT NULL,"
        " effective_from DATE NOT NULL, amount_mode VARCHAR(8) NOT NULL DEFAULT 'rule',"
        " base_amount NUMERIC(14,2), source_salary_row_id INTEGER)"
    ))
    db.execute(text(
        "INSERT INTO employee_salaries (employee_id, effective_from, amount_mode, source_salary_row_id)"
        " VALUES (7, '2026-01-01', 'dept_row', 1)"
    ))
    db.execute(text(
        "INSERT INTO employee_salaries (employee_id, effective_from, amount_mode, base_amount)"
        " VALUES (8, '2026-01-01', 'manual', 12500000)"
    ))
    db.commit()
    return db


def test_migration_0088_keeps_every_salary_unchanged():
    """YÊU CẦU SỐ 1: sau khi tách bậc khỏi tiền, lương từng người KHÔNG ĐỔI —
    số của dòng bậc được copy xuống chính bản ghi NV, bậc chuyển sang `pay_grade_row_id`."""
    db = _legacy_db_v1()
    _migrate_luong_v2_khung_bac(db)
    insp = inspect(db.get_bind())
    assert {"luong_min", "luong_max"} <= {c["name"] for c in insp.get_columns("department_salary_rows")}
    scols = {c["name"] for c in insp.get_columns("employee_salaries")}
    assert {"luong_vi_tri", "luong_trach_nhiem", "pay_grade_row_id"} <= scols

    row = db.execute(text(
        "SELECT luong_vi_tri, luong_trach_nhiem, pay_grade_row_id, source_salary_row_id"
        " FROM employee_salaries WHERE employee_id = 7")).first()
    assert [float(row[0]), float(row[1])] == [6_000_000, 2_000_000]   # = 8tr y như trước
    assert row[2] == 1 and row[3] == 1                                # bậc giữ, ref cũ giữ
    # NV khai tay: KHÔNG bị đụng (vẫn 12,5tr qua base_amount, không gán bậc).
    manual = db.execute(text(
        "SELECT luong_vi_tri, luong_trach_nhiem, pay_grade_row_id, base_amount"
        " FROM employee_salaries WHERE employee_id = 8")).first()
    assert [float(manual[0]), float(manual[1])] == [0, 0]
    assert manual[2] is None and float(manual[3]) == 12_500_000


def test_migration_0088_idempotent_and_noop_on_missing_tables():
    db = _legacy_db_v1()
    _migrate_luong_v2_khung_bac(db)
    # Admin sửa mức sau migration → chạy lại KHÔNG được đè số.
    db.execute(text("UPDATE employee_salaries SET luong_vi_tri = 9000000 WHERE employee_id = 7"))
    db.commit()
    _migrate_luong_v2_khung_bac(db)
    kept = db.execute(text(
        "SELECT luong_vi_tri FROM employee_salaries WHERE employee_id = 7")).scalar_one()
    assert float(kept) == 9_000_000
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_luong_v2_khung_bac(empty)   # DB trắng → no-op


# --- Chủ đảo phần bậc lương (2026-07-20): bỏ bậc + BH = lương vị trí ---------


def test_muc_dong_bh_gom_ca_luong_trach_nhiem(client):
    """ĐẢO LUẬT 12/08/2026: mức đóng BH = **lương cơ bản + lương trách nhiệm**.

    Chốt cũ (20/07/2026) là "chỉ lương vị trí". Chủ chốt đưa bảng lương thật đang dùng, soi ngược
    ra gốc: BH bắt buộc 1.102.080 ÷ 10,5% = 10.496.000 — và đoàn phí 52.480 ÷ 0,5% ra CÙNG con số
    đó. Tức cả hai bám mức nền ĐẦY ĐỦ, không phải riêng lương cơ bản.

    Đổi luật này làm NLĐ có trách nhiệm **bị trừ BHXH nhiều hơn** — nên phải có test ghim lại,
    đừng để ai "sửa về như cũ" vì tưởng là hồi quy."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        emp = _emp_ns(None)
        sal = _salary_ns(base_amount=None, luong_vi_tri=6_841_000, luong_trach_nhiem=5_000_000)
        v = svc._compute(employee=emp, salary=sal, params=params, actual_cong=26,
                         standard_cong=26, on=date(2026, 6, 1))
        assert v["insurance_base"] == 11_841_000, "phải là vị trí + trách nhiệm"
        assert v["bhxh"] == round(11_841_000 * (0.08 + 0.015 + 0.01))

        # ĐỔI TRÁCH NHIỆM LÀ ĐỔI MỨC ĐÓNG — vế ngược hẳn với luật cũ.
        sal2 = _salary_ns(base_amount=None, luong_vi_tri=6_841_000, luong_trach_nhiem=9_000_000)
        v2 = svc._compute(employee=emp, salary=sal2, params=params, actual_cong=26,
                          standard_cong=26, on=date(2026, 6, 1))
        assert v2["insurance_base"] == 15_841_000
        assert v2["bhxh"] > v["bhxh"], "tăng trách nhiệm mà mức đóng đứng yên là luật chưa đổi"

        # Thử việc vẫn KHÔNG đóng BH — luật cũ giữ nguyên.
        tv = svc._compute(employee=_emp_ns(None, status="probation"), salary=sal, params=params,
                          actual_cong=26, standard_cong=26, on=date(2026, 6, 1))
        assert tv["insurance_base"] == 0 and tv["bhxh"] == 0
    finally:
        db.close()


def test_don_gia_tang_ca_chi_bam_luong_co_ban(client):
    """Tăng ca tính trên LƯƠNG CƠ BẢN, bỏ lương trách nhiệm (chủ chốt 12/08/2026).

    Chủ chốt xác nhận premium ca đêm và premium làm ngày nghỉ/lễ **giảm cả** — ba khoản dùng
    chung một đơn giá giờ. Nhưng LƯƠNG THEO CÔNG thì KHÔNG đổi: nó vẫn ăn mức nền đầy đủ."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        emp = _emp_ns(None)
        chung = dict(employee=emp, params=params, actual_cong=26, standard_cong=26,
                     on=date(2026, 6, 1), ot_minutes=600)          # 10 giờ tăng ca ngày thường

        # 8tr vị trí + 2tr trách nhiệm: đơn giá giờ phải bám 8tr.
        v = svc._compute(salary=_salary_ns(base_amount=None, luong_vi_tri=8_000_000,
                                           luong_trach_nhiem=2_000_000), **chung)
        gio = 8_000_000 / 26 / float(params.standard_hours_per_day)
        assert v["ot_pay"] == round(gio * 10 * float(params.ot_multiplier))

        # Cùng mức nền 10tr nhưng KHÔNG có trách nhiệm ⇒ tăng ca CAO HƠN hẳn.
        v_khong_tn = svc._compute(salary=_salary_ns(base_amount=None, luong_vi_tri=10_000_000,
                                                    luong_trach_nhiem=0), **chung)
        assert v_khong_tn["ot_pay"] > v["ot_pay"], (
            "hai người cùng mức nền 10tr mà tăng ca bằng nhau ⇒ đơn giá vẫn bám mức nền"
        )
        # Lương theo công thì KHÔNG được đổi — cùng mức nền, cùng công ⇒ bằng nhau.
        assert v["luong_cong"] == v_khong_tn["luong_cong"]
    finally:
        db.close()


def test_doan_phi_giam_thu_nhap_tinh_thue(client):
    """Đoàn phí TRỪ TRƯỚC THUẾ (chủ chốt 12/08/2026, theo bảng lương thật của công ty).

    Ghi để khỏi bàn lại: TT 111/2013 Đ9 không liệt đoàn phí vào danh sách giảm trừ — đây là cố ý
    làm theo cách công ty hạch toán."""
    client
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        svc.update_params(actor=SimpleNamespace(id=1), cong_doan_rate=0.005)
        params = svc.get_params()
        emp = _emp_ns(None)
        chung = dict(employee=emp, params=params, actual_cong=26, standard_cong=26,
                     on=date(2026, 6, 1))
        luong = dict(base_amount=None, luong_vi_tri=60_000_000, luong_trach_nhiem=0)

        doan_vien = svc._compute(salary=_salary_ns(union_member=True, **luong), **chung)
        khong = svc._compute(salary=_salary_ns(union_member=False, **luong), **chung)

        assert doan_vien["cong_doan"] > 0 and khong["cong_doan"] == 0
        # Cùng gross, cùng BH ⇒ chênh thu nhập TÍNH thuế đúng bằng đoàn phí.
        assert (khong["pit_taxable"] - doan_vien["pit_taxable"]) == doan_vien["cong_doan"]
        assert doan_vien["pit"] < khong["pit"], "trừ đoàn phí rồi mà thuế không giảm"
    finally:
        db.close()


def _legacy_db_v2():
    """DB có hệ thống bậc lương (TRƯỚC 0092): department_salary_rows + các cột bậc/policy_note."""
    eng = create_engine("sqlite://", connect_args={"check_same_thread": False},
                        poolclass=StaticPool)
    db = sessionmaker(bind=eng)()
    db.execute(text(
        "CREATE TABLE department_salary_rows (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " department_id INTEGER NOT NULL, label VARCHAR(120) NOT NULL)"))
    db.execute(text(
        "CREATE TABLE employee_salaries (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " employee_id INTEGER NOT NULL, luong_vi_tri NUMERIC(14,2) NOT NULL DEFAULT 0,"
        " pay_grade_row_id INTEGER, phu_cap_trach_nhiem NUMERIC(14,2) NOT NULL DEFAULT 0)"))
    db.execute(text("INSERT INTO employee_salaries (employee_id, luong_vi_tri, pay_grade_row_id,"
                    " phu_cap_trach_nhiem) VALUES (5, 9000000, 3, 500000)"))
    db.execute(text(
        "CREATE TABLE payroll_lines (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " gross NUMERIC(14,2) NOT NULL DEFAULT 0, phu_cap_trach_nhiem NUMERIC(14,2) NOT NULL DEFAULT 0)"))
    db.execute(text(
        "CREATE TABLE departments (id INTEGER PRIMARY KEY AUTOINCREMENT, name VARCHAR(255),"
        " salary_policy_note VARCHAR(500))"))
    db.execute(text("INSERT INTO departments (name, salary_policy_note) VALUES ('Tổ A', 'ghi chú')"))
    db.commit()
    return db


def test_migration_0092_drops_grade_system():
    """Bỏ hệ bậc lương: DROP table department_salary_rows + 4 cột. Idempotent, DB trắng no-op.
    Lương vị trí của NV KHÔNG bị đụng (chỉ bỏ cột bậc/phu_cap_trach_nhiem)."""
    db = _legacy_db_v2()
    _migrate_luong_bo_bac_luong(db)
    _migrate_luong_bo_bac_luong(db)     # chạy lại không lỗi
    insp = inspect(db.get_bind())
    tables = set(insp.get_table_names())
    assert "department_salary_rows" not in tables
    es = {c["name"] for c in insp.get_columns("employee_salaries")}
    assert "pay_grade_row_id" not in es and "phu_cap_trach_nhiem" not in es
    assert "luong_vi_tri" in es                                   # tiền của NV giữ nguyên
    assert float(db.execute(text("SELECT luong_vi_tri FROM employee_salaries WHERE employee_id=5"))
                 .scalar_one()) == 9_000_000
    assert "phu_cap_trach_nhiem" not in {c["name"] for c in insp.get_columns("payroll_lines")}
    assert "salary_policy_note" not in {c["name"] for c in insp.get_columns("departments")}
    # DB trắng → no-op
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_luong_bo_bac_luong(empty)


# =============================================================================
# Đợt 1 — thâm niên trước khi vào (employees) · phụ cấp cơm/ca đêm dời sang CA (work_shifts)
# =============================================================================


def test_params_no_longer_expose_company_allowances(client):
    """Đợt 1b: phụ cấp cơm/ca đêm KHÔNG còn ở cấp công ty — `/api/luong/params` không phơi 2
    field này nữa (đã dời sang khai theo từng CA ở `work_shifts`)."""
    token = _admin_token(client)
    p = client.get("/api/luong/params", headers=_h(token)).json()
    assert "com_allowance" not in p and "night_allowance" not in p
    # Gửi lên cũng không được ghi (đã bỏ khỏi whitelist + schema) — không vỡ, chỉ bị bỏ qua.
    upd = client.put("/api/luong/params", json={"com_allowance": 30000}, headers=_h(token)).json()
    assert "com_allowance" not in upd


def test_create_employee_stores_prior_seniority(client):
    """Wizard: `prior_seniority_months` (thâm niên đã có TRƯỚC khi vào) khai lúc TẠO thì lưu được;
    engine không dùng số này tính tiền (chỉ để hiển thị / chính sách sau)."""
    token = _admin_token(client)
    body = {"full_name": "NV Thâm niên", "department_id": _dept_id("Hành chính nhân sự"),
            "hire_date": "2020-01-01", "status": "active", "prior_seniority_months": 66}
    res = client.post("/api/employees", json=body, headers=_h(token))
    assert res.status_code == 201, res.text
    eid = res.json()["employee"]["id"]
    db = SessionLocal()
    try:
        assert EmployeeRepository(db).get_by_id(eid).prior_seniority_months == 66
    finally:
        db.close()
    # Không khai → mặc định 0 (không vỡ tạo NV).
    res0 = client.post("/api/employees", json={
        "full_name": "NV Không khai TN", "department_id": _dept_id("Hành chính nhân sự"),
        "hire_date": "2021-01-01", "status": "active"}, headers=_h(token))
    assert res0.status_code == 201, res0.text
    db = SessionLocal()
    try:
        assert EmployeeRepository(db).get_by_id(res0.json()["employee"]["id"]).prior_seniority_months == 0
    finally:
        db.close()


def test_migration_0093_adds_allowance_and_seniority_columns():
    """0093: payroll_params.com_allowance/night_allowance (25k/50k) + employees.prior_seniority_months (0).
    Idempotent (chạy 2 lần không lỗi), hàng cũ nhận default, DB trắng no-op."""
    db = _legacy_db()   # có payroll_params (1 hàng) + payroll_lines
    # employees kiểu cũ, CHƯA có prior_seniority_months.
    db.execute(text(
        "CREATE TABLE employees (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " code VARCHAR(20), full_name VARCHAR(255))"))
    db.execute(text("INSERT INTO employees (code, full_name) VALUES ('NV001', 'A')"))
    db.commit()

    _migrate_luong_phu_cap_com_ca_dem(db)
    _migrate_luong_phu_cap_com_ca_dem(db)      # chạy lại không lỗi
    insp = inspect(db.get_bind())
    pcols = {c["name"] for c in insp.get_columns("payroll_params")}
    assert {"com_allowance", "night_allowance"} <= pcols
    assert "prior_seniority_months" in {c["name"] for c in insp.get_columns("employees")}
    # hàng CŨ nhận mặc định, không NULL (NOT NULL DEFAULT).
    row = db.execute(text("SELECT com_allowance, night_allowance FROM payroll_params")).first()
    assert [float(x) for x in row] == [25000, 50000]
    assert int(db.execute(text("SELECT prior_seniority_months FROM employees")).scalar_one()) == 0
    # DB trắng (chưa có bảng nào) → no-op, không ném lỗi.
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_luong_phu_cap_com_ca_dem(empty)


def test_migration_0094_moves_allowances_from_params_to_work_shifts():
    """0094: gỡ com/night_allowance khỏi payroll_params (best-effort) + thêm meal/night_allowance
    vào work_shifts (25k/50k). Idempotent (chạy 2 lần không lỗi), hàng ca cũ nhận default, DB
    trắng no-op."""
    import sqlite3

    db = _legacy_db()   # payroll_params (1 hàng) + payroll_lines
    # Trạng thái SAU 0093: payroll_params đã có 2 cột cấp công ty.
    db.execute(text("ALTER TABLE payroll_params ADD COLUMN com_allowance NUMERIC(14,2) NOT NULL DEFAULT 25000"))
    db.execute(text("ALTER TABLE payroll_params ADD COLUMN night_allowance NUMERIC(14,2) NOT NULL DEFAULT 50000"))
    # work_shifts kiểu cũ, CHƯA có 2 cột phụ cấp theo ca.
    db.execute(text(
        "CREATE TABLE work_shifts (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " name VARCHAR(100) NOT NULL, start_minute INTEGER NOT NULL, end_minute INTEGER NOT NULL)"))
    db.execute(text("INSERT INTO work_shifts (name, start_minute, end_minute) VALUES ('Ca 1', 480, 1020)"))
    db.commit()

    _migrate_ca_phu_cap_com_ca_dem(db)
    _migrate_ca_phu_cap_com_ca_dem(db)      # chạy lại không lỗi
    insp = inspect(db.get_bind())
    wcols = {c["name"] for c in insp.get_columns("work_shifts")}
    assert {"meal_allowance", "night_allowance"} <= wcols
    # hàng ca CŨ nhận default 25k/50k (NOT NULL DEFAULT), không NULL.
    row = db.execute(text("SELECT meal_allowance, night_allowance FROM work_shifts")).first()
    assert [float(x) for x in row] == [25000, 50000]
    # payroll_params: 2 cột cấp công ty đã gỡ (SQLite ≥ 3.35 hỗ trợ DROP COLUMN; cũ hơn để dormant).
    if sqlite3.sqlite_version_info >= (3, 35, 0):
        pcols = {c["name"] for c in insp.get_columns("payroll_params")}
        assert "com_allowance" not in pcols and "night_allowance" not in pcols
    # DB trắng (chưa có bảng nào) → no-op, không ném lỗi.
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_ca_phu_cap_com_ca_dem(empty)


def test_migration_0095_renames_shift_allowance_and_drops_night_shift():
    """0095: work_shifts đổi tên `night_allowance` → `shift_allowance` (giữ default 50000) +
    gỡ `night_shift`. Idempotent (chạy 2 lần không lỗi), giữ giá trị hàng cũ, DB trắng no-op."""
    import sqlite3

    eng = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    db = sessionmaker(bind=eng)()
    # Trạng thái SAU 0094: work_shifts có night_shift + meal_allowance + night_allowance.
    db.execute(text(
        "CREATE TABLE work_shifts (id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " name VARCHAR(100) NOT NULL, start_minute INTEGER NOT NULL, end_minute INTEGER NOT NULL,"
        " is_overnight BOOLEAN NOT NULL DEFAULT 0, night_shift BOOLEAN NOT NULL DEFAULT 0,"
        " meal_allowance NUMERIC(14,2) NOT NULL DEFAULT 25000,"
        " night_allowance NUMERIC(14,2) NOT NULL DEFAULT 50000,"
        " grace_minutes INTEGER NOT NULL DEFAULT 5, is_active BOOLEAN NOT NULL DEFAULT 1)"))
    db.execute(text(
        "INSERT INTO work_shifts (name, start_minute, end_minute, night_shift, night_allowance) "
        "VALUES ('Ca 3', 1320, 360, 1, 77000)"))
    db.commit()

    _migrate_ca_rename_shift_allowance_go_night_shift(db)
    _migrate_ca_rename_shift_allowance_go_night_shift(db)      # chạy lại không lỗi
    insp = inspect(db.get_bind())
    wcols = {c["name"] for c in insp.get_columns("work_shifts")}
    assert "shift_allowance" in wcols
    # SQLite ≥ 3.35 hỗ trợ RENAME/DROP COLUMN → night_allowance đổi tên, night_shift bị gỡ.
    if sqlite3.sqlite_version_info >= (3, 35, 0):
        assert "night_allowance" not in wcols
        assert "night_shift" not in wcols
        # giá trị hàng cũ theo nguyên qua RENAME (77000).
        assert float(db.execute(text("SELECT shift_allowance FROM work_shifts")).scalar_one()) == 77000

    # DB trắng (chưa có bảng nào) → no-op, không ném lỗi.
    empty = sessionmaker(bind=create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool))()
    _migrate_ca_rename_shift_allowance_go_night_shift(empty)


# --- Trần khấu trừ kỷ luật thành THAM SỐ (chủ 29/07/2026) -------------------
# "Bỏ cái 30% đang fix cứng trong code." KHÔNG xoá trần (đó là mức LUẬT Đ102), chỉ bỏ chỗ viết
# cứng: mặc định 0.30 giữ nguyên hành vi cũ, 0 = tắt trần.
#
# Dùng `_compute` TRỰC TIẾP (như các test Đ102 sẵn có) thay vì đi qua API: NV không có chấm công
# thì `actual_cong = 0` ⇒ lương = 0 ⇒ không có gì để kẹp, test thành vô nghĩa.


def _phat_v(cap_pct, *, phat=100_000_000, luong=20_000_000):
    """Chạy `_compute` với một mức trần, trả về dict kết quả."""
    db = SessionLocal()
    try:
        svc = PayrollService(PayrollRepository(db), EmployeeRepository(db), attendance=None)
        params = svc.get_params()
        params.phat_cap_pct = cap_pct          # đổi trong bộ nhớ, không đụng DB
        emp = SimpleNamespace(status="active", hire_date=date(2020, 1, 1), gender="male",
                              payroll_group=None, pay_grade_key=None, dependents_count=0)
        return svc._compute(
            employee=emp, salary=_salary_ns(luong_vi_tri=luong), params=params,
            actual_cong=26, standard_cong=26, phat_bien_ban=phat, on=date(2026, 6, 1),
        )
    finally:
        db.close()


def test_tran_phat_mac_dinh_van_la_30_phan_tram(client):
    """⭐ Không khai gì ⇒ hành vi Y HỆT trước khi đưa thành tham số."""
    client
    v = _phat_v(0.30)
    income = (v["luong_cong"] + v["chuyen_can"] + v["allowance"] + v["khoan"]
              + v["ot_pay"] + v["night_pay"])
    phat_eff = income - v["gross"]
    base = income - v["bhxh"] - v["pit"]
    assert v["phat_bien_ban"] == 100_000_000, "cột phạt LƯU RAW, không lưu số đã kẹp"
    assert abs(phat_eff - round(0.30 * base)) <= 1, "mặc định phải vẫn đúng 30%"
    assert v["gross"] > 0


def test_tran_phat_doi_duoc_thanh_50(client):
    """Đặt 50% ⇒ kẹp ở 50%, trừ được NHIỀU hơn ⇒ gross thấp hơn."""
    client
    v30, v50 = _phat_v(0.30), _phat_v(0.50)
    assert v50["gross"] < v30["gross"], "nới trần lên 50% phải trừ được nhiều hơn"
    income = (v50["luong_cong"] + v50["chuyen_can"] + v50["allowance"] + v50["khoan"]
              + v50["ot_pay"] + v50["night_pay"])
    base = income - v50["bhxh"] - v50["pit"]
    assert abs((income - v50["gross"]) - round(0.50 * base)) <= 1


def test_tran_phat_bang_0_la_TAT_tran_va_gross_khong_am(client):
    """⭐ Đặt 0 ⇒ ghi phạt bao nhiêu trừ bấy nhiêu.

    Và `gross` phải có SÀN 0: chính trần 30% vốn là thứ ngăn gross xuống âm (phạt ≤ 30% của
    chính thu nhập). Tắt trần mà không có sàn thì phạt 100tr trên lương 20tr ra gross ÂM —
    in ra phiếu lương là số vô nghĩa."""
    client
    v = _phat_v(0)
    assert v["gross"] == 0, f"tắt trần: phạt 100tr phải ăn hết lương, gross={v['gross']}"
    assert v["gross"] >= 0, "gross KHÔNG được âm"
    assert v["phat_bien_ban"] == 100_000_000, "cột phạt vẫn LƯU RAW"


def test_sua_1_o_va_tinh_lai_ra_CUNG_so_khi_doi_tran(client):
    """⭐ `_capped_penalty` có HAI chỗ gọi (`_compute` và `update_line`). Sót một chỗ thì
    "Tính lại" và "Sửa 1 ô" ra hai số khác nhau — lỗi đã tái phát nhiều lần ở file này."""
    token = _admin_token(client)
    try:
        client.put("/api/luong/params", json={"phat_cap_pct": 0.50}, headers=_h(token))
        eid = _make_emp(client, token, name="NV So Hai Đường Trần")
        client.post(f"/api/luong/salaries/{eid}", json={"effective_from": "2026-01-01",
                    "luong_vi_tri": 20_000_000, "union_member": True}, headers=_h(token))
        gen = client.post("/api/luong/generate", json={"year": 2026, "month": 11},
                          headers=_h(token)).json()
        line = next(l for l in gen["lines"] if l["employee_id"] == eid)

        sua = client.put(f"/api/luong/lines/{line['id']}",
                         json={"vi_pham": 30_000_000, "di_tre": 500_000},
                         headers=_h(token)).json()
        gen2 = client.post("/api/luong/generate", json={"year": 2026, "month": 11},
                           headers=_h(token)).json()
        tinh_lai = next(l for l in gen2["lines"] if l["employee_id"] == eid)

        for k in ("gross", "net_pay", "pit", "cong_doan", "bhxh"):
            assert sua[k] == tinh_lai[k], f"lệch ở {k}: sửa 1 ô {sua[k]} vs tính lại {tinh_lai[k]}"
    finally:
        client.put("/api/luong/params", json={"phat_cap_pct": 0.30}, headers=_h(token))


# --- Phụ cấp CƠM CA + PHỤ CẤP CA theo CA THỰC LÀM ---------------------------
#
# Chủ chốt 03/08/2026. Trước đó hai ô này khai trên `work_shifts` nhưng engine KHÔNG đọc — form
# hứa "nhân viên được gán ca này sẽ tự cộng khi tính lương" mà thực trả 0đ. Nay nối thật.
#
# Luật: mỗi NGÀY THỰC LÀM ca đó, nếu công của ngày ĐỦ NGƯỠNG thì hưởng TRỌN mức của ca; dưới
# ngưỡng thì KHÔNG có gì. Cố ý KHÔNG nhân theo tỷ lệ — một suất ăn là có hoặc không.


def _ca_ns(shift_id, *, com, ca):
    return SimpleNamespace(id=shift_id, meal_allowance=com, shift_allowance=ca)


def _svc_pc_ca(db, ten):
    """Mỗi test một tên tổ RIÊNG — `departments.name` là unique, dùng chung là vỡ ở test thứ hai."""
    return _cfg_svc(db, f"Tổ phụ cấp ca {ten}")


def test_phu_cap_theo_ca_cong_dung_tung_ca():
    """⭐ Ca chính: 20 ngày ca ngày + 6 ngày ca đêm, mỗi ca một mức riêng."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "cong_dung")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        sal = _salary_ns(base_amount=None, luong_vi_tri=10_000_000, luong_trach_nhiem=0)
        ca_ngay, ca_dem = _ca_ns(1, com=25_000, ca=50_000), _ca_ns(2, com=30_000, ca=80_000)

        v = svc._compute(
            employee=emp, salary=sal, params=params, standard_cong=26, on=on, actual_cong=26,
            ca_lam={1: [1.0] * 20, 2: [1.0] * 6},
            shift_by_id={1: ca_ngay, 2: ca_dem},
        )
        assert v["meal_allowance_pay"] == 25_000 * 20 + 30_000 * 6      # 500k + 180k
        assert v["shift_allowance_pay"] == 50_000 * 20 + 80_000 * 6     # 1.000k + 480k


    finally:
        db.close()


def test_nguong_cong_du_thi_TRON_duoi_thi_KHONG_CO():
    """⭐ Chốt chủ đã cân nhắc kỹ (03/08/2026) — canh CẢ BA mốc.

    Không nhân tỷ lệ: đi muộn 15 phút (công 0,97) vẫn ăn trọn suất; nghỉ nửa buổi (0,5 = ngưỡng
    mặc định) vẫn ăn trọn; vắng quá nửa ca (0,25) thì không có gì. Đây là chỗ dễ bị ai đó 'tiện
    tay' đổi sang nhân tỷ lệ nhất."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "nguong")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        sal = _salary_ns(base_amount=None, luong_vi_tri=10_000_000, luong_trach_nhiem=0)
        kw = dict(employee=emp, salary=sal, params=params, standard_cong=26, on=on,
                  actual_cong=26, shift_by_id={1: _ca_ns(1, com=25_000, ca=50_000)})

        # 1 ngày đủ công · 1 ngày nghỉ nửa buổi · 1 ngày đi muộn · 1 ngày vắng quá nửa
        v = svc._compute(ca_lam={1: [1.0, 0.5, 0.97, 0.25]}, **kw)
        assert v["meal_allowance_pay"] == 25_000 * 3, "0,25 phai bi loai; 3 ngay kia an TRON"
        assert v["shift_allowance_pay"] == 50_000 * 3

        # Đổi ngưỡng lên 1,0 ⇒ chỉ ngày đủ công mới được.
        params.phu_cap_ca_min_cong = 1.0
        v2 = svc._compute(ca_lam={1: [1.0, 0.5, 0.97, 0.25]}, **kw)
        assert v2["meal_allowance_pay"] == 25_000 and v2["shift_allowance_pay"] == 50_000
    finally:
        db.close()


def test_khong_co_ca_lam_thi_khong_co_phu_cap():
    """Nghỉ phép / nghỉ lễ / không đi làm ⇒ Chấm công không đếm ngày nào ⇒ không đồng nào."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "khong_ca")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        sal = _salary_ns(base_amount=None, luong_vi_tri=10_000_000, luong_trach_nhiem=0)
        v = svc._compute(employee=emp, salary=sal, params=params, standard_cong=26, on=on,
                         actual_cong=0, ca_lam={}, shift_by_id={1: _ca_ns(1, com=25_000, ca=50_000)})
        assert v["meal_allowance_pay"] == 0 and v["shift_allowance_pay"] == 0
    finally:
        db.close()


def test_ca_da_xoa_khoi_danh_muc_thi_BO_QUA_khong_doan_muc():
    """Ca không còn trong danh mục ⇒ bỏ qua, KHÔNG đoán mức. Đoán bừa là đẻ tiền từ hư không."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "ca_xoa")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        sal = _salary_ns(base_amount=None, luong_vi_tri=10_000_000, luong_trach_nhiem=0)
        v = svc._compute(employee=emp, salary=sal, params=params, standard_cong=26, on=on,
                         actual_cong=26, ca_lam={99: [1.0] * 10}, shift_by_id={})
        assert v["meal_allowance_pay"] == 0 and v["shift_allowance_pay"] == 0
    finally:
        db.close()


def test_phu_cap_theo_ca_MIEN_thue_TNCN():
    """⭐ Canh chốt 04/08/2026: cơm ca + phụ cấp ca **MIỄN** TNCN.

    Bản 03/08 để hai khoản này CHỊU thuế với lý do "muốn miễn thì khai ở danh mục khoản thu nhập"
    — nhưng khai ở đó nữa là TRẢ HAI LẦN, nên thực tế không có đường nào miễn. Kế toán đang xếp
    "Tiền ăn ca/CN/GH" vào nhóm miễn (`docs/prd-thu-nhap-chiu-thue.md §1`). Chủ chốt miễn cả hai.

    Mức lương để cao (30tr) cho vượt giảm trừ gia cảnh, nếu không thuế = 0 ở cả hai vế và test
    không có răng.
    """
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "mien_thue")
        params, emp, on = svc.get_params(), _emp_ns(dept.id), date(2026, 6, 1)
        sal = _salary_ns(base_amount=None, luong_vi_tri=30_000_000, luong_trach_nhiem=0)
        kw = dict(employee=emp, salary=sal, params=params, standard_cong=26, on=on,
                  actual_cong=26, shift_by_id={1: _ca_ns(1, com=100_000, ca=200_000)})

        khong = svc._compute(ca_lam={}, **kw)
        co = svc._compute(ca_lam={1: [1.0] * 10}, **kw)
        them = (100_000 + 200_000) * 10
        assert khong["pit"] > 0, "lương phải đủ cao để có thuế, không thì test vô nghĩa"

        assert co["gross"] == khong["gross"] + them, "tiền vẫn phải cộng vào tổng lương"
        # MIỄN thuế ⇒ thu nhập chịu thuế KHÔNG đổi, phần miễn tăng đúng bằng khoản vừa cộng,
        # và thuế phải GIỮ NGUYÊN — đây là vế có răng nhất.
        assert co["thu_nhap_chiu_thue"] == khong["thu_nhap_chiu_thue"]
        assert co["thu_nhap_mien_thue"] == khong["thu_nhap_mien_thue"] + them
        assert co["pit"] == khong["pit"], "miễn thuế mà thuế vẫn tăng ⇒ `ca_exempt` chưa vào _auto_pit"
    finally:
        db.close()


# --- BHXH: luật 14 ngày (QĐ 595 Đ42.4) --------------------------------------


def _bhxh_theo_cong(svc, dept_id, *, actual_cong, plain_cong=0.0):
    return svc._compute(
        employee=_emp_ns(dept_id), salary=_salary_ns(base_amount=None, luong_vi_tri=10_000_000,
                                                     luong_trach_nhiem=0),
        params=svc.get_params(), standard_cong=26, on=date(2026, 6, 1),
        actual_cong=actual_cong, plain_cong=plain_cong,
    )


def test_bhxh_mien_khi_nghi_khong_luong_tu_14_ngay():
    """⭐ QĐ 595 Đ42.4: nghỉ không làm việc & không hưởng lương từ 14 ngày làm việc trở lên trong
    tháng thì tháng đó KHÔNG đóng BHXH. Phủ luôn người vào/nghỉ việc giữa tháng."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "bhxh_14_ngay")

        du_cong = _bhxh_theo_cong(svc, dept.id, actual_cong=26)
        assert du_cong["bhxh"] > 0, "làm đủ tháng thì phải đóng BHXH"

        # 26 − 13 = 13 ngày không lương → CHƯA tới ngưỡng, vẫn đóng.
        assert _bhxh_theo_cong(svc, dept.id, actual_cong=13)["bhxh"] > 0

        # 26 − 12 = 14 ngày → chạm ngưỡng, không đóng. Nhưng vẫn hiện mức đóng để phiếu lương
        # không trông như chưa khai lương cơ bản.
        it_cong = _bhxh_theo_cong(svc, dept.id, actual_cong=12)
        assert it_cong["bhxh"] == 0
        assert it_cong["insurance_base"] == 10_000_000
    finally:
        db.close()


def test_nguong_bhxh_khai_duoc_va_so_0_la_TAT_luat():
    """⭐ Ngưỡng 14 ngày là THAM SỐ, không phải số cắm cứng (chủ 04/08/2026: *"đang hard code à,
    vậy sao đổi luật thì sao"*). Luật đổi thì gõ lại ở Cấu hình lương, không phải sửa code.

    🔴 Vế `0` = TẮT luật là vế có răng nhất: engine phải kiểm `nguong > 0` TRƯỚC khi so, không thì
    `ngay_khong_luong >= 0` luôn đúng và CẢ XƯỞNG mất sạch BHXH mà bảng lương trông vẫn bình thường.
    """
    db = SessionLocal()
    svc, dept = _svc_pc_ca(db, "nguong_bhxh_khai_duoc")
    try:
        assert svc.get_params().bhxh_mien_tu_so_ngay == 14, "mặc định phải là mức luật"
        # 26 − 15 = 11 ngày không lương → dưới mức 14, vẫn đóng.
        assert _bhxh_theo_cong(svc, dept.id, actual_cong=15)["bhxh"] > 0

        # Hạ ngưỡng xuống 10 → CHÍNH người đó mất BHXH. Đây là vế chứng minh "đổi luật gõ được".
        svc.update_params(bhxh_mien_tu_so_ngay=10)
        assert _bhxh_theo_cong(svc, dept.id, actual_cong=15)["bhxh"] == 0

        # 0 = TẮT luật ⇒ người KHÔNG có công nào VẪN đóng BHXH (hành vi trước 04/08/2026).
        svc.update_params(bhxh_mien_tu_so_ngay=0)
        v = _bhxh_theo_cong(svc, dept.id, actual_cong=0)
        assert v["bhxh"] > 0, "số 0 phải là TẮT LUẬT, không phải 'miễn cho tất cả mọi người'"
    finally:
        svc.update_params(bhxh_mien_tu_so_ngay=14)   # trả mặc định cho các test sau
        db.close()


def test_params_nhan_2_tham_so_luat_qua_api(client):
    """`update_params` lọc theo ALLOWLIST — tên nào không có trong rổ đó thì PUT trả 200 nhưng số
    KHÔNG đổi, tức một ô cấu hình giả. `phu_cap_ca_min_cong` (thêm 03/08) đã bị sót đúng kiểu đó."""
    token = _admin_token(client)
    goc = client.get("/api/luong/params", headers=_h(token)).json()
    try:
        upd = client.put("/api/luong/params",
                         json={"bhxh_mien_tu_so_ngay": 10, "phu_cap_ca_min_cong": 0.75},
                         headers=_h(token)).json()
        assert upd["bhxh_mien_tu_so_ngay"] == 10
        assert upd["phu_cap_ca_min_cong"] == 0.75
        doc_lai = client.get("/api/luong/params", headers=_h(token)).json()
        assert doc_lai["bhxh_mien_tu_so_ngay"] == 10, "PUT xong đọc lại phải còn — không thì ô giả"
        assert doc_lai["phu_cap_ca_min_cong"] == 0.75
    finally:
        client.put("/api/luong/params",
                   json={"bhxh_mien_tu_so_ngay": goc["bhxh_mien_tu_so_ngay"],
                         "phu_cap_ca_min_cong": goc["phu_cap_ca_min_cong"]},
                   headers=_h(token))


def test_ngay_off1x_khong_bi_dem_la_nghi_khong_luong():
    """Ngày off1x là ngày CÓ đi làm và CÓ trả 1×, nhưng Chấm công đã trừ nó khỏi `total_cong`
    (`attendance_service.py:1195`). Không cộng trả lại thì người làm ngày off1x bị đếm thành nghỉ
    không lương và MẤT BHXH oan."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "off1x_bhxh")
        # 10 công nền + 4 ngày off1x = thực tế nghỉ 12 ngày → dưới ngưỡng, PHẢI còn đóng BHXH.
        v = _bhxh_theo_cong(svc, dept.id, actual_cong=10, plain_cong=4)
        assert v["bhxh"] > 0, "ngày off1x bị đếm nhầm thành nghỉ không lương"
        # Bỏ 4 ngày off1x ra thì thành 16 ngày nghỉ → vượt ngưỡng, không đóng.
        assert _bhxh_theo_cong(svc, dept.id, actual_cong=10)["bhxh"] == 0
    finally:
        db.close()


def test_nv_khong_lam_ca_nao_khong_bi_anh_huong():
    """Hồi quy: người KHÔNG làm ca nào (không có `ca_lam`) thì cả 3 thay đổi 04/08 đều không được
    đụng tới một đồng — cơm/phụ cấp ca = 0, phần miễn thuế chỉ còn OT/ca đêm."""
    db = SessionLocal()
    try:
        svc, dept = _svc_pc_ca(db, "hoi_quy_khong_ca")
        v = svc._compute(
            employee=_emp_ns(dept.id),
            salary=_salary_ns(base_amount=None, luong_vi_tri=30_000_000, luong_trach_nhiem=0),
            params=svc.get_params(), standard_cong=26, on=date(2026, 6, 1), actual_cong=26,
        )
        assert v["meal_allowance_pay"] == 0 and v["shift_allowance_pay"] == 0
        assert v["thu_nhap_mien_thue"] == 0
        assert v["thu_nhap_chiu_thue"] == v["gross"]
        assert v["bhxh"] > 0
    finally:
        db.close()
